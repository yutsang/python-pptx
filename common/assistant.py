import json, os, httpx, time
import pandas as pd
from tabulate import tabulate
from pathlib import Path
import re
from tqdm import tqdm
from typing import Dict, List, Optional
import numpy as np
import openpyxl

import logging
import streamlit as st

# Suppress httpx logging
logging.getLogger("httpx").setLevel(logging.WARNING)

# AI-related imports (DeepSeek only)
try:
    from openai import OpenAI
    AI_AVAILABLE = True
except ImportError:
    OpenAI = None
    AI_AVAILABLE = False

# --- Config and AI Service Helpers ---
def load_config(file_path):
    """Load configuration from a JSON file."""
    with open(file_path) as config_file:
        config_details = json.load(config_file)
    return config_details

def initialize_ai_services(config_details, use_local=False, use_openai=False):
    """Initialize AI client using config details - supports DeepSeek, OpenAI, and local AI."""
    if not AI_AVAILABLE:
        raise RuntimeError("AI services not available on this machine.")
    httpx_client = httpx.Client(verify=False)
    if OpenAI is None:
        raise RuntimeError("AI modules not available.")
    
    # Allow session-level override of provider selection
    try:
        import streamlit as st
        selected_provider = st.session_state.get('selected_provider')
        if selected_provider == 'Open AI':
            use_openai = True
            use_local = False
        elif selected_provider == 'Local AI':
            use_local = True
            use_openai = False
        elif selected_provider == 'Server AI':
            # Treat Server AI the same as Local AI (same OpenAI-compatible API format)
            # Just different base URL and key configured in config.json
            use_local = True
            use_openai = False
    except Exception:
        pass

    if use_local:
        # Initialize Local/Server AI client (same OpenAI-compatible format)
        # Prefer LOCAL_*; if SERVER_* are provided, they can be mapped externally by updating config.
        local_base = config_details.get('LOCAL_AI_API_BASE') or config_details.get('SERVER_AI_API_BASE')
        local_key = config_details.get('LOCAL_AI_API_KEY') or config_details.get('SERVER_AI_API_KEY') or 'local-key'
        local_enabled = config_details.get('LOCAL_AI_ENABLED', True)
        if local_base and local_enabled:
            oai_client = OpenAI(
                api_key=local_key,
                base_url=local_base,
                http_client=httpx_client
            )
        else:
            raise RuntimeError("Local/Server AI configuration not found or not enabled. Please check LOCAL_AI_API_BASE (or SERVER_AI_API_BASE) and LOCAL_AI_ENABLED in config.")
    elif use_openai:
        # Initialize OpenAI client
        if config_details.get('OPENAI_API_KEY') and config_details.get('OPENAI_API_BASE'):
            oai_client = OpenAI(
                api_key=config_details['OPENAI_API_KEY'],
                base_url=config_details['OPENAI_API_BASE'],
                http_client=httpx_client
            )
            print("🤖 Using OpenAI GPT-4o-mini")
        else:
            raise RuntimeError("OpenAI configuration not found. Please check OPENAI_API_KEY and OPENAI_API_BASE in config.")
    else:
        # Initialize DeepSeek client (server)
        if config_details.get('DEEPSEEK_API_KEY') and config_details.get('DEEPSEEK_API_BASE'):
            oai_client = OpenAI(
                api_key=config_details['DEEPSEEK_API_KEY'],
                base_url=config_details['DEEPSEEK_API_BASE'],
                http_client=httpx_client
            )
            # DeepSeek AI initialized (silent for cleaner progress bar)
        else:
            raise RuntimeError("DeepSeek configuration not found. Please check DEEPSEEK_API_KEY and DEEPSEEK_API_BASE in config.")
    
    return oai_client, None  # No search client needed

def get_openai_client(config_details=None, use_local=False, use_openai=False):
    """Get the appropriate OpenAI client based on model selection."""
    if config_details is None:
        # Load config if not provided
        config_path = os.path.join(os.path.dirname(__file__), '..', 'fdd_utils', 'config.json')
        config_details = load_config(config_path)
    
    client, _ = initialize_ai_services(config_details, use_local=use_local, use_openai=use_openai)
    return client

def get_chat_model(config_details=None, use_local=False, use_openai=False):
    """Get the appropriate chat model name based on selection."""
    if config_details is None:
        # Load config if not provided
        config_path = os.path.join(os.path.dirname(__file__), '..', 'fdd_utils', 'config.json')
        config_details = load_config(config_path)
    
    # Allow session-level override of provider and model
    try:
        import streamlit as st
        selected_provider = st.session_state.get('selected_provider')
        selected_model = st.session_state.get('selected_model')
        if selected_provider == 'Local AI':
            use_local = True
            use_openai = False
            if selected_model:
                return selected_model
        elif selected_provider == 'Open AI':
            use_openai = True
            use_local = False
            if selected_model:
                return selected_model
        elif selected_provider == 'Server AI':
            # Same format as Local AI; use the same model selection/defaults
            use_local = True
            use_openai = False
            if selected_model:
                return selected_model
    except Exception:
        pass

    if use_local:
        return config_details.get('LOCAL_AI_CHAT_MODEL', 'local-model')
    elif use_openai:
        return config_details.get('OPENAI_CHAT_MODEL', 'gpt-4o-mini-2024-07-18')
    else:
        return config_details.get('DEEPSEEK_CHAT_MODEL', 'deepseek-chat')

def generate_response(user_query, system_prompt, oai_client, context_content, openai_chat_model, entity_name="default", use_local_ai=False):
    """Generate a response from the AI model given a user query and system prompt."""
    
    # Include context data in the user query instead of as a separate assistant message
    enhanced_user_query = f"Context data:\n{context_content}\n\nUser query:\n{user_query}"
    
    # Validate content before creating conversation
    if not system_prompt:
        raise ValueError("System prompt is None or empty")
    if not enhanced_user_query:
        raise ValueError("User query is None or empty")
    
    conversation = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": enhanced_user_query}
    ]
    
    print(f"🔍 DEBUG: Conversation created with {len(conversation)} messages")
    print(f"🔍 DEBUG: System prompt length: {len(system_prompt)}")
    print(f"🔍 DEBUG: User query length: {len(enhanced_user_query)}")
    
    try:
        # For local AI with thinking support (DeepSeek 671B), enable reasoning
        extra_params = {}
        if use_local_ai:
            # Do NOT enable reasoning; always request direct answers only
            try:
                pass
            except Exception:
                pass
        
        # Use selected chat model if provided; fall back to session/config selection
        model_name = openai_chat_model or get_chat_model()
        # Remove unsupported params defensively for non-DeepSeek models (e.g., qwen2)
        # Ensure no reasoning param is sent for compatibility across models
        safe_params = {k: v for k, v in extra_params.items() if k != 'reasoning'}
        response = oai_client.chat.completions.create(model=model_name, messages=conversation, **safe_params)
    except Exception as e:
        print(f"❌ API call failed with model '{openai_chat_model}': {e}")
        raise
    
    response_content = response.choices[0].message.content
    
    # If model returns thinking tags anyway, strip them and keep final answer only
    if "<thinking>" in response_content and "</thinking>" in response_content:
        response_content = response_content.split("</thinking>")[-1].strip()
    
    return response_content

# --- Excel and Data Processing ---
def parse_table_to_structured_format(df, entity_name, table_name):
    """
    Parse a DataFrame into structured format for financial tables.
    Extracts table name, entity, date, currency, multiplier, and items.
    """
    try:
        import re
        import os
        from datetime import datetime

        # 🚨 SERVER DEBUG: Comprehensive Application Flow Logging
        print(f"\n🚨 SERVER DEBUG: === APPLICATION START ===")
        print(f"🚨 SERVER DEBUG: Function: parse_table_to_structured_format")
        print(f"🚨 SERVER DEBUG: Table: '{table_name}' | Entity: '{entity_name}'")
        print(f"🚨 SERVER DEBUG: DataFrame shape: {df.shape}")
        print(f"🚨 SERVER DEBUG: DataFrame columns: {list(df.columns)}")
        print(f"🚨 SERVER DEBUG: Python version: {os.sys.version}")
        print(f"🚨 SERVER DEBUG: Current working directory: {os.getcwd()}")
        print(f"🚨 SERVER DEBUG: Available modules: pandas={__import__('pandas', fromlist=['']).__version__ if 'pandas' in str(__import__('sys').modules) else 'Not loaded'}")

        # DEBUG: Log what we're reading from the Excel
        print(f"\n{'='*80}")
        print(f"📊 TABLE PROCESSING: '{table_name}' for entity '{entity_name}'")
        print(f"📊 DataFrame shape: {df.shape}")
        print(f"📊 DataFrame columns: {list(df.columns)}")
        print(f"📊 TABLE CONTENTS:")
        print(f"{'='*80}")

        # Show all rows of the table
        print(f"\n{'='*100}")
        print(f"📊 DETAILED EXCEL CONTENT ANALYSIS FOR TABLE: '{table_name}'")
        print(f"{'='*100}")

        print(f"🔍 FULL RAW DATAFRAME ({len(df)} rows x {len(df.columns)} columns):")
        print(f"   DataFrame shape: {df.shape}")
        print(f"   Column names: {list(df.columns)}")
        print(f"   Column types: {[str(df[col].dtype) for col in df.columns]}")

        print(f"\n🔍 COMPLETE CELL-BY-CELL CONTENT:")
        for i in range(len(df)):
            print(f"\n--- ROW {i} ---")
            for j, col in enumerate(df.columns):
                cell_value = df.iloc[i, j]
                cell_type = type(cell_value).__name__
                cell_str = str(cell_value) if cell_value is not None else "None"

                # Check for RMB-related patterns in this cell
                rmb_found = False
                rmb_patterns = ["人民币", "人民幣", "千元", "CNY", "RMB", "万元", "万", "'000", '"000', "000", "thousands", "THOUSANDS", "Thousands"]
                found_patterns = []
                for pattern in rmb_patterns:
                    if cell_str and pattern in cell_str:
                        found_patterns.append(pattern)
                        rmb_found = True

                if rmb_found:
                    print(f"   Col {j} ({col}): [{cell_type}] '{cell_str}' 💰 RMB-PATTERNS: {found_patterns}")
                else:
                    print(f"   Col {j} ({col}): [{cell_type}] '{cell_str}'")

        print(f"\n🔍 ROW-BY-ROW SUMMARY (for RMB detection):")
        for i in range(len(df)):
            row_data = df.iloc[i].values
            row_str = " | ".join([str(cell) if cell is not None else "None" for cell in row_data])

            # Count RMB patterns in this row
            rmb_count = 0
            pattern_counts = {}
            for cell in row_data:
                cell_str = str(cell) if cell is not None else ""
                for pattern in rmb_patterns:
                    if pattern in cell_str:
                        rmb_count += 1
                        pattern_counts[pattern] = pattern_counts.get(pattern, 0) + 1

            if rmb_count > 0:
                print(f"Row {i:2d}: {row_str} 💰 ({rmb_count} RMB patterns: {pattern_counts})")
            else:
                print(f"Row {i:2d}: {row_str}")

        # Also show raw Excel values for RMB detection verification
        print(f"\n🔍 RAW EXCEL VALUES FOR RMB DETECTION:")
        for i, row in enumerate(df.values.tolist()):
            row_str = " | ".join([str(cell) if cell is not None else "None" for cell in row])
            print(f"Row {i:2d}: {row_str}")
            # Highlight any cells containing RMB keywords
            for j, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    rmb_patterns = ["人民币", "人民幣", "千元", "CNY", "RMB", "万元", "万", "'000", '"000', "000", "thousands", "THOUSANDS", "Thousands"]
                    if any(pattern in cell for pattern in rmb_patterns):
                        print(f"         💰 RMB-RELATED CELL [{i},{j}]: '{cell}'")

        print(f"{'='*100}")
        print(f"🔍 Starting detailed processing of table '{table_name}'...")
        print(f"{'='*100}")
        print(f"🔍 DEBUG: First 5 rows of data:")
        for i, row in enumerate(df.head(5).values.tolist()):
            print(f"   Row {i}: {row}")

        # Initialize structured data
        structured_data = {
            'table_name': table_name,
            'entity': entity_name,
            'date': None,
            'currency': 'CNY',
            'multiplier': 1,
            'items': [],
            'total': None
        }

        # Convert DataFrame to list of rows for easier processing
        rows = df.values.tolist()
        if not rows:
            print(f"⚠️ DEBUG: No rows found in table '{table_name}'")
            return None

        print(f"🔍 DEBUG: Total rows to process: {len(rows)}")
        print(f"🔍 DEBUG: Sample row data:")
        for i, row in enumerate(rows[:3]):  # Show first 3 rows
            print(f"   Row {i}: {row}")

        # 🚨 SERVER DEBUG: Comprehensive RMB Detection Logging
        print(f"\n🚨 SERVER DEBUG: === RMB DETECTION SCAN STARTING ===")
        print(f"🚨 SERVER DEBUG: Table: '{table_name}' | Entity: '{entity_name}'")
        print(f"🚨 SERVER DEBUG: DataFrame shape: {df.shape} | Searching {len(rows)} rows x {len(rows[0]) if rows else 0} columns")
        print(f"🚨 SERVER DEBUG: Looking for patterns: 人民币千元, 人民幣千元, CNY'000, 000")
        print(f"🚨 SERVER DEBUG: Current working directory: {os.getcwd()}")

        rmb_thousands_found = False
        rmb_locations = []
        all_rmb_related_cells = []

        print(f"🚨 SERVER DEBUG: Initialized detection - thousands_found: {rmb_thousands_found}")
        print(f"🚨 SERVER DEBUG: Starting cell-by-cell scan...")

        for row_idx, row in enumerate(rows):
            print(f"🚨 SERVER DEBUG: Scanning row {row_idx}/{len(rows)-1}")
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip()
                # Check for RMB thousand matches (more flexible - handles extra characters)
                if "人民币千元" in cell_str:
                    print(f"🎯 FOUND EXACT: '人民币千元' in row {row_idx}, col {col_idx}: '{cell_str}'")
                    rmb_thousands_found = True
                    rmb_locations.append(f"人民币千元@[{row_idx},{col_idx}]")
                elif "人民幣千元" in cell_str:
                    print(f"🎯 FOUND EXACT: '人民幣千元' in row {row_idx}, col {col_idx}: '{cell_str}'")
                    rmb_thousands_found = True
                    rmb_locations.append(f"人民幣千元@[{row_idx},{col_idx}]")
                # More flexible detection - RMB thousand patterns with extra characters
                elif "人民币" in cell_str and "千元" in cell_str:
                    print(f"🎯 FOUND FLEXIBLE: '人民币...千元' pattern in row {row_idx}, col {col_idx}: '{cell_str}'")
                    rmb_thousands_found = True
                    rmb_locations.append(f"人民币千元(pattern)@[{row_idx},{col_idx}]")
                elif "人民幣" in cell_str and "千元" in cell_str:
                    print(f"🎯 FOUND FLEXIBLE: '人民幣...千元' pattern in row {row_idx}, col {col_idx}: '{cell_str}'")
                    rmb_thousands_found = True
                    rmb_locations.append(f"人民幣千元(pattern)@[{row_idx},{col_idx}]")

                # Also track any RMB-related content for debugging - expanded detection
                rmb_keywords = [
                    "人民币", "人民幣", "千元", "CNY", "RMB",
                    "万元", "万", "十万元", "百万元", "千万元",
                    "'000", '"000', "000",
                    "thousands", "THOUSANDS", "Thousands"
                ]

                # Check for exact RMB thousand patterns (now more flexible)
                exact_thousand_patterns = [
                    "人民币千元", "人民幣千元", "CNY'000", 'CNY"000',
                    "人民币千", "人民幣千", "CNY千",
                    "千人民币", "千人民幣", "千CNY",
                    # Add patterns that might have extra characters
                    "人民币", "人民幣", "千元"
                ]

                if any(keyword in cell_str for keyword in rmb_keywords):
                    all_rmb_related_cells.append(f"[{row_idx},{col_idx}]: '{cell_str}'")
                    print(f"💰 RMB-RELATED: '{cell_str}' in row {row_idx}, col {col_idx}")

                # Check for exact thousand patterns
                for pattern in exact_thousand_patterns:
                    if pattern in cell_str:
                        print(f"🎯 EXACT THOUSAND PATTERN: '{pattern}' found in '{cell_str}' at [{row_idx},{col_idx}]")

        if not rmb_thousands_found:
            print(f"⚠️ DEBUG: No RMB thousand patterns found in table '{table_name}'")
            print(f"   💰 RMB SCAN: Searched {len(rows)} rows, {len(rows[0]) if rows else 0} columns per row")
            print(f"   💰 LOOKED FOR: 人民币千元, 人民幣千元, CNY'000, and flexible patterns")
            if all_rmb_related_cells:
                print(f"   💰 OTHER RMB CONTENT FOUND: {len(all_rmb_related_cells)} RMB-related cells:")
                for cell_info in all_rmb_related_cells[:10]:  # Show first 10
                    print(f"      {cell_info}")
                if len(all_rmb_related_cells) > 10:
                    print(f"      ... and {len(all_rmb_related_cells) - 10} more RMB-related cells")
        else:
            print(f"✅ DEBUG: Found RMB thousands notation in table '{table_name}' at locations: {', '.join(rmb_locations)}")
            print(f"   💰 RMB SCAN: Successfully detected {len(rmb_locations)} RMB thousand instances")
            if all_rmb_related_cells:
                print(f"   💰 ADDITIONAL RMB CONTENT: {len(all_rmb_related_cells)} total RMB-related cells")
        
        # Find the two most important columns (description and amount)
        # Usually the first two columns, but let's be smart about it
        desc_col = 0
        amount_col = 1
        
        # Look for columns with numbers in the amount column
        print(f"🔍 DEBUG: Analyzing columns for numeric content...")
        for col_idx in range(min(2, len(df.columns))):
            numeric_count = 0
            for row in rows:
                if col_idx < len(row):
                    cell_value = str(row[col_idx]).strip()
                    # Check if it's a number (including with commas, decimals, etc.)
                    if re.match(r'^[\d,]+\.?\d*$', cell_value.replace(',', '')):
                        numeric_count += 1

            print(f"🔍 DEBUG: Column {col_idx}: {numeric_count}/{len(rows)} numeric values")
            if numeric_count > len(rows) * 0.3:  # At least 30% of rows have numbers
                amount_col = col_idx
                desc_col = 1 if col_idx == 0 else 0
                print(f"✅ DEBUG: Identified amount column as {col_idx}, desc column as {desc_col}")
                print(f"📋 COLUMN MAPPING: Description='{df.columns[desc_col]}', Amount='{df.columns[amount_col]}'")
                break
        
        # Process rows to extract information
        print(f"🚀 STARTING ROW PROCESSING: {len(rows)} rows to process")
        for row_idx, row in enumerate(rows):
            print(f"🔄 ROW {row_idx}: Starting processing, row length: {len(row)}")
            if len(row) < 2:
                print(f"⏭️ ROW {row_idx}: Skipping due to insufficient columns")
                continue

            desc_cell = str(row[desc_col]).strip() if desc_col < len(row) else ""
            amount_cell = str(row[amount_col]).strip() if amount_col < len(row) else ""

            # DEBUG: Log every row we're processing
            print(f"🔍 DEBUG: Processing row {row_idx}: desc='{desc_cell}', amount='{amount_cell}'")

            # Skip empty rows
            if not desc_cell and not amount_cell:
                continue
            
            # Extract date
            if not structured_data['date']:
                # Look for date patterns in any cell
                for cell in row:
                    cell_str = str(cell).strip()
                    # Common date patterns (English and Chinese)
                    date_patterns = [
                        # English formats
                        r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
                        r'\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
                        r'\d{2}-\d{2}-\d{4}',  # DD-MM-YYYY
                        r'\d{4}/\d{2}/\d{2}',  # YYYY/MM/DD

                        # Chinese formats (prioritize longer matches first)
                        r'\d{4}年\d{1,2}月\d{1,2}日[^\d年月]*',  # Chinese with any non-date suffix: 2024年5月31日...
                        r'\d{4}年\d{1,2}月\d{1,2}日[？?\s]*',  # Chinese with question mark/spaces: 2024年5月31日?
                        r'\d{4}年\d{1,2}月\d{1,2}日',  # Chinese: 2024年5月31日
                        r'\d{4}年\d{1,2}月',  # Chinese: 2024年5月

                        # Chinese traditional/simplified variants
                        r'\d{4}年\d{1,2}月\d{1,2}號[^\d年月]*',  # Chinese Traditional with suffix
                        r'\d{4}年\d{1,2}月\d{1,2}号[^\d年月]*',  # Chinese Simplified with suffix
                        r'\d{4}年\d{1,2}月\d{1,2}號[？?\s]*',  # Chinese Traditional with question mark/spaces
                        r'\d{4}年\d{1,2}月\d{1,2}号[？?\s]*',  # Chinese Simplified with question mark/spaces
                        r'\d{4}年\d{1,2}月\d{1,2}號',  # Chinese Traditional: 2024年5月31號
                        r'\d{4}年\d{1,2}月\d{1,2}号',  # Chinese Simplified: 2024年5月31号

                        # 2-digit year formats
                        r'\d{2}年\d{1,2}月\d{1,2}日[^\d年月]*',  # Chinese 2-digit year with suffix
                        r'\d{2}年\d{1,2}月\d{1,2}日',  # Chinese 2-digit year: 24年5月31日
                        r'\d{2}年\d{1,2}月',  # Chinese 2-digit year: 24年5月

                        # Month-day only (less specific, put last)
                        r'\d{1,2}月\d{1,2}日[^\d年月]*',  # Chinese month-day with suffix
                        r'\d{1,2}月\d{1,2}日',  # Chinese month-day only: 5月31日
                    ]
                    for pattern in date_patterns:
                        match = re.search(pattern, cell_str)
                        if match:
                            try:
                                # Try to parse the date
                                date_str = match.group()
                                print(f"DEBUG: Found date pattern '{pattern}' in cell '{cell_str}', extracted '{date_str}'")
                            except Exception as e:
                                print(f"DEBUG: Date parsing failed for pattern '{pattern}' with '{date_str}': {e}")
                                continue

                                if '年' in date_str and '月' in date_str:
                                    # Chinese date format: 2024年5月31日, 2024年5月31號, 2024年5月31号 or 2024年5月
                                    if '日' in date_str or '號' in date_str or '号' in date_str:
                                        # Full date with day: 2024年5月31日, 2024年5月31號, or 2024年5月31号
                                        # Replace all possible day markers and handle question marks/other suffixes
                                        cleaned_date = date_str.replace('年', '-').replace('月', '-')
                                        # Remove day markers and any trailing non-numeric characters (including question marks)
                                        # Handle both traditional (號) and simplified (号) characters
                                        cleaned_date = re.sub(r'[日號号][^\d-]*$', '', cleaned_date)
                                        # Also handle cases where the marker might be at the end with punctuation
                                        cleaned_date = re.sub(r'[^\d-]$', '', cleaned_date)
                                        parts = cleaned_date.split('-')
                                        if len(parts) == 3:
                                            try:
                                                year, month, day = map(int, parts)
                                                # Handle 2-digit years (assume 2000s)
                                                if year < 100:
                                                    year += 2000
                                                parsed_date = datetime(year, month, day)
                                            except ValueError:
                                                # If parsing fails, try to extract just the numeric parts
                                                numeric_parts = re.findall(r'\d+', cleaned_date)
                                                if len(numeric_parts) >= 3:
                                                    year, month, day = map(int, numeric_parts[:3])
                                                    # Handle 2-digit years (assume 2000s)
                                                    if year < 100:
                                                        year += 2000
                                                    parsed_date = datetime(year, month, day)
                                    else:
                                        # Month only: 2024年5月 (assume last day of month)
                                        parts = date_str.replace('年', '-').replace('月', '').split('-')
                                        if len(parts) == 2:
                                            year, month = map(int, parts)
                                            # Handle 2-digit years (assume 2000s)
                                            if year < 100:
                                                year += 2000
                                            # Assume last day of the month for month-only dates
                                            if month == 2:
                                                # Check for leap year
                                                if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
                                                    day = 29
                                                else:
                                                    day = 28
                                            elif month in [4, 6, 9, 11]:
                                                day = 30
                                            else:
                                                day = 31
                                            parsed_date = datetime(year, month, day)
                                elif '月' in date_str and '日' in date_str and '年' not in date_str:
                                    # Month-day only format: 5月31日 (assume current year)
                                    parts = re.findall(r'\d+', date_str)
                                    if len(parts) >= 2:
                                        month, day = map(int, parts[:2])
                                        current_year = datetime.now().year
                                        parsed_date = datetime(current_year, month, day)
                                elif '-' in date_str:
                                    if len(date_str.split('-')[0]) == 4:  # YYYY-MM-DD
                                        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
                                    else:  # DD-MM-YYYY
                                        parsed_date = datetime.strptime(date_str, '%d-%m-%Y')
                                elif '/' in date_str:
                                    parts = date_str.split('/')
                                    if len(parts) == 3:
                                        # Try both MM/DD/YYYY and DD/MM/YYYY formats
                                        try:
                                            # First try MM/DD/YYYY (US format)
                                            parsed_date = datetime.strptime(date_str, '%m/%d/%Y')
                                        except ValueError:
                                            try:
                                                # Then try DD/MM/YYYY (European format)
                                                parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
                                            except ValueError:
                                                # Finally try YYYY/MM/DD
                                                parsed_date = datetime.strptime(date_str, '%Y/%m/%d')
                                    else:
                                        raise ValueError(f"Invalid date format: {date_str}")

                                structured_data['date'] = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                                break
                            except Exception as e:
                                print(f"Warning: Could not parse date '{date_str}': {e}")
                                continue
            
            # Extract currency and multiplier (English and Chinese)
            currency_detected = False
            currency_source = ""

            if ('CNY' in desc_cell.upper() or 'CNY' in amount_cell.upper()):
                structured_data['currency'] = 'CNY'
                currency_detected = True
                currency_source = "CNY"
            elif ('人民币' in desc_cell or '人民币' in amount_cell):
                structured_data['currency'] = 'CNY'
                currency_detected = True
                currency_source = "人民币 (traditional)"
            elif ('人民幣' in desc_cell or '人民幣' in amount_cell):
                structured_data['currency'] = 'CNY'
                currency_detected = True
                currency_source = "人民幣 (simplified)"
            elif ('RMB' in desc_cell.upper() or 'RMB' in amount_cell.upper()):
                structured_data['currency'] = 'CNY'
                currency_detected = True
                currency_source = "RMB"

            if currency_detected:
                print(f"DEBUG: Currency detected as CNY ({currency_source}) - desc='{desc_cell}', amount='{amount_cell}'")

            # Enhanced check for thousands notation (English and Chinese)
            thousands_detected = (
                "'000" in desc_cell or "'000" in amount_cell or
                "千元" in desc_cell or "千元" in amount_cell or
                "千人民币" in desc_cell or "千人民币" in amount_cell or
                "人民币千" in desc_cell or "人民币千" in amount_cell or
                "千元人民币" in desc_cell or "千元人民币" in amount_cell or
                "人民币千元" in desc_cell or "人民币千元" in amount_cell or
                "千人民幣" in desc_cell or "千人民幣" in amount_cell or
                "人民幣千" in desc_cell or "人民幣千" in amount_cell or
                "千元人民幣" in desc_cell or "千元人民幣" in amount_cell or
                "人民幣千元" in desc_cell or "人民幣千元" in amount_cell or  # 人民幣千元 (simplified Chinese)
                "千RMB" in desc_cell or "千RMB" in amount_cell or
                "RMB千" in desc_cell or "RMB千" in amount_cell or
                # Also check for standalone "千" in amount column
                amount_cell.strip() == "千" or
                # Check for patterns like "100千" meaning 100 thousand
                bool(re.search(r'\d+\s*千', amount_cell)) or
                bool(re.search(r'千\d+', amount_cell)) or
                # Check for "000" without quotes (common in Chinese databooks)
                "000" in amount_cell and not amount_cell.strip().startswith("'") or
                # Check for "千" anywhere in the amount cell
                "千" in amount_cell
            )

            # Special detection for various RMB thousand patterns - check ALL cells in the row
            thousand_patterns = [
                "人民币千元", "人民幣千元", "CNY'000", 'CNY"000',
                "人民币千", "人民幣千", "CNY千",
                "千人民币", "千人民幣", "千CNY",
                "000", "'000", '"000',
                "thousands", "THOUSANDS", "Thousands"
            ]

            # Check for thousand indicators in all cells
            thousands_detected = any(
                any(pattern in str(cell) for pattern in thousand_patterns)
                for cell in row
            )

            # Debug: Show detection of thousand patterns in this row
            detected_patterns = []
            for cell in row:
                cell_str = str(cell)
                for pattern in thousand_patterns:
                    if pattern in cell_str:
                        detected_patterns.append(f"'{pattern}' in '{cell_str}'")

            if detected_patterns:
                print(f"🔍 THOUSAND PATTERNS DETECTED in row {row_idx}: {', '.join(detected_patterns)}")
                print(f"   Row content: {[str(cell) for cell in row]}")

            # Debug logging for multiplier detection
            if thousands_detected or currency_detected:
                print(f"DEBUG: Multiplier detection - desc='{desc_cell}', amount='{amount_cell}', thousands_detected={thousands_detected}, currency_detected={currency_detected}")

            # Set multiplier based on detection
            if thousands_detected:
                old_multiplier = structured_data['multiplier']
                structured_data['multiplier'] = 1000
                print(f"💰 MULTIPLIER SET: Thousand pattern detected - setting multiplier to 1000x")
                print(f"💰 MULTIPLIER SET: Changed from {old_multiplier}x to {structured_data['multiplier']}x")
                print(f"   Detected patterns in row: {detected_patterns if 'detected_patterns' in locals() else 'N/A'}")
            elif currency_detected and ("000" in desc_cell or "000" in amount_cell):
                # Fallback: if we have currency and "000", still apply multiplier
                structured_data['multiplier'] = 1000
                print(f"DEBUG: Set multiplier to 1000 (currency+000) for cell: desc='{desc_cell}', amount='{amount_cell}'")
            elif "million" in desc_cell.lower() or "million" in amount_cell.lower():
                structured_data['multiplier'] = 1000000
                print(f"DEBUG: Set multiplier to 1000000 for cell: desc='{desc_cell}', amount='{amount_cell}'")
            elif "000" in desc_cell or "000" in amount_cell:
                # Check if it's a standalone "000" indicating thousands
                if re.match(r'^0*000$', desc_cell.replace("'", "")) or re.match(r'^0*000$', amount_cell.replace("'", "")):
                        structured_data['multiplier'] = 1000
                        print(f"DEBUG: Set multiplier to 1000 (standalone 000) for cell: desc='{desc_cell}', amount='{amount_cell}'")

            # Final confirmation logging
            if structured_data['multiplier'] == 1000 and thousands_detected:
                print(f"✅ CONFIRMED: Multiplier set to 1000x - thousands notation detected")
                if detected_patterns:
                    print(f"   Detected patterns: {detected_patterns}")
            elif structured_data['multiplier'] == 1000000:
                print(f"✅ CONFIRMED: Multiplier set to 1000000x - million notation detected")

            # Extract items (skip header rows and totals) - INSIDE THE ROW PROCESSING LOOP
            # Be more careful about filtering - don't filter out valid Chinese descriptions
            skip_row = False

            # Skip obvious headers and totals
            if desc_cell.lower() in ['total', 'nan', '']:
                skip_row = True
                print(f"⏭️ SKIP ROW {row_idx}: desc is header/total ('{desc_cell}')")
            elif re.match(r'^[A-Z\s]{3,}$', desc_cell) and not any('\u4e00' <= c <= '\u9fff' for c in desc_cell):
                # Skip all-caps English headers, but allow Chinese text
                skip_row = True
                print(f"⏭️ SKIP ROW {row_idx}: desc is all-caps English header ('{desc_cell}')")
            elif not amount_cell or amount_cell == 'nan':
                skip_row = True
                print(f"⏭️ SKIP ROW {row_idx}: amount is empty or nan ('{amount_cell}')")
            else:
                print(f"✅ PROCESS ROW {row_idx}: desc='{desc_cell}', amount='{amount_cell}'")

            if not skip_row:
                
                # Try to extract numeric amount (support Chinese multipliers)
                amount_cell_clean = amount_cell.replace(',', '').strip()

                # Check for Chinese multiplier patterns like "100千", "千100", "100 千", etc.
                chinese_multiplier_match = None
                amount = None

                # Pattern 1: "100千" or "100 千"
                match1 = re.search(r'(\d+(?:\.\d+)?)\s*千', amount_cell_clean)
                if match1:
                    chinese_multiplier_match = match1
                    base_amount = float(match1.group(1))
                    amount = base_amount * 1000
                    print(f"DEBUG: Found Chinese multiplier pattern 1 '{match1.group(0)}' -> {base_amount} * 1000 = {amount}")

                # Pattern 2: "千100" or "千 100"
                match2 = re.search(r'千\s*(\d+(?:\.\d+)?)', amount_cell_clean)
                if match2 and not chinese_multiplier_match:
                    chinese_multiplier_match = match2
                    base_amount = float(match2.group(1))
                    amount = base_amount * 1000
                    print(f"DEBUG: Found Chinese multiplier pattern 2 '{match2.group(0)}' -> {base_amount} * 1000 = {amount}")

                # Pattern 3: Just "千" in the cell (standalone)
                if not chinese_multiplier_match and amount_cell_clean == "千":
                    amount = 1000
                    print(f"DEBUG: Found standalone '千' -> 1000")

                if amount is None:
                    # Regular numeric extraction
                    amount_match = re.search(r'[\d,]+\.?\d*', amount_cell_clean)
                    if amount_match:
                        amount_str = amount_match.group()
                        try:
                            amount = float(amount_str.replace(',', ''))
                            # Apply multiplier if needed (but don't double-apply for Chinese patterns)
                            if structured_data['multiplier'] > 1 and not chinese_multiplier_match:
                                amount *= structured_data['multiplier']
                        except:
                            amount = None

                if amount is not None:
                    item_data = {
                        'description': desc_cell,
                        'amount': int(amount) if amount.is_integer() else amount
                    }
                    structured_data['items'].append(item_data)
                    print(f"📝 ADDED ITEM: {desc_cell} = {amount} (multiplier: {structured_data['multiplier']}x)")
                    print(f"   Final amount after multiplier: {amount * structured_data['multiplier']}")
                else:
                    print(f"❌ NO AMOUNT: Could not parse amount from '{amount_cell}' for desc '{desc_cell}'")
            
            # Extract total
            if desc_cell.lower() == 'total' and amount_cell and amount_cell != 'nan':
                amount_cell_clean = amount_cell.replace(',', '').strip()

                # Check for Chinese multiplier patterns like "100千", "千100", "100 千", etc.
                chinese_multiplier_match = None
                total_amount = None

                # Pattern 1: "100千" or "100 千"
                match1 = re.search(r'(\d+(?:\.\d+)?)\s*千', amount_cell_clean)
                if match1:
                    chinese_multiplier_match = match1
                    base_amount = float(match1.group(1))
                    total_amount = base_amount * 1000

                # Pattern 2: "千100" or "千 100"
                match2 = re.search(r'千\s*(\d+(?:\.\d+)?)', amount_cell_clean)
                if match2 and not chinese_multiplier_match:
                    chinese_multiplier_match = match2
                    base_amount = float(match2.group(1))
                    total_amount = base_amount * 1000

                # Pattern 3: Just "千" in the cell (standalone)
                if not chinese_multiplier_match and amount_cell_clean == "千":
                    total_amount = 1000

                if total_amount is None:
                    # Regular numeric extraction
                    amount_match = re.search(r'[\d,]+\.?\d*', amount_cell_clean)
                    if amount_match:
                        amount_str = amount_match.group()
                        try:
                            total_amount = float(amount_str.replace(',', ''))
                            # Apply multiplier if needed (but don't double-apply for Chinese patterns)
                            if structured_data['multiplier'] > 1 and not chinese_multiplier_match:
                                total_amount *= structured_data['multiplier']
                        except:
                            total_amount = None

                if total_amount is not None:
                    structured_data['total'] = int(total_amount) if total_amount.is_integer() else total_amount

        # Final summary of multiplier detection (OUTSIDE the row processing loop)
        if structured_data['multiplier'] > 1:
            print(f"🎯 FINAL MULTIPLIER: Table '{table_name}' multiplier set to {structured_data['multiplier']}x")
            if structured_data['multiplier'] == 1000:
                print(f"   💰 RMB THOUSANDS: Multiplier set to 1000x due to detected thousand patterns")
        else:
            print(f"⚠️ FINAL MULTIPLIER: Table '{table_name}' multiplier remains at {structured_data['multiplier']}x (no thousands/million notation detected)")
            print(f"   💰 RMB CHECK: Check if RMB thousand patterns were detected in the table")

        # Extract entity name from table content if not found
        if structured_data['entity'] == entity_name:
            # Look for more specific entity names in the data
            for item in structured_data['items']:
                desc = item['description'].lower()
                if 'haining' in desc:
                    # Extract the full entity name
                    entity_match = re.search(r'haining\s+[a-zA-Z]+', desc, re.IGNORECASE)
                    if entity_match:
                        structured_data['entity'] = entity_match.group()
                        break

        # DEBUG: Final summary of what was detected
        print(f"\n📊 FINAL RESULTS for table '{table_name}':")
        print(f"   - Currency: {structured_data['currency']}")
        print(f"   - Multiplier: {structured_data['multiplier']}x")
        print(f"   - Date: {structured_data['date']}")
        print(f"   - Items found: {len(structured_data['items'])}")
        if structured_data['items']:
            print(f"   - Items details:")
            for item in structured_data['items']:
                print(f"     * {item['description']}: {item['amount']}")
        print(f"   - Total: {structured_data['total']}")
        print(f"{'='*80}")

        # Summary of processing
        if structured_data['items']:
            print(f"✅ TABLE '{table_name}' PROCESSED SUCCESSFULLY")
            print(f"   → {len(structured_data['items'])} financial items extracted")
            print(f"   → Multiplier applied: {structured_data['multiplier']}x")
            print(f"   → Currency detected: {structured_data['currency']}")
        else:
            print(f"❌ TABLE '{table_name}' SKIPPED - No valid items found")

        print(f"{'='*80}\n")

        # 🚨 SERVER DEBUG: Final Results Logging
        print(f"🚨 SERVER DEBUG: === FINAL RESULTS ===")
        print(f"🚨 SERVER DEBUG: Table: '{table_name}' | Entity: '{entity_name}'")
        print(f"🚨 SERVER DEBUG: Multiplier: {structured_data['multiplier']}x")
        print(f"🚨 SERVER DEBUG: Currency: {structured_data['currency']}")
        print(f"🚨 SERVER DEBUG: Items found: {len(structured_data['items'])}")
        print(f"🚨 SERVER DEBUG: Date: {structured_data['date']}")
        print(f"🚨 SERVER DEBUG: Returning data: {'YES' if structured_data['items'] else 'NO (empty items)'}")

        if structured_data['items']:
            print(f"🚨 SERVER DEBUG: Sample items:")
            for i, item in enumerate(structured_data['items'][:3]):  # Show first 3 items
                print(f"   {i+1}. {item['description']}: {item['amount']}")

        print(f"🚨 SERVER DEBUG: === APPLICATION END ===\n")

        return structured_data if structured_data['items'] else None
        
    except Exception as e:
        print(f"Error parsing table to structured format: {e}")
        return None

def find_dense_blocks(df, min_rows=2, min_cols=3, density_threshold=0.6):
    blocks = []
    nrows, ncols = df.shape
    for row_start in range(nrows - min_rows + 1):
        for col_start in range(ncols - min_cols + 1):
            for row_end in range(row_start + min_rows, nrows + 1):
                for col_end in range(col_start + min_cols, ncols + 1):
                    block = df.iloc[row_start:row_end, col_start:col_end]
                    total_cells = block.size
                    non_empty_cells = block.notnull().values.sum()
                    if total_cells > 0 and (non_empty_cells / total_cells) >= density_threshold:
                        # Avoid duplicates
                        if not any((row_start >= b[0] and row_end <= b[1] and col_start >= b[2] and col_end <= b[3]) for b in blocks):
                            blocks.append((row_start, row_end, col_start, col_end))
    return blocks

def extract_tables_robust(worksheet, entity_keywords):
    """
    Robust table extraction using the original method from utils.py
    """
    tables = []

    print(f"🔍 EXTRACT_TABLES_ROBUST: Processing worksheet '{worksheet.title}' with entity_keywords: {entity_keywords}")

    try:
        # Method 1: Try to extract from openpyxl tables (works for individually formatted tables)
        if hasattr(worksheet, '_tables') and worksheet._tables:
            print(f"🔍 Found {len(worksheet._tables)} formal Excel tables")
            for tbl in worksheet._tables.values():
                try:
                    ref = tbl.ref
                    from openpyxl.utils import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(ref)
                    data = []
                    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                        data.append(row)
                    if data and len(data) >= 2:
                        tables.append({
                            'data': data,
                            'method': 'openpyxl_table',
                            'name': tbl.name,
                            'range': ref
                        })
                except Exception as e:
                    print(f"Failed to extract table {tbl.name}: {e}")
                    continue
        
        # Method 2: Original method from utils.py - DataFrame splitting on empty rows
        print(f"🔍 Trying Method 2: DataFrame splitting on empty rows")
        try:
            # Convert worksheet to DataFrame
            all_data = []
            for row in worksheet.iter_rows(values_only=True):
                all_data.append(row)

            print(f"🔍 Raw worksheet data: {len(all_data)} rows")

            if all_data:
                df = pd.DataFrame(all_data)
                df = df.dropna(how='all').dropna(axis=1, how='all')

                print(f"🔍 After cleaning: DataFrame shape {df.shape}")

                if len(df) >= 2:
                    print(f"🔍 DataFrame has {len(df)} rows, proceeding with splitting...")
                    # Split dataframes on empty rows (original method)
                    empty_rows = df.index[df.isnull().all(1)]
                    start_idx = 0
                    dataframes = []
                    
                    for end_idx in empty_rows:
                        if end_idx > start_idx:
                            split_df = df[start_idx:end_idx]
                            if not split_df.dropna(how='all').empty:
                                dataframes.append(split_df)
                            start_idx = end_idx + 1
                    
                    if start_idx < len(df):
                        dataframes.append(df[start_idx:])
                    
                    # Filter dataframes by entity keywords (original method)
                    combined_pattern = '|'.join(re.escape(kw) for kw in entity_keywords)
                    
                    for i, data_frame in enumerate(dataframes):
                        print(f"🔍 Checking dataframe {i}: shape {data_frame.shape}, combined_pattern: '{combined_pattern}'")

                        # Check if dataframe contains entity keywords
                        mask = data_frame.apply(
                            lambda row: row.astype(str).str.contains(
                                combined_pattern, case=False, regex=True, na=False
                            ).any(),
                            axis=1
                        )

                        print(f"🔍 Dataframe {i} entity match: {mask.any()}")

                        if mask.any():
                            # Convert DataFrame to list format for consistency
                            table_data = [data_frame.columns.tolist()] + data_frame.values.tolist()
                            
                            # Check if table has meaningful content (not empty)
                            if table_data and len(table_data) > 1:
                                # Check if there's actual data beyond headers
                                has_data = False
                                for row in table_data[1:]:  # Skip header row
                                    if any(cell and str(cell).strip() for cell in row):
                                        has_data = True
                                        break
                                
                                if has_data:
                                    print(f"✅ ADDED TABLE: original_table_{i} with {len(table_data)} rows")
                                    tables.append({
                                        'data': table_data,
                                        'method': 'original_split',
                                        'name': f'original_table_{i}',
                                        'range': f'dataframe_{i}'
                                    })
                            
        except Exception as e:
            print(f"Error in original table detection: {e}")
        
        return tables
        
    except Exception as e:
        print(f"Error in robust table extraction: {e}")
        return []



def process_and_filter_excel(filename, tab_name_mapping, entity_name, entity_suffixes):
    """Process and filter Excel file"""
    try:
        # 🚨 SERVER DEBUG: Main Application Entry Point
        print(f"\n🚨 SERVER DEBUG: === MAIN APPLICATION START ===")
        print(f"🚨 SERVER DEBUG: Function: process_and_filter_excel")
        print(f"🚨 SERVER DEBUG: Filename: '{filename}'")
        print(f"🚨 SERVER DEBUG: Entity: '{entity_name}'")
        print(f"🚨 SERVER DEBUG: Entity suffixes: {entity_suffixes}")
        print(f"🚨 SERVER DEBUG: Tab mapping type: {type(tab_name_mapping)}")
        print(f"🚨 SERVER DEBUG: Tab mapping keys: {list(tab_name_mapping.keys()) if isinstance(tab_name_mapping, dict) else 'Not a dict'}")
        print(f"🚨 SERVER DEBUG: Tab mapping sample: {list(tab_name_mapping.items())[:2] if isinstance(tab_name_mapping, dict) else 'Not a dict'}")
        print(f"🚨 SERVER DEBUG: Current working directory: {os.getcwd()}")

        main_dir = Path(__file__).parent.parent
        file_path = main_dir / filename
        print(f"🚨 SERVER DEBUG: Looking for file at: {file_path}")
        print(f"🚨 SERVER DEBUG: File exists: {file_path.exists()}")

        wb = None
        markdown_content = ""
        entity_keywords = [entity_name] + list(entity_suffixes)
        entity_keywords = [kw.strip().lower() for kw in entity_keywords if kw.strip()]
        print(f"🚨 SERVER DEBUG: Generated entity keywords: {entity_keywords}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Create reverse mapping from sheet names to keys
            reverse_mapping = {}
            for key, sheet_names in tab_name_mapping.items():
                for sheet_name in sheet_names:
                    reverse_mapping[sheet_name] = key
            
            print(f"🚨 SERVER DEBUG: Reverse mapping created: {reverse_mapping}")
            
            for ws in wb.worksheets:
                if ws.title not in reverse_mapping:
                    print(f"⏭️ SKIPPING WORKSHEET: {ws.title} (not in mapping)")
                    continue

                print(f"\n🔍 PROCESSING WORKSHEET: {ws.title}")
                print(f"🔍 Mapped to: {reverse_mapping[ws.title]}")

                # Processing worksheet: {ws.title}

                # Use robust table extraction
                tables = extract_tables_robust(ws, entity_keywords)
                print(f"🔍 Found {len(tables)} tables in worksheet {ws.title}")
                
                for table_idx, table_info in enumerate(tables):
                    try:
                        data = table_info['data']
                        method = table_info['method']
                        print(f"📊 PROCESSING TABLE {table_idx + 1}/{len(tables)} in {ws.title}")
                        print(f"📊 Table method: {method}")
                        print(f"📊 Table size: {len(data)} rows x {len(data[0]) if data else 0} columns")

                        # Show first few rows of the table
                        if data:
                            print("📊 TABLE PREVIEW (ALL ROWS):")
                            for i, row in enumerate(data):  # Show ALL rows
                                print(f"   Row {i}: {row}")
                            print(f"   📊 TOTAL: {len(data)} rows in this table")
                        table_name = table_info['name']
                        
                        if not data or len(data) < 2:
                            continue
                        
                        # Create DataFrame
                        df = pd.DataFrame(data[1:], columns=data[0])
                        df = df.dropna(how='all').dropna(axis=1, how='all')
                        df = df.map(lambda x: str(x) if x is not None else "")
                        df = df.reset_index(drop=True)
                        
                        # Check for entity keywords - handle mixed data types safely
                        all_cells = [str(cell).lower().strip() for cell in df.values.flatten()]
                        match_found = any(any(kw in cell for cell in all_cells) for kw in entity_keywords)
                        
                        if match_found:
                            print(f"✅ ENTITY MATCH FOUND in table {table_idx + 1} of {ws.title}")
                            print(f"   Table name: {table_name}")

                            # Include all rows that contain any entity information
                            # This allows the AI to see all entity names in the table
                            filtered_rows = []
                            for idx, row in df.iterrows():
                                row_cells = [str(cell).lower().strip() for cell in row.values]
                                # Include rows that contain any entity information
                                # This will help the AI identify the correct entity names
                                if any(cell for cell in row_cells if cell and cell != 'nan'):
                                    filtered_rows.append(row)

                            print(f"   Filtered to {len(filtered_rows)} rows from original {len(df)} rows")

                            # Create filtered DataFrame and parse it into structured format
                            if filtered_rows:
                                filtered_df = pd.DataFrame(filtered_rows)
                                print(f"   Filtered DataFrame shape: {filtered_df.shape}")
                                print(f"   Filtered DataFrame columns: {list(filtered_df.columns)}")

                                # Parse the table into structured format
                                print(f"🚨 SERVER DEBUG: === ABOUT TO CALL TABLE PARSER ===")
                                print(f"🚨 SERVER DEBUG: Table name: '{table_name}' | Entity: '{entity_name}'")
                                print(f"🚨 SERVER DEBUG: Filtered DataFrame: {filtered_df.shape} rows x {filtered_df.shape[1]} columns")
                                print(f"🚨 SERVER DEBUG: First few rows of filtered data:")
                                for i in range(min(3, len(filtered_df))):
                                    row_data = filtered_df.iloc[i].fillna('').values
                                    print(f"   Row {i}: {row_data}")

                                print(f"🔄 CALLING parse_table_to_structured_format for table: {table_name}")
                                structured_table = parse_table_to_structured_format(filtered_df, entity_name, table_name)

                                print(f"🚨 SERVER DEBUG: === TABLE PARSER RETURNED ===")
                                print(f"🚨 SERVER DEBUG: Result: {'SUCCESS' if structured_table else 'FAILED/NONE'}")
                                if structured_table:
                                    print(f"🚨 SERVER DEBUG: Parsed table has {len(structured_table.get('items', []))} items")
                                    print(f"🚨 SERVER DEBUG: Multiplier: {structured_table.get('multiplier', 'N/A')}x")
                                else:
                                    print(f"🚨 SERVER DEBUG: No structured table returned - likely no valid items found")
                            else:
                                print(f"❌ NO FILTERED ROWS for table {table_name}")
                                structured_table = None
                        else:
                            print(f"⏭️ NO ENTITY MATCH in table {table_idx + 1} of {ws.title}")
                            structured_table = None

                        if structured_table:
                            # Add structured table to markdown content
                            markdown_content += f"## {structured_table['table_name']}\n"
                            markdown_content += f"**Entity:** {structured_table['entity']}\n"
                            markdown_content += f"**Date:** {structured_table['date']}\n"
                            markdown_content += f"**Currency:** {structured_table['currency']}\n"
                            markdown_content += f"**Multiplier:** {structured_table['multiplier']}\n\n"

                            # Add items with properly formatted values
                            for item in structured_table['items']:
                                # Format numeric values with commas and 2 decimal places
                                try:
                                    if isinstance(item['amount'], (int, float)):
                                        formatted_amount = ",.2f"
                                    else:
                                        formatted_amount = item['amount']
                                except:
                                    formatted_amount = item['amount']

                                markdown_content += f"- {item['description']}: {formatted_amount}\n"

                            # Format total value
                            try:
                                if isinstance(structured_table['total'], (int, float)):
                                    formatted_total = ",.2f"
                                else:
                                    formatted_total = structured_table['total']
                            except:
                                formatted_total = structured_table['total']

                            markdown_content += f"\n**Total:** {formatted_total}\n\n"
                        else:
                            # Fallback to original format if parsing fails
                            try:
                                markdown_content += tabulate(filtered_df, headers='keys', tablefmt='pipe') + '\n\n'
                            except Exception:
                                markdown_content += filtered_df.to_markdown(index=False) + '\n\n'
                            
                    except Exception as e:
                        print(f"Error processing table {table_info.get('name', 'unknown')}: {e}")
                        continue
        finally:
            # Ensure workbook is closed to release file handle on Windows
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

        # 🚨 SERVER DEBUG: Final Application Results
        print(f"\n🚨 SERVER DEBUG: === MAIN APPLICATION END ===")
        print(f"🚨 SERVER DEBUG: Processing complete for file: '{filename}'")
        print(f"🚨 SERVER DEBUG: Entity: '{entity_name}'")
        print(f"🚨 SERVER DEBUG: Final markdown content length: {len(markdown_content)} characters")
        print(f"🚨 SERVER DEBUG: Returning result: {'SUCCESS' if markdown_content else 'EMPTY'}")
        print(f"🚨 SERVER DEBUG: === APPLICATION FINISHED ===\n")

        return markdown_content
        
    except Exception as e:
        print("An error occurred while processing the Excel file:", e)
        return ""

def detect_latest_date_column(df, sheet_name="Sheet", entity_keywords=None):
    """Detect the latest date column from a DataFrame, including xMxx format dates."""
    import re
    from datetime import datetime
    
    def parse_date(date_str):
        """Parse date string in various formats including xMxx."""
        if not date_str or pd.isna(date_str):
            return None
        
        date_str = str(date_str).strip()
        
        # Handle xMxx format (e.g., 9M22, 12M23) - END OF MONTH
        xmxx_match = re.match(r'^(\d+)M(\d{2})$', date_str)
        if xmxx_match:
            month = int(xmxx_match.group(1))
            year = 2000 + int(xmxx_match.group(2))  # Assume 20xx for 2-digit years
            # Use end of month, not beginning (last day of the month)
            if month == 12:
                return datetime(year, 12, 31)  # December 31st
            elif month in [1, 3, 5, 7, 8, 10]:
                return datetime(year, month, 31)  # 31-day months
            elif month in [4, 6, 9, 11]:
                return datetime(year, month, 30)  # 30-day months
            elif month == 2:
                # February - handle leap years
                if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
                    return datetime(year, 2, 29)  # Leap year
                else:
                    return datetime(year, 2, 28)  # Non-leap year
        
        # Handle standard date formats
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y',
            '%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y',
            '%d/%b/%Y', '%d-%b-%Y', '%b/%d/%Y', '%b-%d-%Y',
            '%d/%B/%Y', '%d-%B-%Y', '%B/%d/%Y', '%B-%d-%Y',
            # Chinese date formats
            '%Y年%m月%d日', '%Y年%m月', '%m月%d日', '%Y/%m/%d',
            '%Y.%m.%d', '%Y年%m月%d号',
            # Additional flexible formats
            '%Y%m%d', '%d%m%Y', '%m%d%Y'
        ]
        
        for fmt in date_formats:
            try:
                result = datetime.strptime(date_str, fmt)
                if '年' in date_str or '月' in date_str or '日' in date_str:
                    print(f"✅ Successfully parsed Chinese date '{date_str}' using format '{fmt}' -> {result.strftime('%Y-%m-%d')}")
                return result
            except ValueError:
                continue
        
        return None
    
    # Get column names
    columns = df.columns.tolist()
    latest_date = None
    latest_column = None

    print(f"🔍 {sheet_name}: Searching for latest date column...")
    print(f"   Available columns: {columns}")
    print(f"   Entity keywords provided: {entity_keywords}")
    print(f"   ✅ detect_latest_date_column called successfully with entity_keywords: {entity_keywords}")

    # Strategy 1: Extract entity-related tables first, then find "Indicative adjusted" columns
    print(f"   🎯 STEP 1: Extracting entity-related tables from {sheet_name}")
    
    # Extract entity-related tables using the same logic as the main processing
    entity_tables = []
    entity_keywords_list = entity_keywords or []
    
    # Split dataframe into sections based on empty rows
    empty_rows = df.index[df.isnull().all(1)]
    start_idx = 0
    sections = []
    
    for end_idx in empty_rows:
        if end_idx > start_idx:
            section_df = df[start_idx:end_idx]
            if not section_df.dropna(how='all').empty:
                sections.append((start_idx, end_idx, section_df))
            start_idx = end_idx + 1
    
    if start_idx < len(df):
        sections.append((start_idx, len(df), df[start_idx:]))
    
    # Filter sections to only those containing entity keywords
    for start_row, end_row, section_df in sections:
        # Check if this section contains entity keywords
        section_has_entity = False
        all_cells = [str(cell).lower() for cell in section_df.values.flatten()]
        
        for keyword in entity_keywords_list:
            # More flexible matching: check if keyword is contained in any cell (case-insensitive)
            keyword_lower = keyword.lower()
            if any(keyword_lower in cell for cell in all_cells):
                section_has_entity = True
                entity_tables.append((start_row, end_row, section_df))
                print(f"   ✅ Entity table found: Rows {start_row}-{end_row} (contains '{keyword}')")
                break
    
    if not entity_tables:
        print(f"   ⚠️  No entity-specific tables found, using all sections as fallback")
        if sections:
            entity_tables = sections  # Use all sections instead of just the first one
    
    print(f"   📊 Found {len(entity_tables)} entity-related tables")
    
    # Strategy 2: Within entity tables, find "Indicative adjusted" and get the correct column
    print(f"   🎯 STEP 2: Searching for 'Indicative adjusted' in {len(entity_tables)} tables")
    indicative_positions = []
    all_found_dates = []
    
    for start_row, end_row, entity_df in entity_tables:
        print(f"   🔍 STEP 2: Searching entity table (rows {start_row}-{end_row}) for 'Indicative adjusted'")
        
        # Find "Indicative adjusted" positions within this entity table
        for local_row_idx in range(len(entity_df)):
            global_row_idx = start_row + local_row_idx
            for col_idx, col in enumerate(columns):
                val = entity_df.iloc[local_row_idx, col_idx]
                if pd.notna(val) and 'indicative' in str(val).lower() and 'adjusted' in str(val).lower():
                    indicative_positions.append((global_row_idx, col_idx))
                    print(f"     📋 Found 'Indicative adjusted' at Row {global_row_idx}, Col {col_idx} ({col})")
        
        # Find dates within this entity table
        print(f"     🔍 Searching for dates in table (rows {start_row}-{end_row})...")
        for local_row_idx in range(len(entity_df)):
            global_row_idx = start_row + local_row_idx
            for col_idx, col in enumerate(columns):
                val = entity_df.iloc[local_row_idx, col_idx]
                
                if isinstance(val, (pd.Timestamp, datetime)):
                    date_val = val if isinstance(val, datetime) else val.to_pydatetime()
                    all_found_dates.append((date_val, col, global_row_idx, col_idx, "datetime"))
                    print(f"     📅 Found datetime in {col}[{global_row_idx}]: {date_val.strftime('%Y-%m-%d')}")
                elif pd.notna(val):
                                    parsed_date = parse_date(str(val))
                if parsed_date:
                    all_found_dates.append((parsed_date, col, global_row_idx, col_idx, "parsed"))
                    print(f"     📅 Parsed date in {col}[{global_row_idx}]: '{val}' -> {parsed_date.strftime('%Y-%m-%d')}")
                    if '年' in str(val) or '月' in str(val) or '日' in str(val):
                        print(f"     🎯 Chinese date detected and parsed: '{val}'")
    
    # Strategy 3: Prioritize "Indicative adjusted" columns with latest dates
    if indicative_positions and all_found_dates:
        print(f"   🎯 STEP 3: Prioritizing 'Indicative adjusted' columns with latest dates")
        
        # Find the latest date
        max_date = max(all_found_dates, key=lambda x: x[0])[0]
        latest_date_columns = [item for item in all_found_dates if item[0] == max_date]
        
        print(f"   📊 Latest date found: {max_date.strftime('%Y-%m-%d')}")
        if len(latest_date_columns) > 1:
            print(f"   📊 Multiple columns with latest date:")
            for date_val, col, row, col_idx, source in latest_date_columns:
                print(f"      • {col} (col {col_idx})")
        
        # Find which "Indicative adjusted" positions have the latest date
        selected_column = None
        
        for indic_row, indic_col in indicative_positions:
            print(f"   🔍 Checking 'Indicative adjusted' at col {indic_col}")
            
            # Check if this exact column has the latest date
            exact_match = None
            for date_val, col, row, col_idx, source in latest_date_columns:
                if col_idx == indic_col:
                    exact_match = (date_val, col, row, col_idx, source)
                    print(f"     ✅ EXACT match: {col} (col {col_idx}) has latest date")
                    break
            
            if exact_match:
                selected_column = exact_match
                print(f"   🎯 SELECTED: {selected_column[1]} (EXACT 'Indicative adjusted' column with latest date)")
                break
            else:
                # Check merged range for this "Indicative adjusted"
                print(f"     🔍 EXACT column doesn't have latest date, checking merged range...")
                
                # Detect merged range
                merge_start = indic_col
                merge_end = indic_col
                
                for check_col in range(indic_col + 1, len(columns)):
                    val = df.iloc[indic_row, check_col]
                    if pd.isna(val):
                        merge_end = check_col
                    else:
                        merge_end = check_col - 1
                        break
                else:
                    merge_end = len(columns) - 1
                
                print(f"     📍 Merged range: columns {merge_start}-{merge_end}")
                
                # Find latest date within this merged range
                range_matches = []
                for date_val, col, row, col_idx, source in latest_date_columns:
                    if merge_start <= col_idx <= merge_end:
                        range_matches.append((date_val, col, row, col_idx, source))
                        print(f"     ✅ {col} (col {col_idx}) is in merged range with latest date")
                
                if range_matches:
                    selected_column = range_matches[0]  # Use first match in range
                    print(f"   🎯 SELECTED: {selected_column[1]} (latest date in 'Indicative adjusted' merged range)")
                    break
        
        if selected_column:
            latest_date, latest_column = selected_column[0], selected_column[1]
        else:
            # No "Indicative adjusted" column found with latest date, use first latest
            selected_column = latest_date_columns[0]
            latest_date, latest_column = selected_column[0], selected_column[1]
            print(f"   ⚠️  No 'Indicative adjusted' match, using: {latest_column}")
    
    # Strategy 2: Fallback to simple logic if no "Indicative adjusted" found
    else:
        print(f"   🔍 No 'Indicative adjusted' found, using simple date detection...")
        
        # First, try to find dates in column names
        column_dates_found = []
        for col in columns:
            col_str = str(col)
            parsed_date = parse_date(col_str)
            if parsed_date:
                column_dates_found.append((parsed_date, col, "column_name"))
                print(f"   📅 Found date in column name '{col}': {parsed_date.strftime('%Y-%m-%d')}")
                if latest_date is None or parsed_date > latest_date:
                    latest_date = parsed_date
                    latest_column = col
                    print(f"   ✅ New latest: {col} ({parsed_date.strftime('%Y-%m-%d')})")
        
        # If no dates found in column names, check the first few rows for datetime values
        if latest_column is None and len(df) > 0:
            print(f"   🔍 No dates in column names, checking row values...")
            cell_dates_found = []
            
            # Check first 5 rows for date values (dates can be in different rows)
            for row_idx in range(min(5, len(df))):
                row = df.iloc[row_idx]
                for col in columns:
                    val = row[col]
                    
                    # Check if it's already a datetime object
                    if isinstance(val, (pd.Timestamp, datetime)):
                        date_val = val if isinstance(val, datetime) else val.to_pydatetime()
                        cell_dates_found.append((date_val, col, f"row_{row_idx}"))
                        print(f"   📅 Found datetime in {col}[{row_idx}]: {date_val.strftime('%Y-%m-%d')}")
                        if latest_date is None or date_val > latest_date:
                            latest_date = date_val
                            latest_column = col
                            print(f"   ✅ New latest: {col} ({date_val.strftime('%Y-%m-%d')}) from row {row_idx}")
                    # Check if it's a string that can be parsed as a date
                    elif pd.notna(val):
                        parsed_date = parse_date(str(val))
                        if parsed_date:
                            cell_dates_found.append((parsed_date, col, f"row_{row_idx}_parsed"))
                            print(f"   📅 Parsed date in {col}[{row_idx}]: '{val}' -> {parsed_date.strftime('%Y-%m-%d')}")
                            if latest_date is None or parsed_date > latest_date:
                                latest_date = parsed_date
                                latest_column = col
                                print(f"   ✅ New latest: {col} ({parsed_date.strftime('%Y-%m-%d')}) from row {row_idx}")
            
            if not cell_dates_found:
                print(f"   ❌ No dates found in cell values")
    
    # Summary of selection
    if latest_column:
        print(f"   🎯 FINAL SELECTION: Column '{latest_column}' with date {latest_date.strftime('%Y-%m-%d')}")
        
        # Show comparison if multiple dates were found
        if 'all_found_dates' in locals() and len(all_found_dates) > 1:
            print(f"   📊 All dates found (for comparison):")
            for date_val, col, row, col_idx, source in sorted(all_found_dates, key=lambda x: x[0], reverse=True):
                marker = "👑" if col == latest_column else "  "
                print(f"   {marker} {col}: {date_val.strftime('%Y-%m-%d')} (from row {row})")
    else:
        print(f"   ❌ No date column detected")
    
    return latest_column

def find_financial_figures_with_context_check(filename, sheet_name, date_str, convert_thousands=False, entity_keywords=None):
    try:
        file_path = Path(filename)
        with pd.ExcelFile(file_path) as xl:
            # Handle both single sheet name and list of possible names
            if isinstance(sheet_name, list):
                found_sheet = None
                for possible_sheet in sheet_name:
                    if possible_sheet in xl.sheet_names:
                        found_sheet = possible_sheet
                        break

                if found_sheet is None:
                    print(f"None of the sheet names {sheet_name} found in the file. Available sheets: {xl.sheet_names}")
                    return {}
                sheet_name = found_sheet
                print(f"Found matching sheet: '{sheet_name}'")
            elif sheet_name not in xl.sheet_names:
                print(f"Sheet '{sheet_name}' not found in the file. Available sheets: {xl.sheet_names}")
                return {}
            df = xl.parse(sheet_name)
        if not isinstance(df, pd.DataFrame):
            return {}
        
        # Detect latest date column automatically
        latest_date_col = detect_latest_date_column(df, sheet_name, entity_keywords)
        if latest_date_col:
            # Use the latest date column instead of the requested date
            date_column = latest_date_col
            print(f"Using latest date column: {latest_date_col}")
        else:
            # Fallback to original logic only if date_str is provided
            if date_str is not None:
                # Handle different sheet formats
                if sheet_name == 'BSHN':
                    # BSHN sheet has different column structure
                    df.columns = ['Description', 'Column1', 'Column2', 'Column3']
                    date_column_map = {
                        '31/12/2020': 'Column1',
                        '31/12/2021': 'Column2', 
                        '30/09/2022': 'Column3'
                    }
                else:
                    # Standard sheet format
                    df.columns = ['Description', 'Date_2020', 'Date_2021', 'Date_2022']
                    date_column_map = {
                        '31/12/2020': 'Date_2020',
                        '31/12/2021': 'Date_2021',
                        '30/09/2022': 'Date_2022'
                    }
                if date_str not in date_column_map:
                    print(f"Date '{date_str}' not recognized.")
                    return {}
                date_column = date_column_map[date_str]
            else:
                print("No date column detected and no fallback date provided.")
                return {}
        # If convert_thousands and '000' in columns or first row, multiply numeric values by 1000 for AI processing
        # For BSHN sheet, always apply scale factor since it's in '000 format
        if sheet_name == 'BSHN':
            scale_factor = 1000  # BSHN sheet is always in '000 format
        else:
            # Check for both English and Chinese thousands notation
            has_thousands_notation = (
                convert_thousands and (
                    any("'000" in str(col) for col in df.columns) or
                    any("千元" in str(col) for col in df.columns) or
                    any("千人民币" in str(col) for col in df.columns) or
                    any("人民币千" in str(col) for col in df.columns) or
                    any("千元人民币" in str(col) for col in df.columns) or
                    any("人民币千元" in str(col) for col in df.columns) or
                    any("千人民幣" in str(col) for col in df.columns) or
                    any("人民幣千" in str(col) for col in df.columns) or
                    any("千元人民幣" in str(col) for col in df.columns) or
                    any("人民幣千元" in str(col) for col in df.columns)
                )
            )
            scale_factor = 1000 if has_thousands_notation else 1
        financial_figure_map = {
            "Cash": "Cash at bank",
            "AR": "Accounts receivable",
            "Prepayments": "Prepayments",
            "OR": "Other receivables",
            "Other CA": "Other current assets",
            "IP": "Investment properties",
            "Other NCA": "Other non-current assets",
            "AP": "Accounts payable",
            "Taxes payable": "Taxes payable",
            "OP": "Other payables",
            "Capital": "Paid-in capital",
            "Reserve": "Surplus reserve"
        }
        financial_figures = {}
        
        # Find the description column (usually the first column)
        desc_column = None
        if 'Description' in df.columns:
            desc_column = 'Description'
        elif len(df.columns) > 0:
            desc_column = df.columns[0]  # Use first column as description column
        
        for key, desc in financial_figure_map.items():
            if desc_column and date_column in df.columns:
                # Convert description column to string to avoid .str accessor error
                desc_series = df[desc_column].astype(str)
                value = df.loc[desc_series.str.contains(desc, case=False, na=False), date_column].values
                if value.size > 0:
                    # Apply scale factor: multiply by 1000 if '000 notation detected
                    financial_figures[key] = float(value[0]) * scale_factor
        return financial_figures
    except FileNotFoundError:
        print(f"❌ Excel file not found: {filename}")
        print(f"   Expected path: {file_path}")
        return {}
    except PermissionError:
        print(f"❌ Permission denied accessing Excel file: {filename}")
        print(f"   Check file permissions: {file_path}")
        return {}
    except Exception as e:
        print(f"❌ Unexpected error processing Excel file '{filename}': {e}")
        print(f"   File path: {file_path}")
        print(f"   Sheet name: {sheet_name}")
        return {}

def get_tab_name(project_name):
    if not project_name:
        return None

    project_name = project_name.strip()

    # Hardcoded mappings for known entities
    if project_name.lower() == 'haining':
        return "BSHN"
    elif project_name.lower() == 'nanjing':
        return "BSNJ"
    elif project_name.lower() == 'ningbo':
        return "BSNB"

    # For other entities, try to extract a meaningful sheet name
    # Remove common suffixes and use first word
    clean_name = project_name.split()[0] if project_name else None
    if clean_name:
        # Try different sheet name patterns
        possible_names = [
            f"BS{clean_name.upper()[:3]}",  # BSCLE, BSHAI, etc.
            f"{clean_name.upper()[:3]}",     # CLE, HAI, etc.
            clean_name.upper(),             # CLEANTECH, HAINING, etc.
            f"BS_{clean_name.upper()[:3]}",  # BS_CLE, etc.
            project_name                     # Original name as last resort
        ]
        return possible_names  # Return list of possible names

    # Fallback: return the project name itself to avoid None
    print(f"Warning: Could not extract sheet name from '{project_name}', using default sheet name")
    return [project_name]

def get_financial_figure(financial_figures, key):
    """Get financial figure with proper K/M formatting and 1 decimal place"""
    figure = financial_figures.get(key, None)
    if figure is None:
        return f"{key} not found in the financial figures."
    
    # Ensure 1 decimal place for all conversions
    if figure >= 1000000:
        return f"{figure / 1000000:.1f}M"
    elif figure >= 1000:
        return f"{figure / 1000:.1f}K"  # Changed to 1dp for K as well
    else:
        return f"{figure:.1f}"

def detect_string_in_file(file_content, target_string):
    try:
        return target_string in file_content
    except Exception:
        return False

# Global cache for loaded JSON files to avoid repeated I/O
_json_cache = {}

def clear_json_cache():
    """Clear the JSON cache to ensure fresh data loading"""
    global _json_cache
    _json_cache.clear()
    print("🧹 JSON cache cleared")

def load_ip(file, key=None):
    """Load JSON file with caching to improve performance"""
    try:
        # Use file path as cache key
        cache_key = file

        # Check cache first
        if cache_key in _json_cache:
            cached_data = _json_cache[cache_key]
        else:
            with open(file, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            # Cache the loaded data
            _json_cache[cache_key] = cached_data

        if key is not None and key in cached_data:
            return cached_data[key]
        return cached_data
    except FileNotFoundError:
        print(f"File {file} not found.")
    except json.JSONDecodeError:
        print(f"Error decoding JSON from file {file}.")
    return {}

# --- Pattern Filling and Main Processing ---
def process_keys(keys, entity_name, entity_helpers, input_file, mapping_file, pattern_file, config_file='utils/config.json', prompts_file='utils/prompts.json', use_ai=True, convert_thousands=False, progress_callback=None, processed_table_data=None, use_local_ai=False, use_openai=False, language='english'):
    # AI is required - no fallback mode
    if not use_ai:
        raise RuntimeError("AI processing is required. Cannot run in offline mode.")
    
    if not AI_AVAILABLE:
        raise RuntimeError("AI services are not available. Please check your configuration and internet connection.")

    print(f"🚀 Starting AI processing for {len(keys)} keys (Offline mode removed)")
    
    # Load prompts from prompts.json file (no hardcoded fallback)
    with open(prompts_file, 'r', encoding='utf-8') as f:
        prompts_config = json.load(f)

    # Use the passed language parameter (defaults to 'english')
    print(f"🔍 DEBUG: Language parameter: '{language}'")
    print(f"🔍 DEBUG: Available languages in prompts: {list(prompts_config.get('system_prompts', {}).keys())}")
    
    system_prompts = prompts_config.get('system_prompts', {}).get(language, {})
    print(f"🔍 DEBUG: System prompts for language '{language}': {list(system_prompts.keys())}")

    system_prompt = system_prompts.get('Agent 1')
    if not system_prompt:
        # Fallback to direct access if nested structure fails
        system_prompt = prompts_config.get('system_prompts', {}).get('Agent 1')
        print(f"🔍 DEBUG: Using fallback system prompt")
    
    if not system_prompt:
        print(f"❌ ERROR: No system prompt found for language '{language}' or fallback")
        raise RuntimeError(f"No system prompt found for language '{language}'")
    
    print(f"✅ DEBUG: System prompt loaded successfully (length: {len(system_prompt)})")
    
    # No need to initialize financial figures - we're using cached data
    print(f"🔍 DEBUG: Using cached data - no need to call find_financial_figures_with_context_check")
    results = {}
    
    # Use Streamlit progress instead of tqdm for better UI integration
    use_streamlit_progress = progress_callback is not None
    if not use_streamlit_progress:
        # Fallback to tqdm for console-only usage
        pbar = tqdm(keys, desc="🤖 AI Processing", unit="key", total=len(keys),
                    bar_format='{desc}: {percentage:3.0f}%|{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]')
    else:
        pbar = None

    # Track start time for final summary
    start_time = time.time()

    for key_index, key in enumerate(pbar):
        # Determine AI model being used for display
        print(f"🔍 DEBUG: Processing key {key_index}: {repr(key)} (type: {type(key)})")
        config_details = load_config(config_file)
        if use_local_ai:
            ai_model = "Local AI"
            openai_model = config_details.get('LOCAL_AI_CHAT_MODEL', 'local-model')
        elif use_openai:
            ai_model = "OpenAI"
            openai_model = config_details.get('OPENAI_CHAT_MODEL', 'gpt-4o-mini-2024-07-18')
        else:
            ai_model = "DeepSeek"
            openai_model = config_details.get('DEEPSEEK_CHAT_MODEL', 'deepseek-chat')

        # Initial progress update with key info
        if not use_streamlit_progress:
            progress_desc = f"🔄 {key} ({ai_model})"
            pbar.set_description(progress_desc)

        # Enhanced Streamlit progress with detailed status
        if progress_callback:
            detailed_message = f"🔄 Processing {key} • {ai_model} • Key {key_index + 1}/{len(keys)}"
            progress_callback((key_index + 1) / len(keys), detailed_message)

        # Update progress: Data loading phase
        if not use_streamlit_progress:
            pbar.set_postfix_str("📊 Loading data...")
        if progress_callback:
            progress_callback((key_index + 0.1) / len(keys), f"📊 Loading data for {key}...")

        # Initialize AI services - required for processing
        oai_client, search_client = initialize_ai_services(config_details, use_local=use_local_ai, use_openai=use_openai)

        pattern = load_ip(pattern_file, key)
        # No need to load mapping since we're using cached data
        print(f"🔍 DEBUG: Using cached data - no need to load mapping")

        # Use processed table data - this should always be available
        print(f"🔍 DEBUG: processed_table_data keys: {list(processed_table_data.keys()) if processed_table_data else 'None'}")
        print(f"🔍 DEBUG: Looking for key: {key}")
        print(f"🔍 DEBUG: Key '{key}' in processed_table_data: {key in processed_table_data if processed_table_data else False}")
        
        if processed_table_data and key in processed_table_data:
            excel_tables = processed_table_data[key]
            data_source = "cached"
            print(f"✅ DEBUG: Using cached data for key: {key}")
            print(f"✅ DEBUG: Cached data has {len(excel_tables) if excel_tables else 0} tables")
        else:
            print(f"❌ ERROR: Key '{key}' not found in processed_table_data")
            print(f"❌ ERROR: Available keys: {list(processed_table_data.keys()) if processed_table_data else 'None'}")
            print(f"❌ ERROR: This should not happen - AI processing requires cached data")
            # Return empty data - this will cause the AI processing to skip this key
            excel_tables = []
            data_source = "empty"

        # Update progress: Data processing phase
        if not use_streamlit_progress:
            pbar.set_postfix_str(f"📈 Processing data ({data_source})...")
        if progress_callback:
            progress_callback((key_index + 0.2) / len(keys), f"📈 Processing {key} data...")

        # Check if '000 notation is detected
        # Check for both English and Chinese thousands notation
        has_thousands_notation = (
            detect_string_in_file(excel_tables, "'000") or
            detect_string_in_file(excel_tables, "千元") or
            detect_string_in_file(excel_tables, "千人民币") or
            detect_string_in_file(excel_tables, "人民币千") or
            detect_string_in_file(excel_tables, "千元人民币") or
            detect_string_in_file(excel_tables, "人民币千元")
        )

        # Process data for AI: multiply figures by 1000 if '000 notation detected
        excel_tables_for_ai = multiply_figures_for_ai_processing(excel_tables) if has_thousands_notation else excel_tables

        # No need to process financial_figures since we're using cached data
        print(f"🔍 DEBUG: Using cached data - no need to process financial_figures")

        # Update progress: AI processing phase
        if not use_streamlit_progress:
            pbar.set_postfix_str(f"🤖 AI generating ({openai_model})...")
        if progress_callback:
            progress_callback((key_index + 0.3) / len(keys), f"🤖 AI generating {key} content...")
        
        # Update prompt to reflect the data processing
        detect_zeros = """IMPORTANT: The numerical figures in the DATA SOURCE have been adjusted for analysis (multiplied by 1000 from the original '000 notation). 
        Express all figures with proper K/M conversion with 1 decimal place:
        - Figures ≥ 1,000,000: express in M (millions) with 1 decimal place (e.g., 2.3M)
        - Figures ≥ 1,000: express in K (thousands) with 1 decimal place (e.g., 1.5K)
        - Figures < 1,000: express with 1 decimal place (e.g., 123.0)""" if has_thousands_notation else """Express all figures with proper K/M conversion with 1 decimal place:
        - Figures ≥ 1,000,000: express in M (millions) with 1 decimal place (e.g., 2.3M)
        - Figures ≥ 1,000: express in K (thousands) with 1 decimal place (e.g., 1.5K)
        - Figures < 1,000: express with 1 decimal place (e.g., 123.0)"""
        
        # User query construction using f-strings for better prompt maintainability
        pattern_json = json.dumps(pattern, indent=2)
        # No need for financial_figure_info since we're using cached data
        financial_figure_info = f"{key}: Using cached table data"
        
        # Template for output requirements - language aware
        if language == 'chinese':
            output_requirements = f"""
            必需的输出格式：
            - 仅包含完成的模式文本
            - 没有模式名称或标签
            - 没有模板结构
            - 没有JSON格式
            - 将所有'xxx'或占位符替换为实际数据值
            - 将所有[ENTITY_NAME]占位符替换为数据源中的具体实体名称
            - 关键：使用表数据中的具体实体名称（如'第三方应收款'、'公司#1'），而不是报告实体名称
            - 不要使用项目符号列出
            - 用正确的K/M转换表示所有数字（如9,076,000 = 9.1M，1,500 = 1.5K）
            - 坚持模板格式，不要添加额外解释或注释
            - 对于要填入模板的实体名称，使用数据源表中的具体实体名称
            - 对于所有列出的数字，请检查总数，应该与财务数据大致相同或构成大部分
            - 确保提到的所有财务数字与数据源中的实际值匹配
            - 重要：查看表数据以识别正确的实体名称（如'第三方应收款'、'公司#1'等）
            """
        else:
            output_requirements = f"""
            REQUIRED OUTPUT FORMAT:
            - Only the completed pattern text
            - No pattern names or labels
            - No template structure
            - No JSON formatting
            - Replace ALL 'xxx' or placeholders with actual data values
            - Replace ALL [ENTITY_NAME] placeholders with the SPECIFIC entity name from the DATA SOURCE
            - CRITICAL: Use the SPECIFIC entity names from the table data (e.g., 'Third-party receivables', 'Company #1') NOT the reporting entity name
            - Do not use bullet point for listing
            - Express all figures with proper K/M conversion (e.g., 9,076,000 = 9.1M, 1,500 = 1.5K)
            - No foreign contents, if any, translate to English
            - Stick to Template format, no extra explanations or comments
            - For entity name to be filled into template, use the specific entity names from the DATA SOURCE table
            - For all listing figures, please check the total, together should be around the same or constituting majority of FINANCIAL FIGURE
            - Ensure all financial figures mentioned match the actual values from the DATA SOURCE
            - IMPORTANT: Look at the table data to identify the correct entity names (e.g., 'Third-party receivables', 'Company #1', etc.)
            """

        # Example formats for consistent output - language aware
        if language == 'chinese':
            examples = f"""
            正确输出格式示例：
            "银行存款包括截至2022年9月30日存放在主要金融机构的CNY9.1M存款。"

            错误输出格式示例：
            "模式1：银行存款包括截至xxx存放在xxx的xxx存款。"
            """
        else:
            examples = f"""
            Example of CORRECT output format:
            "Cash at bank comprises deposits of CNY9.1M held with major financial institutions as at 30/09/2022."

            Example of INCORRECT output format:
            "Pattern 1: Cash at bank comprises deposits of xxx held with xxx as at xxx."
            """
        
        # User query construction - language aware
        if language == 'chinese':
            user_query = f"""
            任务：选择一个模式并用实际数据完成它

            可用模式：{pattern_json}

            财务数据：{financial_figure_info}

            数据源：{excel_tables_for_ai}

            选择标准：
            - 选择数据覆盖最完整的模式
            - 优先选择匹配主要账户类别的模式
            - 使用最新数据：最新的可用数据
            - {detect_zeros}

            {output_requirements}

            {examples}

            重要要求：
            1. 始终指定确切的金额和货币（如"CNY9.1M"、"$2.3M"、"CNY687K"）
            2. 始终识别并提及数据源表中的具体实体名称
            3. 关键：对于实体名称，使用财务数据表中的具体实体名称（如'第三方应收款'、'公司#1'），而不是报告实体名称
            4. 查看表数据以识别正确的实体名称和金额
            5. 当表格显示'第三方应收款'时，使用确切的名称，而不是报告实体
            """
        else:
            user_query = f"""
            TASK: Select ONE pattern and complete it with actual data

            AVAILABLE PATTERNS: {pattern_json}

            FINANCIAL FIGURE: {financial_figure_info}

            DATA SOURCE: {excel_tables_for_ai}

            SELECTION CRITERIA:
            - Choose the pattern with the most complete data coverage
            - Prioritize patterns that match the primary account category
            - Use most recent data: latest available
            - {detect_zeros}

            {output_requirements}

            {examples}

            IMPORTANT REQUIREMENTS:
            1. ALWAYS specify exact dollar amounts and currency (e.g., "CNY9.1M", "$2.3M", "CNY687K")
            2. ALWAYS identify and mention the specific entity names from the DATA SOURCE table
            3. CRITICAL: For entity names, use the SPECIFIC entity names from the financial data table (e.g., 'Third-party receivables', 'Company #1') NOT the reporting entity name
            4. Look at the table data to identify the correct entity names and amounts
            5. When the table shows 'Third-party receivables', use that exact name, not the reporting entity
            """
        
        # Update progress: AI request phase
        if not use_streamlit_progress:
            pbar.set_postfix_str(f"📤 Sending to {openai_model}...")
        if progress_callback:
            progress_callback((key_index + 0.7) / len(keys), f"📤 Sending {key} to {ai_model}...")

        response_txt = generate_response(user_query, system_prompt, oai_client, excel_tables, openai_model, entity_name, use_local_ai)

        # Update progress: Response processing phase
        if not use_streamlit_progress:
            pbar.set_postfix_str("📥 Processing response...")
        if progress_callback:
            progress_callback((key_index + 0.8) / len(keys), f"📥 Processing {key} response...")

        # Clean up response: remove outer quotation marks and translate Chinese (only for English)
        response_txt = clean_response_text(response_txt, language)

        # Store result with pattern information for logging
        results[key] = {
            'content': response_txt,
            'pattern_used': 'Pattern 1',  # Default, will be updated based on actual pattern
            'table_data': excel_tables_for_ai,
            'financial_figure': 0,  # Using cached data, no need for financial_figures
            'entity_name': entity_name
        }

        # Update progress bar with completion info and response preview
        if not use_streamlit_progress:
            completion_status = f"✅ {key}: {response_txt[:15]}..." if len(response_txt) > 15 else f"✅ {key}: {response_txt}"
            pbar.set_postfix_str(completion_status)

        # Enhanced Streamlit progress with completion details
        if progress_callback:
            completion_msg = f"✅ Completed {key} • Generated {len(response_txt)} chars"
            progress_callback((key_index + 1) / len(keys), completion_msg)

    if not use_streamlit_progress and pbar:
        pbar.close()

    # Enhanced final progress update with summary
    total_keys = len(keys)
    successful_keys = len([k for k in results.keys() if results[k].get('content')])
    success_rate = (successful_keys / total_keys * 100) if total_keys > 0 else 0

    final_status = f"🎉 AI processing completed! {successful_keys}/{total_keys} keys processed ({success_rate:.1f}% success)"

    if progress_callback:
        progress_callback(1.0, final_status)

    # Print summary to console as well
    print(f"\n{'='*60}")
    print(f"🤖 AI PROCESSING SUMMARY")
    print(f"{'='*60}")
    print(f"📊 Total keys processed: {total_keys}")
    print(f"✅ Successful: {successful_keys}")
    print(f"❌ Failed: {total_keys - successful_keys}")
    print(f"📈 Success rate: {success_rate:.1f}%")
    print(f"⏱️  Total time: {time.time() - start_time:.2f}s")
    print(f"{'='*60}\n")

    return results

def generate_test_results(keys):
    # This function is deprecated - AI is now required
    raise RuntimeError("AI processing is required. Cannot generate test results.")

# --- QA Agent ---
class QualityAssuranceAgent:
    def __init__(self):
        self.excellent_threshold = 90
        self.good_threshold = 80
        self.acceptable_threshold = 70
        self.template_artifacts = [
            'Pattern 1:', 'Pattern 2:', 'Pattern 3:', '[', ']', '{', '}', 'xxx', 'XXX',
            'template', 'placeholder', 'PLACEHOLDER', 'TBD', 'TODO', 'FIXME'
        ]
        self.professional_terms = [
            'comprised', 'represented', 'indicated', 'demonstrated', 'reflected',
            'maintained', 'established', 'confirmed', 'verified', 'assessed',
            'evaluated', 'analyzed'
        ]
        self.risk_indicators = [
            'provision', 'impairment', 'restricted', 'covenant', 'collateral',
            'mortgage', 'guarantee', 'contingent'
        ]
    def validate_content(self, content: str) -> Dict:
        # Simple QA: check for template artifacts, paragraph structure, and number formatting
        issues = []
        score = 100
        for artifact in self.template_artifacts:
            if artifact.lower() in content.lower():
                issues.append(f"Template artifact found: '{artifact}'")
                score -= 5
        if not re.search(r'^##\s+\w+', content, re.MULTILINE):
            issues.append("Missing proper markdown headers")
            score -= 10
        if re.search(r'###\s+[^\n]+\n\s*\n', content):
            issues.append("Empty content sections detected")
            score -= 5
        paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
        if len(paragraphs) < 3:
            issues.append("Insufficient content paragraphs")
            score -= 5
        return {"score": max(0, score), "issues": issues}
    def auto_correct(self, content: str) -> str:
        # Remove template artifacts and fix paragraph structure
        for artifact in self.template_artifacts:
            content = re.sub(re.escape(artifact), '', content, flags=re.IGNORECASE)
        # Ensure double newlines between paragraphs
        content = re.sub(r'\n{2,}', '\n\n', content)
        return content.strip()

# --- Data Validation Agent ---
class DataValidationAgent:
    def __init__(self):
        self.config_file = 'fdd_utils/config.json'
        self.financial_figure_map = {
            "Cash": "Cash at bank",
            "AR": "Accounts receivable",
            "Prepayments": "Prepayments",
            "OR": "Other receivables",
            "Other CA": "Other current assets",
            "IP": "Investment properties",
            "Other NCA": "Other non-current assets",
            "AP": "Accounts payable",
            "Taxes payable": "Taxes payable",
            "OP": "Other payables",
            "Capital": "Paid-in capital",
            "Reserve": "Surplus reserve"
        }
    
    def validate_financial_data(self, content: str, excel_file: str, entity: str, key: str) -> Dict:
        """Validate that financial figures in content match the Excel data"""
        try:
            import json
            
            # Extract financial figures from Excel
            financial_figures = find_financial_figures_with_context_check(
                excel_file, 
                get_tab_name(entity), 
                None
            )
            expected_figure = financial_figures.get(key)
            
            if not AI_AVAILABLE:
                raise RuntimeError("AI services are required for data validation. Please check your configuration.")
            
            # Use AI to validate data accuracy
            config_details = load_config(self.config_file)
            oai_client, _ = initialize_ai_services(config_details)
            
            # Load system prompt from prompts.json
            try:
                with open('fdd_utils/prompts.json', 'r', encoding='utf-8') as f:
                    prompts_config = json.load(f)
                system_prompt = prompts_config.get('system_prompts', {}).get('Agent 2', '')
                
                if not system_prompt:
                    # Fallback to hardcoded prompt
                    system_prompt = """
                    You are AI2, a financial data validation specialist. Your task is to double-check each response 
                    by key to ensure figures match the balance sheet and verify data accuracy including K/M conversions.
                    
                    CRITICAL REQUIREMENTS:
                    1. Extract all financial figures from AI1 content
                    2. Compare with expected balance sheet figures for accuracy
                    3. Verify proper K/M conversion with 1 decimal place (e.g., 2.3M, 1.5K, 123.0)
                    4. Check entity names match data source (not reporting entity)
                    5. Identify ONLY top 2 most critical data accuracy issues
                    6. Remove unnecessary quotation marks around sections
                    7. Ensure no data inconsistencies or conversion errors
                    8. Verify figures are properly adjusted for '000 notation if applicable
                    """
            except (FileNotFoundError, json.JSONDecodeError):
                # Fallback to hardcoded prompt
                system_prompt = """
                You are AI2, a financial data validation specialist. Your task is to double-check each response 
                by key to ensure figures match the balance sheet and verify data accuracy including K/M conversions.
                
                CRITICAL REQUIREMENTS:
                1. Extract all financial figures from AI1 content
                2. Compare with expected balance sheet figures for accuracy
                3. Verify proper K/M conversion with 1 decimal place (e.g., 2.3M, 1.5K, 123.0)
                4. Check entity names match data source (not reporting entity)
                5. Identify ONLY top 2 most critical data accuracy issues
                6. Remove unnecessary quotation marks around sections
                7. Ensure no data inconsistencies or conversion errors
                8. Verify figures are properly adjusted for '000 notation if applicable
                """
            
            user_query = f"""
            AI2 DATA VALIDATION TASK:
            
            CONTENT TO VALIDATE: {content}
            
            EXPECTED FIGURE FOR {key}: {expected_figure}
            
            COMPLETE BALANCE SHEET DATA: {json.dumps(financial_figures, indent=2)}
            
            ENTITY: {entity}
            
            DETAILED VALIDATION CHECKLIST:
            1. Extract all financial figures from AI1 content and list them
            2. Compare each extracted figure with expected balance sheet figure for {key}: {expected_figure}
            3. Verify proper K/M conversion accuracy (should be 1 decimal place: 2.3M, 1.5K, 123.0)
            4. Check entity names match data source (should NOT be reporting entity "{entity}")
            5. Verify mathematical accuracy of any calculations or breakdowns
            6. Check for proper currency formatting and consistency
            7. Ensure dates are accurate (should be 30/09/2022 or Sep-22 format)
            8. Validate that component figures sum to expected total where applicable
            9. Remove unnecessary quotation marks around full sections
            10. Check for template artifacts that shouldn't be in final content
            
            SPECIFIC FIGURE ANALYSIS FOR {key}:
            - Expected total: {expected_figure}
            - Check if content figures match or reasonably approximate this total
            - Verify thousand/million notation is correct
            - Ensure component breakdowns add up properly
            
            CRITICAL: You MUST respond with ONLY valid JSON in this exact format:
            {{
                "is_valid": true,
                "issues": ["issue 1", "issue 2"],
                "score": 95,
                "corrected_content": "corrected content here if needed",
                "extracted_figures": ["figure1", "figure2"],
                "figure_validation": "detailed analysis of figure accuracy"
            }}
            
            Do not include any text before or after the JSON. Only return the JSON object.
            """
            
            response = generate_response(user_query, system_prompt, oai_client, content, config_details['DEEPSEEK_CHAT_MODEL'], entity, False)
            
            # Clean response and ensure it's valid JSON
            response = response.strip()
            
            # Remove any markdown formatting if present
            if response.startswith('```json'):
                response = response.replace('```json', '').replace('```', '').strip()
            elif response.startswith('```'):
                response = response.replace('```', '').strip()
            
            # Parse AI response
            try:
                result = json.loads(response)
                # Ensure all required fields are present with defaults
                result.setdefault('is_valid', True)
                result.setdefault('issues', [])
                result.setdefault('score', 100)
                result.setdefault('corrected_content', content)
                result.setdefault('needs_correction', False)
                result.setdefault('suggestions', [])
                return result
            except (json.JSONDecodeError, ValueError) as parse_error:
                print(f"Failed to parse AI response: {parse_error}")
                print(f"Raw AI response: {response}")
                # Return structured fallback result
                return {
                    "is_valid": True,
                    "issues": [f"AI response parsing failed: {str(parse_error)}"],
                    "score": 75,
                    "corrected_content": content,
                    "needs_correction": False,
                    "suggestions": ["Check AI response format"]
                }
                
        except Exception as e:
            print(f"Data validation error: {e}")
            return {"needs_correction": False, "issues": [f"Validation error: {e}"], "score": 50, "suggestions": []}
    
# Removed _fallback_data_validation function - offline mode eliminated
    
    def correct_financial_data(self, content: str, issues: List[str]) -> str:
        """Correct financial data issues using AI"""
        try:
            if not AI_AVAILABLE:
                raise RuntimeError("AI services are required for data correction. Please check your configuration.")
            
            config_details = load_config(self.config_file)
            oai_client, _ = initialize_ai_services(config_details)
            
            system_prompt = """
            You are a financial data correction specialist. Your task is to fix financial data 
            accuracy issues in the content while maintaining the original structure and tone.
            
            REQUIREMENTS:
            1. Fix all identified data accuracy issues
            2. Ensure figures match financial statements exactly
            3. Maintain proper formatting (K/M notation)
            4. Keep the original writing style and structure
            5. Only correct the identified issues, don't rewrite unnecessarily
            """
            
            user_query = f"""
            CORRECT FINANCIAL DATA ISSUES:
            
            ORIGINAL CONTENT: {content}
            IDENTIFIED ISSUES: {issues}
            
            TASK: Fix the identified issues while maintaining the original content structure.
            
            REQUIREMENTS:
            - Fix all data accuracy issues
            - Ensure proper figure formatting
            - Maintain original writing style
            - Keep the same paragraph structure
            - Only correct what needs fixing
            
            RETURN: Only the corrected content text, no explanations or JSON.
            """
            
            corrected_content = generate_response(user_query, system_prompt, oai_client, "", config_details['DEEPSEEK_CHAT_MODEL'], entity, False)
            return corrected_content.strip()
            
        except Exception as e:
            print(f"Data correction error: {e}")
            return content

# --- Pattern Validation Agent ---
class PatternValidationAgent:
    def __init__(self, use_local_ai=False, use_openai=False):
        self.config_file = 'fdd_utils/config.json'
        self.pattern_file = 'fdd_utils/pattern.json'
        self.use_local_ai = use_local_ai
        self.use_openai = use_openai
    
    def validate_pattern_compliance(self, content: str, key: str) -> Dict:
        """Validate that content follows the expected pattern structure"""
        try:
            import json
            
            # Load patterns for the key
            patterns = load_ip(self.pattern_file, key)
            
            if not AI_AVAILABLE:
                raise RuntimeError("AI services are required for pattern validation. Please check your configuration.")
            
            # Use AI to validate pattern compliance
            config_details = load_config(self.config_file)
            oai_client, _ = initialize_ai_services(config_details, use_local=self.use_local_ai, use_openai=self.use_openai)
            
            # Load system prompt from prompts.json
            try:
                with open('fdd_utils/prompts.json', 'r', encoding='utf-8') as f:
                    prompts_config = json.load(f)
                system_prompt = prompts_config.get('system_prompts', {}).get('Agent 3', '')
                
                if not system_prompt:
                    # Fallback to hardcoded prompt
                    system_prompt = """
                    You are AI3, a pattern compliance validation specialist. Your task is to check if content 
                    follows patterns correspondingly and clean up excessive items.
                    
                    CRITICAL REQUIREMENTS:
                    1. Compare AI1 content against available pattern templates
                    2. Check proper pattern structure and professional formatting
                    3. Verify all placeholders are filled with actual data
                    4. If AI1 lists too many items, limit to top 2 most important
                    5. Remove quotation marks quoting full sections
                    6. Check for anything that shouldn't be there (template artifacts)
                    7. Ensure content follows pattern structure consistently
                    8. Verify proper K/M conversion with 1 decimal place formatting
                    """
            except (FileNotFoundError, json.JSONDecodeError):
                # Fallback to hardcoded prompt
                system_prompt = """
                You are AI3, a pattern compliance validation specialist. Your task is to check if content 
                follows patterns correspondingly and clean up excessive items.
                
                CRITICAL REQUIREMENTS:
                1. Compare AI1 content against available pattern templates
                2. Check proper pattern structure and professional formatting
                3. Verify all placeholders are filled with actual data
                4. If AI1 lists too many items, limit to top 2 most important
                5. Remove quotation marks quoting full sections
                6. Check for anything that shouldn't be there (template artifacts)
                7. Ensure content follows pattern structure consistently
                8. Verify proper K/M conversion with 1 decimal place formatting
                """
            
            user_query = f"""
            AI3 PATTERN COMPLIANCE CHECK:
            
            CONTENT TO ANALYZE: {content}
            
            KEY: {key}
            
            AVAILABLE PATTERNS FOR {key}: {json.dumps(patterns, indent=2)}
            
            PATTERN COMPLIANCE VALIDATION TASKS:
            1. Analyze AI1 content structure against available pattern templates above
            2. Verify all placeholders (xxx, [Amount], [Entity], etc.) are filled with actual data
            3. Check if content follows the narrative flow of selected pattern
            4. If AI1 content lists too many items/entities, limit to top 2 most important
            5. Remove quotation marks around full sections or paragraphs  
            6. Check for template artifacts that shouldn't appear in final content
            7. Ensure professional financial writing style and tone
            8. Verify pattern structure is maintained (intro → details → conclusion format)
            9. Check that selected pattern matches the type of data provided
            10. Ensure consistent tense and formatting throughout
            
            PATTERN ANALYSIS FOR {key}:
            - Which pattern from above appears to be most suitable?
            - Are all pattern elements properly filled?
            - Does the content maintain professional audit report language?
            - Are there any deviations from expected pattern structure?
            
            CONTENT OPTIMIZATION REQUIREMENTS:
            - Keep only essential information (top 2 items if listing multiple)
            - Remove redundant or verbose explanations
            - Ensure dates, amounts, and entities are accurate
            - Maintain consistent formatting with other sections
            
            CRITICAL: You MUST respond with ONLY valid JSON in this exact format:
            {{
                "is_compliant": true,
                "issues": ["issue 1", "issue 2"],
                "corrected_content": "cleaned content with top 2 items if needed",
                "pattern_used": "Pattern 1 or Pattern 2",
                "compliance_analysis": "detailed analysis of pattern compliance"
            }}
            
            Do not include any text before or after the JSON. Only return the JSON object.
            """
            
            # Use appropriate model based on selection
            if self.use_local_ai:
                model = config_details.get('LOCAL_AI_CHAT_MODEL', 'local-model')
            elif self.use_openai:
                model = config_details.get('OPENAI_CHAT_MODEL', 'gpt-4o-mini-2024-07-18')
            else:
                model = config_details.get('DEEPSEEK_CHAT_MODEL', 'deepseek-chat')
            
            response = generate_response(user_query, system_prompt, oai_client, content, model, entity, self.use_local_ai)
            
            # Clean response and ensure it's valid JSON
            response = response.strip()
            
            # Remove any markdown formatting if present
            if response.startswith('```json'):
                response = response.replace('```json', '').replace('```', '').strip()
            elif response.startswith('```'):
                response = response.replace('```', '').strip()
            
            # Parse AI response
            try:
                result = json.loads(response)
                # Ensure all required fields are present with defaults
                result.setdefault('is_compliant', True)
                result.setdefault('issues', [])
                result.setdefault('corrected_content', content)
                result.setdefault('needs_correction', False)
                result.setdefault('score', 100)
                result.setdefault('pattern_match', 'compliant')
                result.setdefault('missing_elements', [])
                result.setdefault('suggestions', [])
                return result
            except (json.JSONDecodeError, ValueError) as parse_error:
                print(f"Failed to parse AI3 response: {parse_error}")
                print(f"Raw AI3 response: {response}")
                # Return structured fallback result
                return {
                    "is_compliant": True,
                    "issues": [f"AI response parsing failed: {str(parse_error)}"],
                    "corrected_content": content,
                    "needs_correction": False,
                    "score": 75,
                    "pattern_match": "unknown",
                    "missing_elements": [],
                    "suggestions": ["Check AI response format"]
                }
                
        except Exception as e:
            print(f"Pattern validation error: {e}")
            return {"needs_correction": False, "issues": [f"Validation error: {e}"], "score": 50, "suggestions": []}
    
# Removed _fallback_pattern_validation function - offline mode eliminated
    
    def correct_pattern_compliance(self, content: str, issues: List[str]) -> str:
        """Correct pattern compliance issues using AI"""
        try:
            if not AI_AVAILABLE:
                raise RuntimeError("AI services are required for pattern correction. Please check your configuration.")
            
            config_details = load_config(self.config_file)
            oai_client, _ = initialize_ai_services(config_details)
            
            system_prompt = """
            You are a pattern compliance correction specialist. Your task is to fix pattern 
            compliance issues in the content while maintaining accuracy and professionalism.
            
            REQUIREMENTS:
            1. Fix all identified pattern compliance issues
            2. Ensure content follows expected pattern structure
            3. Fill any missing placeholders appropriately
            4. Maintain professional financial language
            5. Keep the original meaning and accuracy
            """
            
            user_query = f"""
            CORRECT PATTERN COMPLIANCE ISSUES:
            
            ORIGINAL CONTENT: {content}
            IDENTIFIED ISSUES: {issues}
            
            TASK: Fix the identified pattern compliance issues while maintaining content accuracy.
            
            REQUIREMENTS:
            - Fix all pattern compliance issues
            - Ensure proper pattern structure
            - Fill missing placeholders appropriately
            - Maintain professional language
            - Keep original meaning intact
            
            RETURN: Only the corrected content text, no explanations or JSON.
            """
            
            corrected_content = generate_response(user_query, system_prompt, oai_client, "", config_details['DEEPSEEK_CHAT_MODEL'], entity, False)
            return corrected_content.strip()
            
        except Exception as e:
            print(f"Pattern correction error: {e}")
            return content

class ProofreadingAgent:
    """AI Proofreader for compliance, figure formatting, entity matching, and grammar/style."""
    def __init__(self, use_local_ai: bool = False, use_openai: bool = False, language: str = 'English'):
        self.config_file = 'fdd_utils/config.json'
        self.pattern_file = 'fdd_utils/pattern.json'
        self.use_local_ai = use_local_ai
        self.use_openai = use_openai
        self.language = language

    def proofread(self, content: str, key: str, tables_markdown: str, entity: str, progress_bar: Optional[tqdm] = None) -> Dict:
        try:
            import json
            import re

            # Debug: Proofreading start (remove after testing)
            # print(f"🔍 DEBUG: Proofreading {key} with language: {self.language}")

            # Update progress bar if provided
            if progress_bar:
                progress_bar.set_description(f"Proofreading {key}")
                progress_bar.update(1)
            
            # Load patterns for the key
            patterns = load_ip(self.pattern_file, key)

            if not AI_AVAILABLE:
                raise RuntimeError("AI services are required for proofreading. Please check your configuration.")

            # Initialize AI
            config_details = load_config(self.config_file)
            oai_client, _ = initialize_ai_services(config_details, use_local=self.use_local_ai, use_openai=self.use_openai)

            # Load system prompt based on language
            try:
                with open('fdd_utils/prompts.json', 'r', encoding='utf-8') as f:
                    prompts_config = json.load(f)

                # Convert language to the format used in prompts.json
                language_key = 'chinese' if self.language == '中文' else 'english'

                # Get the appropriate prompt for the language
                system_prompt = prompts_config.get('system_prompts', {}).get(language_key, {}).get('AI Proofreader', '')

                # Debug: Log which prompt was loaded (remove after testing)
                # print(f"🔍 DEBUG: Loaded system prompt for language '{language_key}': {system_prompt[:100]}...")

                # Fallback if language-specific prompt not found
                if not system_prompt:
                    system_prompt = prompts_config.get('system_prompts', {}).get('AI Proofreader', '')
                    # print(f"⚠️ DEBUG: Using fallback prompt for language '{language_key}'")

            except (FileNotFoundError, json.JSONDecodeError):
                # Fallback to language-appropriate default prompt
                if self.language == '中文':
                    system_prompt = (
                        "您是财务尽职调查叙述的人工智能校对员。审查Agent 1内容是否符合：模式要求、数据格式化（K/M 1位小数；处理'000）、按表格的实体/细节正确性（不是报告实体）、语法/专业语气（移除外引号）、语言规范化（确保使用简体中文；使用VAT、CIT、WHT、附加税）。规则：不要发明数据；保持简洁；如果列表太长，保留前2项；确保所有内容都是简体中文。返回JSON格式：is_compliant (bool), issues (array), corrected_content (string), figure_checks (array), entity_checks (array), grammar_notes (array), pattern_used (string)。corrected_content必须是最终清理后的简体中文文本。"
                    )
                else:
                    system_prompt = (
                        "You are an AI proofreader for financial due diligence narratives. Focus on pattern compliance, "
                        "K/M figure formatting, entity correctness, grammar/pro tone, and language normalization (keep original language). Return JSON."
                    )

            # Truncate content and tables to fit within context limits
            max_content_length = 8000  # Leave room for other parts
            max_tables_length = 4000

            truncated_content = content[:max_content_length] + ("..." if len(content) > max_content_length else "")
            truncated_tables = tables_markdown[:max_tables_length] + ("..." if len(tables_markdown) > max_tables_length else "")

            # Build user query with truncated content (language-aware)
            if language_key == 'chinese':
                user_query = f"""
                AI PROOFREADING TASK (中文内容)

                KEY: {key}
                REPORTING ENTITY: {entity}

                CONTENT TO REVIEW:
                {truncated_content}

                DATA TABLES (for entity/details and figure source):
                {truncated_tables}

                TASK: Analyze the content for compliance, figure formatting, entity correctness, and grammar.
                IMPORTANT: Keep ALL content in 简体中文 (Simplified Chinese). Do NOT translate to English.
                Return ONLY a JSON object with these exact fields:
                {{
                  "is_compliant": true,
                  "issues": ["list of issues found"],
                  "corrected_content": "the corrected content text in Chinese",
                  "figure_checks": ["figure validation notes"],
                  "entity_checks": ["entity validation notes"],
                  "grammar_notes": ["grammar and style notes"],
                  "pattern_used": "Pattern X"
                }}

                IMPORTANT: Ensure the JSON is valid and properly formatted. Keep corrected_content in Chinese.
                """
            else:
                user_query = f"""
                AI PROOFREADING TASK

                KEY: {key}
                REPORTING ENTITY: {entity}

                CONTENT TO REVIEW:
                {truncated_content}

                DATA TABLES (for entity/details and figure source):
                {truncated_tables}

                TASK: Analyze the content for compliance, figure formatting, entity correctness, and grammar.
                Return ONLY a JSON object with these exact fields:
                {{
                  "is_compliant": true,
                  "issues": ["list of issues found"],
                  "corrected_content": "the corrected content text",
                  "figure_checks": ["figure validation notes"],
                  "entity_checks": ["entity validation notes"],
                  "grammar_notes": ["grammar and style notes"],
                  "pattern_used": "Pattern X"
                }}

                IMPORTANT: Ensure the JSON is valid and properly formatted.
                """

            # Model selection
            if self.use_local_ai:
                model = config_details.get('LOCAL_AI_CHAT_MODEL', 'local-model')
            elif self.use_openai:
                model = config_details.get('OPENAI_CHAT_MODEL', 'gpt-4o-mini-2024-07-18')
            else:
                model = config_details.get('DEEPSEEK_CHAT_MODEL', 'deepseek-chat')

            response = generate_response(user_query, system_prompt, oai_client, tables_markdown, model, entity, self.use_local_ai)
            response = response.strip()
            if response.startswith('```json'):
                response = response.replace('```json', '').replace('```', '').strip()
            elif response.startswith('```'):
                response = response.replace('```', '').strip()

            try:
                result = json.loads(response)
                # Fill defaults
                result.setdefault('is_compliant', True)
                result.setdefault('issues', [])
                result.setdefault('corrected_content', content)
                result.setdefault('figure_checks', [])
                result.setdefault('entity_checks', [])
                result.setdefault('grammar_notes', [])
                result.setdefault('pattern_used', '')
                result.setdefault('translation_runs', 0)

                # Only apply heuristic translation for English language (not for Chinese)
                def contains_cjk(txt: str) -> bool:
                    try:
                        return bool(re.search(r"[\u4e00-\u9fff]", txt or ''))
                    except Exception:
                        return False

                corrected = result.get('corrected_content') or content
                runs = 0  # Initialize runs counter

                # Only translate Chinese to English if we're using English language
                if language_key == 'english':
                    # Heuristic translation loop for English language
                    while contains_cjk(corrected) and runs < 2:
                        trans_system = (
                            "You are a professional financial translator. Translate ALL non-English text to clear business English. "
                            "Keep numbers/currency intact, use standard tax abbreviations (VAT, CIT, WHT, LUT), remove pinyin, "
                            "no brackets or explanations, output final English text only."
                        )
                        trans_user = f"Translate to English (final text only):\n{corrected}"
                        trans = generate_response(trans_user, trans_system, oai_client, tables_markdown, model, entity, self.use_local_ai)
                        corrected = clean_response_text(trans, 'english')  # Explicitly pass english for translation
                        runs += 1
                        if not contains_cjk(corrected):
                            break

                if runs > 0:
                    result['corrected_content'] = corrected
                    result['translation_runs'] = runs

                # Post-process: Replace any remaining xxx placeholders
                corrected = result.get('corrected_content') or content
                if corrected:
                    # Replace common xxx patterns with appropriate content
                    corrected = corrected.replace('xxx', 'the relevant amount')
                    corrected = corrected.replace('XXX', 'the relevant amount')
                    corrected = corrected.replace('xxx', 'applicable')
                    corrected = corrected.replace('XXX', 'applicable')
                    result['corrected_content'] = corrected

                # Debug: Log final result (remove after testing)
                # print(f"🔍 DEBUG: Proofreading completed for {key}. Translation runs: {runs}")
                # print(f"🔍 DEBUG: Final corrected content (first 200 chars): {result.get('corrected_content', '')[:200]}...")

                return result
            except Exception as parse_error:
                print(f"Failed to parse AI Proofreader response: {parse_error}")
                print(f"Raw response (first 500 chars): {response[:500]}")

                # Try to extract content from malformed response
                corrected_content = content
                if '"corrected_content"' in response:
                    try:
                        # Try to extract corrected_content from malformed JSON
                        import re
                        content_match = re.search(r'"corrected_content"\s*:\s*"([^"]*(?:\\.[^"]*)*)"', response, re.DOTALL)
                        if content_match:
                            corrected_content = content_match.group(1).replace('\\n', '\n').replace('\\t', '\t').replace('\\"', '"')
                    except Exception:
                        pass

                return {
                    'is_compliant': False,
                    'issues': [f"AI response parsing failed: {str(parse_error)}", "Used original content due to parsing error"],
                    'corrected_content': corrected_content,
                    'figure_checks': ["Unable to validate - parsing error"],
                    'entity_checks': ["Unable to validate - parsing error"],
                    'grammar_notes': ["Unable to validate - parsing error"],
                    'pattern_used': 'Unknown',
                    'translation_runs': 0
                }
        except Exception as e:
            print(f"Proofreading error: {e}")
            return {
                'is_compliant': False,
                'issues': [f"Proofreading error: {e}"],
                'corrected_content': content,
                'figure_checks': [],
                'entity_checks': [],
                'grammar_notes': [],
                'pattern_used': '',
                'translation_runs': 0
            }

def clean_response_text(text: str, language: str = 'english') -> str:
    """Clean up AI response text: remove outer quotes, translate Chinese, etc."""
    if not text:
        return text
    
    # Remove outer quotation marks
    text = text.strip()
    if (text.startswith('"') and text.endswith('"')) or (text.startswith("'") and text.endswith("'")):
        text = text[1:-1]
    
    # Only apply Chinese translation logic if we're using English prompts
    # If language is Chinese, we expect Chinese output and should NOT translate it
    if language == 'english':
        # Translate Chinese to English (extended common terms)
        chinese_translations = {
        '進項稅金中轉': 'Input Tax Transfer',
        '自來水有限公司': 'Water Supply Co., Ltd.',
        '仲量聯行北京諮詢公司': 'Jones Lang LaSalle Beijing Consulting Co.',
        '信用中和會計師事務所': 'Credit Zhonghe Accounting Firm',
        '企業管理諮詢公司': 'Business Management Consulting Co.',
        '科技有限公司': 'Technology Co., Ltd.',
        '物業服務有限公司': 'Property Services Co., Ltd.',
        '物流發展有限公司': 'Logistics Development Co., Ltd.',
        '股權投資基金合夥企業': 'Equity Investment Fund Partnership',
        '物流倉儲有限公司': 'Logistics Storage Co., Ltd.',
        '特殊普通合夥': 'Special General Partnership',
        '戰區運營': 'Regional Operations',
        '已認證未入帳': 'Certified but Not Recorded',
        '有限合夥': 'Limited Partnership',
        '增值稅': 'Value-Added Tax (VAT)',
        '企業所得稅': 'Corporate Income Tax (CIT)',
        '個人所得稅': 'Individual Income Tax (IIT)',
        '附加稅費': 'Surtaxes',
        '房產稅': 'Property Tax',
        '土地使用稅': 'LUT',
        '印花稅': 'Stamp Tax'
    }
        for chinese, english in chinese_translations.items():
            text = text.replace(chinese, english)

        # Fallback: detect any residual CJK characters and annotate/remedy
        try:
            import re
            if re.search(r"[\u4e00-\u9fff]", text):
                # As a minimal remediation, wrap unknown Chinese segments in brackets with a note
                # while keeping the rest English; prevents raw Chinese leaking to final output
                text = re.sub(r"([\u4e00-\u9fff]+)", r"[Translate: \1]", text)
        except Exception:
            pass

    # For Chinese language, we expect Chinese output, so don't apply any translation logic
    # Just return the cleaned text as-is
    
    return text.strip()

def multiply_figures_for_ai_processing(excel_content: str) -> str:
    """
    Multiply all numerical figures by 1000 in Excel content for AI processing when '000 notation is detected.
    This function processes the markdown table content to adjust figures for AI analysis.
    """
    import re
    
    # Check for both English and Chinese thousands notation
    has_chinese_thousands = (
        "千元" in excel_content or
        "千人民币" in excel_content or
        "人民币千" in excel_content or
        "千元人民币" in excel_content or
        "人民币千元" in excel_content
    )

    if "'000" not in excel_content and not has_chinese_thousands:
        return excel_content
    
    lines = excel_content.split('\n')
    processed_lines = []
    
    for line in lines:
        # Skip header lines and separator lines
        if '|' not in line or line.strip().startswith('|---') or 'Description' in line:
            processed_lines.append(line)
            continue
        
        # Process table rows with numerical data
        cells = line.split('|')
        processed_cells = []
        
        for cell in cells:
            cell = cell.strip()
            
            # Look for numerical patterns and multiply by 1000
            # Match various number formats: 123, 1,234, 1.23, (123), etc.
            number_pattern = r'(\(?)(-?\d{1,3}(?:,\d{3})*\.?\d*)(\)?)'
            
            def multiply_number(match):
                opening_paren = match.group(1)
                number_str = match.group(2)
                closing_paren = match.group(3)
                
                try:
                    # Remove commas and convert to float
                    clean_number = number_str.replace(',', '')
                    number = float(clean_number)
                    
                    # Multiply by 1000
                    adjusted_number = number * 1000
                    
                    # Format back with commas for large numbers
                    if adjusted_number == int(adjusted_number):
                        formatted = f"{int(adjusted_number):,}"
                    else:
                        formatted = f"{adjusted_number:,.1f}"
                    
                    return f"{opening_paren}{formatted}{closing_paren}"
                except ValueError:
                    # If conversion fails, return original
                    return match.group(0)
            
            # Apply multiplication to numbers in the cell
            processed_cell = re.sub(number_pattern, multiply_number, cell)
            processed_cells.append(processed_cell)
        
        processed_lines.append('|'.join(processed_cells))
    
    return '\n'.join(processed_lines) 