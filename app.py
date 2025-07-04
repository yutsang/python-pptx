import streamlit as st
import pandas as pd
from common import assistant
import tempfile
import os
import json
import re
from utils.utils import name_mapping
import sys
from common.pptx_export import export_pptx
import datetime
import numpy as np
# from pptx_export import ReportGenerator  # Uncomment after renaming 2.pptx.py
sys.path.append('.')

st.set_page_config(page_title="Due Diligence Report Generator", layout="wide")
st.title("Real Estate Due Diligence Report Generator")

# --- Upload Excel File ---
st.sidebar.header("Upload Excel Databook")
uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx"]) 

# --- Entity and Helpers ---
st.sidebar.header("Entity & Helpers")
entity = st.sidebar.selectbox("Select Entity", ["Haining", "Nanjing", "Ningbo"])

# Predefined entity suffixes for better UX
entity_suffixes = {
    "Haining": [" Wanpu", " Wanpu Limited", " Limited", " Co.", " Ltd", " Corp"],
    "Nanjing": [" Wanpu", " Wanpu Limited", " Limited", " Co.", " Ltd", " Corp"],
    "Ningbo": [" Wanpu", " Wanpu Limited", " Limited", " Co.", " Ltd", " Corp"]
}

# Use predefined suffixes or allow custom input
use_custom_helpers = st.sidebar.checkbox("Use custom entity helpers", value=False)
if use_custom_helpers:
    helpers_input = st.sidebar.text_input("Entity Helpers (comma separated)", value="Wanpu, Limited, ...", placeholder="Wanpu, Limited, ...")
    helpers = [h.strip() for h in helpers_input.split(",") if h.strip()]
else:
    helpers = entity_suffixes.get(entity, ["Wanpu", "Limited", "Co.", "Ltd", "Corp"])
    st.sidebar.caption(f"Using default helpers: {', '.join(helpers)}")

# --- AI or Local Mode ---
st.sidebar.header("Generation Mode")
use_ai = st.sidebar.radio("Use AI for text generation?", ("No (Test Mode)", "Yes (AI Mode)")) == "Yes (AI Mode)"

# --- Template Upload ---
st.sidebar.header("Upload PPTX Template")
template_file = st.sidebar.file_uploader("Choose a PPTX template", type=["pptx"])
template_path = None
if template_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pptx') as tmp:
        tmp.write(template_file.read())
        template_path = tmp.name

# --- '000 Scaling Toggle ---
st.sidebar.header("Scaling Option")
convert_thousands = st.sidebar.radio("Convert figures in '000 to base units?", ("Yes", "No")) == "Yes"
st.sidebar.caption("If Yes, convert figures in '000 to base units (multiply by 1,000).")

# --- Sheet Type Selector ---
sheet_type = st.sidebar.selectbox("Sheet Type", ["BS", "IS"], index=0)

# --- Table Filtering Options ---
st.sidebar.header("Table Filtering")
enable_filtering = st.sidebar.checkbox("Enable table filtering by entity", value=True)
show_all_tables = st.sidebar.checkbox("Show all tables (not just matching ones)", value=False)

# --- State ---
if 'results' not in st.session_state:
    st.session_state['results'] = None
if 'keys' not in st.session_state:
    st.session_state['keys'] = []
if 'excel_tables' not in st.session_state:
    st.session_state['excel_tables'] = {}

# --- Output File Name ---
default_output_name = f"{entity}_{sheet_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pptx"
output_file_name = st.sidebar.text_input("Output PPTX File Name", value=default_output_name)

# --- Helper Functions ---
def extract_tables_from_worksheet_robust(excel_path, sheet_name, entity_keywords):
    """
    Robust table extraction for worksheet view that works with both individually formatted tables and upslide smart format tables.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
        
        tables = []
        
        # Method 1: Try to extract from openpyxl tables (works for individually formatted tables)
        if hasattr(ws, '_tables') and ws._tables:
            for tbl in ws._tables.values():
                try:
                    ref = tbl.ref
                    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
                    data = []
                    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                        data.append(row)
                    if data and len(data) >= 2:
                        tables.append({
                            'data': data,
                            'method': 'openpyxl_table',
                            'name': tbl.name,
                            'range': ref
                        })
                except Exception as e:
                    print(f"Failed to extract table {tbl.name}: {e}")
                    continue
        
        # Method 2: Smart table detection for upslide smart format and other table-like structures
        all_data = []
        for row in ws.iter_rows(values_only=True):
            all_data.append(row)
        
        if all_data:
            df = pd.DataFrame(all_data)
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            if len(df) >= 2:
                # Look for table patterns in the data
                table_regions = find_table_regions_smart(df, entity_keywords)
                
                for region in table_regions:
                    row_start, row_end, col_start, col_end = region
                    table_data = df.iloc[row_start:row_end+1, col_start:col_end+1]
                    
                    # Convert to list format
                    table_list = [table_data.columns.tolist()] + table_data.values.tolist()
                    
                    tables.append({
                        'data': table_list,
                        'method': 'smart_detection',
                        'name': f'smart_table_{row_start}_{col_start}',
                        'range': f'{row_start}:{row_end},{col_start}:{col_end}'
                    })
        
        return tables
        
    except Exception as e:
        print(f"Error in robust table extraction for worksheet view: {e}")
        return []

def find_table_regions_smart(df, entity_keywords):
    """
    Smart table detection that works with upslide smart format tables.
    Looks for patterns that indicate table structures.
    """
    table_regions = []
    nrows, ncols = df.shape
    
    # Look for table headers and data patterns
    for i in range(nrows - 2):  # Need at least 3 rows for a table
        # Check if current row could be a table header
        current_row = df.iloc[i]
        next_row = df.iloc[i + 1]
        next_next_row = df.iloc[i + 2] if i + 2 < nrows else None
        
        # Criteria for table header:
        # 1. Contains entity keywords
        # 2. Has mostly text content
        # 3. Followed by rows with data
        
        # Check for entity keywords in current row
        row_text = ' '.join([str(cell).lower() for cell in current_row if pd.notna(cell)])
        has_entity = any(kw.lower() in row_text for kw in entity_keywords)
        
        if has_entity:
            # Count text vs numeric content in current and next rows
            current_text_count = sum(1 for cell in current_row if isinstance(cell, str) and str(cell).strip())
            next_numeric_count = sum(1 for cell in next_row if is_numeric_like(cell))
            
            # If current row has mostly text and next row has some numbers, this might be a table
            if current_text_count >= len(current_row) * 0.3 and next_numeric_count >= 1:
                # Find the table boundaries
                table_start = i
                table_end = i
                
                # Look for where the table ends
                for j in range(i + 1, nrows):
                    row = df.iloc[j]
                    
                    # Check if this row continues the table pattern
                    if is_table_data_row(row):
                        table_end = j
                    else:
                        # Check if it's an empty row or different pattern
                        if is_empty_row(row) or not is_table_continuation(row, df.iloc[j-1] if j > 0 else None):
                            break
                        else:
                            table_end = j
                
                # Ensure we have a reasonable table size
                if table_end > table_start and (table_end - table_start + 1) >= 2:
                    # Find column boundaries
                    col_start, col_end = find_column_boundaries(df.iloc[table_start:table_end+1])
                    
                    # Only add if we have a reasonable table size
                    if col_end > col_start and (col_end - col_start + 1) >= 2:
                        table_regions.append((table_start, table_end, col_start, col_end))
    
    return table_regions

def is_numeric_like(cell):
    """Check if a cell contains numeric-like data."""
    if pd.isna(cell):
        return False
    if isinstance(cell, (int, float)):
        return True
    if isinstance(cell, str):
        # Remove common formatting characters
        cleaned = str(cell).replace(',', '').replace('.', '').replace('-', '').replace('(', '').replace(')', '')
        return cleaned.replace(' ', '').isdigit()
    return False

def is_table_data_row(row):
    """Check if a row looks like table data."""
    if len(row) == 0:
        return False
    
    # Count non-empty cells
    non_empty = sum(1 for cell in row if pd.notna(cell) and str(cell).strip())
    
    # Count numeric cells
    numeric = sum(1 for cell in row if is_numeric_like(cell))
    
    # Row should have some content and preferably some numbers
    return non_empty >= len(row) * 0.2 and numeric >= 1

def is_empty_row(row):
    """Check if a row is essentially empty."""
    return all(pd.isna(cell) or str(cell).strip() == '' for cell in row)

def is_table_continuation(row, prev_row):
    """Check if this row continues a table pattern."""
    if prev_row is None:
        return False
    
    # Check if the structure is similar to previous row
    current_structure = [is_numeric_like(cell) for cell in row]
    prev_structure = [is_numeric_like(cell) for cell in prev_row]
    
    # If structures are similar, it might be table continuation
    return sum(current_structure) > 0 and len(set(current_structure) - set(prev_structure)) <= 1

def find_column_boundaries(table_df):
    """Find the column boundaries for a table."""
    if table_df.empty:
        return 0, 0
    
    # Find columns with actual content
    non_empty_cols = []
    for col_idx in range(table_df.shape[1]):
        col = table_df.iloc[:, col_idx]
        if col.notna().sum() > 0:
            non_empty_cols.append(col_idx)
    
    if not non_empty_cols:
        return 0, 0
    
    return min(non_empty_cols), max(non_empty_cols)

def save_results_to_markdown(results, entity: str):
    """Save AI-generated results back to the markdown file"""
    try:
        # Read existing markdown content
        with open('utils/bs_content.md', 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Update each section with new content
        for key, new_content in results.items():
            # Get the heading name from name_mapping
            heading = name_mapping.get(key, key)
            if heading is None:
                heading = key
            
            # Create the new section content
            new_section = f'### {heading}\n{new_content.strip()}\n\n'
            
            # Find and replace the existing section
            pattern = re.compile(r'(###\s+' + re.escape(heading) + r'.*?)(?=\n###|\Z)', re.DOTALL | re.IGNORECASE)
            
            if pattern.search(content):
                # Replace existing section
                content = pattern.sub(new_section, content)
            else:
                # Add new section at the end
                content += '\n' + new_section
        
        # Write back to file
        with open('utils/bs_content.md', 'w', encoding='utf-8') as f:
            f.write(content)
            
    except Exception as e:
        print(f"Error saving to markdown: {e}")
        raise

# --- Main Workflow ---
if uploaded_file:
    # Save uploaded file to a temp location
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(uploaded_file.read())
        excel_path = tmp.name
    xl = pd.ExcelFile(excel_path)
    st.success(f"Excel file loaded: {uploaded_file.name}")
    # Load mapping.json for worksheet mapping
    with open('utils/mapping.json', 'r') as f:
        mapping = json.load(f)
    # Define keys (can be made dynamic)
    keys = [
        "Cash", "AR", "Prepayments", "OR", "Other CA", "IP", "Other NCA",
        "AP", "Taxes payable", "OP", "Capital", "Reserve"
    ]
    if entity in ['Ningbo', 'Nanjing']:
        keys = [key for key in keys if key != "Reserve"]
    st.session_state['keys'] = keys
    # Build a set of all possible worksheet names from mapping.json for the selected keys
    mapped_sheet_names = set()
    for key in keys:
        mapped_sheet_names.update(mapping.get(key, []))
    valid_sheets = [sheet for sheet in xl.sheet_names if sheet in mapped_sheet_names or sheet in name_mapping.values()]
    # --- View Worksheets Tabs (independent) ---
    tabs = st.tabs(valid_sheets)
    for i, tab in enumerate(tabs):
        with tab:
            sheet = valid_sheets[i]
            # Map worksheet name to key
            selected_key = None
            for key in keys:
                mapped_names = mapping.get(key, [])
                if sheet in mapped_names or sheet == name_mapping.get(key, key):
                    selected_key = key
                    break
            if selected_key is not None:
                # Use robust table extraction for better compatibility with different table formats
                # Create entity keywords that match the actual table content
                entity_keywords = [entity] + [f"{entity}{suffix}" for suffix in helpers]
                tables = extract_tables_from_worksheet_robust(excel_path, sheet, entity_keywords)
                
                if not tables:
                    st.info("No tables found in this sheet.")
                else:
                    st.success(f"Found {len(tables)} table(s) using robust detection")
                    
                    if enable_filtering:
                        matching_tables = []
                        for i, table_info in enumerate(tables):
                            try:
                                data = table_info['data']
                                method = table_info['method']
                                table_name = table_info['name']
                                
                                if not data or len(data) < 2:
                                    continue
                                
                                # Create DataFrame
                                df = pd.DataFrame(data[1:], columns=data[0])
                                df = df.dropna(how='all').dropna(axis=1, how='all')
                                
                                # Convert all columns to string to avoid ArrowTypeError with mixed data types
                                for col in df.columns:
                                    try:
                                        if df[col].dtype == 'object':
                                            df[col] = df[col].apply(lambda x: str(x) if pd.notna(x) else '')
                                        elif pd.api.types.is_datetime64_any_dtype(df[col]):
                                            df[col] = df[col].astype(str)
                                        elif pd.api.types.is_numeric_dtype(df[col]):
                                            df[col] = df[col].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else '')
                                        else:
                                            df[col] = df[col].astype(str)
                                    except Exception as e:
                                        df[col] = df[col].apply(lambda x: str(x) if pd.notna(x) else '')
                                
                                # Drop Unnamed columns - handle mixed data types safely
                                df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed')]
                                df = df.reset_index(drop=True)
                                
                                # Check for entity keywords - handle mixed data types safely
                                # Include both the original data (which may contain headers) and the processed DataFrame
                                all_cells = [str(cell).lower().strip() for cell in df.values.flatten()]
                                # Also check the original data for headers/titles
                                original_cells = [str(cell).lower().strip() for cell in data[0]] if data and len(data) > 0 else []
                                all_cells.extend(original_cells)
                                
                                # Also check the table name itself for entity keywords
                                table_name_cells = [str(cell).lower().strip() for cell in table_name.split() if cell]
                                all_cells.extend(table_name_cells)
                                
                                # Check ALL rows of the original data for entity keywords
                                for row in data:
                                    row_cells = [str(cell).lower().strip() for cell in row if cell]
                                    all_cells.extend(row_cells)
                                
                                # Check for entity keywords in all collected cells
                                match_found = any(any(kw.lower() in cell for cell in all_cells) for kw in entity_keywords)
                                
                                if match_found:
                                    matching_tables.append((i+1, method, df))
                                        
                            except Exception as e:
                                st.error(f"Error processing table {i+1}: {str(e)}")
                                continue
                        
                        # Only show matching tables
                        if matching_tables:
                            for table_num, method, df in matching_tables:
                                st.write(f"**Table {table_num}** (Detected by: {method})")
                                try:
                                    st.table(df)
                                except Exception as e:
                                    st.text(str(df.to_string()))
                        else:
                            st.info("No tables found matching the selected entity.")
                    else:
                        # Show all tables without filtering
                        for i, table_info in enumerate(tables):
                            try:
                                data = table_info['data']
                                method = table_info['method']
                                table_name = table_info['name']
                                
                                if not data or len(data) < 2:
                                    continue
                                
                                # Create DataFrame
                                df = pd.DataFrame(data[1:], columns=data[0])
                                df = df.dropna(how='all').dropna(axis=1, how='all')
                                
                                # Convert all columns to string
                                for col in df.columns:
                                    df[col] = df[col].astype(str)
                                
                                # Drop Unnamed columns - handle mixed data types safely
                                df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed')]
                                df = df.reset_index(drop=True)
                                
                                st.write(f"**Table {i+1}** (Detected by: {method})")
                                try:
                                    st.table(df)
                                except Exception as e:
                                    st.text(str(df.to_string()))
                                    
                            except Exception as e:
                                st.error(f"Error processing table {i+1}: {str(e)}")
                                continue
                
                # Store the first table for session state (for compatibility)
                if tables:
                    first_table_data = tables[0]['data']
                    if first_table_data and len(first_table_data) >= 2:
                        df_session = pd.DataFrame(first_table_data[1:], columns=first_table_data[0])
                        df_session = df_session.dropna(how='all').dropna(axis=1, how='all')
                        st.session_state['excel_tables'][sheet] = df_session
    # --- AI Generation Button ---
    st.subheader("Generate Report Text")
    if st.button("Generate Text (AI/Test)"):
        if use_ai:
            # Create progress bar for AI processing
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Initialize progress tracking with more granular steps
            total_steps = len(keys) * 4  # 4 steps per key: generation, data validation, pattern validation, final QA
            current_step = 0
            
            try:
                status_text.text("🚀 Initializing AI services and loading data...")
                progress_bar.progress(0)
                
                # Step 1: Generate initial content
                status_text.text("🤖 Generating initial content with AI...")
                results = assistant.process_keys(
                    keys=keys,
                    entity_name=entity,
                    entity_helpers=helpers,
                    input_file=excel_path,
                    mapping_file='utils/mapping.json',
                    pattern_file='utils/pattern.json',
                    config_file='utils/config.json',
                    use_ai=use_ai,
                    convert_thousands=convert_thousands
                )
                
                current_step += len(keys)
                progress_bar.progress(current_step / total_steps)
                
                # Step 2: Data Accuracy Validation Agent
                status_text.text("🔍 Running data accuracy validation...")
                data_validation_agent = assistant.DataValidationAgent()
                validation_issues = []
                for i, key in enumerate(results):
                    status_text.text(f"📊 Validating data accuracy for {key} ({i+1}/{len(keys)})...")
                    try:
                        validation_result = data_validation_agent.validate_financial_data(
                            results[key], 
                            excel_path, 
                            entity, 
                            key
                        )
                        if validation_result['needs_correction']:
                            status_text.text(f"🔧 Correcting data accuracy issues for {key}...")
                            results[key] = data_validation_agent.correct_financial_data(
                                results[key], 
                                validation_result['issues']
                            )
                            validation_issues.append(f"{key}: {len(validation_result['issues'])} issues fixed")
                    except Exception as e:
                        st.warning(f"⚠️ Data validation failed for {key}: {str(e)}")
                    current_step += 1
                    progress_bar.progress(current_step / total_steps)
                
                # Step 3: Pattern Validation Agent
                status_text.text("📋 Running pattern compliance validation...")
                pattern_validation_agent = assistant.PatternValidationAgent()
                pattern_issues = []
                for i, key in enumerate(results):
                    status_text.text(f"📝 Validating pattern compliance for {key} ({i+1}/{len(keys)})...")
                    try:
                        pattern_result = pattern_validation_agent.validate_pattern_compliance(
                            results[key], 
                            key
                        )
                        if pattern_result['needs_correction']:
                            status_text.text(f"🔧 Correcting pattern compliance for {key}...")
                            results[key] = pattern_validation_agent.correct_pattern_compliance(
                                results[key], 
                                pattern_result['issues']
                            )
                            pattern_issues.append(f"{key}: {len(pattern_result['issues'])} issues fixed")
                    except Exception as e:
                        st.warning(f"⚠️ Pattern validation failed for {key}: {str(e)}")
                    current_step += 1
                    progress_bar.progress(current_step / total_steps)
                
                # Step 4: Final QA Review
                status_text.text("✨ Performing final quality assurance review...")
                qa_agent = assistant.QualityAssuranceAgent()
                qa_issues = []
                for i, key in enumerate(results):
                    status_text.text(f"🎯 Final QA review for {key} ({i+1}/{len(keys)})...")
                    try:
                        qa_result = qa_agent.validate_content(results[key])
                        if qa_result['score'] < 90:
                            results[key] = qa_agent.auto_correct(results[key])
                            qa_issues.append(f"{key}: QA score {qa_result['score']} improved")
                    except Exception as e:
                        st.warning(f"⚠️ QA review failed for {key}: {str(e)}")
                    current_step += 1
                    progress_bar.progress(current_step / total_steps)
                
                progress_bar.progress(1.0)
                status_text.text("🎉 AI processing completed successfully!")
                
                # Show summary of issues fixed
                if validation_issues or pattern_issues or qa_issues:
                    with st.expander("📋 Quality Assurance Summary", expanded=True):
                        if validation_issues:
                            st.write("**Data Accuracy Issues Fixed:**")
                            for issue in validation_issues:
                                st.write(f"• {issue}")
                        if pattern_issues:
                            st.write("**Pattern Compliance Issues Fixed:**")
                            for issue in pattern_issues:
                                st.write(f"• {issue}")
                        if qa_issues:
                            st.write("**QA Improvements:**")
                            for issue in qa_issues:
                                st.write(f"• {issue}")
                
                st.success("✅ AI processing completed with comprehensive validation and corrections!")
                
                # Save updated content back to markdown file
                try:
                    status_text.text("💾 Saving updated content to markdown file...")
                    save_results_to_markdown(results, entity)
                    st.success("✅ Content saved to markdown file!")
                except Exception as e:
                    st.warning(f"⚠️ Could not save to markdown file: {str(e)}")
                
            except Exception as e:
                st.error(f"❌ Error during AI processing: {str(e)}")
                status_text.text("Processing failed")
                # Show more detailed error information
                with st.expander("🔍 Error Details"):
                    st.code(str(e))
            finally:
                # Clear progress indicators after a delay
                import time
                time.sleep(3)
                progress_bar.empty()
                status_text.empty()
            
            st.session_state['results'] = results
        else:
            # Offline/test mode: split bs_content.md into sections for each mapped heading
            with open('utils/bs_content.md', 'r') as f:
                content = f.read()
            # Improved split: each section ends at the next heading of the same or higher level
            section_pattern = r'(### .+?)(?=\n### |\n## |\Z)'
            matches = re.findall(section_pattern, content, flags=re.DOTALL)
            section_map = {}
            for match in matches:
                heading_line = match.split('\n', 1)[0][4:].strip().lower()
                body = match.split('\n', 1)[1] if '\n' in match else ''
                section_map[heading_line] = body.strip('\n')
            # Map each key to its section using name_mapping, robust to case/whitespace
            mapped_results = {}
            for key in keys:
                heading = name_mapping.get(key, key).strip().lower()
                # Try exact match, then case-insensitive, then fallback
                section = section_map.get(heading)
                if section is None:
                    for h in section_map:
                        if h.replace(' ', '').lower() == heading.replace(' ', '').lower():
                            section = section_map[h]
                            break
                if not section:
                    section = f"No content found for section '{heading}'."
                mapped_results[key] = section
            st.session_state['results'] = mapped_results
    # --- Edit & Export Tabs (independent, manual mapping) ---
    if st.session_state['results']:
        st.subheader("Edit & Export")
        edit_tabs = st.tabs(valid_sheets)
        for i, tab in enumerate(edit_tabs):
            with tab:
                sheet = valid_sheets[i]
                # Manually map tab name (worksheet section) to the correct key using mapping.json and name_mapping
                mapped_key = None
                for key in keys:
                    mapped_names = mapping.get(key, [])
                    if sheet in mapped_names or sheet == name_mapping.get(key, key):
                        mapped_key = key
                        break
                if mapped_key is not None:
                    heading = name_mapping.get(mapped_key, mapped_key)
                    orig = st.session_state['results'].get(mapped_key, "")
                    edited = st.text_area(f"Edit text for {heading}", value=orig, height=150, key=f"edit_{mapped_key}")
                    if st.button(f"Save {heading}", key=f"save_{mapped_key}"):
                        # Update the section in bs_content.md
                        with open('utils/bs_content.md', 'r') as f:
                            content = f.read()
                        # Replace the section for this heading
                        pattern = re.compile(r'(###\s+' + re.escape(heading) + r'.*?)(?=\n###|\Z)', re.DOTALL | re.IGNORECASE)
                        new_section = f'### {heading}\n' + edited.strip() + '\n'
                        if pattern.search(content):
                            content = pattern.sub(new_section, content)
                        else:
                            content += '\n' + new_section
                        with open('utils/bs_content.md', 'w') as f:
                            f.write(content)
                        st.session_state['results'][mapped_key] = edited
                    # --- Export Button ---
                    if st.button("Export to PPTX", key=f"export_{mapped_key}"):
                        if template_path:
                            output_path = output_file_name
                            export_pptx(template_path, 'utils/bs_content.md', output_path, project_name=entity)
                            st.success(f"Exported to {output_path}")
                            with open(output_path, "rb") as f:
                                st.download_button("Download PPTX", data=f.read(), file_name=output_file_name)
                        else:
                            st.error("Please upload a PPTX template before exporting.")
    # Clean up temp file only after all processing is done
    if not st.session_state.get('results'):
        try:
            os.unlink(excel_path)
        except Exception:
            pass
else:
    st.info("Please upload an Excel databook to begin.") 