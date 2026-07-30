#!/usr/bin/env python3
"""Puts a hand-built '<entity>-量价桥图' tab next to what this tool computes
from that entity's raw AB- tab, item by item, and localises any difference.

Needed because "the bridge is off by 60" cannot be answered from either
side alone. The manual tab and the computed decomposition use the same
algebra, so a gap comes from one of three places, and this separates them:

  1. the tab's own internal check doesn't close (the manual work has a
     residual before any comparison with us);
  2. an INPUT differs -- unit rent, area or days for one phase/period, which
     is what a stale cross-sheet reference or a different averaging window
     produces;
  3. only the FACTOR SPLIT differs while the start and end totals agree --
     the same total attributed differently between price, area and days.

Case 3 matters: it means neither side is wrong about the movement, only
about its attribution. That is the case a user sees as "the middle
breakdown is off but the rest is fine".

Read-only.

Usage:
    python compare_bridge_manual_vs_computed.py "databooks/x.xlsx" --manual "SJ-量价桥图" --raw AB-SJ
    python compare_bridge_manual_vs_computed.py "databooks/x.xlsx" --manual "成都-量价桥图" --raw AB-CD
"""
import argparse
import re
import sys

sys.path.insert(0, ".")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from fdd_utils.bridge_chart_prototype import find_bridge_blocks
from fdd_utils.generate_bridge_waterfall_batch import build_bridges_for_ab_tab

# Longest marker first per factor: the days label is '运营天数变动', so
# splitting on the bare '天数' would leave '运营' stuck on the phase name and
# a manual '干仓运营天数增加' would not line up with a computed one whose
# phase came from a different source.
_FACTOR_WORDS = (
    ("days", ("运营天数", "運營天數", "运营天", "運營天", "天数", "天數")),
    ("price", ("单价", "單價")),
    ("area", ("出租率", "面积", "面積")),
)


def _classify(label: str):
    """(phase, factor) from a bridge item label like '干仓单价增加'."""
    text = str(label or "").strip()
    for factor, words in _FACTOR_WORDS:
        for w in words:
            if w in text:
                phase = text.split(w)[0].strip()
                # Trailing connective left by a partial marker match.
                phase = re.sub(r"(运营|運營)$", "", phase).strip()
                return phase, factor
    return text, None


def _read_check_row(ws, block, scan: int = 6):
    """The value of the tab's own 'check' cell just below a bridge block.

    That cell compares the bridge's summed end against the ACTUAL revenue
    (e.g. '=E8-N39'), which is the only non-circular test of whether the
    hand-built bridge ties out -- unlike start+factors, where the end cell
    is itself that sum."""
    end_row = block.header_row + len(block.items)
    for r in range(end_row, end_row + scan):
        for c in range(max(1, block.label_col - 1), block.change_col + 2):
            text = ws.cell(row=r, column=c).value
            if isinstance(text, str) and "check" in text.lower():
                for cc in range(c + 1, block.change_col + 2):
                    v = ws.cell(row=r, column=cc).value
                    if isinstance(v, (int, float)):
                        return float(v)
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path")
    ap.add_argument("--manual", required=True, help="the hand-built bridge tab, e.g. 'SJ-量价桥图'")
    ap.add_argument("--raw", required=True, help="that entity's raw data tab, e.g. AB-SJ")
    ap.add_argument("--tol", type=float, default=0.5, help="difference below which items are treated as equal")
    args = ap.parse_args()

    wb = load_workbook(args.path, data_only=True)
    for name in (args.manual, args.raw):
        if name not in wb.sheetnames:
            print(f"❌ sheet {name!r} not found. Available: {wb.sheetnames}")
            return 1

    manual_blocks = find_bridge_blocks(wb[args.manual])
    if not manual_blocks:
        print(f"❌ no Base/Change block found on {args.manual!r}.")
        return 1
    _blocks, computed = build_bridges_for_ab_tab(wb[args.raw], args.raw, log=lambda _m: None)
    if not computed:
        print(f"❌ {args.raw!r} produced no computed bridge.")
        return 1

    print(f"manual tab  : {args.manual!r}  ({len(manual_blocks)} block(s))")
    print(f"computed from: {args.raw!r}  ({len(computed)} transition(s))\n")

    for bi, mblock in enumerate(manual_blocks):
        start_label = mblock.items[0].label
        end_label = mblock.items[-1].label
        print("=" * 78)
        print(f"MANUAL BLOCK {bi + 1}: {start_label} -> {end_label}")
        print("=" * 78)

        # 1. Does the manual tab agree with the ACTUAL revenue?
        m_start = mblock.items[0].value
        m_end = mblock.items[-1].value
        print(f"  [1] MANUAL TAB'S OWN CHECK ROW")
        # Recomputing start+factors and comparing it to the stated end is
        # circular: the tab's end cell IS that sum (=SUM(N29:O38)), so it
        # always closes and says nothing. The real test is the tab's own
        # 'check' row, which compares that sum against the ACTUAL revenue --
        # and on real data it read 60.06 while the tautological version
        # reported a clean zero.
        check_value = _read_check_row(wb[args.manual], mblock)
        if check_value is None:
            m_recon = m_start + sum(it.value for it in mblock.items[1:-1])
            print(f"      no 'check' row found; falling back to start+factors vs stated end")
            print(f"      (this is circular -- the end cell is usually that same sum)")
            print(f"      residual {m_recon - m_end:>+14,.2f}")
        else:
            print(f"      the tab's own check cell reads {check_value:>+14,.2f}")
            if abs(check_value) <= args.tol:
                print(f"      ✅ the hand-built bridge ties to the actual revenue")
            else:
                print(f"      ❌ the hand-built bridge does NOT tie to the actual revenue --")
                print(f"         it is short by this amount BEFORE any comparison with us.")
                print(f"         Whatever gap appears below, this is the tab's own residual;")
                print(f"         fix it there rather than treating our figure as the outlier.")

        # 2. Match to the computed transition with the closest end total.
        best = min(computed, key=lambda r: abs(r.bridge.items[-1].value - m_end))
        c_start = best.bridge.items[0].value
        c_end = best.bridge.items[-1].value
        print(f"\n  [2] TOTALS vs computed ({best.year_a} -> "
              f"{'LTM' if best.is_ltm else best.year_b})")
        print(f"      start : manual {m_start:>13,.2f}   computed {c_start:>13,.2f}   "
              f"diff {m_start - c_start:>+12,.2f}")
        # Absolute tolerance alone is misleading at these magnitudes: a 1.01
        # gap on an 18,746 total is 0.005% -- rounding in the hand-built tab,
        # not a data problem -- yet it exceeds any sensible absolute floor.
        # Report the relative size so the reader can tell the two apart.
        end_rel = abs(m_end - c_end) / abs(c_end) if c_end else None
        rel_note = f"  ({end_rel * 100:.3f}% of total)" if end_rel is not None else ""
        print(f"      end   : manual {m_end:>13,.2f}   computed {c_end:>13,.2f}   "
              f"diff {m_end - c_end:>+12,.2f}{rel_note}")
        totals_agree = abs(m_start - c_start) <= args.tol and abs(m_end - c_end) <= args.tol

        # 3. Factor-by-factor.
        m_items = {}
        for it in mblock.items[1:-1]:
            phase, factor = _classify(it.label)
            if factor:
                m_items[(phase, factor)] = (it.label, it.value)
        c_items = {}
        for it in best.bridge.items[1:-1]:
            phase, factor = _classify(it.label)
            if factor:
                c_items[(phase, factor)] = (it.label, it.value)

        print(f"\n  [3] FACTOR-BY-FACTOR")
        print(f"      {'phase':10s} {'factor':6s} {'manual':>14s} {'computed':>14s} {'diff':>12s}")
        print("      " + "-" * 60)
        keys = sorted(set(m_items) | set(c_items))
        total_abs_diff = 0.0
        biggest = None
        for k in keys:
            phase, factor = k
            mv = m_items.get(k, (None, None))[1]
            cv = c_items.get(k, (None, None))[1]
            if mv is None:
                print(f"      {phase:10s} {factor:6s} {'(absent)':>14s} {cv:>14,.2f}"
                      f" {'-':>12s}   <-- only in computed")
                continue
            if cv is None:
                print(f"      {phase:10s} {factor:6s} {mv:>14,.2f} {'(absent)':>14s}"
                      f" {'-':>12s}   <-- only in manual")
                continue
            d = mv - cv
            total_abs_diff += abs(d)
            if biggest is None or abs(d) > abs(biggest[2]):
                biggest = (phase, factor, d)
            mark = "" if abs(d) <= args.tol else "   <-- DIFFERS"
            print(f"      {phase:10s} {factor:6s} {mv:>14,.2f} {cv:>14,.2f} {d:>+12,.2f}{mark}")

        # Per PHASE, do that phase's three factor differences cancel? If they
        # do, both sides agree on the phase's total movement and differ only
        # in how it is attributed between price, area and days -- a
        # substitution-order question, not a data one. If they do not cancel,
        # that phase's INPUTS differ and its residual is real. Judging this
        # per phase matters: a bridge can be pure attribution on one phase and
        # a genuine input gap on another, and a single verdict for the whole
        # block hides which.
        print(f"\n  [4] PER-PHASE: attribution or input?")
        phases = sorted({p for p, _f in set(m_items) | set(c_items)})
        input_phases, attribution_phases = [], []
        for phase in phases:
            diffs = []
            complete = True
            for factor in ("price", "area", "days"):
                mv = m_items.get((phase, factor), (None, None))[1]
                cv = c_items.get((phase, factor), (None, None))[1]
                if mv is None or cv is None:
                    complete = False
                    continue
                diffs.append(mv - cv)
            if not diffs:
                # Present on one side only -- worth saying so rather than
                # skipping, since a phase missing entirely is a bigger problem
                # than any numeric gap (it usually means the two sides label
                # their phases differently and nothing lined up).
                side = "manual only" if any(k[0] == phase for k in m_items) else "computed only"
                print(f"      {phase:10s} ⚠️ present in {side} -- no counterpart to compare")
                continue
            net = sum(diffs)
            spread = sum(abs(d) for d in diffs)
            if not complete:
                print(f"      {phase:10s} ⚠️ only {len(diffs)}/3 factors present on both sides "
                      f"-- partial net {net:+,.2f}, treat with care")
                continue
            if spread <= args.tol:
                print(f"      {phase:10s} ✅ identical on all three factors")
            elif abs(net) <= args.tol:
                print(f"      {phase:10s} ⇄ ATTRIBUTION ONLY -- factors differ by up to "
                      f"{max(abs(d) for d in diffs):,.2f} but net {net:+,.2f} cancels")
                attribution_phases.append(phase)
            else:
                print(f"      {phase:10s} ❌ INPUT DIFFERENCE -- net {net:+,.2f} does NOT cancel")
                input_phases.append((phase, net))

        print(f"\n  [5] VERDICT")
        m_factor_sum = sum(v for _l, v in m_items.values())
        c_factor_sum = sum(v for _l, v in c_items.values())
        net_all = m_factor_sum - c_factor_sum
        print(f"      manual factors {m_factor_sum:>13,.2f}   computed {c_factor_sum:>13,.2f}"
              f"   net {net_all:>+12,.2f}")
        if end_rel is not None:
            print(f"      end-total gap: {m_end - c_end:+,.2f} "
                  f"({end_rel * 100:.3f}% of the end total)")
        if input_phases:
            print(f"      => The gap is REAL and localised to: "
                  f"{', '.join(f'{p} ({n:+,.2f})' for p, n in input_phases)}")
            print(f"         Those phases' INPUTS differ -- unit rent, area or days. Where the")
            print(f"         days effect is zero on one side and not the other, the two are")
            print(f"         using different day counts for the same window (an LTM window and")
            print(f"         a full year both being ~365 makes that effect zero by definition).")
            if attribution_phases:
                print(f"         {', '.join(attribution_phases)} differ only in attribution and")
                print(f"         need no fix -- do not chase those.")
        elif attribution_phases:
            print(f"      => Every difference is ATTRIBUTION only ({', '.join(attribution_phases)}).")
            print(f"         Both sides agree on each phase's total movement and disagree only")
            print(f"         on how it splits between price, area and days -- that follows from")
            print(f"         the order of sequential substitution (ours is price->area->days).")
            print(f"         Nothing here is wrong; pick one order and apply it consistently.")
        elif end_rel is not None and end_rel > 0.001:
            print(f"      => No single phase accounts for it; the totals themselves differ.")
        else:
            print(f"      => Manual and computed agree.")
        print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
