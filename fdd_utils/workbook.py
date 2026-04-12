from __future__ import annotations

# --- begin workbook/mapping.py ---
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd

from .financial_common import load_yaml_file, package_file_path
from .keyword_registry import (
    BS_END_KEYWORDS,
    BS_HEADER_KEYWORDS,
    INDICATIVE_KEYWORDS,
    IS_END_KEYWORDS,
    IS_HEADER_KEYWORDS,
    REMARK_KEYWORDS,
    SUBTOTAL_KEYWORDS,
    SUMMARY_ACCOUNT_SKIP_KEYWORDS,
    TABLE_END_KEYWORDS,
)


def load_mappings(mappings_path: Optional[str] = None) -> Dict[str, Any]:
    return load_yaml_file(mappings_path or package_file_path("mappings.yml"))


def get_effective_mappings(
    base_mappings: Dict[str, Any],
    resolution: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    effective = dict(base_mappings or {})
    for mapping_key, config in ((resolution or {}).get("dynamic_mappings") or {}).items():
        if mapping_key and isinstance(config, dict):
            effective[mapping_key] = config
    return effective


def should_skip_account_label(account_name: str) -> bool:
    text = str(account_name or "").strip()
    if not text:
        return False

    lowered = text.lower()
    if text.endswith(("合计", "总计", "小计")):
        return True
    if lowered.startswith("total "):
        return True
    return any(keyword in lowered for keyword in SUMMARY_ACCOUNT_SKIP_KEYWORDS)


def normalize_mapping_label(account_name: str) -> str:
    normalized = (account_name or "").strip().lower()
    for suffix in ["：", ":", "（", "）", "(", ")"]:
        normalized = normalized.replace(suffix, "")
    return " ".join(normalized.split())


def iter_account_mappings(mappings: Dict[str, Any]) -> Iterable[tuple[str, Dict[str, Any]]]:
    for mapping_key, config in (mappings or {}).items():
        if str(mapping_key).startswith("_") or not isinstance(config, dict):
            continue
        yield str(mapping_key), config


def find_mapping_key(account_name: str, mappings: Dict[str, Any]) -> str | None:
    """Find the canonical mapping key for an account name or alias."""
    if account_name in mappings:
        return account_name

    normalized_account = normalize_mapping_label(account_name)
    for mapping_key, config in iter_account_mappings(mappings):
        aliases = config.get("aliases", [])
        if account_name in aliases:
            return mapping_key
        normalized_aliases = {normalize_mapping_label(alias) for alias in aliases}
        if normalized_account and normalized_account in normalized_aliases:
            return mapping_key
    return None


def split_accounts_by_type(
    account_names: List[str],
    mappings: Dict[str, Any],
    dfs: Dict[str, "pd.DataFrame"] | None = None,
) -> tuple[List[str], List[str], List[str]]:
    """Preserve account order while grouping by mapping type.

    Falls back to ``df.attrs["integrity"]["statement_type"]`` when the
    account cannot be found in *mappings*, so dynamically-resolved
    accounts are still classified as BS or IS instead of "other".
    """
    bs_accounts: List[str] = []
    is_accounts: List[str] = []
    other_accounts: List[str] = []

    for account_name in account_names:
        account_type = ""
        mapping_key = find_mapping_key(account_name, mappings)
        if mapping_key:
            account_type = mappings[mapping_key].get("type", "")

        # Fallback: read statement_type from the DataFrame attrs
        if account_type not in ("BS", "IS") and dfs and account_name in dfs:
            df = dfs[account_name]
            integrity = getattr(df, "attrs", {}).get("integrity") or {}
            account_type = integrity.get("statement_type", "")

        if account_type == "BS":
            bs_accounts.append(account_name)
        elif account_type == "IS":
            is_accounts.append(account_name)
        else:
            other_accounts.append(account_name)

    return bs_accounts, is_accounts, other_accounts


def build_account_mapping_diagnostics(
    account_names: Iterable[str],
    mappings: Dict[str, Any],
) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []

    for account_name in account_names:
        mapping_key = find_mapping_key(account_name, mappings)
        if not mapping_key:
            rows.append(
                {
                    "account_name": str(account_name),
                    "mapping_key": "",
                    "account_type": "",
                    "classification": "other",
                    "reason": "No mappings.yml key or alias matched this account/tab name.",
                }
            )
            continue

        mapping_config = mappings.get(mapping_key, {})
        account_type = str(mapping_config.get("type", "") or "")
        if account_type in {"BS", "IS"}:
            classification = account_type
            reason = f"Mapped to {account_type} via '{mapping_key}'."
        else:
            classification = "other"
            reason = f"mapped type '{account_type or 'blank'}' is not classified as BS or IS."

        rows.append(
            {
                "account_name": str(account_name),
                "mapping_key": mapping_key,
                "account_type": account_type,
                "classification": classification,
                "reason": reason,
            }
        )

    return pd.DataFrame(
        rows,
        columns=["account_name", "mapping_key", "account_type", "classification", "reason"],
    )
# --- end workbook/mapping.py ---

# --- begin workbook/analysis.py ---
from typing import Any, Dict, List, Optional

import pandas as pd

_TOTAL_KEYWORDS = [
    "total",
    "合计",
    "总计",
    "subtotal",
    "sub-total",
    "sub total",
]


def _trend_direction(values: List[float]) -> str:
    deltas = [curr - prev for prev, curr in zip(values, values[1:])]
    positive = any(delta > 0 for delta in deltas)
    negative = any(delta < 0 for delta in deltas)
    if positive and negative:
        return "volatile"
    if positive:
        return "rising"
    if negative:
        return "falling"
    return "flat"


def _select_trend_focus_row(analysis_df: pd.DataFrame) -> Optional[pd.Series]:
    if analysis_df is None or analysis_df.empty:
        return None
    desc_col = str(analysis_df.columns[0])
    descriptions = analysis_df[desc_col].astype(str)
    total_mask = descriptions.str.lower().str.contains("|".join(_TOTAL_KEYWORDS), regex=True)
    if total_mask.any():
        return analysis_df[total_mask].iloc[0]
    non_zero_mask = analysis_df.iloc[:, 1:].fillna(0).abs().sum(axis=1) > 0
    if non_zero_mask.any():
        return analysis_df[non_zero_mask].iloc[-1]
    return analysis_df.iloc[-1]


def build_trend_summary(analysis_df: pd.DataFrame) -> Dict[str, Any]:
    focus_row = _select_trend_focus_row(analysis_df)
    if focus_row is None or len(analysis_df.columns) < 3:
        return {}

    periods = [str(col) for col in analysis_df.columns[1:]]
    values = [float(focus_row[col] or 0) for col in analysis_df.columns[1:]]
    deltas = [
        {
            "from_period": periods[idx],
            "to_period": periods[idx + 1],
            "delta": values[idx + 1] - values[idx],
        }
        for idx in range(len(values) - 1)
    ]

    largest_increase = max(deltas, key=lambda item: item["delta"]) if deltas else None
    if largest_increase and largest_increase["delta"] <= 0:
        largest_increase = None
    largest_decrease = min(deltas, key=lambda item: item["delta"]) if deltas else None
    if largest_decrease and largest_decrease["delta"] >= 0:
        largest_decrease = None

    return {
        "focus_description": str(focus_row.iloc[0]),
        "series_direction": _trend_direction(values),
        "start_period": periods[0],
        "end_period": periods[-1],
        "start_value": values[0],
        "end_value": values[-1],
        "net_change": values[-1] - values[0],
        "largest_increase": largest_increase,
        "largest_decrease": largest_decrease,
    }


def _change_direction(prev_value: float, curr_value: float) -> str:
    if prev_value == 0 and curr_value > 0:
        return "new_increase"
    if prev_value == 0 and curr_value < 0:
        return "new_decrease"
    if curr_value > prev_value:
        return "increase"
    if curr_value < prev_value:
        return "decrease"
    return "flat"


def build_significant_movements(analysis_df: pd.DataFrame, max_items: int = 3) -> List[Dict[str, Any]]:
    if analysis_df is None or analysis_df.empty or len(analysis_df.columns) < 3:
        return []

    periods = [str(col) for col in analysis_df.columns[1:]]
    movement_candidates: List[Dict[str, Any]] = []
    for _, row in analysis_df.iterrows():
        description = str(row.iloc[0]).strip()
        row_values = [float(row[col] or 0) for col in analysis_df.columns[1:]]
        best_movement = None
        for idx in range(len(row_values) - 1):
            prev_value = row_values[idx]
            curr_value = row_values[idx + 1]
            delta = curr_value - prev_value
            candidate = {
                "description": description,
                "from_period": periods[idx],
                "to_period": periods[idx + 1],
                "from_value": prev_value,
                "to_value": curr_value,
                "delta": delta,
                "abs_delta": abs(delta),
                "direction": _change_direction(prev_value, curr_value),
            }
            if best_movement is None or candidate["abs_delta"] > best_movement["abs_delta"]:
                best_movement = candidate
        if best_movement and best_movement["abs_delta"] > 0:
            movement_candidates.append(best_movement)

    if not movement_candidates:
        return []

    total_change = sum(item["abs_delta"] for item in movement_candidates)
    if total_change <= 0:
        return []

    significant = []
    for item in sorted(movement_candidates, key=lambda entry: entry["abs_delta"], reverse=True):
        percent_of_total_change = (item["abs_delta"] / total_change) * 100
        if percent_of_total_change < 25:
            continue
        significant.append(
            {
                "description": item["description"],
                "from_period": item["from_period"],
                "to_period": item["to_period"],
                "from_value": item["from_value"],
                "to_value": item["to_value"],
                "delta": item["delta"],
                "direction": item["direction"],
                "percent_of_total_change": round(percent_of_total_change, 1),
            }
        )
        if len(significant) >= max_items:
            break
    return significant
# --- end workbook/analysis.py ---

# --- begin workbook/inspector.py ---
"""
Workbook profiling helpers for financial databooks.

This module inspects sheet structure without applying business mappings so later
steps can resolve tabs and normalize values using real workbook metadata.
"""


from functools import lru_cache
import logging
import re
import time
from typing import Any, Callable, Dict, List, Optional

import pandas as pd

from .financial_common import cell_text

logger = logging.getLogger(__name__)

CANONICAL_STAGE_LABELS = (
    ("Indicative adjusted", ("indicative adjusted", "indivative adjusted", "示意性调整后", "示意性調整後")),
    ("Indicative adjustment", ("indicative adjustment", "indivative adjustment", "示意性调整", "示意性調整")),
    ("Audited", ("audited", "审定数", "審定數")),
    ("Audit adjustment", ("audit adjustment", "审计调整", "審計調整")),
    ("Mgt acc", ("mgt acc", "management account", "管理层数", "管理層數")),
)

_UNIT_PATTERNS = (
    "CNY'000",
    "人民币千元",
    "人民幣千元",
    "million",
    "百万",
    "百萬",
)

_TITLE_SKIP_VALUES = {
    "",
    "back",
    "nan",
    "none",
}


@lru_cache(maxsize=8)
def load_workbook_frames(workbook_path: str) -> Dict[str, pd.DataFrame]:
    """Load all workbook sheets once for profiling/normalization reuse."""
    return pd.read_excel(workbook_path, sheet_name=None, header=None, engine="openpyxl")


def _cell_text(value: Any) -> str:
    return cell_text(value)


def _coerce_numeric(value: Any) -> Optional[float]:
    return coerce_numeric(value)


def _normalize_spaces(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def canonical_stage_label(value: Any) -> Optional[str]:
    text = _normalize_spaces(cell_text(value)).lower()
    if not text or len(text) > 35:
        return None
    for canonical, variants in CANONICAL_STAGE_LABELS:
        if any(variant in text for variant in variants):
            return canonical
    return None


def contains_indicative_text(value: Any) -> bool:
    return canonical_stage_label(value) == "Indicative adjusted"


def contains_unit_marker(value: Any, markers: tuple[str, ...] = _UNIT_PATTERNS) -> bool:
    text = _normalize_spaces(cell_text(value)).lower()
    return any(marker.lower() in text for marker in markers)


def stage_row_indices(
    df: pd.DataFrame,
    date_detector: Callable[[Any], Any],
    max_rows: int = 80,
) -> List[int]:
    indices: List[int] = []
    for row_idx in range(min(max_rows, len(df))):
        hits = sum(1 for value in df.iloc[row_idx].tolist() if canonical_stage_label(value))
        if hits >= 2:
            indices.append(row_idx)
            continue
        if hits == 1:
            nearby_date_row = any(
                sum(1 for value in df.iloc[candidate_idx].tolist() if date_detector(value)) >= 1
                for candidate_idx in range(max(0, row_idx - 1), min(len(df), row_idx + 3))
            )
            if nearby_date_row:
                indices.append(row_idx)
    return indices


def primary_stage_row_index(
    df: pd.DataFrame,
    max_rows: int = 12,
) -> Optional[int]:
    best_idx = None
    best_hits = 0
    for row_idx in range(min(max_rows, len(df))):
        hits = sum(1 for value in df.iloc[row_idx].tolist() if canonical_stage_label(value))
        if hits > best_hits:
            best_hits = hits
            best_idx = row_idx
    return best_idx if best_hits > 0 else None


def date_row_index(
    df: pd.DataFrame,
    stage_row_idx: Optional[int],
    date_detector: Callable[[Any], Any],
    max_distance: int = 2,
) -> Optional[int]:
    if stage_row_idx is None:
        return None
    best_idx = None
    best_hits = 0
    for row_idx in range(max(0, stage_row_idx - max_distance), min(len(df), stage_row_idx + max_distance + 1)):
        hits = sum(1 for value in df.iloc[row_idx].tolist() if date_detector(value))
        if hits > best_hits:
            best_hits = hits
            best_idx = row_idx
    return best_idx if best_hits > 0 else None


def _unit_markers(df: pd.DataFrame, max_rows: int = 8) -> List[str]:
    found: List[str] = []
    rows = min(max_rows, len(df))
    for row_idx in range(rows):
        for value in df.iloc[row_idx].tolist():
            text = _cell_text(value)
            for marker in _UNIT_PATTERNS:
                if contains_unit_marker(text, (marker,)) and marker not in found:
                    found.append(marker if marker != "人民幣千元" else "人民币千元")
    return found


def _stage_row_index(df: pd.DataFrame, max_rows: int = 12) -> Optional[int]:
    return primary_stage_row_index(df, max_rows=max_rows)


def _looks_like_entity_heading(text: str) -> bool:
    normalized = _normalize_spaces(text)
    if not normalized:
        return False
    if re.search(r"\s[-–]\s", normalized):
        suffix = re.split(r"\s[-–]\s", normalized, maxsplit=1)[-1].strip()
        if suffix and re.search(r"[A-Za-z\u4e00-\u9fff]", suffix):
            return True
    return any(token in normalized for token in ("公司", "集团", "集團", "company", "Company", "Ltd", "Limited"))


def _stage_block_heading(df: pd.DataFrame, stage_row_idx: int) -> str:
    for row_idx in range(stage_row_idx - 1, max(-1, stage_row_idx - 4), -1):
        values = [_normalize_spaces(_cell_text(value)) for value in df.iloc[row_idx].tolist()]
        texts = [value for value in values if value.lower() not in _TITLE_SKIP_VALUES]
        if not texts:
            continue
        if any(canonical_stage_label(value) for value in texts):
            continue
        if any(_parse_date_label(value) for value in texts):
            continue
        if any(any(marker.lower() in value.lower() for marker in _UNIT_PATTERNS) for value in texts):
            continue
        return " ".join(texts)
    return ""


def _parse_date_label(value: Any) -> Optional[str]:
    parsed = parse_date(value)
    if not parsed:
        return None
    return parsed.strftime("%Y-%m-%d")


def _date_row_index(df: pd.DataFrame, stage_row_idx: Optional[int], max_distance: int = 2) -> Optional[int]:
    return date_row_index(df, stage_row_idx, _parse_date_label, max_distance=max_distance)


def _title_row_index(df: pd.DataFrame, stage_row_idx: Optional[int]) -> Optional[int]:
    search_limit = stage_row_idx if stage_row_idx is not None else min(6, len(df))
    for row_idx in range(min(search_limit + 1, len(df))):
        row = [_normalize_spaces(_cell_text(value)) for value in df.iloc[row_idx].tolist()]
        texts = [value for value in row if value.lower() not in _TITLE_SKIP_VALUES]
        if len(texts) == 1:
            return row_idx
    return None


def _sheet_title(df: pd.DataFrame, stage_row_idx: Optional[int]) -> str:
    title_idx = _title_row_index(df, stage_row_idx)
    if title_idx is not None:
        values = [_normalize_spaces(_cell_text(value)) for value in df.iloc[title_idx].tolist()]
        for value in values:
            if value.lower() not in _TITLE_SKIP_VALUES:
                return value
    for row_idx in range(min(6, len(df))):
        for value in df.iloc[row_idx].tolist():
            text = _normalize_spaces(_cell_text(value))
            lowered = text.lower()
            if lowered in _TITLE_SKIP_VALUES:
                continue
            if canonical_stage_label(text) or _parse_date_label(text):
                continue
            if any(marker.lower() in lowered for marker in ("cny'000", "人民币千元", "人民幣千元")):
                continue
            return text
    return "Sheet"


def _stage_labels(df: pd.DataFrame, stage_row_idx: Optional[int]) -> List[str]:
    return _collect_row_labels(df, stage_row_idx, canonical_stage_label)


def _date_labels(df: pd.DataFrame, date_row_idx: Optional[int]) -> List[str]:
    return _collect_row_labels(df, date_row_idx, _parse_date_label)


def _collect_row_labels(
    df: pd.DataFrame,
    row_idx: Optional[int],
    label_parser,
) -> List[str]:
    if row_idx is None:
        return []
    seen: List[str] = []
    for value in df.iloc[row_idx].tolist():
        label = label_parser(value)
        if label and label not in seen:
            seen.append(label)
    return seen


def _sheet_kind(sheet_name: str, title: str, df: pd.DataFrame, stage_row_idx: Optional[int]) -> str:
    title_lower = title.lower()
    sheet_lower = sheet_name.lower()
    sample = " ".join(_cell_text(value).lower() for value in df.head(min(10, len(df))).fillna("").to_numpy().ravel())
    has_balance = "balance sheet" in sample or "资产负债表" in sample or "資產負債表" in sample
    has_income = "income statement" in sample or "profit and loss" in sample or "利润表" in sample or "利潤表" in sample
    if "financials" in sheet_lower or (has_balance and has_income):
        return "financial_summary"
    if stage_row_idx is not None:
        return "financial_schedule"
    if "ledger" in title_lower or "台账" in sample or "台賬" in sample:
        return "support_schedule"
    return "other"


def _entity_scope(sheet_kind: str, df: pd.DataFrame) -> str:
    if sheet_kind == "financial_summary":
        return "single"
    stage_blocks = stage_row_indices(df, _parse_date_label)
    if len(stage_blocks) >= 2:
        entity_headings = {
            heading
            for heading in (_stage_block_heading(df, stage_row_idx) for stage_row_idx in stage_blocks)
            if _looks_like_entity_heading(heading)
        }
        if len(entity_headings) >= 2:
            return "multiple"
    if len(stage_blocks) >= 3 and any(len(_stage_labels(df, idx)) <= 1 for idx in stage_blocks):
        return "multiple"
    return "single"


def _header_signature(df: pd.DataFrame, stage_row_idx: Optional[int], date_row_idx: Optional[int]) -> Dict[str, Any]:
    signature: Dict[str, Any] = {}
    if stage_row_idx is not None:
        signature["stage_row_idx"] = stage_row_idx
    if date_row_idx is not None:
        signature["date_row_idx"] = date_row_idx
    return signature


def profile_sheet(df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
    stage_row_idx = _stage_row_index(df)
    date_row_idx = _date_row_index(df, stage_row_idx)
    title = _sheet_title(df, stage_row_idx)
    title_row_idx = _title_row_index(df, stage_row_idx)
    sheet_kind = _sheet_kind(sheet_name, title, df, stage_row_idx)
    stage_labels = _stage_labels(df, stage_row_idx)
    date_labels = _date_labels(df, date_row_idx)
    unit_markers = _unit_markers(df)
    entity_scope = _entity_scope(sheet_kind, df)

    return {
        "sheet_name": sheet_name,
        "title": title,
        "title_row_idx": title_row_idx,
        "sheet_kind": sheet_kind,
        "entity_scope": entity_scope,
        "stage_row_idx": stage_row_idx,
        "date_row_idx": date_row_idx,
        "stage_labels": stage_labels,
        "date_labels": date_labels,
        "unit_markers": unit_markers,
        "has_indicative_stage": "Indicative adjusted" in stage_labels,
        "header_signature": _header_signature(df, stage_row_idx, date_row_idx),
    }


@lru_cache(maxsize=8)
def profile_workbook(workbook_path: str) -> Dict[str, Dict[str, Any]]:
    started = time.perf_counter()
    workbook_frames = load_workbook_frames(workbook_path)
    profiles: Dict[str, Dict[str, Any]] = {}
    for sheet_name, df in workbook_frames.items():
        profiles[sheet_name] = profile_sheet(df, sheet_name)
    logger.debug(
        "Workbook profiler scanned %s sheets from %s in %.2fs",
        len(profiles),
        workbook_path,
        time.perf_counter() - started,
    )
    return profiles
# --- end workbook/inspector.py ---

# --- begin workbook/preflight.py ---
from functools import lru_cache
import logging
import os
import re
import time
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import load_workbook

from .financial_common import cell_text

logger = logging.getLogger(__name__)


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _rows_are_blank(rows: Sequence[Sequence[Any]]) -> bool:
    return not any(any(not _is_empty_value(value) for value in row) for row in rows)


def _serialize_rows(rows: Iterable[Sequence[Any]]) -> List[List[Any]]:
    return [list(row) for row in rows]


def _is_likely_entity_name(value: str) -> bool:
    cleaned = value.strip()
    if not 2 < len(cleaned) < 50:
        return False
    if "|" in cleaned:
        return False
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", cleaned):
        return False
    if re.fullmatch(r"[\d\s\-–_/.:]+", cleaned):
        return False
    if cleaned.lower().startswith("project "):
        return False
    lowered = cleaned.lower()
    blocked_phrases = {
        "as at",
        "as of",
        "balance sheet",
        "income statement",
        "statement of financial position",
        "statement of profit or loss",
    }
    if lowered in blocked_phrases or lowered.startswith("as at ") or lowered.startswith("as of "):
        return False
    financial_terms = (
        "vat",
        "tax",
        "receivable",
        "payable",
        "income",
        "expense",
        "asset",
        "liability",
        "增值税",
        "税金",
        "应交",
        "應交",
        "应收",
        "應收",
        "应付",
        "應付",
        "收入",
        "成本",
        "费用",
        "資產",
        "资产",
        "負債",
        "负债",
    )
    if any(term in lowered for term in financial_terms):
        return False
    return True


def _looks_like_schedule_title_prefix(value: str) -> bool:
    lowered = value.lower()
    keywords = (
        "cash",
        "receivable",
        "prepayment",
        "payable",
        "capital",
        "property",
        "tax",
        "other current assets",
        "investment",
        "貨幣資金",
        "应收",
        "應收",
        "预付",
        "預付",
        "应付",
        "應付",
        "股本",
        "税",
    )
    return any(keyword in lowered for keyword in keywords)


def _looks_like_generic_entity_prefix(value: str) -> bool:
    cleaned = value.strip()
    lowered = cleaned.lower()
    if not cleaned or any(char.isdigit() for char in cleaned):
        return False
    return lowered in {"project", "entity", "company"}


def _looks_like_financial_schedule_preview(rows: Sequence[Sequence[Any]]) -> bool:
    flattened = " ".join(_cell_text(value).lower() for row in rows for value in row if not _is_empty_value(value))
    if not flattened:
        return False
    if any(token in flattened for token in ("indicative adjusted", "示意性调整后", "示意性調整後")):
        return True
    return bool(re.search(r"\b20\d{2}-\d{2}-\d{2}\b", flattened))


def _strip_leading_date_fragment(value: str) -> str:
    cleaned = value.strip()
    while True:
        updated = re.sub(
            r"^(?:\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}日?)\s*[-–]\s*",
            "",
            cleaned,
        ).strip()
        if updated == cleaned:
            return cleaned
        cleaned = updated


@lru_cache(maxsize=8)
def _build_workbook_preflight_cached(
    workbook_path: str,
    preview_rows: int = 12,
    entity_rows: int = 20,
    file_mtime_ns: int = 0,
    file_size: int = 0,
) -> Dict[str, Any]:
    del file_mtime_ns, file_size
    started = time.perf_counter()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets: List[Dict[str, Any]] = []

    try:
        max_preview_rows = max(preview_rows, entity_rows)
        for worksheet in workbook.worksheets:
            preview = _serialize_rows(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=max_preview_rows,
                    values_only=True,
                )
            )
            preview_slice = preview[:preview_rows]
            sheet_state = getattr(worksheet, "sheet_state", "visible") or "visible"
            max_row = int(getattr(worksheet, "max_row", 0) or 0)
            max_column = int(getattr(worksheet, "max_column", 0) or 0)
            sheets.append(
                {
                    "name": worksheet.title,
                    "sheet_state": sheet_state,
                    "is_hidden": sheet_state != "visible",
                    "is_blank_preview": _rows_are_blank(preview_slice) and max_row <= preview_rows,
                    "max_row": max_row,
                    "max_column": max_column,
                    "preview_rows": preview,
                }
            )
    finally:
        workbook.close()

    logger.debug(
        "Workbook preflight scanned %s sheets from %s in %.2fs",
        len(sheets),
        workbook_path,
        time.perf_counter() - started,
    )
    return {
        "workbook_path": workbook_path,
        "preview_rows": preview_rows,
        "entity_rows": entity_rows,
        "sheets": sheets,
    }


def build_workbook_preflight(
    workbook_path: str,
    preview_rows: int = 12,
    entity_rows: int = 20,
) -> Dict[str, Any]:
    stat = os.stat(workbook_path)
    return _build_workbook_preflight_cached(
        workbook_path,
        preview_rows,
        entity_rows,
        stat.st_mtime_ns,
        stat.st_size,
    )


def _visible_non_blank_sheets(preflight: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [
        sheet
        for sheet in preflight.get("sheets", [])
        if not sheet.get("is_hidden") and not sheet.get("is_blank_preview")
    ]


def extract_entity_names_from_preflight(preflight: Dict[str, Any]) -> List[str]:
    entity_sources: Dict[str, set[str]] = {}
    entity_counts: Dict[str, int] = {}
    visible_sheets = _visible_non_blank_sheets(preflight)

    def add_candidate(name: str, source: str) -> None:
        cleaned_name = _strip_leading_date_fragment(name)
        if not _is_likely_entity_name(cleaned_name):
            return
        cleaned_name = cleaned_name.strip()
        entity_sources.setdefault(cleaned_name, set()).add(source)
        entity_counts[cleaned_name] = entity_counts.get(cleaned_name, 0) + 1

    for sheet in visible_sheets:
        preview_rows = sheet.get("preview_rows", [])[: preflight.get("entity_rows", 20)]
        is_financial_schedule_preview = _looks_like_financial_schedule_preview(preview_rows)

        for row in preview_rows:
            for value in row:
                if _is_empty_value(value):
                    continue
                value_str = _cell_text(value)
                lowered = value_str.lower()
                if (
                    "示意性调整后" in value_str
                    or "balance sheet" in lowered
                    or "利润表" in value_str
                    or "income statement" in lowered
                ):
                    if " - " in value_str:
                        add_candidate(value_str.split(" - ", 1)[1].strip(), "summary_title")
                    elif " – " in value_str:
                        add_candidate(value_str.split(" – ", 1)[1].strip(), "summary_title")
                elif " - " in value_str or " – " in value_str:
                    parts = re.split(r"\s[-–]\s", value_str, maxsplit=1)
                    if len(parts) > 1:
                        prefix = parts[0].strip()
                        candidate = parts[1].strip()
                        if _looks_like_schedule_title_prefix(prefix):
                            source = "financial_schedule_title" if is_financial_schedule_preview else "schedule_title"
                            add_candidate(candidate, source)
                        elif _looks_like_generic_entity_prefix(prefix):
                            add_candidate(candidate, "generic_dash")

    entity_names = [
        name
        for name, count in entity_counts.items()
        if "summary_title" in entity_sources.get(name, set())
        or "financial_schedule_title" in entity_sources.get(name, set())
        or "generic_dash" in entity_sources.get(name, set())
        or count >= 2
    ]
    return sorted(name for name in entity_names if name and name.strip())


def get_financial_sheet_options(preflight: Dict[str, Any]) -> List[str]:
    def sheet_score(sheet: Dict[str, Any]) -> tuple[int, str]:
        lowered = str(sheet.get("name", "")).lower()
        score = 100
        financial_summary_prefix = (
            lowered == "financial"
            or lowered == "financials"
            or lowered.startswith("financials ")
            or lowered.startswith("financials-")
            or lowered.startswith("financials -")
            or lowered.startswith("financial -")
        )
        if "financial" in lowered:
            score += 60
        if "balance" in lowered or "income" in lowered or "profit" in lowered:
            score += 40
        if lowered.startswith("bs") or lowered.startswith("is"):
            score += 30
        if lowered in {"bshn", "bs"}:
            score += 20
        if "-->" in lowered or lowered == "adj":
            score -= 20
        return (0 if financial_summary_prefix else 1, -score, lowered)

    visible_sheets = _visible_non_blank_sheets(preflight)
    return [sheet["name"] for sheet in sorted(visible_sheets, key=sheet_score)]
# --- end workbook/preflight.py ---

# --- begin workbook/table_debug.py ---
"""
FDD Table Inspector - Understand databook table structure and number flow.
Adapts HR-style island/section detection for financial databooks.
Focus: 示意性調整後 / Indicative adjusted section, header hierarchy, multiplier, calculation flow.
"""

import pandas as pd
import re
from functools import lru_cache
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field



@dataclass
class TableSection:
    """A detected table section (BS or IS) with metadata."""
    name: str
    start_row: int
    end_row: int
    header_row: int
    marker_row: int
    date_row: int
    desc_col_idx: int
    indicative_cols: List[Tuple[int, str, Optional[Any]]]  # (col_idx, date_str, parsed_date)
    multiply_by_1000: bool
    project_name: Optional[str] = None


@dataclass
class RowClassification:
    """Classification of a data row."""
    is_detail: bool
    is_subtotal: bool
    is_total: bool
    description: str
    values: Dict[str, float]
    row_idx: int


@dataclass
class TableInspection:
    """Full inspection result for a sheet."""
    sheet_name: str
    sections: List[TableSection] = field(default_factory=list)
    row_classifications: List[RowClassification] = field(default_factory=list)
    header_hierarchy: Dict[str, Any] = field(default_factory=dict)
    multiplier_note: str = ""


def _cell_contains_indicative(cell) -> bool:
    """Check if cell contains indicative adjusted marker."""
    if pd.isna(cell):
        return False
    return contains_indicative_text(cell)


def _cell_contains_cny000(cell) -> bool:
    """Check if cell indicates CNY'000 (multiply by 1000)."""
    if pd.isna(cell):
        return False
    return contains_unit_marker(cell, ("CNY'000", "人民币千元"))


def _is_subtotal_or_total(desc: str) -> Tuple[bool, bool]:
    """Return (is_subtotal, is_total)."""
    if not desc or pd.isna(desc):
        return False, False
    d = str(desc).strip().lower()
    total_kw = ['总计', '合计', 'total', '小计', 'subtotal', '小計', '合計', '總計']
    is_total = any(k in d for k in ['总计', '合计', 'total', '總計', '合計'])
    is_subtotal = any(k in d for k in ['小计', 'subtotal', 'sub-total', '小計'])
    return is_subtotal, is_total


def _find_header_rows(df: pd.DataFrame) -> List[Tuple[int, str]]:
    """Find all header rows (Indicative adjusted) and their section type. Returns [(row_idx, 'BS'|'IS'|'')]."""
    found = []
    for idx, row in df.iterrows():
        row_str = ' '.join(row.astype(str).values)
        if 'Indicative adjusted' not in row_str and '示意性调整后' not in row_str and '示意性調整後' not in row_str:
            continue
        row_lower = row_str.lower()
        if any(kw.lower() in row_lower for kw in BS_HEADER_KEYWORDS):
            found.append((idx, 'BS'))
        elif any(kw.lower() in row_lower for kw in IS_HEADER_KEYWORDS):
            found.append((idx, 'IS'))
        else:
            found.append((idx, ''))  # Generic indicative row (e.g. first occurrence)
    return sorted(found, key=lambda x: x[0])  # Ensure row order


def inspect_sheet(df: pd.DataFrame, sheet_name: str = "Sheet") -> TableInspection:
    """
    Inspect a financial sheet: detect BS/IS sections, 示意性調整後 columns, header hierarchy, multiplier.
    Uses logic aligned with financial_extraction.py. Handles both BS and IS in same sheet.
    """
    inspection = TableInspection(sheet_name=sheet_name)
    
    if df is None or df.empty:
        return inspection
    
    # ── Find all header rows (may have BS and IS in same sheet) ──
    header_rows = _find_header_rows(df)
    if not header_rows:
        # Fallback: any row with Indicative adjusted
        for idx, row in df.iterrows():
            row_str = ' '.join(row.astype(str).values)
            if 'Indicative adjusted' in row_str or '示意性调整后' in row_str:
                header_rows = [(idx, '')]
                break
    
    if not header_rows:
        inspection.multiplier_note = "No 示意性調整後 / Indicative adjusted header found."
        return inspection
    
    # Use first header row for shared structure (desc col, multiplier)
    header_row_idx = header_rows[0][0]
    
    # ── For each header row, build section ──
    desc_col_idx = _find_description_column(df)
    if desc_col_idx is None:
        inspection.multiplier_note = "No description column found."
        return inspection
    
    all_classifications = []
    header_hierarchies = []
    
    for h_idx, (header_row_idx, section_type) in enumerate(header_rows):
        marker_row_idx = header_row_idx + 2
        date_row_idx = header_row_idx + 1
        
        if marker_row_idx >= len(df) or date_row_idx >= len(df):
            continue
        
        date_row = df.iloc[date_row_idx]
        
        indicative_cols = []
        validated_columns = get_valid_financial_columns(
            df=df,
            desc_col_idx=desc_col_idx,
            header_row_idx=header_row_idx,
        )
        for col_idx, parsed, date_str in validated_columns:
            date_display = parsed.strftime('%Y-%m-%d') if parsed else str(date_str)
            indicative_cols.append((col_idx, date_display, None))
        
        date_row_str = ' '.join(date_row.astype(str).values)
        multiply_by_1000 = "CNY'000" in date_row_str or "人民币千元" in date_row_str
        if not inspection.multiplier_note:
            inspection.multiplier_note = "×1000 (CNY'000 / 人民币千元)" if multiply_by_1000 else "No multiplier"
        
        section_name = "Balance Sheet" if section_type == 'BS' else ("Income Statement" if section_type == 'IS' else "Financial Table")
        data_start = date_row_idx + 1
        data_end = len(df)
        end_keywords = BS_END_KEYWORDS + IS_END_KEYWORDS
        
        # For IS, end at 净利润; for BS, end at 负债及所有者权益总计 or at next section start
        next_section_start = header_rows[h_idx + 1][0] if h_idx + 1 < len(header_rows) else len(df)
        for row_idx in range(data_start, min(next_section_start, len(df))):
            desc = df.iloc[row_idx, desc_col_idx]
            desc_str = str(desc).strip() if pd.notna(desc) else ""
            if any(kw.lower() in desc_str.lower() for kw in end_keywords):
                data_end = row_idx + 1
                break
        
        section = TableSection(
            name=section_name,
            start_row=int(header_row_idx),
            end_row=int(data_end),
            header_row=int(header_row_idx),
            marker_row=int(marker_row_idx),
            date_row=int(date_row_idx),
            desc_col_idx=int(desc_col_idx),
            indicative_cols=indicative_cols,
            multiply_by_1000=multiply_by_1000,
        )
        inspection.sections.append(section)
        
        header_hierarchies.append({
            "section": section_name,
            "header_row": int(header_row_idx),
            "date_row": int(date_row_idx),
            "marker_row": int(marker_row_idx),
            "indicative_columns": [{"col_idx": c[0], "date": c[1]} for c in indicative_cols],
        })
        
        for row_idx in range(data_start, data_end):
            if row_idx >= len(df):
                break
            row = df.iloc[row_idx]
            desc = row.iloc[desc_col_idx]
            desc_str = str(desc).strip() if pd.notna(desc) else ""
            
            values = {}
            for col_idx, date_disp, _ in indicative_cols:
                val = row.iloc[col_idx]
                try:
                    v = float(val)
                    if multiply_by_1000:
                        v *= 1000
                    values[date_disp] = round(v, 0)
                except (ValueError, TypeError):
                    values[date_disp] = 0.0
            
            is_subtotal, is_total = _is_subtotal_or_total(desc_str)
            all_classifications.append(RowClassification(
                is_detail=not (is_subtotal or is_total),
                is_subtotal=is_subtotal,
                is_total=is_total,
                description=desc_str,
                values=values,
                row_idx=int(row_idx),
            ))
    
    inspection.header_hierarchy = {"sections": header_hierarchies} if header_hierarchies else {}
    if header_hierarchies:
        inspection.header_hierarchy["description_col"] = desc_col_idx
        inspection.header_hierarchy["multiply_by_1000"] = inspection.sections[0].multiply_by_1000 if inspection.sections else False
    inspection.row_classifications = all_classifications
    
    return inspection


def inspect_workbook(workbook_path: str, sheet_name: Optional[str] = None) -> Dict[str, TableInspection]:
    """
    Inspect all relevant sheets in a workbook.
    Returns dict of sheet_name -> TableInspection.
    """
    try:
        xls = pd.ExcelFile(workbook_path, engine='openpyxl')
        sheets = [sheet_name] if sheet_name else xls.sheet_names
        results = {}
        for sh in sheets:
            if sh not in xls.sheet_names:
                continue
            df = pd.read_excel(workbook_path, sheet_name=sh, header=None)
            results[sh] = inspect_sheet(df, sh)
        return results
    except Exception as e:
        return {"_error": TableInspection(sheet_name="_error")}  # Placeholder; multiplier_note can hold error


def format_inspection_for_display(inspection: TableInspection) -> str:
    """
    Format TableInspection as readable markdown for UI.
    Shows: header hierarchy, 示意性調整後 section, multiplier, row types, number flow.
    """
    lines = []
    lines.append(f"### Sheet: {inspection.sheet_name}")
    lines.append("")
    
    if inspection.multiplier_note:
        lines.append(f"**Multiplier**: {inspection.multiplier_note}")
        lines.append("")
    
    if inspection.header_hierarchy:
        h = inspection.header_hierarchy
        if "sections" in h:
            for sec_h in h["sections"]:
                lines.append(f"**{sec_h.get('section', 'Section')}**")
                lines.append(f"- Header row: {sec_h.get('header_row', '?')} (Indicative adjusted / 示意性調整後)")
                lines.append(f"- Date row: {sec_h.get('date_row', '?')}")
                lines.append(f"- Marker row: {sec_h.get('marker_row', '?')} (管理层数 | 审定数 | 示意性调整后)")
                cols = sec_h.get('indicative_columns', [])
                if cols:
                    lines.append("  示意性調整後 columns:")
                    for c in cols:
                        lines.append(f"    - Col {c.get('col_idx', '?')}: {c.get('date', '?')}")
                lines.append("")
        else:
            lines.append("**Header Hierarchy**")
            lines.append(f"- Header row: {h.get('header_row', '?')} (Indicative adjusted / 示意性調整後)")
            lines.append(f"- Date row: {h.get('date_row', '?')}")
            lines.append(f"- Marker row: {h.get('marker_row', '?')} (管理层数 | 审定数 | 示意性调整后)")
            lines.append(f"- Description column: {h.get('description_col', '?')}")
            cols = h.get('indicative_columns', [])
            if cols:
                lines.append("**示意性調整後 Columns** (selected for extraction):")
                for c in cols:
                    lines.append(f"  - Col {c.get('col_idx', '?')}: {c.get('date', '?')}")
            lines.append("")
    
    if inspection.sections:
        for sec in inspection.sections:
            lines.append(f"**Section**: {sec.name} (rows {sec.start_row}–{sec.end_row})")
        lines.append("")
    
    if inspection.row_classifications:
        lines.append("**Row Types & Number Flow**")
        detail_count = sum(1 for r in inspection.row_classifications if r.is_detail)
        subtotal_count = sum(1 for r in inspection.row_classifications if r.is_subtotal)
        total_count = sum(1 for r in inspection.row_classifications if r.is_total)
        lines.append(f"- Detail rows: {detail_count} | Subtotals: {subtotal_count} | Totals: {total_count}")
        lines.append("- Preview below is taken from the real uploaded workbook. It shows the first 25 rows and up to the first 3 indicative value columns for readability.")
        lines.append("")
        lines.append("| Row | Type | Description | Indicative values (real workbook preview) |")
        lines.append("|-----|------|-------------|-------------------------------------------|")
        
        for rc in inspection.row_classifications[:25]:  # Show first 25
            t = "Detail" if rc.is_detail else ("Subtotal" if rc.is_subtotal else "Total")
            preview_vals = list(rc.values.values())[:3]
            vals_str = ", ".join(f"{v:,.0f}" if v != 0 else "0" for v in preview_vals)
            desc_short = (rc.description[:40] + "…") if len(rc.description) > 40 else rc.description
            lines.append(f"| {rc.row_idx} | {t} | {desc_short} | {vals_str} |")
        
        if len(inspection.row_classifications) > 25:
            lines.append(f"| … | … | (+{len(inspection.row_classifications) - 25} more rows) | |")
    
    return "\n".join(lines)


@lru_cache(maxsize=32)
def get_table_inspection(workbook_path: str, sheet_name: str) -> Any:
    inspections = inspect_workbook(workbook_path, sheet_name)
    return inspections.get(sheet_name)


def clear_table_inspection_cache():
    get_table_inspection.cache_clear()


def clear_workbook_caches():
    """Clear all lru_cache entries for workbook profiling and loading."""
    load_workbook_frames.cache_clear()
    profile_workbook.cache_clear()
    _build_workbook_preflight_cached.cache_clear()
    get_table_inspection.cache_clear()
# --- end workbook/table_debug.py ---

# --- begin workbook/text_export.py ---
"""
Utilities for exporting selected Excel tabs into compact plain text.
"""


from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

import pandas as pd


@dataclass(frozen=True)
class TrimmedSheet:
    sheet_name: str
    dataframe: pd.DataFrame
    start_row: int  # zero-based
    start_col: int  # zero-based

    @property
    def end_row(self) -> int:
        return self.start_row + len(self.dataframe.index) - 1

    @property
    def end_col(self) -> int:
        return self.start_col + len(self.dataframe.columns) - 1

    @property
    def used_range(self) -> str:
        return (
            f"{_column_letter(self.start_col + 1)}{self.start_row + 1}:"
            f"{_column_letter(self.end_col + 1)}{self.end_row + 1}"
        )


def _column_letter(column_number: int) -> str:
    result = []
    current = column_number
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def _is_blank_or_na_value(value) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _normalize_cell_value(value) -> str:
    if _is_blank_or_na_value(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _trim_sheet(sheet_name: str, df: pd.DataFrame) -> TrimmedSheet:
    if df is None or df.empty:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    non_empty_rows = [
        idx for idx in range(len(df.index)) if not df.iloc[idx].map(_is_empty_value).all()
    ]
    non_empty_cols = [
        idx for idx in range(len(df.columns)) if not df.iloc[:, idx].map(_is_empty_value).all()
    ]

    if not non_empty_rows or not non_empty_cols:
        raise ValueError(f"Sheet '{sheet_name}' is empty after trimming blank borders.")

    start_row = min(non_empty_rows)
    end_row = max(non_empty_rows)
    start_col = min(non_empty_cols)
    end_col = max(non_empty_cols)

    trimmed = df.iloc[start_row : end_row + 1, start_col : end_col + 1].copy()
    trimmed = trimmed.fillna("")

    return TrimmedSheet(
        sheet_name=sheet_name,
        dataframe=trimmed,
        start_row=start_row,
        start_col=start_col,
    )


def _validate_selected_tabs(all_sheet_names: Sequence[str], selected_tabs: Iterable[str]) -> List[str]:
    normalized_tabs = [str(tab).strip() for tab in selected_tabs if str(tab).strip()]
    if not normalized_tabs:
        raise ValueError("At least one sheet tab must be selected.")

    missing_tabs = [tab for tab in normalized_tabs if tab not in all_sheet_names]
    if missing_tabs:
        raise ValueError(
            f"Requested sheet tabs were not found: {', '.join(missing_tabs)}. "
            f"Available sheets: {', '.join(all_sheet_names)}"
        )

    return normalized_tabs


def _render_sheet_table(trimmed_sheet: TrimmedSheet) -> str:
    column_headers = ["ExcelRow"] + [
        f"Col{_column_letter(index + 1)}" for index in range(len(trimmed_sheet.dataframe.columns))
    ]
    lines = [
        f"===== SHEET: {trimmed_sheet.sheet_name} =====",
        f"USED_RANGE: {trimmed_sheet.used_range}",
        "| " + " | ".join(column_headers) + " |",
    ]

    for row_offset, (_, row) in enumerate(trimmed_sheet.dataframe.iterrows()):
        excel_row = trimmed_sheet.start_row + row_offset + 1
        cells = [_normalize_cell_value(value) for value in row.tolist()]
        lines.append("| " + " | ".join([str(excel_row), *cells]) + " |")

    return "\n".join(lines)


def render_selected_tabs_text(workbook_path: str, selected_tabs: Sequence[str]) -> str:
    workbook = Path(workbook_path)
    all_sheets = pd.read_excel(workbook, sheet_name=None, header=None, engine="openpyxl")
    tabs_to_export = _validate_selected_tabs(list(all_sheets.keys()), selected_tabs)

    rendered_sections = []
    for sheet_name in tabs_to_export:
        trimmed_sheet = _trim_sheet(sheet_name, all_sheets[sheet_name])
        rendered_sections.append(_render_sheet_table(trimmed_sheet))

    header = [
        f"WORKBOOK: {workbook.name}",
        f"EXPORTED_AT: {datetime.now().isoformat(timespec='seconds')}",
        f"SHEETS: {', '.join(tabs_to_export)}",
        "",
    ]
    return "\n\n".join(["\n".join(header).rstrip(), *rendered_sections]).strip() + "\n"


def export_selected_tabs_to_file(
    workbook_path: str,
    selected_tabs: Sequence[str],
    output_path: str | None = None,
) -> str:
    workbook = Path(workbook_path)
    destination = (
        Path(output_path)
        if output_path
        else workbook.with_name(f"{workbook.stem}_selected_tabs.txt")
    )
    rendered = render_selected_tabs_text(str(workbook), selected_tabs)
    destination.write_text(rendered, encoding="utf-8")
    return str(destination)
# --- end workbook/text_export.py ---

# --- begin workbook/statements.py ---
"""
Standalone Financial Extraction Helper Module.

Extracts Balance Sheet and Income Statement data from Excel workbooks.
"""

import pandas as pd
import re
import logging
from datetime import datetime
from typing import Any, Dict, Tuple, Optional, List
import warnings

from .financial_common import cell_text, coerce_numeric, normalize_financial_date_label

warnings.simplefilter(action='ignore', category=UserWarning)

logger = logging.getLogger(__name__)


def _contains_indicative_marker(text: str) -> bool:
    """Check whether text marks an indicative-adjusted value column."""
    lowered = text.lower()
    # Handle both "Indicative adjusted" (with space) and "Indicativeadjusted" (no space)
    normalised = lowered.replace(' ', '')
    return '示意性调整后' in lowered or '示意性調整後' in lowered or 'indicativeadjusted' in normalised


def _looks_like_remark_text(text: str) -> bool:
    """Detect note-like text that should not be treated as a value column."""
    lowered = text.lower()
    return any(keyword in lowered for keyword in REMARK_KEYWORDS)


def _column_numeric_profile(series: pd.Series) -> Dict[str, float]:
    """Measure how numeric a candidate value column looks."""
    non_empty_count = 0
    numeric_count = 0
    text_count = 0

    for value in series:
        text = _cell_text(value)
        if not text:
            continue
        non_empty_count += 1
        if _coerce_numeric(value) is not None:
            numeric_count += 1
        else:
            text_count += 1

    numeric_ratio = (numeric_count / non_empty_count) if non_empty_count else 0.0
    return {
        'non_empty_count': non_empty_count,
        'numeric_count': numeric_count,
        'text_count': text_count,
        'numeric_ratio': numeric_ratio,
    }

def parse_date(date_str, debug=False):
    """
    Parse date string in various formats including xMxx and Chinese formats.
    Uses the same preprocessing logic as process_databook.py for consistency.
    
    Args:
        date_str: Date string in various formats
        debug: If True, print debug info
        
    Returns:
        datetime object or None if parsing fails
    """
    if not date_str or pd.isna(date_str):
        return None

    if isinstance(date_str, (int, float)) and not isinstance(date_str, bool):
        serial_value = float(date_str)
        if 20000 <= serial_value <= 60000:
            try:
                result = pd.to_datetime(serial_value, unit='D', origin='1899-12-30', errors='coerce')
                if pd.notna(result):
                    if debug:
                        print(f"      [parse_date]   ✅ Excel serial success: {result.strftime('%Y-%m-%d')}")
                    return result.to_pydatetime(warn=False)
            except Exception:
                pass

    original_str = str(date_str).strip()

    if original_str.isdigit():
        try:
            serial_value = float(original_str)
            if 20000 <= serial_value <= 60000:
                result = pd.to_datetime(serial_value, unit='D', origin='1899-12-30', errors='coerce')
                if pd.notna(result):
                    if debug:
                        print(f"      [parse_date]   ✅ Excel serial string success: {result.strftime('%Y-%m-%d')}")
                    return result.to_pydatetime(warn=False)
        except Exception:
            pass
    
    if debug:
        print(f"      [parse_date] Parsing: '{original_str}'")
    
    # Preprocess the date
    preprocessed = normalize_financial_date_label(original_str)
    
    if debug and preprocessed != original_str:
        print(f"      [parse_date]   Preprocessed: '{preprocessed}'")
    
    # Try to convert preprocessed date to datetime using pandas
    try:
        result = pd.to_datetime(preprocessed, errors='coerce')
        if pd.notna(result):
            if debug:
                print(f"      [parse_date]   ✅ Success: {result.strftime('%Y-%m-%d')}")
            return result.to_pydatetime(warn=False)
    except (TypeError, ValueError, OverflowError):
        pass
    
    if debug:
        print(f"      [parse_date]   ❌ Failed to parse")
    
    return None


def _find_description_column(df: pd.DataFrame, debug: bool = False) -> Optional[int]:
    """Find the most likely description column for a financial table."""
    for col_idx in range(len(df.columns)):
        try:
            col_str = df.iloc[:, col_idx].astype(str)
            if col_str.str.contains(r"CNY'000|人民币千元", case=False, na=False, regex=True).any():
                if debug:
                    print(f"[DEBUG] ✅ Description column found by unit marker at index: {col_idx}")
                return col_idx
        except Exception:
            continue

    tick_symbols = ['✓', '√', '✔', '☑', '■', '□', '●', '○', '★', '☆']
    best_idx = None
    best_score = -1
    rows_to_scan = min(12, len(df))

    for col_idx in range(len(df.columns)):
        sample_values = [_cell_text(df.iloc[row_idx, col_idx]) for row_idx in range(rows_to_scan)]
        non_empty_values = [value for value in sample_values if value]
        if not non_empty_values:
            continue

        has_tick_symbols = sum(any(tick in value for tick in tick_symbols) for value in non_empty_values)
        text_like_count = 0
        remark_count = 0
        date_like_count = 0

        for value in non_empty_values:
            if _looks_like_remark_text(value):
                remark_count += 1
                continue
            if parse_date(value):
                date_like_count += 1
                continue
            if _coerce_numeric(value) is not None:
                continue
            if len(value) > 1:
                text_like_count += 1

        score = (text_like_count * 2) - (has_tick_symbols * 2) - remark_count - date_like_count
        if text_like_count > 0 and score > best_score:
            best_idx = col_idx
            best_score = score

    if best_idx is not None:
        if debug:
            print(f"[DEBUG] ✅ Description column found by text heuristic at index: {best_idx}")
        return best_idx

    if len(df.columns) > 0:
        if debug:
            print("[DEBUG] ⚠️ Falling back to first column as description")
        return 0

    return None


def _get_valid_financial_columns_for_rows(
    df: pd.DataFrame,
    desc_col_idx: int,
    date_row_idx: int,
    marker_row_idx: Optional[int] = None,
    data_start_row: Optional[int] = None,
    data_end_row: Optional[int] = None,
    debug: bool = False,
) -> List[Tuple[int, datetime, str]]:
    """Validate financial columns given explicit date and marker rows."""
    if date_row_idx < 0 or date_row_idx >= len(df):
        return []

    if data_end_row is None:
        data_end_row = len(df)

    marker_row = df.iloc[marker_row_idx] if marker_row_idx is not None and 0 <= marker_row_idx < len(df) else None
    if data_start_row is None:
        last_header_row = max(date_row_idx, marker_row_idx if marker_row_idx is not None else date_row_idx)
        data_start_row = min(last_header_row + 1, len(df))

    date_row = df.iloc[date_row_idx]
    validated_columns: List[Tuple[int, datetime, str]] = []

    for col_idx in range(desc_col_idx + 1, len(df.columns)):
        date_text = _cell_text(date_row.iloc[col_idx])
        marker_text = _cell_text(marker_row.iloc[col_idx]) if marker_row is not None else ''
        parsed_date = parse_date(date_text)
        has_marker = _contains_indicative_marker(marker_text)
        header_has_remark = _looks_like_remark_text(date_text) or _looks_like_remark_text(marker_text)
        body_profile = _column_numeric_profile(df.iloc[data_start_row:data_end_row, col_idx])

        is_numeric_enough = (
            body_profile['numeric_count'] >= 1 and
            body_profile['numeric_ratio'] >= 0.5 and
            body_profile['text_count'] <= 2
        )

        is_valid = bool(
            parsed_date and
            not header_has_remark and
            (has_marker or is_numeric_enough or body_profile['non_empty_count'] == 0)
        )

        if debug:
            print(
                "[DEBUG]   Column {col_idx}: date_row={date_row_idx}, marker_row={marker_row_idx}, "
                "date='{date_text}', marker='{marker_text}', parsed_date={parsed_date}, "
                "numeric_count={numeric_count}, text_count={text_count}, numeric_ratio={numeric_ratio:.2f}, "
                "valid={is_valid}".format(
                    col_idx=col_idx,
                    date_row_idx=date_row_idx,
                    marker_row_idx=marker_row_idx,
                    date_text=date_text,
                    marker_text=marker_text,
                    parsed_date=parsed_date.strftime('%Y-%m-%d') if parsed_date else None,
                    numeric_count=body_profile['numeric_count'],
                    text_count=body_profile['text_count'],
                    numeric_ratio=body_profile['numeric_ratio'],
                    is_valid=is_valid,
                )
            )

        if is_valid:
            validated_columns.append((col_idx, parsed_date, date_text))

    return validated_columns


def _find_best_columns_from_header(
    df: pd.DataFrame,
    desc_col_idx: int,
    header_row_idx: int,
    data_end_row: Optional[int] = None,
    debug: bool = False,
) -> Tuple[Optional[int], Optional[int], List[Tuple[int, datetime, str]]]:
    """Try nearby date/marker layouts and return the best match for a header row."""
    best_date_row_idx = None
    best_marker_row_idx = None
    best_columns: List[Tuple[int, datetime, str]] = []
    best_score = (-1, -1, -9999, -1)

    start_date_row = max(0, header_row_idx - 2)
    end_date_row = min(len(df), header_row_idx + 4)
    for date_row_idx in range(start_date_row, end_date_row):
        for marker_row_idx in (header_row_idx, date_row_idx - 1, date_row_idx + 1):
            if marker_row_idx < 0 or marker_row_idx >= len(df):
                continue
            candidate_columns = _get_valid_financial_columns_for_rows(
                df=df,
                desc_col_idx=desc_col_idx,
                date_row_idx=date_row_idx,
                marker_row_idx=marker_row_idx,
                data_end_row=data_end_row,
                debug=debug,
            )
            marker_row = df.iloc[marker_row_idx] if 0 <= marker_row_idx < len(df) else None
            indicative_count = 0
            if marker_row is not None:
                indicative_count = sum(
                    _contains_indicative_marker(_cell_text(marker_row.iloc[col_idx]))
                    for col_idx, _, _ in candidate_columns
                )
            unique_dates = {
                parsed_date.strftime('%Y-%m-%d')
                for _, parsed_date, _ in candidate_columns
            }
            duplicate_count = len(candidate_columns) - len(unique_dates)
            candidate_score = (
                indicative_count,
                len(unique_dates),
                -duplicate_count,
                len(candidate_columns),
            )
            if candidate_score > best_score:
                best_score = candidate_score
                best_columns = candidate_columns
                best_date_row_idx = date_row_idx
                best_marker_row_idx = marker_row_idx

    return best_date_row_idx, best_marker_row_idx, best_columns


def _select_indicative_cluster(
    date_columns: List[Tuple[int, datetime, str]],
    marker_row: Optional[pd.Series],
) -> List[Tuple[int, datetime, str]]:
    """Prefer the contiguous date block anchored by indicative-adjusted markers."""
    if marker_row is None or not date_columns:
        return date_columns

    indicative_columns = [
        column
        for column in date_columns
        if _contains_indicative_marker(_cell_text(marker_row.iloc[column[0]]))
    ]
    if not indicative_columns:
        return date_columns

    min_col = min(column[0] for column in indicative_columns)
    max_col = max(column[0] for column in indicative_columns)
    clustered_columns = [
        column
        for column in date_columns
        if min_col <= column[0] <= max_col
    ]
    return clustered_columns or indicative_columns


def _dedupe_date_columns(
    date_columns: List[Tuple[int, datetime, str]],
) -> List[Tuple[int, datetime, str]]:
    """Keep the first column for each output date to avoid later overwrites."""
    deduped_columns: List[Tuple[int, datetime, str]] = []
    seen_dates = set()
    for column in date_columns:
        date_key = column[1].strftime('%Y-%m-%d')
        if date_key in seen_dates:
            continue
        seen_dates.add(date_key)
        deduped_columns.append(column)
    return deduped_columns


def _find_relaxed_date_columns(
    df: pd.DataFrame,
    desc_col_idx: int,
    debug: bool = False,
) -> Tuple[Optional[int], List[Tuple[int, datetime, str]]]:
    """Fallback for layouts where headers are messy but date columns are still parseable."""
    return _scan_relaxed_date_columns(
        df=df,
        desc_col_idx=desc_col_idx,
        max_scan_rows=8,
        min_numeric_ratio=0.4,
        max_text_count=2,
        debug=debug,
        label="Relaxed date-column fallback",
    )


def _find_extended_relaxed_date_columns(
    df: pd.DataFrame,
    desc_col_idx: int,
    max_scan_rows: int = 20,
    debug: bool = False,
) -> Tuple[Optional[int], List[Tuple[int, datetime, str]]]:
    """Broader fallback for messy IS layouts with the date row farther down."""
    return _scan_relaxed_date_columns(
        df=df,
        desc_col_idx=desc_col_idx,
        max_scan_rows=max_scan_rows,
        min_numeric_ratio=0.3,
        max_text_count=3,
        debug=debug,
        label="Extended relaxed date fallback",
    )


def _scan_relaxed_date_columns(
    df: pd.DataFrame,
    desc_col_idx: int,
    max_scan_rows: int,
    min_numeric_ratio: float,
    max_text_count: int,
    debug: bool,
    label: str,
) -> Tuple[Optional[int], List[Tuple[int, datetime, str]]]:
    """Shared relaxed date-row scan used by the primary and extended fallbacks."""
    best_date_row_idx = None
    best_columns: List[Tuple[int, datetime, str]] = []
    max_rows = min(max_scan_rows, len(df))

    for date_row_idx in range(max_rows):
        candidate_columns: List[Tuple[int, datetime, str]] = []
        for col_idx in range(desc_col_idx + 1, len(df.columns)):
            date_text = _cell_text(df.iloc[date_row_idx, col_idx])
            parsed_date = parse_date(date_text)
            if not parsed_date or _looks_like_remark_text(date_text):
                continue

            body_profile = _column_numeric_profile(df.iloc[date_row_idx + 1:, col_idx])
            if (
                body_profile['numeric_count'] >= 1
                and body_profile['numeric_ratio'] >= min_numeric_ratio
                and body_profile['text_count'] <= max_text_count
            ):
                candidate_columns.append((col_idx, parsed_date, date_text))

        if len(candidate_columns) > len(best_columns):
            best_columns = candidate_columns
            best_date_row_idx = date_row_idx

    if debug and best_columns:
        print(f"[DEBUG] ✅ {label} selected row {best_date_row_idx} with columns {[col for col, _, _ in best_columns]}")

    return best_date_row_idx, best_columns


def _table_end_keywords(table_name: str) -> List[str]:
    return TABLE_END_KEYWORDS.get(table_name, TABLE_END_KEYWORDS["Income Statement"])


def _find_table_end_row(df: pd.DataFrame, desc_col_idx: int, data_start_row: int, table_name: str) -> int:
    """Locate the logical end of a BS/IS table."""
    data_end_row = len(df)
    end_keywords = _table_end_keywords(table_name)

    for row_idx in range(data_start_row, len(df)):
        desc = _cell_text(df.iloc[row_idx, desc_col_idx])
        if any(keyword.lower() in desc.lower() for keyword in end_keywords):
            return row_idx + 1

    for row_idx in range(data_start_row, len(df)):
        if df.iloc[row_idx].isna().all():
            return row_idx

    return data_end_row


def _build_financial_result(
    df: pd.DataFrame,
    table_name: str,
    desc_col_idx: int,
    date_row_idx: int,
    date_columns: List[Tuple[int, datetime, str]],
    multiply_values: bool = True,
    debug: bool = False,
) -> Optional[pd.DataFrame]:
    """Build output DataFrame once description/date columns are known."""
    if date_row_idx >= len(df) or not date_columns:
        return None

    date_row = df.iloc[date_row_idx]
    header_texts = [' '.join(date_row.astype(str).values)]
    if date_row_idx - 1 >= 0:
        header_texts.append(' '.join(df.iloc[date_row_idx - 1].astype(str).values))
    if date_row_idx + 1 < len(df):
        header_texts.append(' '.join(df.iloc[date_row_idx + 1].astype(str).values))
    header_blob = ' '.join(header_texts)
    multiply_by_1000 = multiply_values and ("CNY'000" in header_blob or "人民币千元" in header_blob)

    data_start_row = date_row_idx + 1
    if data_start_row < len(df):
        row_text = ' '.join(_cell_text(value).lower() for value in df.iloc[data_start_row].values)
        if _contains_indicative_marker(row_text):
            data_start_row += 1

    data_end_row = _find_table_end_row(df, desc_col_idx, data_start_row, table_name)

    result_rows = []
    for row_idx in range(data_start_row, data_end_row):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        if not description or _contains_indicative_marker(description):
            continue

        row_dict = {'Description': description}
        for col_idx, parsed_date, _ in date_columns:
            value = df.iloc[row_idx, col_idx]
            numeric_value = _coerce_numeric(value)
            if numeric_value is None:
                numeric_value = 0
            if multiply_by_1000:
                numeric_value *= 1000
            row_dict[parsed_date.strftime('%Y-%m-%d')] = int(round(numeric_value, 0))

        result_rows.append(row_dict)

    if not result_rows:
        if debug:
            print("[DEBUG] ❌ No rows built in financial result helper")
        return None

    result_df = pd.DataFrame(result_rows)
    date_cols = [col for col in result_df.columns if col != 'Description']
    if date_cols:
        result_df = result_df[result_df[date_cols].ne(0).any(axis=1)]

    if result_df.empty:
        if debug:
            print("[DEBUG] ❌ Financial result helper produced only zero rows")
        return None

    return result_df.reset_index(drop=True)


def _extract_income_statement_directly(
    df: pd.DataFrame,
    debug: bool = False,
    multiply_values: bool = True,
) -> Optional[pd.DataFrame]:
    """Fallback IS extractor for layouts the primary indicative-header logic misses."""
    working_df = df.dropna(how='all').dropna(axis=1, how='all').reset_index(drop=True)
    if working_df.empty:
        return None

    if debug:
        print(f"[DEBUG] 🔄 Direct IS fallback on shape {working_df.shape}")

    desc_col_idx = _find_description_column(working_df, debug=debug)
    if desc_col_idx is None:
        return None

    date_row_idx, date_columns = _find_extended_relaxed_date_columns(
        working_df,
        desc_col_idx,
        debug=debug,
    )
    if not date_columns or date_row_idx is None:
        return None

    return _build_financial_result(
        df=working_df,
        table_name="Income Statement",
        desc_col_idx=desc_col_idx,
        date_row_idx=date_row_idx,
        date_columns=date_columns,
        multiply_values=multiply_values,
        debug=debug,
    )


def get_valid_financial_columns(
    df: pd.DataFrame,
    desc_col_idx: int,
    header_row_idx: int,
    data_start_row: Optional[int] = None,
    data_end_row: Optional[int] = None,
    debug: bool = False,
) -> List[Tuple[int, datetime, str]]:
    """
    Return validated financial value columns to the right of the description column.

    A valid financial column should look like a reporting-date column and should not
    behave like a free-text remark column.
    """
    if data_start_row is not None:
        date_row_idx = max(header_row_idx + 1, data_start_row - 1)
        marker_row_idx = date_row_idx + 1 if date_row_idx + 1 < len(df) else None
        return _get_valid_financial_columns_for_rows(
            df=df,
            desc_col_idx=desc_col_idx,
            date_row_idx=date_row_idx,
            marker_row_idx=marker_row_idx,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
            debug=debug,
        )

    _, _, validated_columns = _find_best_columns_from_header(
        df=df,
        desc_col_idx=desc_col_idx,
        header_row_idx=header_row_idx,
        data_end_row=data_end_row,
        debug=debug,
    )
    return validated_columns


def detect_table_header_row(df: pd.DataFrame, keywords: List[str] = None) -> Optional[int]:
    """
    Detect the header row containing indicative keywords.
    
    Args:
        df: DataFrame to search
        keywords: List of keywords to search for (default: indicative adjusted keywords)
        
    Returns:
        Row index of header or None if not found
    """
    if keywords is None:
        keywords = ['Indicative adjusted', '示意性调整后', '示意性調整後', "CNY'000", "人民币千元"]
    
    for idx, row in df.iterrows():
        row_str = ' '.join(row.astype(str).values)
        if any(keyword in row_str for keyword in keywords):
            return idx
    
    return None


def find_date_columns(df: pd.DataFrame, header_row_idx: int) -> Tuple[List[int], List[datetime], Optional[int]]:
    """
    Find date columns and return the most recent date column index.
    
    Args:
        df: DataFrame to search
        header_row_idx: Index of the header row
        
    Returns:
        Tuple of (date_column_indices, parsed_dates, most_recent_column_index)
    """
    if header_row_idx >= len(df) - 1:
        return [], [], None
    
    date_row_idx = header_row_idx + 1
    date_row = df.iloc[date_row_idx]
    
    parsed_dates = []
    date_indices = []
    
    for col_idx, value in enumerate(date_row):
        parsed_date = parse_date(value)
        if parsed_date:
            parsed_dates.append(parsed_date)
            date_indices.append(col_idx)
    
    if not parsed_dates:
        return [], [], None
    
    # Find most recent date
    most_recent_idx = parsed_dates.index(max(parsed_dates))
    most_recent_col_idx = date_indices[most_recent_idx]
    
    return date_indices, parsed_dates, most_recent_col_idx


def extract_financial_table(
    df: pd.DataFrame,
    table_name: str,
    entity_keywords: Optional[List[str]] = None,
    debug: bool = False,
    multiply_values: bool = True
) -> Optional[pd.DataFrame]:
    """
    Extract financial table (Balance Sheet or Income Statement) from a worksheet.
    Gets ALL columns with "示意性调整后" or "Indicative adjusted".
    
    Args:
        df: DataFrame containing the financial data
        table_name: Name of the table (e.g., "Balance Sheet", "Income Statement")
        entity_keywords: Optional list of entity name components to search for
        debug: If True, print debugging information
        multiply_values: If True, multiply by 1000 if CNY'000 detected
        
    Returns:
        Cleaned DataFrame with Description column and ALL adjusted columns
    """
    if debug:
        print(f"\n[DEBUG] Extracting {table_name}...")
        print(f"[DEBUG] DataFrame shape: {df.shape}")
    
    # Detect header row candidates with "Indicative adjusted" or "示意性调整后"
    header_row_candidates = []
    for idx, row in df.iterrows():
        row_str = ' '.join(row.astype(str).values)
        if _contains_indicative_marker(row_str):
            header_row_candidates.append(idx)

    if debug:
        if header_row_candidates:
            print(f"[DEBUG] ✅ Header row candidates found: {header_row_candidates}")
        else:
            print(f"[DEBUG] ⚠️ No indicative-adjusted header row candidates found")

    # Find description column
    if debug:
        print(f"[DEBUG] Searching for description column...")

    desc_col_idx = _find_description_column(df, debug=debug)

    if desc_col_idx is None:
        if debug:
            print(f"[DEBUG] ❌ No description column found")
        return None
    
    if debug:
        print(f"[DEBUG] Showing first 5 rows of ALL columns (to understand structure):")
        for row_num in range(min(5, len(df))):
            print(f"[DEBUG] Row {row_num}: {df.iloc[row_num].values[:20]}")  # Show first 20 cols
    
    if debug:
        print(f"\n[DEBUG] ========== FILTERING FINANCIAL VALUE COLUMNS ==========")

    best_header_row_idx = None
    best_date_row_idx = None
    best_marker_row_idx = None
    date_columns: List[Tuple[int, datetime, str]] = []

    for candidate_header_row in header_row_candidates:
        candidate_date_row_idx, candidate_marker_row_idx, candidate_columns = _find_best_columns_from_header(
            df=df,
            desc_col_idx=desc_col_idx,
            header_row_idx=candidate_header_row,
            debug=debug,
        )
        if len(candidate_columns) > len(date_columns):
            best_header_row_idx = candidate_header_row
            best_date_row_idx = candidate_date_row_idx
            best_marker_row_idx = candidate_marker_row_idx
            date_columns = candidate_columns

    if not date_columns and table_name == "Income Statement":
        best_date_row_idx, date_columns = _find_relaxed_date_columns(df, desc_col_idx, debug=debug)
        best_header_row_idx = best_date_row_idx - 1 if best_date_row_idx is not None and best_date_row_idx > 0 else best_date_row_idx
        best_marker_row_idx = None

    if not date_columns:
        if debug:
            print(f"[DEBUG] ❌ No columns found!")
        return None

    marker_row = None
    if best_marker_row_idx is not None and 0 <= best_marker_row_idx < len(df):
        marker_row = df.iloc[best_marker_row_idx]
        date_columns = _select_indicative_cluster(date_columns, marker_row)

    date_columns = _dedupe_date_columns(date_columns)

    header_row_idx = best_header_row_idx if best_header_row_idx is not None else 0
    date_row_idx = best_date_row_idx if best_date_row_idx is not None else header_row_idx + 1

    if debug:
        print(f"\n[DEBUG] ========== FINAL SELECTED COLUMNS ==========")
        print(f"[DEBUG] Header row index: {header_row_idx}")
        print(f"[DEBUG] Date row index: {date_row_idx}")
        print(f"[DEBUG] Marker row index: {best_marker_row_idx}")
        print(f"[DEBUG] Total columns selected: {len(date_columns)}")
        print(f"[DEBUG] Column indices: {[col_idx for col_idx, _, _ in date_columns]}")
        print(f"=" * 80)

    # Get date row
    if date_row_idx >= len(df):
        if debug:
            print(f"[DEBUG] ❌ Date row index {date_row_idx} is out of bounds")
        return None
    
    date_row = df.iloc[date_row_idx]
    
    if debug:
        print(f"\n[DEBUG] ========== COLUMNS SELECTED FOR OUTPUT ==========")
        print(f"[DEBUG] Description column: {desc_col_idx}")
        print(f"[DEBUG] Date columns found: {len(date_columns)}")
        for col_idx, parsed_date, date_str in date_columns:
            print(f"[DEBUG]   ✅ Column {col_idx}: '{date_str}' → will be named '{parsed_date.strftime('%Y-%m-%d')}'")
        
        print(f"\n[DEBUG] Final output will have columns:")
        output_cols = ['Description'] + [parsed_date.strftime('%Y-%m-%d') for _, parsed_date, _ in date_columns]
        print(f"[DEBUG]   {output_cols}")
        print(f"[DEBUG] Total: {len(output_cols)} columns ({len(output_cols)-1} date columns)")
        print(f"=" * 80)
    
    # Check if CNY'000 multiplier needed
    header_texts = [' '.join(date_row.astype(str).values)]
    if date_row_idx - 1 >= 0:
        header_texts.append(' '.join(df.iloc[date_row_idx - 1].astype(str).values))
    if date_row_idx + 1 < len(df):
        header_texts.append(' '.join(df.iloc[date_row_idx + 1].astype(str).values))
    header_blob = ' '.join(header_texts)
    multiply_by_1000 = multiply_values and ("CNY'000" in header_blob or "人民币千元" in header_blob)
    
    if debug and multiply_by_1000:
        print(f"[DEBUG] Will multiply values by 1000 (CNY'000 detected)")
    elif debug and not multiply_values:
        print(f"[DEBUG] Multiplication disabled by parameter")
    
    # Determine end row based on table type
    data_start_row = date_row_idx + 1
    data_end_row = len(df)
    
    # For Balance Sheet: end at liabilities/equity total.
    # For Income Statement: end at net profit/(loss)-style rows.
    end_keywords = _table_end_keywords(table_name)
    
    if debug:
        print(f"[DEBUG] Looking for end markers: {end_keywords}")
        print(f"[DEBUG] Searching from row {data_start_row} to {len(df)}")
    
    end_marker_found = False
    for row_idx in range(data_start_row, len(df)):
        row = df.iloc[row_idx]
        desc = str(row.iloc[desc_col_idx]).strip()
        
        if debug and row_idx < data_start_row + 10:  # Show first 10 rows
            print(f"[DEBUG]   Row {row_idx}: '{desc}'")
        
        if any(keyword.lower() in desc.lower() for keyword in end_keywords):
            data_end_row = row_idx + 1  # Include this row
            end_marker_found = True
            if debug:
                print(f"[DEBUG] ✅ Found end marker at row {row_idx}: '{desc}'")
            break
    
    if debug:
        if not end_marker_found:
            print(f"[DEBUG] ⚠️  No end marker found! Will extract to end of dataframe (row {len(df)})")
        print(f"[DEBUG] Data extraction range: rows {data_start_row} to {data_end_row} ({data_end_row - data_start_row} rows)")
        print(f"[DEBUG] Preview of extraction range:")
        for row_idx in range(data_start_row, min(data_start_row + 5, data_end_row)):
            if row_idx < len(df):
                desc = str(df.iloc[row_idx].iloc[desc_col_idx]).strip()
                print(f"[DEBUG]   Row {row_idx}: '{desc}'")
    
    # Build result dataframe with Description + ALL adjusted columns
    if debug:
        print(f"\n[DEBUG] Extracting data from rows {data_start_row} to {data_end_row}...")
        print(f"[DEBUG] Will extract from description column {desc_col_idx} and date columns: {[col for col, _, _ in date_columns]}")
    
    result_rows = []
    skipped_empty_desc = 0
    
    for row_idx in range(data_start_row, data_end_row):
        row = df.iloc[row_idx]
        
        description = row.iloc[desc_col_idx]
        
        # Skip if description is null or empty
        if pd.isna(description) or str(description).strip() == '':
            skipped_empty_desc += 1
            continue
        
        # Build row dict with description and all date values
        row_dict = {'Description': str(description).strip()}
        
        has_any_nonzero_value = False
        conversion_errors = 0
        
        for col_idx, parsed_date, date_str in date_columns:
            value = row.iloc[col_idx]
            col_name = parsed_date.strftime('%Y-%m-%d')
            
            numeric_value = _coerce_numeric(value)
            try:
                if numeric_value is None:
                    raise ValueError("not numeric")
                if multiply_by_1000:
                    numeric_value *= 1000
                numeric_value = round(numeric_value, 0)
                
                row_dict[col_name] = int(numeric_value)
                
                if numeric_value != 0:
                    has_any_nonzero_value = True
                    
            except (ValueError, TypeError) as e:
                conversion_errors += 1
                row_dict[col_name] = 0
        
        # Debug only problematic rows or first few
        if debug and (len(result_rows) < 3 or conversion_errors > 0 or not has_any_nonzero_value):
            values_str = ", ".join([f"{col}: {row_dict[col]}" for col in row_dict if col != 'Description'])
            status = "⚠️ ALL ZEROS" if not has_any_nonzero_value else ("⚠️ ERRORS" if conversion_errors > 0 else "✅")
            print(f"[DEBUG]   Row {row_idx} {status}: '{row_dict['Description'][:50]}' → {values_str}")
            print(f"[DEBUG]     row_dict keys: {list(row_dict.keys())}")
            print(f"[DEBUG]     row_dict values: {list(row_dict.values())}")
        
        # Add row (even if all zeros, we'll filter later)
        result_rows.append(row_dict)
    
    if debug:
        print(f"\n[DEBUG] Extraction complete:")
        print(f"[DEBUG]   - Total rows processed: {data_end_row - data_start_row}")
        print(f"[DEBUG]   - Rows with empty descriptions: {skipped_empty_desc}")
        print(f"[DEBUG]   - Rows extracted: {len(result_rows)}")
        
        # Show which columns had most conversion errors
        if result_rows:
            temp_df = pd.DataFrame(result_rows)
            for col in temp_df.columns:
                if col != 'Description':
                    zero_count = (temp_df[col] == 0).sum()
                    nonzero_count = (temp_df[col] != 0).sum()
                    print(f"[DEBUG]   Column '{col}': {nonzero_count} non-zero, {zero_count} zeros")
    
    if not result_rows:
        if debug:
            print(f"[DEBUG] ❌ No valid data rows found!")
            print(f"[DEBUG] Processed {data_end_row - data_start_row} rows but none had valid data")
        return None
    
    if debug:
        print(f"\n[DEBUG] ========== CREATING DATAFRAME ==========")
        print(f"[DEBUG] Creating DataFrame from {len(result_rows)} rows")
        
        # Show first 3 result_rows as dict
        for i, row_dict in enumerate(result_rows[:3]):
            print(f"[DEBUG] result_rows[{i}]:")
            for k, v in row_dict.items():
                print(f"[DEBUG]   {k}: {v} (type: {type(v).__name__})")
    
    result_df = pd.DataFrame(result_rows)
    
    if debug:
        print(f"\n[DEBUG] DataFrame created successfully!")
        print(f"[DEBUG]   Shape: {result_df.shape}")
        print(f"[DEBUG]   Columns: {list(result_df.columns)}")
        print(f"[DEBUG]   Dtypes: {result_df.dtypes.to_dict()}")
        print(f"\n[DEBUG] DataFrame content (first 3 rows):")
        print(result_df.head(3).to_string())
        
        # Check for any issues with column values
        for col in result_df.columns:
            if col != 'Description':
                print(f"[DEBUG]   Column '{col}' stats: min={result_df[col].min()}, max={result_df[col].max()}, mean={result_df[col].mean():.0f}")
    
    # Remove rows where ALL date column values are 0
    date_cols = [col for col in result_df.columns if col != 'Description']
    
    if debug:
        print(f"\n[DEBUG] ========== FILTERING ZERO ROWS ==========")
        print(f"[DEBUG] Date columns to check: {date_cols}")
    
    if date_cols:
        # Keep rows where at least one date column is non-zero
        rows_before = len(result_df)
        mask = result_df[date_cols].ne(0).any(axis=1)
        
        if debug:
            print(f"[DEBUG] Rows before filtering: {rows_before}")
            print(f"[DEBUG] Mask (True = keep, False = remove):")
            print(f"[DEBUG]   {mask.values[:20]}")  # Show first 20
            
            # Show which rows will be removed
            removed_indices = result_df[~mask].index
            if len(removed_indices) > 0:
                print(f"[DEBUG] Rows to be removed ({len(removed_indices)} total):")
                for idx in list(removed_indices)[:5]:
                    desc = result_df.loc[idx, 'Description']
                    vals = [result_df.loc[idx, col] for col in date_cols]
                    print(f"[DEBUG]   Row {idx}: '{desc}' → {vals}")
        
        result_df = result_df[mask]
        rows_after = len(result_df)
        
        if debug:
            print(f"[DEBUG] Rows after filtering: {rows_after}")
            print(f"[DEBUG] Removed {rows_before - rows_after} rows with all zeros")
    
    if result_df.empty:
        if debug:
            print(f"[DEBUG] ❌ DataFrame is empty after removing zero rows!")
        return None
    
    if debug:
        print(f"\n[DEBUG] ========== FINAL RESULT ==========")
        print(f"[DEBUG] ✅ Final DataFrame: {len(result_df)} rows × {len(result_df.columns)} columns")
        print(f"[DEBUG] Columns: {list(result_df.columns)}")
        
        # Show value statistics for each date column
        for col in result_df.columns:
            if col != 'Description':
                non_zero = (result_df[col] != 0).sum()
                zero = (result_df[col] == 0).sum()
                print(f"[DEBUG]   '{col}': {non_zero} non-zero, {zero} zeros (max: {result_df[col].max():,.0f})")
        
        print(f"\n[DEBUG] First 5 rows:")
        print(result_df.head(5).to_string())
        
        print(f"\n[DEBUG] Last 5 rows:")
        print(result_df.tail(5).to_string())
        
        # Check if there are any rows with all values as 0
        if len(date_cols) > 0:
            all_zero_mask = (result_df[date_cols] == 0).all(axis=1)
            all_zero_count = all_zero_mask.sum()
            if all_zero_count > 0:
                print(f"\n[DEBUG] ⚠️ WARNING: {all_zero_count} rows still have ALL zeros!")
                print(f"[DEBUG] These rows:")
                print(result_df[all_zero_mask].to_string())
    
    return result_df


def extract_balance_sheet_and_income_statement(
    workbook_path: str,
    sheet_name: str,
    debug: bool = False,
    multiply_values: bool = True
) -> Dict[str, Any]:
    """
    Extract Balance Sheet and Income Statement from a SINGLE Excel worksheet.
    Both BS and IS are in the same sheet, separated by header rows.
    
    Args:
        workbook_path: Path to Excel workbook
        sheet_name: Worksheet name containing both BS and IS
        debug: If True, print debugging information
        multiply_values: If True, multiply by 1000 if CNY'000 detected
        
    Returns:
        Dictionary with keys:
        - 'balance_sheet': DataFrame or None
        - 'income_statement': DataFrame or None  
        - 'project_name': String (extracted from headers) or None
        
    Example:
        >>> results = extract_balance_sheet_and_income_statement(
        ...     workbook_path="databook.xlsx",
        ...     sheet_name="Financial Statements",
        ...     debug=True
        ... )
        >>> print(results['balance_sheet'])
        >>> print(results['income_statement'])
        >>> print(results['project_name'])
    """
    results = {
        'balance_sheet': None,
        'income_statement': None,
        'project_name': None
    }
    
    if debug:
        print("=" * 80)
        print("FINANCIAL EXTRACTION - DEBUG MODE")
        print("=" * 80)
        print(f"Workbook: {workbook_path}")
        print(f"Sheet: {sheet_name}")
    
    try:
        # Load Excel file
        df = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        
        if debug:
            print(f"\n[DEBUG] ✅ Sheet loaded: {df.shape}")
        
        # Find Balance Sheet section
        bs_start_row = None
        bs_keywords = [
            "示意性调整后资产负债表",
            "示意性調整後資產負債表",
            "Indicative adjusted balance sheet",
            "Indicative Adjusted Balance Sheet",
            "Balance sheet",
        ]
        
        for idx, row in df.iterrows():
            row_str = ' '.join(row.astype(str).values).lower()
            if any(kw.lower() in row_str for kw in bs_keywords):
                bs_start_row = idx
                if debug:
                    print(f"[DEBUG] ✅ Balance Sheet starts at row {idx}: {df.iloc[idx].values[0]}")
                break
        
        # Find Income Statement section  
        is_start_row = None
        is_keywords = [
            "示意性调整后利润表",
            "示意性調整後利潤表",
            "Indicative adjusted income statement",
            "Indicative Adjusted Income Statement",
            "Income statement",
            "profit and loss",
            "statement of comprehensive income",
        ]
        
        for idx, row in df.iterrows():
            row_str = ' '.join(row.astype(str).values).lower()
            if any(kw.lower() in row_str for kw in is_keywords):
                is_start_row = idx
                if debug:
                    print(f"[DEBUG] ✅ Income Statement starts at row {idx}: {df.iloc[idx].values[0]}")
                break

        if is_start_row is None:
            relaxed_is_keywords = ["利润表", "利潤表", "income statement", "profit and loss"]
            for idx, row in df.iterrows():
                row_str = ' '.join(row.astype(str).values).lower()
                if any(keyword in row_str for keyword in relaxed_is_keywords):
                    is_start_row = idx
                    if debug:
                        print(f"[DEBUG] ✅ Income Statement starts at row {idx} using relaxed detection")
                    break
        
        # Extract project name (from header row pattern)
        # Pattern: "示意性调整后资产负债表 - 东莞联洋" or "Balance Sheet - Project Name"
        # Should appear in both BS and IS headers
        project_name_bs = None
        project_name_is = None
        
        if bs_start_row is not None:
            # Check all cells in BS header row for the pattern
            bs_row = df.iloc[bs_start_row]
            for val in bs_row:
                val_str = str(val)
                if _contains_indicative_marker(val_str) or 'balance sheet' in val_str.lower():
                    if ' - ' in val_str:
                        project_name_bs = val_str.split(' - ', 1)[1].strip()
                    elif '-' in val_str and '调整后' not in val_str.split('-')[-1]:
                        project_name_bs = val_str.split('-')[-1].strip()
                    break
            
            if debug:
                print(f"[DEBUG] BS header project name: '{project_name_bs}'")
        
        if is_start_row is not None:
            # Check all cells in IS header row for the pattern
            is_row = df.iloc[is_start_row]
            for val in is_row:
                val_str = str(val)
                if _contains_indicative_marker(val_str) or 'income statement' in val_str.lower():
                    if ' - ' in val_str:
                        project_name_is = val_str.split(' - ', 1)[1].strip()
                    elif '-' in val_str and '调整后' not in val_str.split('-')[-1]:
                        project_name_is = val_str.split('-')[-1].strip()
                    break
            
            if debug:
                print(f"[DEBUG] IS header project name: '{project_name_is}'")
        
        # Use project name if it appears in both headers (or if only one is found)
        if project_name_bs and project_name_is:
            if project_name_bs == project_name_is:
                project_name = project_name_bs
                if debug:
                    print(f"[DEBUG] ✅ Project name confirmed in both headers: '{project_name}'")
            else:
                if debug:
                    print(f"[DEBUG] ⚠️  Project names don't match! BS: '{project_name_bs}', IS: '{project_name_is}'")
                project_name = project_name_bs  # Use BS name as default
        elif project_name_bs:
            project_name = project_name_bs
        elif project_name_is:
            project_name = project_name_is
        else:
            project_name = None
            if debug:
                print(f"[DEBUG] ❌ No project name found in headers")
        
        results['project_name'] = project_name
        
        # Extract Balance Sheet
        if bs_start_row is not None:
            # Determine end row (either IS start or end of sheet)
            bs_end_row = is_start_row if is_start_row else len(df)
            df_bs = df.iloc[bs_start_row:bs_end_row].copy().reset_index(drop=True)
            
            results['balance_sheet'] = extract_financial_table(
                df_bs, "Balance Sheet", None, debug, multiply_values
            )
        
        # Extract Income Statement
        if is_start_row is not None:
            # IS goes to end of sheet
            df_is = df.iloc[is_start_row:].copy().reset_index(drop=True)
            
            results['income_statement'] = extract_financial_table(
                df_is, "Income Statement", None, debug, multiply_values
            )

            if results['income_statement'] is None:
                if debug:
                    print("[DEBUG] ⚠️ Primary Income Statement extraction returned None; trying direct fallback")
                results['income_statement'] = _extract_income_statement_directly(
                    df_is,
                    debug=debug,
                    multiply_values=multiply_values,
                )
                if debug:
                    status = "✅ succeeded" if results['income_statement'] is not None else "❌ failed"
                    print(f"[DEBUG] Direct Income Statement fallback {status}")
        
        # Post-processing: Remove date columns with all zeros in Income Statement
        if results['income_statement'] is not None:
            is_df = results['income_statement']
            date_cols = [col for col in is_df.columns if col != 'Description']
            
            # Find columns with all zeros in IS
            cols_to_remove = []
            for col in date_cols:
                if (is_df[col] == 0).all():
                    cols_to_remove.append(col)
            
            if cols_to_remove:
                remaining_date_cols = [col for col in date_cols if col not in cols_to_remove]
                if not remaining_date_cols:
                    if debug:
                        print("[DEBUG] ⚠️ Skipping zero-column removal because it would remove all Income Statement date columns")
                    cols_to_remove = []

            if cols_to_remove:
                if debug:
                    print(f"\n[DEBUG] ========== REMOVING ZERO COLUMNS ==========")
                    print(f"[DEBUG] Found {len(cols_to_remove)} date columns with ALL zeros in Income Statement:")
                    print(f"[DEBUG]   {cols_to_remove}")
                    print(f"[DEBUG] Removing these columns from BOTH Balance Sheet and Income Statement...")
                
                # Remove from Income Statement
                results['income_statement'] = is_df.drop(columns=cols_to_remove)
                
                # Remove from Balance Sheet if it exists and has those columns
                if results['balance_sheet'] is not None:
                    bs_df = results['balance_sheet']
                    cols_to_remove_from_bs = [col for col in cols_to_remove if col in bs_df.columns]
                    if cols_to_remove_from_bs:
                        results['balance_sheet'] = bs_df.drop(columns=cols_to_remove_from_bs)
                        if debug:
                            print(f"[DEBUG]   Removed {len(cols_to_remove_from_bs)} columns from Balance Sheet")
                
                if debug:
                    print(f"[DEBUG] ✅ Columns removed successfully")
        
        if debug:
            print("\n" + "=" * 80)
            print("EXTRACTION RESULTS:")
            print("=" * 80)
            print(f"Project Name: {results['project_name'] or '❌ Not found'}")
            print(f"Balance Sheet: {'✅ Extracted' if results['balance_sheet'] is not None else '❌ None'}")
            print(f"Income Statement: {'✅ Extracted' if results['income_statement'] is not None else '❌ None'}")
            if results['balance_sheet'] is not None:
                print(f"  - Balance Sheet: {len(results['balance_sheet'])} rows × {len(results['balance_sheet'].columns)} cols")
                print(f"  - Columns: {list(results['balance_sheet'].columns)}")
            if results['income_statement'] is not None:
                print(f"  - Income Statement: {len(results['income_statement'])} rows × {len(results['income_statement'].columns)} cols")
                print(f"  - Columns: {list(results['income_statement'].columns)}")
        
    except Exception as e:
        logger.error("Error extracting financial data: %s", e)
        if debug:
            import traceback
            logger.debug("Full traceback for financial extraction error:", exc_info=True)
    
    return results




# Example usage and testing
if __name__ == "__main__":
    # Example: Extract BS and IS from single sheet
    print("="*80)
    print("Example: Extract Balance Sheet and Income Statement from Single Sheet")
    print("="*80)
    
    workbook_path = "databook.xlsx"
    sheet_name = "Financial Statements"  # Sheet containing both BS and IS
    
    results = extract_balance_sheet_and_income_statement(
        workbook_path=workbook_path,
        sheet_name=sheet_name,
        debug=True  # Enable debugging
    )
    
    print(f"\n{'='*80}")
    print("EXTRACTION SUMMARY")
    print(f"{'='*80}")
    
    # Show project name
    if results['project_name']:
        print(f"✅ Project Name: {results['project_name']}")
    else:
        print("❌ Project Name: Not found")
    
    # Show Balance Sheet
    if results['balance_sheet'] is not None:
        print(f"\n✅ Balance Sheet Extracted:")
        print(f"   Total rows: {len(results['balance_sheet'])}")
        print(f"   Columns: {list(results['balance_sheet'].columns)}")
        print(f"   Sample data:")
        print(results['balance_sheet'].head(5))
    else:
        print("\n❌ Balance Sheet: Not found")
    
    # Show Income Statement
    if results['income_statement'] is not None:
        print(f"\n✅ Income Statement Extracted:")
        print(f"   Total rows: {len(results['income_statement'])}")
        print(f"   Columns: {list(results['income_statement'].columns)}")
        print(f"   Sample data:")
        print(results['income_statement'].head(5))
    else:
        print("\n❌ Income Statement: Not found")
    
    # Example: Access specific account data
    if results['balance_sheet'] is not None:
        print(f"\n{'='*80}")
        print("Example: Access specific account")
        print(f"{'='*80}")
        
        # Find account with description containing keyword
        cash_data = results['balance_sheet'][
            results['balance_sheet']['Description'].str.contains('货币资金', na=False)
        ]
        
        if not cash_data.empty:
            print("货币资金 (Cash) data:")
            print(cash_data.to_string())
            
            # Get values for each date
            for col in cash_data.columns:
                if col != 'Description':
                    value = cash_data.iloc[0][col]
                    print(f"  {col}: {value:,.0f}")
# --- end workbook/statements.py ---

# --- begin workbook/schedules.py ---
"""
Normalize FDD financial schedules into integrity-aware payloads.
"""


from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional
import math
import re

import pandas as pd

from .financial_common import cell_text, coerce_numeric


PREFERRED_STAGE = "Indicative adjusted"
INTERNAL_ROW_KEY = "__source_row_idx"
_TOTAL_KEYWORDS = ("total", "总计", "合计", "總計", "合計")
_SUBTOTAL_KEYWORDS = ("subtotal", "sub-total", "sub total", "小计", "小計")
_UNIT_MARKERS = ("cny'000", "人民币千元", "人民幣千元")
_WORKING_REMARK_KEYWORDS = ("check", "对账单", "對賬單", "对帐单", "差异", "差異", "difference", "diff", "recon")
_CARRYING_AMOUNT_LABELS = ("carrying amounts", "net book value", "carrying amount", "账面价值", "賬面價值")


def _forward_fill_stage_row(row: pd.Series) -> Dict[int, Optional[str]]:
    labels: Dict[int, Optional[str]] = {}
    current = None
    for col_idx, value in enumerate(row.tolist()):
        text = _cell_text(value)
        detected = canonical_stage_label(value)
        if detected:
            current = detected
        elif text:
            # Stop the stage span when a non-stage header appears so trailing
            # note/recon columns stay available as remarks but do not get
            # treated as additional financial value columns.
            current = None
        labels[col_idx] = current
    return labels


def _stage_row_indices(df: pd.DataFrame) -> List[int]:
    return stage_row_indices(df, parse_date)


def _block_title_for_stage_row(df: pd.DataFrame, stage_row_idx: int, sheet_name: str) -> str:
    for row_idx in range(stage_row_idx - 1, max(-1, stage_row_idx - 5), -1):
        texts = [_cell_text(value) for value in df.iloc[row_idx].tolist()]
        texts = [text for text in texts if text]
        if not texts:
            continue
        if any(canonical_stage_label(text) for text in texts):
            continue
        if any(parse_date(text) for text in texts):
            continue
        lowered = " ".join(text.lower() for text in texts)
        if contains_unit_marker(lowered, _UNIT_MARKERS):
            continue
        return " ".join(texts)
    return sheet_name


def _extract_entity_name_from_block_title(title: str) -> Optional[str]:
    if not title:
        return None
    parts = re.split(r"\s[-–]\s", title, maxsplit=1)
    if len(parts) != 2:
        return None
    candidate = parts[1].strip()
    if not candidate:
        return None
    if candidate.lower().startswith("project "):
        return None
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", candidate):
        return None
    return candidate


def _is_strict_entity_title_match(block_title: str, block_entity_name: Optional[str], entity_name: str) -> bool:
    normalized_entity = str(entity_name or "").strip().lower()
    if not normalized_entity:
        return False

    normalized_block_entity = str(block_entity_name or "").strip().lower()
    if normalized_block_entity:
        return normalized_block_entity == normalized_entity

    normalized_title = str(block_title or "").strip().lower()
    pattern = rf"(^|[\s\-–_/()]){re.escape(normalized_entity)}($|[\s\-–_/()])"
    return bool(re.search(pattern, normalized_title))


def _select_entity_block(
    df: pd.DataFrame,
    sheet_name: str,
    default_stage_row_idx: int,
    entity_name: Optional[str],
) -> Dict[str, Any]:
    candidates = _stage_row_indices(df)
    if not candidates:
        candidates = [default_stage_row_idx]

    blocks: List[Dict[str, Any]] = []
    for index, candidate_stage_row_idx in enumerate(candidates):
        next_stage_row_idx = candidates[index + 1] if index + 1 < len(candidates) else len(df)
        block_title = _block_title_for_stage_row(df, candidate_stage_row_idx, sheet_name)
        block_entity_name = _extract_entity_name_from_block_title(block_title)
        block_context = " ".join(
            _cell_text(value).lower()
            for value in df.iloc[max(0, candidate_stage_row_idx - 5) : min(len(df), candidate_stage_row_idx + 2)].to_numpy(dtype=object, copy=False).ravel()
        )
        blocks.append(
            {
                "stage_row_idx": candidate_stage_row_idx,
                "date_row_idx": _local_date_row_index(df, candidate_stage_row_idx),
                "data_end_row": next_stage_row_idx,
                "block_title": block_title,
                "block_entity_name": block_entity_name,
                "context": block_context,
            }
        )

    default_block = next((block for block in blocks if block["stage_row_idx"] == default_stage_row_idx), blocks[0])
    if not entity_name:
        return {**default_block, "strict_entity_match": False}

    entity_text = str(entity_name).strip().lower()
    strict_matches = [
        block
        for block in blocks
        if _is_strict_entity_title_match(
            block_title=str(block["block_title"]),
            block_entity_name=block.get("block_entity_name"),
            entity_name=entity_text,
        )
    ]
    if strict_matches:
        working_blocks = strict_matches
        strict_entity_match = True
    else:
        working_blocks = blocks
        strict_entity_match = False

    entity_parts = [part for part in re.split(r"\s+", entity_text) if len(part) >= 3]
    best_block = default_block
    best_score = -1
    for block in working_blocks:
        score = 0
        title_lower = str(block["block_title"]).lower()
        entity_lower = str(block.get("block_entity_name") or "").lower()
        if entity_text and entity_text == entity_lower:
            score += 10
        if entity_text and entity_text in title_lower:
            score += 6
        if entity_text and entity_text in block["context"]:
            score += 4
        score += sum(1 for part in entity_parts if part in title_lower)
        score += sum(1 for part in entity_parts if part in block["context"])
        if score > best_score:
            best_score = score
            best_block = block
    return {**best_block, "strict_entity_match": strict_entity_match}


def _local_date_row_index(df: pd.DataFrame, stage_row_idx: int) -> int:
    return date_row_index(df, stage_row_idx, parse_date, max_distance=2) or stage_row_idx


def _rollforward_header_row_index(df: pd.DataFrame, desc_col_idx: int) -> Optional[int]:
    for row_idx in range(min(len(df), 12)):
        desc_cell = _cell_text(df.iloc[row_idx, desc_col_idx]) if desc_col_idx < len(df.columns) else ""
        cells = [_cell_text(value) for value in df.iloc[row_idx, desc_col_idx + 1 :].tolist()]
        if not cells:
            continue
        has_unit = contains_unit_marker(desc_cell, _UNIT_MARKERS) or any(
            contains_unit_marker(cell, _UNIT_MARKERS) for cell in cells if cell
        )
        has_component_header = any(
            cell
            and not parse_date(cell)
            and not contains_unit_marker(cell, _UNIT_MARKERS)
            for cell in cells
        )
        if has_unit and has_component_header:
            return row_idx
    return None


def _standardize_rollforward_schedule_df(
    df: pd.DataFrame,
    sheet_name: str,
    profile: Dict[str, Any],
    desc_col_idx: int,
) -> Optional[pd.DataFrame]:
    carrying_row_idx = None
    for row_idx in range(len(df)):
        description = _cell_text(df.iloc[row_idx, desc_col_idx]).lower()
        if description in _CARRYING_AMOUNT_LABELS:
            carrying_row_idx = row_idx
            break
    if carrying_row_idx is None:
        return None

    header_row_idx = _rollforward_header_row_index(df, desc_col_idx)
    if header_row_idx is None:
        return None

    component_columns: List[Dict[str, Any]] = []
    for col_idx in range(desc_col_idx + 1, len(df.columns)):
        header = _cell_text(df.iloc[header_row_idx, col_idx])
        if not header:
            continue
        if parse_date(header):
            continue
        lowered = header.lower()
        if any(marker in lowered for marker in _UNIT_MARKERS):
            continue
        component_columns.append({"col_idx": col_idx, "header": header})
    if not component_columns:
        return None

    date_rows: List[Dict[str, Any]] = []
    for row_idx in range(carrying_row_idx + 1, len(df)):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        lowered = description.lower()
        if lowered.startswith("check"):
            break
        parsed_date = parse_date(description)
        if not parsed_date:
            continue
        date_rows.append(
            {
                "row_idx": row_idx,
                "date": parsed_date.strftime("%Y-%m-%d"),
            }
        )
    if len(date_rows) < 2:
        return None

    usable_components: List[Dict[str, Any]] = []
    for component in component_columns:
        if any(_coerce_numeric(df.iloc[item["row_idx"], component["col_idx"]]) is not None for item in date_rows):
            usable_components.append(component)
    if not usable_components:
        return None

    block_title = str(profile.get("title") or sheet_name).strip() or sheet_name
    temp_rows: List[List[Any]] = []
    temp_rows.append([block_title])
    temp_rows.append(["", *([PREFERRED_STAGE] * len(date_rows))])
    temp_rows.append([block_title, *[item["date"] for item in date_rows]])
    for component in usable_components:
        temp_rows.append(
            [
                component["header"],
                *[_coerce_numeric(df.iloc[item["row_idx"], component["col_idx"]]) for item in date_rows],
            ]
        )
    return pd.DataFrame(temp_rows)


def _is_numeric_enough(df: pd.DataFrame, col_idx: int, data_start_row: int) -> bool:
    values = [_coerce_numeric(value) for value in df.iloc[data_start_row:, col_idx].tolist()]
    non_empty = [value for value in values if value is not None]
    return len(non_empty) >= 1


def _dedupe_columns_by_key(columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen_keys: set[str] = set()
    for column in columns:
        key = str(column.get("key") or "")
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        deduped.append(column)
    return deduped


def _row_type(description: str) -> str:
    lowered = description.lower()
    if lowered.startswith("of which") or lowered.startswith("其中"):
        return "breakdown"
    if any(keyword in lowered for keyword in _SUBTOTAL_KEYWORDS):
        return "subtotal"
    if any(keyword in lowered for keyword in _TOTAL_KEYWORDS):
        return "total"
    if re.match(r"^\d{6,}", description):
        return "breakdown"
    if any(token in description for token in ("有限公司", "股份", "基金", "公司")):
        return "breakdown"
    return "detail"


def _detect_implicit_breakdowns_from_sum(
    row_entries: List[Dict[str, Any]],
    projection_column_key: str,
    tolerance_ratio: float = 0.005,
) -> None:
    """
    Detect rows that are implicit breakdowns of a parent detail row.

    In some Excel schedules the parent total appears first (without a "total"
    keyword) followed by its component sub-rows, e.g.:

        Cash at bank:  1,000        ← parent detail row
          Bank A CNY:    600        ← sub-row (no "其中" / "of which" prefix)
          Bank B USD:    400        ← sub-row

    When the sum of consecutive following detail rows equals the parent's value
    within a small tolerance AND every child is strictly smaller than the parent,
    those following rows are re-classified as "breakdown" so they are filtered
    out of the AI prompt data and reconciliation totals.

    This only runs when there are NO explicitly-typed total/subtotal rows in the
    schedule (i.e. when keyword-based detection has already handled the standard
    case).  Mutates row_entries in-place.
    """
    n = len(row_entries)
    # Only apply when no explicit total/subtotal rows exist
    if any(r["row_type"] in ("total", "subtotal") for r in row_entries):
        return

    marked: set[int] = set()
    for i in range(n - 1):
        if i in marked:
            continue
        parent = row_entries[i]
        if parent["row_type"] != "detail":
            continue
        parent_val = parent["values"].get(projection_column_key)
        if parent_val is None or abs(parent_val) < 1.0:
            continue

        running_sum = 0.0
        child_candidates: List[int] = []
        for j in range(i + 1, min(i + 20, n)):
            if j in marked:
                break
            child = row_entries[j]
            if child["row_type"] != "detail":
                break
            child_val = child["values"].get(projection_column_key) or 0.0
            # Each child must be strictly smaller than the parent
            if abs(child_val) >= abs(parent_val):
                break
            running_sum += child_val
            child_candidates.append(j)
            if len(child_candidates) >= 2:
                tolerance = max(1.0, abs(parent_val) * tolerance_ratio)
                if abs(running_sum - parent_val) <= tolerance:
                    for idx in child_candidates:
                        row_entries[idx]["row_type"] = "breakdown"
                        marked.add(idx)
                    break


def _fallback_description(description: str, title: str, last_label: Optional[str]) -> str:
    if description:
        return description
    if last_label:
        return last_label
    return title


def _looks_like_supporting_note(text: str) -> bool:
    lowered = text.lower()
    if not text or len(text) < 8:
        return False
    if canonical_stage_label(text) or parse_date(text):
        return False
    if any(marker in lowered for marker in _UNIT_MARKERS):
        return False
    return True


def _looks_like_working_remark(text: str) -> bool:
    lowered = str(text or "").strip().lower()
    if not lowered:
        return False
    return any(keyword in lowered for keyword in _WORKING_REMARK_KEYWORDS)


def _build_working_remark_note(description: str, values: Dict[str, Optional[float]]) -> Optional[str]:
    label = str(description or "").strip()
    if not _looks_like_working_remark(label):
        return None

    non_zero_parts: List[str] = []
    seen_period_values: set[tuple[str, float]] = set()
    for key, value in values.items():
        if value is None:
            continue
        if abs(float(value)) <= 0:
            continue
        period = str(key).split("|", 1)[-1]
        dedupe_key = (period, float(value))
        if dedupe_key in seen_period_values:
            continue
        seen_period_values.add(dedupe_key)
        non_zero_parts.append(f"{period}: {value:,.0f}")

    if non_zero_parts:
        return f"{label} | " + " | ".join(non_zero_parts)
    return label


def _trim_block_end_row(
    df: pd.DataFrame,
    desc_col_idx: int,
    data_start_row: int,
    data_end_row: int,
    stage_row_idx: Optional[int],
) -> int:
    trimmed_end_row = int(data_end_row or len(df))
    if trimmed_end_row <= data_start_row:
        return trimmed_end_row

    while trimmed_end_row > data_start_row:
        row_idx = trimmed_end_row - 1
        row_values = df.iloc[row_idx].tolist()
        text_values = [_cell_text(value) for value in row_values if _cell_text(value)]
        if not text_values:
            trimmed_end_row -= 1
            continue

        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        has_numeric = any(_coerce_numeric(value) is not None for value in row_values)
        if has_numeric:
            break

        if stage_row_idx is not None and row_idx == stage_row_idx - 1:
            trimmed_end_row -= 1
            continue

        if description and (_looks_like_supporting_note(description) or _looks_like_working_remark(description)):
            break

        if len(text_values) <= 2 and not any(parse_date(text) or canonical_stage_label(text) for text in text_values):
            trimmed_end_row -= 1
            continue

        break

    return trimmed_end_row


def _extract_supporting_notes(
    df: pd.DataFrame,
    desc_col_idx: int,
    columns: List[Dict[str, Any]],
    data_start_row: int,
    data_end_row: int,
) -> List[str]:
    if not columns:
        return []

    notes: List[str] = []
    seen: set[str] = set()
    max_numeric_col_idx = max(column["col_idx"] for column in columns)

    for row_idx in range(data_start_row, data_end_row):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        numeric_hits = [
            _coerce_numeric(df.iloc[row_idx, column["col_idx"]]) is not None
            for column in columns
        ]
        extra_text_cells = [
            _cell_text(value)
            for value in df.iloc[row_idx, max_numeric_col_idx + 1 :].tolist()
            if _looks_like_supporting_note(_cell_text(value))
        ]

        if extra_text_cells:
            note = " | ".join(
                [part for part in [description, *extra_text_cells] if part and not parse_date(part)]
            ).strip()
            if note and note not in seen:
                notes.append(note)
                seen.add(note)
                continue

        if description and not any(numeric_hits) and _looks_like_supporting_note(description):
            if description not in seen:
                notes.append(description)
                seen.add(description)
            continue

        if not any(numeric_hits):
            row_text_cells = [
                _cell_text(value)
                for value in df.iloc[row_idx].tolist()
                if _looks_like_supporting_note(_cell_text(value))
            ]
            if row_text_cells:
                note = " | ".join(dict.fromkeys(row_text_cells))
                if note not in seen:
                    notes.append(note)
                    seen.add(note)

    return notes[:8]


def _extract_auxiliary_check_totals(
    df: pd.DataFrame,
    columns: List[Dict[str, Any]],
    multiplier: int,
) -> Dict[str, float]:
    if not columns:
        return {}

    max_financial_col_idx = max(column["col_idx"] for column in columns)
    trailing_start_col_idx = max_financial_col_idx + 1
    if trailing_start_col_idx >= len(df.columns):
        return {}

    for desc_col_idx in range(trailing_start_col_idx, len(df.columns)):
        for row_idx in range(len(df)):
            label = _cell_text(df.iloc[row_idx, desc_col_idx]).lower()
            if label != "check":
                continue

            header_row_idx = None
            for candidate_row_idx in range(row_idx - 1, -1, -1):
                date_hits = sum(
                    1
                    for col_idx in range(desc_col_idx + 1, len(df.columns))
                    if parse_date(df.iloc[candidate_row_idx, col_idx])
                )
                if date_hits >= 1:
                    header_row_idx = candidate_row_idx
                    break
            if header_row_idx is None:
                continue

            totals_by_date: Dict[str, float] = {}
            for col_idx in range(desc_col_idx + 1, len(df.columns)):
                parsed_date = parse_date(df.iloc[header_row_idx, col_idx])
                if not parsed_date:
                    continue
                value = _coerce_numeric(df.iloc[row_idx, col_idx])
                if value is None:
                    continue
                totals_by_date[parsed_date.strftime("%Y-%m-%d")] = round(value * multiplier, 0)

            if totals_by_date:
                return totals_by_date

    return {}


def _auxiliary_header_context(
    df: pd.DataFrame,
    block_title: str,
    stage_row_idx: int,
    date_row_idx: int,
    col_idx: int,
) -> Dict[str, str]:
    stage_text = _cell_text(df.iloc[stage_row_idx, col_idx]) if stage_row_idx is not None else ""
    date_text = _cell_text(df.iloc[date_row_idx, col_idx]) if date_row_idx is not None else ""

    header_text = ""
    for row_idx in (date_row_idx, stage_row_idx):
        text = _cell_text(df.iloc[row_idx, col_idx])
        lowered = text.lower()
        if not text:
            continue
        if canonical_stage_label(text) or parse_date(text):
            continue
        if any(marker in lowered for marker in _UNIT_MARKERS):
            continue
        header_text = text
        break

    return {
        "table_header": str(block_title or "").strip(),
        "stage_header": stage_text,
        "date_header": date_text,
        "header": header_text or f"Detail {col_idx}",
    }


def _extract_adjacent_detail_columns(
    df: pd.DataFrame,
    block_title: str,
    desc_col_idx: int,
    columns: List[Dict[str, Any]],
    stage_row_idx: int,
    date_row_idx: int,
    data_start_row: int,
    data_end_row: int,
    max_columns: int = 5,
) -> List[Dict[str, Any]]:
    if not columns:
        return []

    numeric_col_indices = {column["col_idx"] for column in columns}
    max_numeric_col_idx = max(numeric_col_indices)
    candidates: List[Dict[str, Any]] = []
    for col_idx in range(max_numeric_col_idx + 1, min(len(df.columns), max_numeric_col_idx + max_columns + 1)):
        if col_idx in numeric_col_indices:
            continue
        text_values = []
        for row_idx in range(data_start_row, data_end_row):
            value = df.iloc[row_idx, col_idx]
            text = _cell_text(value)
            if text:
                text_values.append(text)
        if not text_values:
            continue
        candidates.append(
            {
                "col_idx": col_idx,
                **_auxiliary_header_context(df, block_title, stage_row_idx, date_row_idx, col_idx),
            }
        )
    unique_candidates: List[Dict[str, Any]] = []
    seen_headers: Dict[str, int] = {}
    for candidate in candidates[:max_columns]:
        base_header = str(candidate.get("header") or f"Detail {candidate.get('col_idx')}").strip()
        if not base_header:
            base_header = f"Detail {candidate.get('col_idx')}"
        seen_headers[base_header] = seen_headers.get(base_header, 0) + 1
        header_occurrence = seen_headers[base_header]
        unique_header = base_header if header_occurrence == 1 else f"{base_header} ({header_occurrence})"
        unique_candidates.append(
            {
                **candidate,
                "header": unique_header,
            }
        )
    return unique_candidates


def _build_table_linked_remarks(
    supporting_notes: List[str],
    adjacent_detail_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    linked_remarks: List[Dict[str, Any]] = []
    seen: set[str] = set()

    for note in supporting_notes:
        text = str(note or "").strip()
        if not text:
            continue
        dedupe_key = f"row_note::{text}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        linked_remarks.append(
            {
                "source": "row_note",
                "summary": text,
            }
        )

    for row in adjacent_detail_rows:
        if not isinstance(row, dict):
            continue
        description = str(row.get("Description") or "").strip()
        remarks: List[Dict[str, str]] = []
        for key, value in row.items():
            key_text = str(key)
            if (
                key_text == INTERNAL_ROW_KEY
                or key_text == "Description"
                or key_text.endswith("| table_header")
                or key_text.endswith("| indicative_adjusted_row")
                or key_text.endswith("| date_row")
            ):
                continue
            text = str(value or "").strip()
            if not text:
                continue
            remarks.append(
                {
                    "header": key_text,
                    "value": text,
                    "table_header": str(row.get(f"{key_text} | table_header") or "").strip(),
                    "indicative_adjusted_row": str(row.get(f"{key_text} | indicative_adjusted_row") or "").strip(),
                    "date_row": str(row.get(f"{key_text} | date_row") or "").strip(),
                }
            )
        if not remarks:
            continue
        summary = " | ".join(
            part for part in [description, "; ".join(f"{item['header']}: {item['value']}" for item in remarks)] if part
        ).strip()
        dedupe_key = f"rhs::{summary}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        linked_remarks.append(
            {
                "source": "rhs_columns",
                "description": description,
                "summary": summary,
                "remarks": remarks,
            }
        )

    return linked_remarks


def _build_prompt_analysis_df(
    block_title: str,
    columns: List[Dict[str, Any]],
    row_entries: List[Dict[str, Any]],
    analysis_stage: str,
) -> pd.DataFrame:
    analysis_columns = sorted(
        [column for column in columns if column["stage"] == analysis_stage],
        key=lambda column: column["date"],
    )
    if not analysis_columns:
        return pd.DataFrame(columns=[block_title])

    prompt_rows: List[Dict[str, Any]] = []
    for row in row_entries:
        if row["row_type"] == "breakdown":
            continue
        row_values = {
            column["date"]: row["values"].get(column["key"])
            for column in analysis_columns
        }
        if row["row_type"] == "detail" and not any(value is not None for value in row_values.values()):
            continue
        prompt_rows.append(
            {
                block_title: row["description"],
                INTERNAL_ROW_KEY: row["row_idx"],
                **{key: (0 if value is None else value) for key, value in row_values.items()},
            }
        )

    if not prompt_rows:
        return pd.DataFrame(columns=[block_title, INTERNAL_ROW_KEY, *[column["date"] for column in analysis_columns]])
    return pd.DataFrame(prompt_rows)


def _multiply_factor(profile: Dict[str, Any]) -> int:
    markers = [str(marker).lower() for marker in profile.get("unit_markers") or []]
    if any("cny'000" in marker or "千元" in marker for marker in markers):
        return 1000
    return 1


def _choose_projection(columns: List[Dict[str, Any]], row_entries: List[Dict[str, Any]]) -> Dict[str, Any]:
    def non_zero_score(column: Dict[str, Any]) -> float:
        total = 0.0
        for row in row_entries:
            value = row["values"].get(column["key"])
            if value is not None:
                total += abs(value)
        return total

    stage_priority = [PREFERRED_STAGE, "Audited", "Mgt acc", "Audit adjustment", "Indicative adjustment"]
    sorted_columns = sorted(
        columns,
        key=lambda column: (
            stage_priority.index(column["stage"]) if column["stage"] in stage_priority else len(stage_priority),
            column["date"],
        ),
    )

    preferred_candidates = [column for column in sorted_columns if column["stage"] == PREFERRED_STAGE]
    preferred_candidates.sort(key=lambda column: column["date"], reverse=True)
    for column in preferred_candidates:
        if non_zero_score(column) > 0:
            return {
                "preferred_stage": PREFERRED_STAGE,
                "effective_stage": column["stage"],
                "effective_date": column["date"],
                "column": column,
            }

    fallback_candidates = sorted(sorted_columns, key=lambda column: column["date"], reverse=True)
    for column in fallback_candidates:
        if non_zero_score(column) > 0:
            return {
                "preferred_stage": PREFERRED_STAGE,
                "effective_stage": column["stage"],
                "effective_date": column["date"],
                "column": column,
            }

    column = preferred_candidates[0] if preferred_candidates else fallback_candidates[0]
    return {
        "preferred_stage": PREFERRED_STAGE,
        "effective_stage": column["stage"],
        "effective_date": column["date"],
        "column": column,
    }


def normalize_financial_schedule(
    workbook_path: str,
    sheet_name: str,
    profile: Optional[Dict[str, Any]] = None,
    entity_name: Optional[str] = None,
    sheet_df: Optional[pd.DataFrame] = None,
    statement_type: Optional[str] = None,
) -> Dict[str, Any]:
    df = sheet_df if sheet_df is not None else pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, engine="openpyxl")
    profile = profile or profile_sheet(df, sheet_name)

    desc_col_idx = _find_description_column(df)
    if desc_col_idx is None:
        raise ValueError(f"Unable to detect description column for sheet: {sheet_name}")

    stage_row_idx = profile.get("stage_row_idx")
    date_row_idx = profile.get("date_row_idx")
    if stage_row_idx is None or date_row_idx is None:
        standardized_rollforward_df = _standardize_rollforward_schedule_df(
            df=df,
            sheet_name=sheet_name,
            profile=profile,
            desc_col_idx=desc_col_idx,
        )
        if standardized_rollforward_df is not None:
            standardized_profile = profile_sheet(standardized_rollforward_df, sheet_name)
            standardized_profile["unit_markers"] = profile.get("unit_markers") or standardized_profile.get("unit_markers")
            standardized_profile["sheet_kind"] = "financial_schedule"
            return normalize_financial_schedule(
                workbook_path=workbook_path,
                sheet_name=sheet_name,
                profile=standardized_profile,
                entity_name=entity_name,
                sheet_df=standardized_rollforward_df,
                statement_type=statement_type,
            )
        raise ValueError(f"Unable to detect stage/date rows for sheet: {sheet_name}")
    selected_block = _select_entity_block(df, sheet_name, stage_row_idx, entity_name)
    stage_row_idx = selected_block["stage_row_idx"]
    date_row_idx = selected_block["date_row_idx"]
    block_title = str(selected_block.get("block_title") or profile.get("title") or sheet_name)
    block_entity_name = selected_block.get("block_entity_name")
    strict_entity_match = bool(selected_block.get("strict_entity_match"))

    stage_map = _forward_fill_stage_row(df.iloc[stage_row_idx])
    data_start_row = max(stage_row_idx, date_row_idx) + 1
    raw_data_end_row = int(selected_block.get("data_end_row") or len(df))
    data_end_row = _trim_block_end_row(
        df=df,
        desc_col_idx=desc_col_idx,
        data_start_row=data_start_row,
        data_end_row=raw_data_end_row,
        stage_row_idx=raw_data_end_row if raw_data_end_row < len(df) else None,
    )
    columns: List[Dict[str, Any]] = []
    for col_idx in range(desc_col_idx + 1, len(df.columns)):
        stage = stage_map.get(col_idx)
        parsed_date = parse_date(df.iloc[date_row_idx, col_idx])
        if not stage or not parsed_date:
            continue
        if not _is_numeric_enough(df.iloc[:data_end_row], col_idx, data_start_row):
            continue
        columns.append(
            {
                "col_idx": col_idx,
                "stage": stage,
                "date": parsed_date.strftime("%Y-%m-%d"),
                "key": f"{stage}|{parsed_date.strftime('%Y-%m-%d')}",
            }
        )
    columns = _dedupe_columns_by_key(columns)

    if not columns:
        raise ValueError(f"No financial value columns detected for sheet: {sheet_name}")

    adjacent_detail_columns = _extract_adjacent_detail_columns(
        df=df,
        block_title=block_title,
        desc_col_idx=desc_col_idx,
        columns=columns,
        stage_row_idx=stage_row_idx,
        date_row_idx=date_row_idx,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
    )

    multiplier = _multiply_factor(profile)
    auxiliary_check_totals_by_date = _extract_auxiliary_check_totals(
        df=df,
        columns=columns,
        multiplier=multiplier,
    )
    row_entries: List[Dict[str, Any]] = []
    working_remark_notes: List[str] = []
    last_label = None
    for row_idx in range(data_start_row, data_end_row):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        numeric_values = {column["key"]: _coerce_numeric(df.iloc[row_idx, column["col_idx"]]) for column in columns}
        has_numeric = any(value is not None for value in numeric_values.values())
        if not description and not has_numeric:
            continue
        if description:
            last_label = description
        effective_description = _fallback_description(description, block_title, last_label)
        if not effective_description:
            continue
        values = {
            key: (round(value * multiplier, 0) if value is not None else None)
            for key, value in numeric_values.items()
        }
        working_remark_note = _build_working_remark_note(effective_description, values)
        if working_remark_note:
            working_remark_notes.append(working_remark_note)
            continue
        row_entries.append(
            {
                "row_idx": row_idx,
                "description": effective_description,
                "row_type": _row_type(effective_description),
                "values": values,
            }
        )

    if not row_entries:
        raise ValueError(f"No data rows detected for sheet: {sheet_name}")

    projection = _choose_projection(columns, row_entries)
    # Detect implicit breakdown rows (parent-first structure without total keywords)
    # Must be called after _choose_projection so we know the projection column key.
    _detect_implicit_breakdowns_from_sum(row_entries, projection["column"]["key"])
    projection_column = projection["column"]
    analysis_stage = PREFERRED_STAGE if any(column["stage"] == PREFERRED_STAGE for column in columns) else projection["effective_stage"]
    prompt_analysis_df = _build_prompt_analysis_df(
        block_title=block_title,
        columns=columns,
        row_entries=row_entries,
        analysis_stage=analysis_stage,
    )
    trend_summary = build_trend_summary(prompt_analysis_df)
    significant_movements = build_significant_movements(prompt_analysis_df)
    supporting_notes = _extract_supporting_notes(df, desc_col_idx, columns, data_start_row, data_end_row)
    for note in working_remark_notes:
        if note not in supporting_notes:
            supporting_notes.append(note)
    annualization = infer_partial_year_annualization(
        statement_type=statement_type or "",
        available_dates=[column["date"] for column in columns],
        effective_date=projection["effective_date"],
    )
    original_column_label = projection_column["date"]
    annualized_column_label = (
        f"{projection_column['date']} annualised"
        if annualization.get("annualized")
        else projection_column["date"]
    )

    projection_rows_original: List[Dict[str, Any]] = []
    projection_rows_annualized: List[Dict[str, Any]] = []
    adjacent_detail_rows: List[Dict[str, Any]] = []
    projection_original_values_by_description: Dict[str, float] = {}
    projection_totals_by_date: Dict[str, float] = {}
    non_zero_rows = 0
    for row in row_entries:
        original_value = row["values"].get(projection_column["key"])
        annualized_value = original_value
        if annualized_value is not None and annualization.get("annualized") and annualization.get("factor"):
            annualized_value = round(annualized_value * float(annualization["factor"]), 0)
        if original_value is None and row["row_type"] == "detail":
            continue
        effective_value = annualized_value if annualization.get("annualized") else original_value
        if effective_value is not None and abs(effective_value) > 0:
            non_zero_rows += 1
        if original_value is not None:
            projection_original_values_by_description[row["description"]] = original_value
        projection_rows_original.append(
            {
                block_title: row["description"],
                INTERNAL_ROW_KEY: row["row_idx"],
                original_column_label: original_value if original_value is not None else 0,
            }
        )
        projection_rows_annualized.append(
            {
                block_title: row["description"],
                INTERNAL_ROW_KEY: row["row_idx"],
                annualized_column_label: annualized_value if annualized_value is not None else 0,
            }
        )
        adjacent_row = {
            INTERNAL_ROW_KEY: row["row_idx"],
            "Description": row["description"],
            annualized_column_label if annualization.get("annualized") else original_column_label: effective_value if effective_value is not None else 0,
        }
        has_adjacent_text = False
        for detail_column in adjacent_detail_columns:
            detail_value = _cell_text(df.iloc[row["row_idx"], detail_column["col_idx"]])
            adjacent_row[f"{detail_column['header']} | table_header"] = detail_column.get("table_header", "")
            adjacent_row[f"{detail_column['header']} | indicative_adjusted_row"] = detail_column.get("stage_header", "")
            adjacent_row[f"{detail_column['header']} | date_row"] = detail_column.get("date_header", "")
            adjacent_row[detail_column["header"]] = detail_value
            if detail_value:
                has_adjacent_text = True
        if has_adjacent_text:
            adjacent_detail_rows.append(adjacent_row)
        if row["row_type"] == "total":
            if original_value is not None:
                projection_totals_by_date[original_column_label] = float(original_value)
            if annualized_value is not None:
                projection_totals_by_date[annualized_column_label] = float(annualized_value)

    table_linked_remarks = _build_table_linked_remarks(
        supporting_notes=supporting_notes,
        adjacent_detail_rows=adjacent_detail_rows,
    )

    projection_df_original = pd.DataFrame(projection_rows_original)
    if projection_df_original.empty:
        projection_df_original = pd.DataFrame(columns=[block_title, original_column_label])

    projection_df_annualized = pd.DataFrame(projection_rows_annualized)
    if projection_df_annualized.empty:
        projection_df_annualized = pd.DataFrame(columns=[block_title, annualized_column_label])

    projection_df = projection_df_annualized.copy() if annualization.get("annualized") else projection_df_original.copy()
    if projection_df.empty:
        projection_df = pd.DataFrame(columns=[block_title, annualized_column_label if annualization.get("annualized") else original_column_label])

    shared_attrs = {
        "preferred_stage": projection["preferred_stage"],
        "effective_stage": projection["effective_stage"],
        "effective_date": projection["effective_date"],
        "statement_type": statement_type or "",
        "sheet_name": sheet_name,
        "non_zero_rows": non_zero_rows,
        "row_type_counts": dict(Counter(row["row_type"] for row in row_entries)),
        "block_title": block_title,
        "block_entity_name": block_entity_name,
        "block_start_row": data_start_row,
        "block_end_row": data_end_row,
        "strict_entity_match": strict_entity_match,
        "annualized": bool(annualization.get("annualized")),
        "annualization_factor": annualization.get("factor"),
        "annualization_months": annualization.get("months"),
        "raw_effective_date": annualization.get("raw_effective_date"),
        "fiscal_year_end_month": annualization.get("fiscal_year_end_month"),
        "fiscal_year_end_day": annualization.get("fiscal_year_end_day"),
    }
    row_types_by_description = {
        row["description"]: row["row_type"] for row in row_entries
    }
    common_attrs = {
        "integrity": shared_attrs,
        "row_types_by_description": row_types_by_description,
        "supporting_notes": supporting_notes,
        "normalized_columns": columns,
        "source_multiplier": multiplier,
        "sheet_kind": profile.get("sheet_kind"),
        "entity_name": block_entity_name,
        "block_title": block_title,
        "annualization": annualization,
        "projection_original_column_label": original_column_label,
        "projection_annualized_column_label": annualized_column_label,
        "projection_original_values_by_description": projection_original_values_by_description,
        "projection_totals_by_date": projection_totals_by_date,
        "auxiliary_check_totals_by_date": auxiliary_check_totals_by_date,
        "adjacent_detail_columns": adjacent_detail_columns,
        "adjacent_detail_rows": adjacent_detail_rows,
        "table_linked_remarks": table_linked_remarks,
        "prompt_analysis_df": prompt_analysis_df,
        "trend_summary": trend_summary,
        "significant_movements": significant_movements,
        "prompt_analysis_label": (
            "All indicative adjusted periods"
            if analysis_stage == PREFERRED_STAGE
            else f"All {analysis_stage} periods"
        ),
        "prompt_analysis_stage": analysis_stage,
    }
    projection_df.attrs.update(common_attrs)
    projection_df_original.attrs.update(common_attrs)
    projection_df_annualized.attrs.update(common_attrs)
    # Exclude prompt_analysis_df from its own attrs to avoid circular reference
    # (deepcopy during .copy() would recurse infinitely on Python 3.13+)
    prompt_analysis_attrs = {k: v for k, v in common_attrs.items() if k != "prompt_analysis_df"}
    prompt_analysis_df.attrs.update(prompt_analysis_attrs)
    projection_df_original.attrs["selected_variant"] = "original"
    projection_df_annualized.attrs["selected_variant"] = "annualized"
    projection_df.attrs["selected_variant"] = "annualized" if annualization.get("annualized") else "original"
    prompt_analysis_df.attrs["selected_variant"] = "analysis"

    return {
        "sheet_name": sheet_name,
        "title": block_title,
        "entity_name": block_entity_name,
        "block_title": block_title,
        "profile": profile,
        "columns": columns,
        "row_entries": row_entries,
        "projection_df": projection_df,
        "projection_df_original": projection_df_original,
        "projection_df_annualized": projection_df_annualized,
        "prompt_analysis_df": prompt_analysis_df,
        "integrity": projection_df.attrs["integrity"],
    }
def _annualization_factor(column_name: str) -> float | None:
    match = re.match(r"^(\d+)M\d{2}$", str(column_name).strip(), flags=re.IGNORECASE)
    if not match:
        return None
    months = int(match.group(1))
    if months <= 0:
        return None
    return 12.0 / months


def _parse_statement_date_label(value: str) -> datetime | None:
    try:
        return datetime.strptime(str(value), "%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def infer_partial_year_annualization(
    statement_type: str,
    available_dates: list[str],
    effective_date: str,
) -> dict[str, float | int | bool | str | None]:
    metadata: dict[str, float | int | bool | str | None] = {
        "annualized": False,
        "factor": None,
        "months": None,
        "raw_effective_date": effective_date,
        "fiscal_year_end_month": None,
        "fiscal_year_end_day": None,
    }
    if statement_type != "IS":
        return metadata

    period_factor = _annualization_factor(str(effective_date))
    if period_factor is not None:
        months = round(12.0 / period_factor)
        metadata.update({"annualized": months < 12, "factor": period_factor, "months": months})
        return metadata

    parsed_effective = _parse_statement_date_label(effective_date)
    if parsed_effective is None:
        return metadata

    parsed_dates = sorted(
        {
            parsed
            for parsed in (_parse_statement_date_label(value) for value in available_dates)
            if parsed is not None
        }
    )
    if len(parsed_dates) < 2:
        return metadata
    if parsed_dates and parsed_effective != max(parsed_dates):
        return metadata
    previous_date = parsed_dates[-2]
    if (parsed_effective.month, parsed_effective.day) == (previous_date.month, previous_date.day):
        return metadata
    anchor_candidates = [date for date in parsed_dates if (date.month, date.day) != (parsed_effective.month, parsed_effective.day)]
    if anchor_candidates:
        dominant_anchor = max(
            {(date.month, date.day) for date in anchor_candidates},
            key=lambda month_day: sum((candidate.month, candidate.day) == month_day for candidate in anchor_candidates),
        )
    else:
        dominant_anchor = (12, 31)
    metadata.update(
        {
            "fiscal_year_end_month": dominant_anchor[0],
            "fiscal_year_end_day": dominant_anchor[1],
        }
    )

    if (parsed_effective.month, parsed_effective.day) == dominant_anchor:
        return metadata

    months = (parsed_effective.month - dominant_anchor[0]) % 12
    if months <= 0:
        months = parsed_effective.month
    if months >= 12:
        return metadata

    factor = 12.0 / float(months)
    metadata.update({"annualized": True, "factor": factor, "months": months})
    return metadata


def annualize_income_statement_df(df: pd.DataFrame) -> pd.DataFrame:
    annualized_columns = [df.columns[0]] if len(df.columns) > 0 else []
    annualized_data = {df.columns[0]: df.iloc[:, 0]} if len(df.columns) > 0 else {}

    for column in list(df.columns)[1:]:
        factor = _annualization_factor(str(column))
        if factor is None or not pd.api.types.is_numeric_dtype(df[column]):
            annualized_columns.append(column)
            annualized_data[column] = df[column]
            continue
        annualized_columns.append(f"{column} annualised")
        annualized_data[f"{column} annualised"] = df[column] * factor

    return pd.DataFrame(annualized_data, columns=annualized_columns)
# --- end workbook/schedules.py ---

# --- begin workbook/resolver.py ---
"""
Resolve workbook tabs to FDD mapping keys using workbook metadata and fuzzy aliases.
"""


from difflib import SequenceMatcher
import json
import os
import re
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import pandas as pd


def _normalize_label(text: str) -> str:
    if not text:
        return ""
    normalized = str(text).strip().lower()
    normalized = normalized.replace("&", " and ")
    normalized = re.sub(r"[\W_]+", " ", normalized, flags=re.UNICODE)
    normalized = re.sub(r"\bexpenses\b", "expense", normalized)
    normalized = re.sub(r"\bpayables\b", "payable", normalized)
    normalized = re.sub(r"\breceivables\b", "receivable", normalized)
    normalized = re.sub(r"\bproperties\b", "property", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _token_set(text: str) -> set[str]:
    normalized = _normalize_label(text)
    return {token for token in normalized.split(" ") if token}


def _is_compact_cjk_label(text: str) -> bool:
    normalized = _normalize_label(text)
    if not normalized or " " in normalized:
        return False
    return bool(re.fullmatch(r"[\u4e00-\u9fff]+", normalized))


def _candidate_strings(profile: Dict[str, Any]) -> List[str]:
    values = [profile.get("sheet_name", ""), profile.get("title", "")]
    return [value for value in values if value]


def _is_exact_alias_match(sheet_name: str, title: str, alias: Optional[str]) -> bool:
    alias_norm = _normalize_label(alias or "")
    if not alias_norm:
        return False
    return any(
        _normalize_label(candidate) == alias_norm
        for candidate in (sheet_name, title)
        if candidate
    )


def _score_candidate(candidate: str, alias: str) -> float:
    candidate_norm = _normalize_label(candidate)
    alias_norm = _normalize_label(alias)
    if not candidate_norm or not alias_norm:
        return 0.0
    if candidate_norm == alias_norm:
        return 100.0
    candidate_tokens = _token_set(candidate)
    alias_tokens = _token_set(alias)
    overlap = len(candidate_tokens & alias_tokens)
    score = 0.0
    if overlap:
        score += overlap * 18.0
        if alias_tokens and alias_tokens.issubset(candidate_tokens):
            score += 24.0
        if candidate_tokens and candidate_tokens.issubset(alias_tokens):
            score += 12.0
    if _is_compact_cjk_label(candidate_norm) and _is_compact_cjk_label(alias_norm) and candidate_norm != alias_norm:
        if alias_norm in candidate_norm or candidate_norm in alias_norm:
            score += 40.0
        else:
            return min(score, 24.0)
    ratio = SequenceMatcher(None, candidate_norm, alias_norm).ratio()
    if ratio >= 0.70:
        score += ratio * 40.0
    score += _semantic_alignment_adjustment(candidate_norm, alias_norm)
    return min(score, 95.0)


def _semantic_alignment_adjustment(candidate_norm: str, alias_norm: str) -> float:
    adjustment = 0.0

    def contains(text: str, phrase: str) -> bool:
        return phrase in text

    candidate_non_operating = contains(candidate_norm, "non operating")
    alias_non_operating = contains(alias_norm, "non operating")
    if candidate_non_operating != alias_non_operating and (
        contains(candidate_norm, "operating") or contains(alias_norm, "operating")
    ):
        adjustment -= 42.0

    candidate_non_current = contains(candidate_norm, "non current")
    alias_non_current = contains(alias_norm, "non current")
    candidate_current = contains(candidate_norm, "current")
    alias_current = contains(alias_norm, "current")
    if candidate_non_current != alias_non_current:
        adjustment -= 34.0
    elif candidate_current != alias_current and not candidate_non_current and not alias_non_current:
        adjustment -= 16.0

    candidate_other = bool(re.search(r"\bother\b", candidate_norm))
    alias_other = bool(re.search(r"\bother\b", alias_norm))
    if candidate_other != alias_other:
        adjustment -= 18.0

    candidate_payable = bool(re.search(r"\bpayable\b", candidate_norm))
    alias_payable = bool(re.search(r"\bpayable\b", alias_norm))
    if candidate_payable != alias_payable:
        adjustment -= 16.0

    candidate_surcharge = bool(re.search(r"\bsurcharge\b", candidate_norm))
    alias_surcharge = bool(re.search(r"\bsurcharge\b", alias_norm))
    if candidate_surcharge != alias_surcharge:
        adjustment -= 16.0

    candidate_by_customer = contains(candidate_norm, "by customer") or contains(candidate_norm, "customer")
    alias_by_customer = contains(alias_norm, "by customer") or contains(alias_norm, "customer")
    if candidate_by_customer and not alias_by_customer:
        adjustment -= 24.0

    candidate_receivable = bool(re.search(r"\breceivable\b", candidate_norm))
    alias_receivable = bool(re.search(r"\breceivable\b", alias_norm))
    if candidate_receivable != alias_receivable:
        adjustment -= 16.0

    candidate_notes = bool(re.search(r"\bnotes?\b", candidate_norm))
    alias_notes = bool(re.search(r"\bnotes?\b", alias_norm))
    if candidate_receivable and alias_receivable and candidate_notes != alias_notes:
        adjustment -= 20.0
    elif candidate_notes != alias_notes:
        adjustment -= 12.0

    return adjustment


def _sheet_type_bonus(profile: Dict[str, Any], mapping_type: str) -> float:
    sheet_kind = profile.get("sheet_kind")
    if sheet_kind == "financial_summary":
        return -100.0
    if sheet_kind == "support_schedule":
        return -20.0
    if profile.get("has_indicative_stage"):
        return 8.0
    return 0.0


def _best_sheet_for_mapping(mapping_key: str, config: Dict[str, Any], profiles: Dict[str, Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    aliases = list(config.get("aliases") or [])
    aliases.append(mapping_key)
    best: Optional[Dict[str, Any]] = None
    for sheet_name, profile in profiles.items():
        base_score = _sheet_type_bonus(profile, config.get("type", ""))
        if base_score <= -100:
            continue
        candidate_score = base_score
        candidate_alias = None
        for alias in aliases:
            for candidate in _candidate_strings(profile):
                score = _score_candidate(candidate, alias)
                if score > candidate_score:
                    candidate_score = score
                    candidate_alias = alias
        if candidate_score <= 0:
            continue
        current = {
            "sheet_name": sheet_name,
            "title": profile.get("title"),
            "score": round(candidate_score, 2),
            "matched_alias": candidate_alias,
            "sheet_kind": profile.get("sheet_kind"),
            "entity_scope": profile.get("entity_scope", "single"),
        }
        if best is None or current["score"] > best["score"]:
            best = current
    return best


def _candidate_sheets_for_mapping(mapping_key: str, config: Dict[str, Any], profiles: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    aliases = list(config.get("aliases") or [])
    aliases.append(mapping_key)
    candidates: List[Dict[str, Any]] = []
    for sheet_name, profile in profiles.items():
        base_score = _sheet_type_bonus(profile, config.get("type", ""))
        if base_score <= -100:
            continue
        best_score = 0.0
        matched_alias = None
        for alias in aliases:
            for candidate in _candidate_strings(profile):
                score = _score_candidate(candidate, alias)
                if score > best_score:
                    best_score = score
                    matched_alias = alias
        if best_score <= 0:
            continue
        total_score = best_score + base_score
        if total_score <= 0:
            continue
        candidates.append(
            {
                "sheet_name": sheet_name,
                "title": profile.get("title"),
                "score": round(total_score, 2),
                "matched_alias": matched_alias,
                "exact_alias_match": _is_exact_alias_match(sheet_name, profile.get("title"), matched_alias),
                "sheet_kind": profile.get("sheet_kind"),
                "entity_scope": profile.get("entity_scope", "single"),
                "mapping_key": mapping_key,
                "category": config.get("category"),
                "type": config.get("type"),
            }
        )
    return sorted(candidates, key=lambda item: item["score"], reverse=True)


def should_use_ai_for_candidates(
    candidates: List[Dict[str, Any]],
    score_gap_threshold: float = 3.0,
) -> bool:
    if len(candidates) < 2:
        return False
    top = candidates[0]
    second = candidates[1]
    gap = float(top.get("score", 0)) - float(second.get("score", 0))
    names_blob = " ".join(str(candidate.get("sheet_name", "")) for candidate in candidates[:3]).lower()
    has_version_signal = bool(re.search(r"\bv\d+\b|version|\bcopy\b", names_blob))
    same_alias = top.get("matched_alias") == second.get("matched_alias")
    return gap <= score_gap_threshold or (has_version_signal and same_alias)


def _extract_sheet_names_from_ai_response(content: str, candidates: List[Dict[str, Any]]) -> List[str]:
    if not content:
        return []
    candidate_names = [str(candidate["sheet_name"]) for candidate in candidates]
    stripped = content.strip()
    try:
        parsed = json.loads(stripped)
        if isinstance(parsed, dict):
            chosen = parsed.get("sheet_name") or parsed.get("selected_sheet")
            if chosen in candidate_names:
                return [chosen]
            chosen_list = parsed.get("sheet_names") or parsed.get("candidate_sheets")
            if isinstance(chosen_list, list):
                return [name for name in chosen_list if name in candidate_names]
        if isinstance(parsed, list):
            return [name for name in parsed if name in candidate_names]
    except Exception:
        pass

    matched_names: List[str] = []
    for candidate_name in candidate_names:
        if candidate_name in stripped and candidate_name not in matched_names:
            matched_names.append(candidate_name)
    return matched_names


def _pick_financial_summary_sheet(profiles: Dict[str, Dict[str, Any]]) -> Optional[str]:
    summary_candidates = [
        (sheet_name, profile)
        for sheet_name, profile in profiles.items()
        if profile.get("sheet_kind") == "financial_summary"
    ]
    if not summary_candidates:
        return None

    def score(item: Tuple[str, Dict[str, Any]]) -> tuple[int, str]:
        sheet_name, profile = item
        text = f"{sheet_name} {profile.get('title', '')}".lower()
        points = 0
        if "financial" in text:
            points += 3
        if "balance" in text or "income" in text or "profit" in text:
            points += 2
        if "bs" in text or "is" in text:
            points += 1
        return (-points, sheet_name.lower())

    return sorted(summary_candidates, key=score)[0][0]


def _statement_values_for_mapping(statement_df: Any, mapping_key: str, config: Dict[str, Any]) -> Dict[str, float]:
    if statement_df is None or getattr(statement_df, "empty", True):
        return {}
    aliases = [mapping_key, *(config.get("aliases") or [])]
    desc_col = statement_df.columns[0]
    best_row = None
    best_score = 0.0
    for _, row in statement_df.iterrows():
        description = str(row[desc_col]).strip()
        if not description:
            continue
        row_score = max(_score_candidate(description, alias) for alias in aliases if alias)
        if row_score > best_score:
            best_score = row_score
            best_row = row
    if best_row is None or best_score <= 0:
        return {}

    values: Dict[str, float] = {}
    for col in statement_df.columns[1:]:
        try:
            numeric_value = float(best_row[col])
        except Exception:
            continue
        if col:
            values[str(col)] = numeric_value
    return values


def _normalized_total_values(normalized: Dict[str, Any]) -> Dict[str, float]:
    columns = normalized.get("columns") or []
    row_entries = normalized.get("row_entries") or []
    total_entry = next((row for row in row_entries if row.get("row_type") == "total"), None)
    if total_entry is None:
        return {}
    return {
        column["date"]: float(total_entry["values"][column["key"]])
        for column in columns
        if total_entry["values"].get(column["key"]) is not None
    }


def _financial_context_for_workbook(
    workbook_path: str,
    profiles: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    summary_sheet = _pick_financial_summary_sheet(profiles)
    if not summary_sheet:
        return {}
    try:
        financial_results = extract_balance_sheet_and_income_statement(workbook_path, summary_sheet, debug=False)
    except Exception:
        return {}
    return {
        "summary_sheet": summary_sheet,
        "financial_results": financial_results,
    }


def _build_financial_reference_context(
    workbook_path: str,
    profiles: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    summary_sheet = _pick_financial_summary_sheet(profiles)
    if not summary_sheet:
        return {
            "summary_sheet": None,
            "financial_results": {},
            "reference_available": False,
            "reference_stage": "missing_financial_summary",
        }
    try:
        financial_results = extract_balance_sheet_and_income_statement(workbook_path, summary_sheet, debug=False)
    except Exception as exc:
        return {
            "summary_sheet": summary_sheet,
            "financial_results": {},
            "reference_available": False,
            "reference_stage": "financial_summary_error",
            "reference_error": str(exc),
        }
    return {
        "summary_sheet": summary_sheet,
        "financial_results": financial_results,
        "reference_available": True,
        "reference_stage": "financial_summary_loaded",
    }


def _is_summary_account_candidate(description: str) -> bool:
    text = str(description or "").strip()
    if not text:
        return False
    return not should_skip_account_label(text)


def _iter_financial_reference_rows(financial_context: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    financial_results = financial_context.get("financial_results") or {}
    for statement_type, dataframe in (
        ("BS", financial_results.get("balance_sheet")),
        ("IS", financial_results.get("income_statement")),
    ):
        if dataframe is None or getattr(dataframe, "empty", True):
            continue
        desc_col = dataframe.columns[0]
        for _, row in dataframe.iterrows():
            description = str(row.get(desc_col) or "").strip()
            if not _is_summary_account_candidate(description):
                continue
            values: Dict[str, float] = {}
            for col in dataframe.columns[1:]:
                try:
                    values[str(col)] = float(row[col])
                except Exception:
                    continue
            rows.append(
                {
                    "account_name": description,
                    "statement_type": statement_type,
                    "values": values,
                }
            )
    return rows


def _infer_accounting_category(account_name: str, statement_type: str) -> str:
    normalized = _normalize_label(account_name)
    if statement_type == "IS":
        if any(token in normalized for token in ("income", "revenue", "sale", "gain", "收益", "收入")):
            return "Revenue"
        return "Expenses"

    if any(token in normalized for token in ("capital", "reserve", "earnings", "equity", "股本", "资本", "留存", "未分配利润")):
        return "Equity"
    if any(token in normalized for token in ("loan", "borrowing", "payable", "liability", "tax", "借款", "应付", "负债", "負債")):
        if any(token in normalized for token in ("long term", "long-term", "non current", "non-current", "长期", "非流动", "非流動")):
            return "Non-current liabilities"
        return "Current liabilities"
    if any(token in normalized for token in ("property", "fixed asset", "intangible", "deferred", "investment", "长期", "非流动", "非流動", "固定资产", "固定資產", "无形资产", "無形資產", "投资")):
        return "Non-current assets"
    return "Current assets"


def _build_dynamic_mapping_config(account_name: str, statement_type: str) -> Dict[str, Any]:
    return {
        "type": statement_type,
        "category": _infer_accounting_category(account_name, statement_type),
        "aliases": [account_name],
        "dynamic_mapping": True,
        "accounting_nature": _infer_accounting_category(account_name, statement_type),
    }


def _build_sheet_candidate_for_account(
    account_name: str,
    config: Dict[str, Any],
    sheet_name: str,
    profiles: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    profile = profiles[sheet_name]
    return {
        "sheet_name": sheet_name,
        "title": profile.get("title"),
        "score": 108.0 if _is_exact_alias_match(sheet_name, profile.get("title"), account_name) else 95.0,
        "matched_alias": account_name,
        "exact_alias_match": _is_exact_alias_match(sheet_name, profile.get("title"), account_name),
        "sheet_kind": profile.get("sheet_kind"),
        "entity_scope": profile.get("entity_scope", "single"),
        "mapping_key": account_name,
        "category": config.get("category"),
        "type": config.get("type"),
        "dynamic_mapping": True,
        "accounting_nature": config.get("accounting_nature"),
    }


def _candidate_passes_dynamic_confirmation(candidate: Dict[str, Any], materiality_threshold: float = 0.005) -> bool:
    compared_dates = int(candidate.get("financial_dates_compared", 0))
    matched_dates = int(candidate.get("financial_match_dates", 0))
    avg_pct_diff = candidate.get("financial_avg_pct_diff")
    if compared_dates <= 0:
        return False
    if matched_dates > 0:
        return True
    return avg_pct_diff is not None and float(avg_pct_diff) <= materiality_threshold


def _summary_row_exact_matches_sheet(account_name: str, profile: Dict[str, Any]) -> bool:
    normalized_account = _normalize_label(account_name)
    if not normalized_account:
        return False
    return any(
        _normalize_label(candidate) == normalized_account
        for candidate in _candidate_strings(profile)
    )


def _resolve_dynamic_sheet_mapping(
    workbook_path: str,
    account_name: str,
    config: Dict[str, Any],
    sheet_name: str,
    profiles: Dict[str, Dict[str, Any]],
    financial_context: Dict[str, Any],
    workbook_frames: Dict[str, Any],
    normalized_totals_cache: Dict[Tuple[str, str], Dict[str, float]],
    resolution_method: str,
) -> Optional[Dict[str, Any]]:
    candidate = _build_sheet_candidate_for_account(
        account_name=account_name,
        config=config,
        sheet_name=sheet_name,
        profiles=profiles,
    )
    ranked = _rank_candidates_with_financial_signals(
        workbook_path=workbook_path,
        mapping_key=account_name,
        config=config,
        candidates=[candidate],
        profiles=profiles,
        financial_context=financial_context,
        workbook_frames=workbook_frames,
        normalized_totals_cache=normalized_totals_cache,
    )
    if not ranked:
        return None
    top = ranked[0]
    if not _candidate_passes_dynamic_confirmation(top):
        return None
    return {
        **top,
        "resolution_method": resolution_method,
        "dynamic_mapping": True,
        "accounting_nature": config.get("accounting_nature"),
    }


def _discover_dynamic_sheet_resolutions(
    workbook_path: str,
    profiles: Dict[str, Dict[str, Any]],
    financial_context: Dict[str, Any],
    workbook_frames: Dict[str, Any],
    normalized_totals_cache: Dict[Tuple[str, str], Dict[str, float]],
    used_sheets: set[str],
    mappings: Dict[str, Any],
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    resolved: Dict[str, Dict[str, Any]] = {}
    dynamic_mappings: Dict[str, Dict[str, Any]] = {}
    for row in _iter_financial_reference_rows(financial_context):
        account_name = row["account_name"]
        if account_name in mappings or account_name in resolved:
            continue
        config = _build_dynamic_mapping_config(account_name, row["statement_type"])
        for sheet_name, profile in profiles.items():
            if sheet_name in used_sheets:
                continue
            if profile.get("sheet_kind") != "financial_schedule":
                continue
            if not _summary_row_exact_matches_sheet(account_name, profile):
                continue
            discovered = _resolve_dynamic_sheet_mapping(
                workbook_path=workbook_path,
                account_name=account_name,
                config=config,
                sheet_name=sheet_name,
                profiles=profiles,
                financial_context=financial_context,
                workbook_frames=workbook_frames,
                normalized_totals_cache=normalized_totals_cache,
                resolution_method="dynamic_exact_name",
            )
            if discovered is None:
                continue
            resolved[account_name] = discovered
            dynamic_mappings[account_name] = config
            used_sheets.add(sheet_name)
            break
    return resolved, dynamic_mappings


def _resolve_manual_override_target(
    workbook_path: str,
    mapping_key: str,
    override_sheet: str,
    mappings: Dict[str, Any],
    profiles: Dict[str, Dict[str, Any]],
    financial_context: Dict[str, Any],
    workbook_frames: Dict[str, Any],
    normalized_totals_cache: Dict[Tuple[str, str], Dict[str, float]],
) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    if mapping_key in mappings:
        profile = profiles[override_sheet]
        return (
            {
                "sheet_name": override_sheet,
                "title": profile.get("title"),
                "score": 999.0,
                "matched_alias": "manual_override",
                "sheet_kind": profile.get("sheet_kind"),
                "entity_scope": profile.get("entity_scope", "single"),
                "mapping_key": mapping_key,
                "category": mappings.get(mapping_key, {}).get("category"),
                "type": mappings.get(mapping_key, {}).get("type"),
                "resolution_method": "manual_override",
            },
            None,
        )

    summary_rows = _iter_financial_reference_rows(financial_context)
    matched_row = next(
        (row for row in summary_rows if _normalize_label(row["account_name"]) == _normalize_label(mapping_key)),
        None,
    )
    if matched_row is None:
        return None, None
    config = _build_dynamic_mapping_config(matched_row["account_name"], matched_row["statement_type"])
    resolved = _resolve_dynamic_sheet_mapping(
        workbook_path=workbook_path,
        account_name=matched_row["account_name"],
        config=config,
        sheet_name=override_sheet,
        profiles=profiles,
        financial_context=financial_context,
        workbook_frames=workbook_frames,
        normalized_totals_cache=normalized_totals_cache,
        resolution_method="manual_dynamic_override",
    )
    if resolved is None:
        return None, None
    return resolved, config


def _candidate_total_values(
    workbook_path: str,
    candidate: Dict[str, Any],
    config: Dict[str, Any],
    profiles: Dict[str, Dict[str, Any]],
    workbook_frames: Dict[str, Any],
    normalized_totals_cache: Dict[Tuple[str, str], Dict[str, float]],
) -> Dict[str, float]:
    cache_key = (str(candidate.get("sheet_name", "")), str(config.get("type", "")))
    if cache_key in normalized_totals_cache:
        return normalized_totals_cache[cache_key]
    try:
        normalized = normalize_financial_schedule(
            workbook_path=workbook_path,
            sheet_name=candidate["sheet_name"],
            profile=profiles.get(candidate["sheet_name"]),
            sheet_df=workbook_frames.get(candidate["sheet_name"]),
            statement_type=config.get("type"),
        )
    except Exception:
        normalized_totals_cache[cache_key] = {}
        return {}
    values = _normalized_total_values(normalized)
    normalized_totals_cache[cache_key] = values
    return values


def _rank_candidates_with_financial_signals(
    workbook_path: str,
    mapping_key: str,
    config: Dict[str, Any],
    candidates: List[Dict[str, Any]],
    profiles: Dict[str, Dict[str, Any]],
    financial_context: Optional[Dict[str, Any]] = None,
    workbook_frames: Optional[Dict[str, Any]] = None,
    normalized_totals_cache: Optional[Dict[Tuple[str, str], Dict[str, float]]] = None,
    materiality_threshold: float = 0.005,
    absolute_tolerance: float = 1.0,
) -> List[Dict[str, Any]]:
    if not candidates:
        return []
    financial_context = financial_context or {}
    workbook_frames = workbook_frames or load_workbook_frames(workbook_path)
    normalized_totals_cache = normalized_totals_cache if normalized_totals_cache is not None else {}
    financial_results = financial_context.get("financial_results") or {}
    statement_df = (
        financial_results.get("balance_sheet")
        if config.get("type") == "BS"
        else financial_results.get("income_statement")
    )
    summary_values = _statement_values_for_mapping(statement_df, mapping_key, config)
    ranked_candidates: List[Dict[str, Any]] = []
    compare_as_absolute = bool(
        str(config.get("type") or "").strip().upper() == "IS"
        and _should_compare_income_statement_as_absolute(mapping_key, config.get("category"))
    )
    for candidate in candidates:
        candidate_values = _candidate_total_values(
            workbook_path=workbook_path,
            candidate=candidate,
            config=config,
            profiles=profiles,
            workbook_frames=workbook_frames,
            normalized_totals_cache=normalized_totals_cache,
        )
        pct_diffs: List[float] = []
        matched_dates = 0
        for date_label, candidate_value in candidate_values.items():
            summary_value = summary_values.get(date_label)
            if summary_value is None:
                continue
            candidate_value_for_compare = abs(candidate_value) if compare_as_absolute else candidate_value
            summary_value_for_compare = abs(summary_value) if compare_as_absolute else summary_value
            difference = abs(candidate_value_for_compare - summary_value_for_compare)
            pct_diff = (
                0.0
                if abs(summary_value_for_compare) <= absolute_tolerance
                else difference / abs(summary_value_for_compare)
            )
            if difference <= absolute_tolerance or pct_diff <= materiality_threshold:
                matched_dates += 1
            pct_diffs.append(pct_diff)
        compared_dates = len(pct_diffs)
        avg_pct_diff = (sum(pct_diffs) / compared_dates) if compared_dates else None
        ranked_candidates.append(
            {
                **candidate,
                "exact_alias_match": bool(candidate.get("exact_alias_match")),
                "financial_match_dates": matched_dates,
                "financial_dates_compared": compared_dates,
                "financial_avg_pct_diff": (round(avg_pct_diff, 6) if avg_pct_diff is not None else None),
                "financial_values_available": bool(candidate_values),
                "summary_values_available": bool(summary_values),
            }
        )
    ranked_candidates.sort(
        key=lambda item: (
            1 if item.get("exact_alias_match") else 0,
            1 if item.get("financial_dates_compared", 0) > 0 else 0,
            int(item.get("financial_match_dates", 0)),
            -(float(item["financial_avg_pct_diff"]) if item.get("financial_avg_pct_diff") is not None else float("inf")),
            float(item.get("score", 0)),
        ),
        reverse=True,
    )
    return ranked_candidates


def _build_candidate_map(
    mappings: Dict[str, Any],
    profiles: Dict[str, Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    candidate_map: Dict[str, List[Dict[str, Any]]] = {}
    for mapping_key, config in mappings.items():
        if mapping_key.startswith("_") or not isinstance(config, dict):
            continue
        candidate_map[mapping_key] = _candidate_sheets_for_mapping(mapping_key, config, profiles)
    return candidate_map


def _available_candidates(
    candidates: List[Dict[str, Any]],
    used_sheets: set[str],
) -> List[Dict[str, Any]]:
    return [
        candidate for candidate in candidates
        if candidate["sheet_name"] not in used_sheets
    ]


def _rank_mapping_candidates(
    workbook_path: str,
    mapping_key: str,
    config: Dict[str, Any],
    candidates: List[Dict[str, Any]],
    profiles: Dict[str, Dict[str, Any]],
    financial_context: Dict[str, Any],
    workbook_frames: Dict[str, Any],
    normalized_totals_cache: Dict[Tuple[str, str], Dict[str, float]],
) -> List[Dict[str, Any]]:
    ranked_candidates = _rank_candidates_with_financial_signals(
        workbook_path=workbook_path,
        mapping_key=mapping_key,
        config=config,
        candidates=candidates,
        profiles=profiles,
        financial_context=financial_context,
        workbook_frames=workbook_frames,
        normalized_totals_cache=normalized_totals_cache,
    )
    return ranked_candidates or candidates


def _should_accept_hybrid_top_candidate(candidates: List[Dict[str, Any]]) -> bool:
    if not candidates:
        return False
    if len(candidates) == 1:
        return True
    top = candidates[0]
    second = candidates[1]
    top_match_dates = int(top.get("financial_match_dates", 0))
    second_match_dates = int(second.get("financial_match_dates", 0))
    top_compared = int(top.get("financial_dates_compared", 0))
    second_compared = int(second.get("financial_dates_compared", 0))
    top_avg = top.get("financial_avg_pct_diff")
    second_avg = second.get("financial_avg_pct_diff")
    top_score = float(top.get("score", 0))
    second_score = float(second.get("score", 0))
    top_exact = bool(top.get("exact_alias_match"))
    second_exact = bool(second.get("exact_alias_match"))

    if top_exact and not second_exact:
        return True

    if top_compared and top_match_dates > second_match_dates:
        return True
    if top_compared and not second_compared and top_match_dates > 0:
        return True
    if top_match_dates > 0 and second_match_dates == 0 and top_compared >= second_compared:
        return True
    if top_avg is not None and second_avg is not None and top_match_dates == second_match_dates:
        if (second_avg - top_avg) >= 0.02:
            return True
        if (second_avg - top_avg) >= 0.005 and (top_score - second_score) >= 5:
            return True
        if abs(second_avg - top_avg) <= 0.001 and (top_score - second_score) >= 20:
            return True
    if top_compared == 0 and second_compared == 0:
        return not should_use_ai_for_candidates(candidates)
    return False


def _resolve_top_ranked_candidate(candidates: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not candidates:
        raise ValueError("No candidates provided")
    top = candidates[0]
    if int(top.get("financial_match_dates", 0)) > 0:
        return {**top, "resolution_method": "financial_validated"}
    return {**top, "resolution_method": "deterministic"}


def _default_ai_decider(
    mapping_key: str,
    candidates: List[Dict[str, Any]],
    model_type: str = "deepseek",
    language: str = "Eng",
) -> List[str]:
    try:
        from .ai import AIClient
    except Exception:
        return []

    try:
        helper = AIClient(
            model_type=model_type,
            agent_name="sheet_resolver",
            language=language,
        )
    except Exception:
        return []

    candidate_payload = [
        {
            "sheet_name": candidate.get("sheet_name"),
            "title": candidate.get("title"),
            "score": candidate.get("score"),
            "matched_alias": candidate.get("matched_alias"),
            "sheet_kind": candidate.get("sheet_kind"),
            "entity_scope": candidate.get("entity_scope", "single"),
            "financial_match_dates": candidate.get("financial_match_dates"),
            "financial_dates_compared": candidate.get("financial_dates_compared"),
            "financial_avg_pct_diff": candidate.get("financial_avg_pct_diff"),
        }
        for candidate in candidates[:4]
    ]
    system_prompt = (
        "You are resolving ambiguous Excel tab mappings for a financial databook. "
        "Pick the best candidate sheet or shortlist of candidate sheets for the requested mapping key. "
        "Prefer the tabs whose title and schedule semantics best match the mapping key. "
        "Return JSON only in one of these forms: "
        "{\"sheet_name\": \"exact candidate name\"} "
        "or {\"sheet_names\": [\"candidate one\", \"candidate two\"]}."
    )
    user_prompt = (
        f"Mapping key: {mapping_key}\n"
        f"Candidates:\n{json.dumps(candidate_payload, ensure_ascii=False, indent=2)}\n\n"
        "Choose the exact candidate sheet_name that best matches the mapping key, or return a shortlist if multiple tabs still look plausible."
    )
    try:
        response = helper.get_response(user_prompt=user_prompt, system_prompt=system_prompt, temperature=0.0, max_tokens=120)
    except Exception:
        return []
    return _extract_sheet_names_from_ai_response(response.get("content", ""), candidates)


def resolve_ambiguous_candidate(
    workbook_path: str,
    mapping_key: str,
    config: Dict[str, Any],
    candidates: List[Dict[str, Any]],
    profiles: Dict[str, Dict[str, Any]],
    financial_context: Optional[Dict[str, Any]] = None,
    workbook_frames: Optional[Dict[str, Any]] = None,
    normalized_totals_cache: Optional[Dict[Tuple[str, str], Dict[str, float]]] = None,
    ai_decider: Optional[Callable[[str, List[Dict[str, Any]]], Any]] = None,
) -> Dict[str, Any]:
    if not candidates:
        raise ValueError("No candidates provided")
    if ai_decider is None:
        ai_decider = _default_ai_decider

    chosen_sheets: List[str] = []
    try:
        ai_choice = ai_decider(mapping_key, candidates)
        if isinstance(ai_choice, str) and ai_choice:
            chosen_sheets = [ai_choice]
        elif isinstance(ai_choice, list):
            chosen_sheets = [sheet for sheet in ai_choice if isinstance(sheet, str)]
    except Exception:
        chosen_sheets = []

    if len(chosen_sheets) == 1:
        for candidate in candidates:
            if candidate.get("sheet_name") == chosen_sheets[0]:
                return {**candidate, "resolution_method": "ai_fallback"}
    if len(chosen_sheets) > 1:
        shortlisted_candidates = [
            candidate for candidate in candidates if candidate.get("sheet_name") in chosen_sheets
        ]
        reranked_shortlist = _rank_candidates_with_financial_signals(
            workbook_path=workbook_path,
            mapping_key=mapping_key,
            config=config,
            candidates=shortlisted_candidates,
            profiles=profiles,
            financial_context=financial_context,
            workbook_frames=workbook_frames,
            normalized_totals_cache=normalized_totals_cache,
        )
        if reranked_shortlist:
            best = reranked_shortlist[0]
            best["ai_candidate_sheets"] = chosen_sheets
            best["resolution_method"] = "ai_financial_validated" if int(best.get("financial_match_dates", 0)) > 0 else "ai_shortlist_fallback"
            return best
        return {
            **shortlisted_candidates[0],
            "resolution_method": "ai_shortlist_fallback",
            "ai_candidate_sheets": chosen_sheets,
        }
    fallback = candidates[0]
    if int(fallback.get("financial_match_dates", 0)) > 0:
        return {**fallback, "resolution_method": "financial_fallback"}
    return {**fallback, "resolution_method": "deterministic_fallback"}


def resolve_workbook_mappings(
    workbook_path: str,
    profiles: Optional[Dict[str, Dict[str, Any]]] = None,
    workbook_frames: Optional[Dict[str, pd.DataFrame]] = None,
    mappings_path: Optional[str] = None,
    use_ai_for_ambiguity: bool = True,
    ai_decider: Optional[Callable[[str, List[Dict[str, Any]]], Optional[str]]] = None,
    model_type: str = "deepseek",
    language: str = "Eng",
    mapping_overrides: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    mappings = load_mappings(mappings_path)
    if profiles is None:
        profiles = profile_workbook(workbook_path)
    mapping_overrides = mapping_overrides or {}
    workbook_frames = workbook_frames or load_workbook_frames(workbook_path)
    financial_context = _build_financial_reference_context(workbook_path, profiles)
    normalized_totals_cache: Dict[Tuple[str, str], Dict[str, float]] = {}

    resolved: Dict[str, Dict[str, Any]] = {}
    candidate_map = _build_candidate_map(mappings, profiles)
    used_sheets: set[str] = set()
    ambiguities: Dict[str, List[Dict[str, Any]]] = {}
    override_issues: List[Dict[str, Any]] = []
    dynamic_mappings: Dict[str, Dict[str, Any]] = {}

    ranked_keys = sorted(
        candidate_map,
        key=lambda key: candidate_map[key][0]["score"] if candidate_map[key] else -1,
        reverse=True,
    )
    for mapping_key in ranked_keys:
        override_sheet = mapping_overrides.get(mapping_key)
        if override_sheet:
            if override_sheet not in profiles:
                override_issues.append(
                    {
                        "mapping_key": mapping_key,
                        "sheet_name": override_sheet,
                        "issue_type": "invalid_override",
                        "details": "Requested override sheet is not present in the workbook profile.",
                    }
                )
            elif override_sheet in used_sheets:
                override_issues.append(
                    {
                        "mapping_key": mapping_key,
                        "sheet_name": override_sheet,
                        "issue_type": "override_conflict",
                        "details": "Requested override sheet is already assigned to another mapping key.",
                    }
                )
            else:
                resolved_override, dynamic_config = _resolve_manual_override_target(
                    workbook_path=workbook_path,
                    mapping_key=mapping_key,
                    override_sheet=override_sheet,
                    mappings=mappings,
                    profiles=profiles,
                    financial_context=financial_context,
                    workbook_frames=workbook_frames,
                    normalized_totals_cache=normalized_totals_cache,
                )
                if resolved_override is not None:
                    resolved[mapping_key] = resolved_override
                    if dynamic_config is not None:
                        dynamic_mappings[mapping_key] = dynamic_config
                    used_sheets.add(override_sheet)
                    continue

        available_candidates = _available_candidates(candidate_map[mapping_key], used_sheets)
        if not available_candidates:
            continue

        ranked_available_candidates = _rank_mapping_candidates(
            workbook_path=workbook_path,
            mapping_key=mapping_key,
            config=mappings.get(mapping_key, {}),
            candidates=available_candidates,
            profiles=profiles,
            financial_context=financial_context,
            workbook_frames=workbook_frames,
            normalized_totals_cache=normalized_totals_cache,
        )

        if _should_accept_hybrid_top_candidate(ranked_available_candidates):
            resolved_candidate = _resolve_top_ranked_candidate(ranked_available_candidates)
        else:
            ambiguities[mapping_key] = ranked_available_candidates[:3]
            if use_ai_for_ambiguity:
                resolved_candidate = resolve_ambiguous_candidate(
                    workbook_path=workbook_path,
                    mapping_key=mapping_key,
                    config=mappings.get(mapping_key, {}),
                    candidates=ranked_available_candidates,
                    profiles=profiles,
                    financial_context=financial_context,
                    workbook_frames=workbook_frames,
                    normalized_totals_cache=normalized_totals_cache,
                    ai_decider=ai_decider or (
                        lambda key, candidate_list: _default_ai_decider(
                            key,
                            candidate_list,
                            model_type=model_type,
                            language=language,
                        )
                    ),
                )
            else:
                resolved_candidate = _resolve_top_ranked_candidate(ranked_available_candidates)
        resolved[mapping_key] = resolved_candidate
        used_sheets.add(resolved_candidate["sheet_name"])

    dynamic_override_keys = [
        key for key in mapping_overrides.keys()
        if key not in candidate_map
    ]
    for mapping_key in dynamic_override_keys:
        override_sheet = mapping_overrides.get(mapping_key)
        if not override_sheet:
            continue
        if override_sheet not in profiles:
            override_issues.append(
                {
                    "mapping_key": mapping_key,
                    "sheet_name": override_sheet,
                    "issue_type": "invalid_override",
                    "details": "Requested override sheet is not present in the workbook profile.",
                }
            )
            continue
        if override_sheet in used_sheets:
            override_issues.append(
                {
                    "mapping_key": mapping_key,
                    "sheet_name": override_sheet,
                    "issue_type": "override_conflict",
                    "details": "Requested override sheet is already assigned to another mapping key.",
                }
            )
            continue
        resolved_override, dynamic_config = _resolve_manual_override_target(
            workbook_path=workbook_path,
            mapping_key=mapping_key,
            override_sheet=override_sheet,
            mappings=mappings,
            profiles=profiles,
            financial_context=financial_context,
            workbook_frames=workbook_frames,
            normalized_totals_cache=normalized_totals_cache,
        )
        if resolved_override is None:
            override_issues.append(
                {
                    "mapping_key": mapping_key,
                    "sheet_name": override_sheet,
                    "issue_type": "unknown_override_target",
                    "details": "Override target does not match an existing mapping key or an exact Financials account name.",
                }
            )
            continue
        resolved[mapping_key] = resolved_override
        if dynamic_config is not None:
            dynamic_mappings[mapping_key] = dynamic_config
        used_sheets.add(override_sheet)

    discovered_resolved, discovered_dynamic_mappings = _discover_dynamic_sheet_resolutions(
        workbook_path=workbook_path,
        profiles=profiles,
        financial_context=financial_context,
        workbook_frames=workbook_frames,
        normalized_totals_cache=normalized_totals_cache,
        used_sheets=used_sheets,
        mappings={**mappings, **dynamic_mappings},
    )
    resolved.update(discovered_resolved)
    dynamic_mappings.update(discovered_dynamic_mappings)

    unresolved_sheets = sorted(
        sheet_name
        for sheet_name, profile in profiles.items()
        if profile.get("sheet_kind") == "financial_schedule" and sheet_name not in used_sheets
    )

    return {
        "profiles": profiles,
        "resolved": resolved,
        "candidate_map": candidate_map,
        "unresolved_sheets": unresolved_sheets,
        "ambiguities": ambiguities,
        "override_issues": override_issues,
        "dynamic_mappings": dynamic_mappings,
        "financial_reference": {
            "summary_sheet": financial_context.get("summary_sheet"),
            "reference_available": bool(financial_context.get("reference_available")),
            "reference_stage": financial_context.get("reference_stage"),
            "reference_error": financial_context.get("reference_error"),
        },
    }


def build_resolution_report_tables(resolution: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    resolved = resolution.get("resolved", {}) or {}
    ambiguities = resolution.get("ambiguities", {}) or {}
    normalization_errors = resolution.get("normalization_errors", {}) or {}
    unresolved_sheets = resolution.get("unresolved_sheets", []) or []
    override_issues = resolution.get("override_issues", []) or []

    ambiguous_rows = []
    fallback_rows = []
    issue_rows = []

    for mapping_key, candidates in ambiguities.items():
        chosen = resolved.get(mapping_key, {})
        candidate_sheets = ", ".join(
            f"{candidate.get('sheet_name')} ({candidate.get('score')})" for candidate in candidates
        )
        ambiguous_rows.append(
            {
                "mapping_key": mapping_key,
                "chosen_sheet": chosen.get("sheet_name", ""),
                "resolution_method": chosen.get("resolution_method", ""),
                "entity_scope": chosen.get("entity_scope", ""),
                "candidate_sheets": candidate_sheets,
            }
        )

    for mapping_key, chosen in resolved.items():
        if chosen.get("resolution_method") in {
            "deterministic_fallback",
            "manual_override",
            "ai_primary",
            "financial_tiebreak",
            "ai_shortlist_fallback",
        }:
            fallback_rows.append(
                {
                    "mapping_key": mapping_key,
                    "chosen_sheet": chosen.get("sheet_name", ""),
                    "resolution_method": chosen.get("resolution_method", ""),
                    "entity_scope": chosen.get("entity_scope", ""),
                    "matched_alias": chosen.get("matched_alias", ""),
                    "score": chosen.get("score", ""),
                    "ai_candidate_sheets": ", ".join(chosen.get("ai_candidate_sheets", []) or []),
                    "financial_match_dates": chosen.get("financial_match_dates", ""),
                }
            )

    for sheet_name in unresolved_sheets:
        issue_rows.append(
            {
                "sheet_name": sheet_name,
                "issue_type": "unresolved_sheet",
                "details": "No mapping key was assigned to this financial schedule.",
            }
        )

    for sheet_name, details in normalization_errors.items():
        issue_rows.append(
            {
                "sheet_name": sheet_name,
                "issue_type": "normalization_error",
                "details": details,
            }
        )

    for issue in override_issues:
        issue_rows.append(
            {
                "sheet_name": issue.get("sheet_name", ""),
                "issue_type": issue.get("issue_type", "override_issue"),
                "details": issue.get("details", ""),
            }
        )

    return {
        "ambiguous": pd.DataFrame(
            ambiguous_rows,
            columns=["mapping_key", "chosen_sheet", "resolution_method", "entity_scope", "candidate_sheets"],
        ),
        "fallbacks": pd.DataFrame(
            fallback_rows,
            columns=[
                "mapping_key",
                "chosen_sheet",
                "resolution_method",
                "entity_scope",
                "matched_alias",
                "score",
                "ai_candidate_sheets",
                "financial_match_dates",
            ],
        ),
        "issues": pd.DataFrame(
            issue_rows,
            columns=["sheet_name", "issue_type", "details"],
        ),
    }
# --- end workbook/resolver.py ---

# --- begin workbook/databook.py ---
import pandas as pd
import json
import warnings
import os
import re
import logging
import time
 
warnings.simplefilter(action='ignore', category=UserWarning)
logger = logging.getLogger(__name__)
from .financial_common import clean_english_placeholders, load_yaml_file, normalize_financial_date_label, package_file_path
from .financial_display_format import add_language_display_columns
 
def load_mapping(filename):
    return load_yaml_file(filename)
 
def filter_worksheets_by_mode(worksheets, mode, mapping):
    result = []
    for ws in worksheets:
        for key, value in mapping.items():
            # Skip non-account entries (e.g., _default_subagent_1)
            if key.startswith('_') or not isinstance(value, dict):
                continue
            
            # Skip if required keys are missing
            if 'type' not in value or 'aliases' not in value:
                continue
            
            # Ensure correct matching against mode
            if mode == "All":
                if ws in value['aliases']:
                    result.append(ws)
                    break
            elif value['type'] == mode and ws in value['aliases']:
                result.append(ws)
                break
    return result
 
def process_excel_data(dfs, sheet_name, entity_name, keep_zero_rows=False):
    def detect_entities_and_extract(df, entity_name):
        # Detect presence of indicative keywords
        indicative_keywords = ['Indicative adjusted', '示意性调整后', "CNY'000", "人民币千元"]
        keyword_occurrences = sum([df.to_string().count(keyword) for keyword in indicative_keywords])
 
        # Check if multiple entities are likely based on the occurrences of keywords
        is_multiple = keyword_occurrences > 1
 
        #print('detect_entities_and_extract:', 'multiple' if is_multiple else 'single', df)
 
        # Extract the entity table with the logic adjusted
        return ('multiple' if is_multiple else 'single', extract_entity_table(df, entity_name, single=not is_multiple))
 
    def extract_entity_table(df, entity_name, single=False):
        # Find the first and last non-empty column
        first_column_with_data = None
        last_column_with_data = None
 
        for i in range(len(df.columns)):
            if df.iloc[:, i].notnull().any():
                if first_column_with_data is None:
                    first_column_with_data = i
                last_column_with_data = i
            else:
                if first_column_with_data is not None:
                    break
 
        if first_column_with_data is None:
            return None
 
        # Slice the dataframe up to the last column with data
        df = df.iloc[:, :last_column_with_data + 1]
 
        result_df = None
        if single:
            result_df = df.reset_index(drop=True)
        else:
            # Try finding entity name related rows or default start
            start_idx = None
            if entity_name:
                # Prepare combinations from the entity name
                components = entity_name.split()
                combinations = {entity_name}
 
                # Create combinations
                for i in range(len(components)):
                    for j in range(i, len(components)):
                        combination = ' '.join(components[i:j+1])
                        combinations.add(combination)
 
                for i in range(len(df)):
                    if df.iloc[i].astype(str).str.contains('|'.join(combinations), na=False).any():
                        start_idx = i
                        break
           
            # If entity not found, consider a logical start point to slice the table
            if start_idx is None:
                start_idx = 0  # Fallback if entity_name not found
       
            # Collect rows until an empty row is encountered
            result_rows = []
            for i in range(start_idx, len(df)):
                if df.iloc[i].isnull().all():
                    break
                result_rows.append(df.iloc[i])
           
            result_df = pd.DataFrame(result_rows).reset_index(drop=True)
 
        # Drop empty columns, if any
        #if result_df is not None:
        #    result_df.dropna(axis=1, how='all', inplace=True)
 
        return result_df
   
    def preprocess_date(date_str):
        return normalize_financial_date_label(date_str)
   
    def find_columns(result_table, keep_zero_rows=False):
        result_table = result_table.copy()
        object_cols = result_table.select_dtypes(include=['object']).columns
        if len(object_cols) > 0:
            result_table[object_cols] = result_table[object_cols].apply(
                lambda col: col.map(lambda value: value.strip() if isinstance(value, str) else value)
            )
 
        value_column_name = None
        description_column_name = None
        value_column_number = None
 
        indicative_row_idx = None
        indicative_col_index = None
 
        for idx, row in result_table.iterrows():
            if row.apply(lambda x: bool(re.match(r'^\s*(Indicative adjusted|示意性调整后)\s*$', str(x), re.IGNORECASE))).any():
                indicative_row_idx = idx
                indicative_col_name = row[row.apply(lambda x: bool(re.match(r'^\s*(Indicative adjusted|示意性调整后)\s*$', str(x), re.IGNORECASE)))].index[0]
                indicative_col_index = result_table.columns.get_loc(indicative_col_name)
                break
        #print(indicative_col_index)
 
        if indicative_row_idx is not None and indicative_col_index is not None:
            date_row_idx = indicative_row_idx + 1
            if date_row_idx < len(result_table.index):
                description_col_original = None
                for col in result_table.columns:
                    col_series = result_table[col].astype(str)
                    if col_series.str.contains(r"CNY'000|人民币千元", case=False, na=False).any():
                        description_col_original = col
                        break

                if description_col_original is not None:
                    desc_col_idx = result_table.columns.get_loc(description_col_original)
                    date_columns = get_valid_financial_columns(
                        df=result_table,
                        desc_col_idx=desc_col_idx,
                        header_row_idx=indicative_row_idx,
                    )
                    date_columns = [item for item in date_columns if item[0] >= indicative_col_index]

                    if date_columns:
                        most_recent_col_idx, most_recent_date, _ = max(date_columns, key=lambda item: item[1])
                        value_column_name = most_recent_date.strftime('%Y-%m-%d')
                        value_column_index = most_recent_col_idx
                        value_column_number = value_column_index + 1

        description_col_original = None
        for col in result_table.columns:
            if result_table[col].astype(str).str.contains(r"CNY'000|人民币千元", case=False, na=False).any():
                description_col_original = col
                description_column_name = result_table.iloc[0, result_table.columns.get_loc(col)]
                break
 
        if value_column_name and description_col_original is not None:
            result_df = result_table.iloc[date_row_idx + 1:, [result_table.columns.get_loc(description_col_original), value_column_index]]
            result_df.columns = [description_column_name, value_column_name]
            result_df[value_column_name] = pd.to_numeric(result_df[value_column_name], errors='coerce')
 
            # Check all values in a specific column or set of columns
            contains_cny = False
            contains_chinese_cny = False
 
            # Iterate over all elements in the row to find the key phrase
            for element in result_table.iloc[date_row_idx]:
                # Convert each element to a string and check
                if isinstance(element, str):
                    if "CNY'000" in element:
                        contains_cny = True
                    if "人民币千元" in element:
                        contains_chinese_cny = True
 
            if contains_cny or contains_chinese_cny:
                #print('Triggered')
                result_df[value_column_name] *= 1000

            # Round all decimal values to integers to avoid decimal issues with AI
            # First, handle NaN and infinite values by filtering them out or replacing with 0
            result_df = result_df[~result_df[value_column_name].isin([float('inf'), float('-inf')])]  # Remove inf
            numeric_mask = result_df[value_column_name].notna()
            result_df.loc[numeric_mask, value_column_name] = (
                result_df.loc[numeric_mask, value_column_name].round(0).astype(int)
            )

            #print("After multiplying and rounding:")
            #print(result_df[value_column_name])

            if not keep_zero_rows:
                result_df = filter_zero_value_rows(result_df)

            return result_df.reset_index(drop=True), value_column_index
        else:
            return None, None
 
    # Fetch the dataframe from the dictionary
    df = dfs.get(sheet_name)
    if df is None:
        raise ValueError(f"No data found for sheet: {sheet_name}")
 
    result_type, result_table = detect_entities_and_extract(df, entity_name)
    #print('result_table:', result_table)
    if result_table is not None:
        extracted_df, value_col_num = find_columns(result_table, keep_zero_rows=keep_zero_rows)
        return result_type, extracted_df, value_col_num
    else:
        return result_type, None, None
   
def determine_result_type(sheet_data):
    # Check the number of occurrences of the keywords in the sheet
    indicative_keywords = ['Indicative adjusted', '示意性调整后', "CNY'000", "人民币千元"]
    occurrences = sum([sheet_data.to_string().count(keyword) for keyword in indicative_keywords])
   
    # Determine if the sheet is single or multiple based on occurrences
    return 'multiple' if occurrences > 1 else 'single'
 
def _update_display_description_map(df: pd.DataFrame, mapping: dict[str, str]) -> None:
    existing = dict(df.attrs.get("display_description_map") or {})
    for source, target in mapping.items():
        source_text = str(source or "").strip()
        target_text = str(target or "").strip()
        if not source_text or not target_text:
            continue
        existing[source_text] = target_text
    df.attrs["display_description_map"] = existing


def filter_detail_accounts(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filter out detail sub-account rows, keeping only main account totals.
    Removes rows with patterns like "应付利息_借款利息" or containing "  " (indentation).
    Also cleans English placeholders from descriptions.
    
    Args:
        df: DataFrame with account descriptions
        
    Returns:
        Filtered DataFrame with cleaned descriptions
    """
    if df is None or df.empty:
        return df
    
    df_filtered = df.copy()
    
    # Get the first column name (description column)
    desc_col = df_filtered.columns[0]

    row_types_by_description = df_filtered.attrs.get("row_types_by_description", {})
    if row_types_by_description:
        df_filtered = df_filtered[
            ~df_filtered[desc_col].astype(str).map(
                lambda value: row_types_by_description.get(str(value), "") == "breakdown"
            )
        ]
    
    # Filter patterns that indicate detail sub-accounts
    filter_patterns = [
        r'_',  # Sub-account separator
        r'^\s{2,}',  # Multiple spaces at start (indentation)
        r'其中[：:]',  # "Including:" markers in Chinese
    ]
    
    for pattern in filter_patterns:
        df_filtered = df_filtered[~df_filtered[desc_col].astype(str).str.contains(pattern, regex=True, na=False)]
    
    original_descriptions = df_filtered[desc_col].astype(str).tolist()
    cleaned_descriptions = [clean_english_placeholders(value) for value in original_descriptions]

    # Clean English placeholders from descriptions
    df_filtered[desc_col] = cleaned_descriptions
    
    # Remove rows that became empty after cleaning
    df_filtered = df_filtered[df_filtered[desc_col].astype(str).str.strip() != '']

    cleaned_mapping = {}
    for original, cleaned in zip(original_descriptions, cleaned_descriptions):
        original_text = str(original or "").strip()
        cleaned_text = str(cleaned or "").strip()
        if not original_text or not cleaned_text:
            continue
        cleaned_mapping[original_text] = cleaned_text
        cleaned_mapping[cleaned_text] = cleaned_text
    _update_display_description_map(df_filtered, cleaned_mapping)
    
    return df_filtered


def filter_zero_value_rows(df: pd.DataFrame, tolerance: float = 0.01) -> pd.DataFrame:
    """
    Remove rows where the numeric value column is zero/insignificant or missing.

    Args:
        df: DataFrame with description in the first column and value in the second column
        tolerance: Absolute value threshold treated as zero

    Returns:
        Filtered DataFrame containing only meaningful numeric rows
    """
    if df is None or df.empty or len(df.columns) < 2:
        return df

    filtered_df = df.copy()
    desc_col = filtered_df.columns[0]
    value_col = filtered_df.columns[1]
    numeric_values = pd.to_numeric(filtered_df[value_col], errors='coerce')
    row_types_by_description = filtered_df.attrs.get("row_types_by_description", {})
    preserved_zero_rows = filtered_df[desc_col].astype(str).map(
        lambda value: row_types_by_description.get(str(value), "") in {"subtotal", "total"}
    )
    keep_mask = preserved_zero_rows | (numeric_values.notna() & (numeric_values.abs() >= tolerance))
    return filtered_df.loc[keep_mask].reset_index(drop=True)


def _detect_report_language_from_profiles(profiles):
    english_count = 0
    chinese_count = 0
    for profile in profiles.values():
        sample = " ".join(
            str(value or "")
            for value in (
                profile.get("sheet_name"),
                profile.get("title"),
                *profile.get("stage_labels", []),
            )
        )
        if re.search(r'[\u4e00-\u9fff]', sample):
            chinese_count += 1
        else:
            english_count += 1
    if english_count + chinese_count == 0:
        return None
    return 'Eng' if english_count >= chinese_count else 'Chi'


def build_dataframes_from_normalized_results(
    normalized_results,
    workbook_list,
    report_language,
    filter_details=True,
    keep_zero_rows=False,
    variant: str = "default",
):
    """Build legacy dataframe outputs from normalized schedule payloads."""
    variants = build_dataframe_variants_from_normalized_results(
        normalized_results=normalized_results,
        workbook_list=workbook_list,
        report_language=report_language,
        variant_specs=[
            {
                "name": variant,
                "variant": variant,
                "filter_details": filter_details,
                "keep_zero_rows": keep_zero_rows,
            }
        ],
    )
    result = variants.get(variant, {})
    return result.get("dfs", {}), result.get("workbook_list", [])


def build_dataframe_variants_from_normalized_results(
    normalized_results,
    workbook_list,
    report_language,
    variant_specs,
):
    """Build one or more dataframe variants while traversing normalized sheets once."""
    if not variant_specs:
        return {}

    prepared_specs = []
    for spec in variant_specs:
        spec_name = str(spec.get("name") or spec.get("variant") or "").strip()
        if not spec_name:
            continue
        prepared_specs.append(
            {
                "name": spec_name,
                "variant": str(spec.get("variant") or "default"),
                "filter_details": bool(spec.get("filter_details", True)),
                "keep_zero_rows": bool(spec.get("keep_zero_rows", False)),
                "dfs": {},
                "workbook_list": [],
            }
        )

    if not prepared_specs:
        return {}

    variant_key_map = {
        "default": "projection_df",
        "original": "projection_df_original",
        "annualized": "projection_df_annualized",
        "analysis": "prompt_analysis_df",
    }

    for sheet in workbook_list:
        normalized = normalized_results.get(sheet)
        if not normalized:
            continue
        display_key = str(normalized.get("display_key") or sheet).strip() or str(sheet)
        for spec in prepared_specs:
            projection_key = variant_key_map.get(spec["variant"], "projection_df")
            source_df = normalized.get(projection_key)
            if source_df is None:
                source_df = normalized.get("projection_df")
            if source_df is None:
                continue

            extracted_df = source_df.copy()
            if extracted_df is None or extracted_df.empty:
                continue

            if not spec["keep_zero_rows"]:
                extracted_df = filter_zero_value_rows(extracted_df)

            if spec["filter_details"]:
                extracted_df = filter_detail_accounts(extracted_df)

            if extracted_df is None or extracted_df.empty:
                continue

            derived_attrs = dict(extracted_df.attrs)
            extracted_df.attrs.update(source_df.attrs)
            extracted_df.attrs.update(derived_attrs)
            extracted_df.attrs["report_language"] = report_language
            extracted_df.attrs["source_sheet_name"] = str(sheet)
            extracted_df.attrs["display_key"] = display_key

            if report_language and len(extracted_df.columns) > 1:
                extracted_df = add_language_display_columns(extracted_df, report_language)
                post_format_attrs = dict(extracted_df.attrs)
                extracted_df.attrs.update(source_df.attrs)
                extracted_df.attrs.update(post_format_attrs)
                extracted_df.attrs["report_language"] = report_language
                extracted_df.attrs["selected_variant"] = source_df.attrs.get(
                    "selected_variant",
                    spec["variant"],
                )
                extracted_df.attrs["source_sheet_name"] = str(sheet)
                extracted_df.attrs["display_key"] = display_key

            spec["dfs"][display_key] = extracted_df.reset_index(drop=True)
            spec["workbook_list"].append(display_key)

    return {
        spec["name"]: {
            "dfs": spec["dfs"],
            "workbook_list": spec["workbook_list"],
        }
        for spec in prepared_specs
    }


def extract_normalized_data_from_excel(databook_path, mode="All", entity_name=None, mapping_overrides=None):
    """
    Build integrity-aware normalized schedule payloads for the workbook.

    Returns:
        Tuple of (normalized_results, workbook_list, overall_result_type, report_language, resolution)
    """
    overall_started = time.perf_counter()
    profiles_started = time.perf_counter()
    profiles = profile_workbook(databook_path)
    workbook_frames = load_workbook_frames(databook_path)
    logger.debug("Profiled workbook %s in %.2fs", os.path.basename(databook_path), time.perf_counter() - profiles_started)
    resolver_language = _detect_report_language_from_profiles(profiles) or "Eng"

    resolution_started = time.perf_counter()
    resolution = resolve_workbook_mappings(
        databook_path,
        profiles=profiles,
        workbook_frames=workbook_frames,
        mapping_overrides=mapping_overrides,
        language=resolver_language,
    )
    logger.debug("Resolved workbook mappings for %s in %.2fs", os.path.basename(databook_path), time.perf_counter() - resolution_started)
    mappings = load_mapping(package_file_path('mappings.yml'))
    dynamic_mappings = resolution.get("dynamic_mappings") or {}
    resolution.setdefault("normalization_errors", {})

    normalized_results = {}
    workbook_list = []
    entity_scopes = []

    for mapping_key, resolved in resolution.get("resolved", {}).items():
        mapping_config = mappings.get(mapping_key, {}) or dynamic_mappings.get(mapping_key, {}) or {}
        mapping_type = mapping_config.get("type") or resolved.get("type")
        if mode != "All" and mapping_type != mode:
            continue

        sheet_name = resolved["sheet_name"]
        if sheet_name in normalized_results:
            continue

        statement_type = mapping_type
        if mapping_type == "IS" and not (
            resolved.get("matched_alias") or str(resolved.get("resolution_method") or "").startswith("manual_")
        ):
            statement_type = None

        try:
            normalized = normalize_financial_schedule(
                workbook_path=databook_path,
                sheet_name=sheet_name,
                profile=profiles.get(sheet_name),
                entity_name=entity_name,
                sheet_df=workbook_frames.get(sheet_name),
                statement_type=statement_type,
            )
        except Exception as exc:
            resolution["normalization_errors"][sheet_name] = str(exc)
            continue

        normalized_results[sheet_name] = {
            **normalized,
            "mapping_key": mapping_key,
            "category": mapping_config.get("category") or resolved.get("category"),
            "type": mapping_type,
            "display_key": (
                mapping_key
                if (
                    str(resolved.get("resolution_method") or "").startswith("manual_")
                    or bool(resolved.get("dynamic_mapping") or mapping_config.get("dynamic_mapping"))
                )
                else sheet_name
            ),
            "dynamic_mapping_context": {
                "dynamic_mapping": bool(resolved.get("dynamic_mapping") or mapping_config.get("dynamic_mapping")),
                "accounting_nature": (
                    mapping_config.get("accounting_nature")
                    or resolved.get("accounting_nature")
                    or mapping_config.get("category")
                    or resolved.get("category")
                    or ""
                ),
                "category": mapping_config.get("category") or resolved.get("category"),
                "type": mapping_type,
            },
        }
        workbook_list.append(sheet_name)
        entity_scopes.append((profiles.get(sheet_name) or {}).get("entity_scope", "single"))

    overall_result_type = 'multiple' if any(scope == 'multiple' for scope in entity_scopes) else 'single'
    report_language = _detect_report_language_from_profiles(profiles)
    logger.debug(
        "Normalized %s sheets from %s in %.2fs",
        len(workbook_list),
        os.path.basename(databook_path),
        time.perf_counter() - overall_started,
    )
    return normalized_results, workbook_list, overall_result_type, report_language, resolution


def extract_data_from_excel(databook_path, entity_name, mode="All", filter_details=True, keep_zero_rows=False, return_resolution=False, mapping_overrides=None):
    """
    Extract data from Excel file and determine language.
    
    Args:
        databook_path: Path to Excel file
        entity_name: Name of entity to extract
        mode: Filter mode ('All', 'Assets', 'Liabilities', 'Equity', 'Income', 'Expenses')
        filter_details: Whether to filter out detail sub-accounts (default: True)
        keep_zero_rows: Whether to preserve zero-value and header/detail rows (default: False)
    
    Returns:
        Tuple of (final_dfs, final_workbook_list, overall_result_type, report_language)
    """
    normalized_results, workbook_list, overall_result_type, report_language, resolution = extract_normalized_data_from_excel(
        databook_path=databook_path,
        mode=mode,
        entity_name=entity_name,
        mapping_overrides=mapping_overrides,
    )

    final_dfs, final_workbook_list = build_dataframes_from_normalized_results(
        normalized_results=normalized_results,
        workbook_list=workbook_list,
        report_language=report_language,
        filter_details=filter_details,
        keep_zero_rows=keep_zero_rows,
    )

    if not final_workbook_list:
        overall_result_type = 'None'

    if return_resolution:
        return final_dfs, final_workbook_list, overall_result_type, report_language, resolution
    return final_dfs, final_workbook_list, overall_result_type, report_language
# --- end workbook/databook.py ---

# --- begin workbook/reconcile.py ---
"""
Reconciliation Module
Compares financial data from two sources to verify accuracy
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Tuple, Optional

should_skip_mapping = should_skip_account_label


def _normalize_account_name(account_name: str) -> str:
    """Normalize account labels for tolerant matching."""
    return normalize_mapping_label(account_name)


def find_reconciliation_example(repo_root: Optional[Path] = None) -> Optional[dict]:
    """Return the first readable local workbook/sheet combination for the reconciliation demo."""
    base_dir = Path(repo_root) if repo_root is not None else Path(__file__).resolve().parents[2]
    candidates = [
        {
            'workbook_path': base_dir / 'databook-Foshan Wanyuan_rebuilt.xlsx',
            'sheet_name': 'Financials',
            'entity_name': 'Foshan Wanyuan',
        },
        {
            'workbook_path': base_dir / '[DEMO]240627.东莞岭南-databook.xlsx',
            'sheet_name': 'Financials',
            'entity_name': '',
        },
        {
            'workbook_path': base_dir / 'databook.xlsx',
            'sheet_name': 'Financials',
            'entity_name': '',
        },
    ]

    for candidate in candidates:
        workbook_path = candidate['workbook_path']
        if not workbook_path.exists():
            continue
        try:
            workbook = pd.ExcelFile(workbook_path)
        except (OSError, ValueError, ImportError):
            continue
        if candidate['sheet_name'] in workbook.sheet_names:
            return {
                'workbook_path': str(workbook_path),
                'sheet_name': candidate['sheet_name'],
                'entity_name': candidate['entity_name'],
            }

    return None


def _resolve_mapping_alias(account_name: str, mappings: dict) -> Tuple[Optional[str], Optional[dict], Optional[str]]:
    account_clean = account_name.strip()
    account_normalized = _normalize_account_name(account_clean)
    for mapping_key, config in iter_account_mappings(mappings):
        aliases = config.get('aliases', [])
        normalized_aliases = {_normalize_account_name(alias) for alias in aliases}
        if (
            account_name in aliases
            or account_clean in aliases
            or account_normalized in normalized_aliases
        ):
            return mapping_key, config, config.get('category', None)
    return None, None, None


def find_account_in_dfs(
    account_name: str,
    dfs: Dict[str, pd.DataFrame],
    mappings: dict,
    debug: bool = False,
) -> Tuple[Optional[str], Optional[pd.DataFrame], Optional[str], Optional[str], str, str]:
    """
    Find an account in dfs by:
    1. Finding which mapping KEY the source account belongs to (via aliases)
    2. Preferring a direct normalized match against the available dfs keys
    3. Falling back to a normalized alias match if no direct key exists
    
    Args:
        account_name: Account name from BS/IS
        dfs: Dictionary of DataFrames from extract_data_from_excel
        mappings: Mappings configuration
        debug: Enable debug output
        
    Returns:
        Tuple of (dfs_key, dfs_df, category, mapping_key, mapping_status, mapping_note)
        where dfs_key is the actual matched key in dfs.
    """
    if debug:
        print(f"    [MATCH] Source account: '{account_name}'")
    
    # Check if this account should skip mapping
    if should_skip_mapping(account_name):
        if debug:
            print(f"    [MATCH]   ⏭️  Skipped (total/profit line)")
        return 'SKIP', None, None, None, 'Skipped', 'Skipped total/subtotal/profit line.'
    
    # Remove common suffixes for better matching
    account_clean = account_name.strip()
    account_normalized = _normalize_account_name(account_clean)
    
    if debug:
        print(f"    [MATCH]   Cleaned: '{account_clean}'")
        print(f"    [MATCH]   DFS keys: {list(dfs.keys())}")

    mapping_key, mapping_config, category = _resolve_mapping_alias(account_name, mappings)

    # Helper: collect the normalised identifiers for a dfs entry (tab key + block_title).
    def _dfs_names(key: str, df: pd.DataFrame) -> List[str]:
        names = [_normalize_account_name(key)]
        block_title = df.attrs.get("block_title") or ""
        if block_title:
            norm_bt = _normalize_account_name(block_title)
            if norm_bt and norm_bt != names[0]:
                names.append(norm_bt)
        return names

    # STEP 0: Always honor an exact source-account to DFS-key/block_title match first.
    for dfs_key, dfs_df_candidate in dfs.items():
        if account_normalized in _dfs_names(dfs_key, dfs_df_candidate):
            if debug:
                print(f"    [MATCH]   Step 0: ✅ Exact source-to-tab match '{dfs_key}'")
            if mapping_key:
                return (
                    dfs_key,
                    dfs_df_candidate,
                    category,
                    mapping_key,
                    'Mapped',
                    f"Mapped in mappings.yml as '{mapping_key}' and matched the workbook tab directly.",
                )
            return (
                dfs_key,
                dfs_df_candidate,
                None,
                None,
                'Tab-only match',
                "Matched by workbook tab name only. Add this account to mappings.yml if it should be classified into BS or IS.",
            )

    # STEP 1: Find which mapping KEY this source account belongs to
    if mapping_key and isinstance(mapping_config, dict):
        aliases = mapping_config.get('aliases', [])
        normalized_aliases = {_normalize_account_name(alias) for alias in aliases}
        if debug:
            print(f"    [MATCH]   Step 1: Found in mappings.yml")
            print(f"    [MATCH]     Mapping key: '{mapping_key}'")
            print(f"    [MATCH]     Category: '{category}'")
            print(f"    [MATCH]     Aliases: {aliases}")

        # STEP 2b: Fall back to any alias match against DFS keys or block_titles.
        for dfs_key, dfs_df_candidate in dfs.items():
            if normalized_aliases & set(_dfs_names(dfs_key, dfs_df_candidate)):
                if debug:
                    print(f"    [MATCH]   Step 2b: ✅ DFS key '{dfs_key}' matches alias!")
                return (
                    dfs_key,
                    dfs_df_candidate,
                    category,
                    mapping_key,
                    'Mapped',
                    f"Mapped in mappings.yml as '{mapping_key}'.",
                )

        # No alias matched any dfs key
        if debug:
            print(f"    [MATCH]   Step 2: ❌ No alias matches any dfs key")
            print(f"    [MATCH]     Aliases: {aliases}")
            print(f"    [MATCH]     DFS keys: {list(dfs.keys())}")
        return (
            None,
            None,
            category,
            mapping_key,
            'Mapped but missing tab',
            f"Mapped in mappings.yml as '{mapping_key}', but no matching workbook tab was found.",
        )

    # Not found in any mapping aliases
    if debug:
        print(f"    [MATCH]   ❌ '{account_name}' not in any mappings.yml aliases")

    return (
        None,
        None,
        None,
        None,
        'Missing mapping',
        "This account is not covered by mappings.yml. Add a mapping or alias if it should reconcile to a schedule tab.",
    )


def get_total_from_dfs(dfs_df: pd.DataFrame, date_col: str, debug: bool = False) -> Optional[float]:
    """
    Get total value from DFS dataframe.
    ONLY looks for rows with 'Total', '合计', '总计'.
    Skips subtotal rows ('Subtotal', '小计').
    No fallback - returns None if no total row found.
    
    Args:
        dfs_df: DataFrame from dfs
        date_col: Date column to get value from
        debug: Enable debug output
        
    Returns:
        Total value or None if no total row found
    """
    if dfs_df is None or dfs_df.empty:
        return None

    attrs = dfs_df.attrs or {}

    projection_totals_by_date = attrs.get('projection_totals_by_date') or {}
    if isinstance(projection_totals_by_date, dict):
        projection_total = projection_totals_by_date.get(date_col)
        if projection_total is not None:
            if debug:
                print(f"      Using normalized projection total for '{date_col}': {projection_total:,.0f}")
            return projection_total

    def _scan_table_total(target_date_col: str) -> Optional[float]:
        if target_date_col not in dfs_df.columns:
            return None
        desc_col = dfs_df.columns[0]
        total_keywords = ['合计', '总计', 'total']
        skip_keywords = ['小计', 'subtotal', 'sub-total', 'sub total']
        for _, row in dfs_df.iterrows():
            desc = str(row[desc_col])
            desc_lower = desc.lower()
            if any(skip_kw in desc_lower for skip_kw in skip_keywords):
                continue
            if any(keyword in desc_lower for keyword in total_keywords):
                return row[target_date_col]
        return None

    auxiliary_check_totals_by_date = attrs.get('auxiliary_check_totals_by_date') or {}
    if isinstance(auxiliary_check_totals_by_date, dict):
        auxiliary_total = auxiliary_check_totals_by_date.get(date_col)
        if auxiliary_total is not None:
            main_total = _scan_table_total(date_col)
            if isinstance(main_total, (int, float)) and main_total not in (0, 0.0):
                if abs(auxiliary_total) > 0 and (main_total > 0) != (auxiliary_total > 0):
                    auxiliary_total = abs(auxiliary_total) * (1 if main_total > 0 else -1)
            if debug:
                print(f"      Using auxiliary check total for '{date_col}': {auxiliary_total:,.0f}")
            return auxiliary_total
    
    if date_col not in dfs_df.columns:
        original_column_label = attrs.get('projection_original_column_label')
        original_values_by_description = attrs.get('projection_original_values_by_description') or {}
        if original_column_label == date_col and isinstance(original_values_by_description, dict):
            desc_col = dfs_df.columns[0]
            for _, row in dfs_df.iterrows():
                desc = str(row[desc_col])
                desc_lower = desc.lower()
                if any(skip_kw in desc_lower for skip_kw in ['小计', 'subtotal', 'sub-total', 'sub total']):
                    continue
                if any(keyword in desc_lower for keyword in ['合计', '总计', 'total']):
                    original_total = original_values_by_description.get(desc)
                    if original_total is not None:
                        if debug:
                            print(
                                f"      Using original projection total for '{date_col}' "
                                f"from annualized dataframe row '{desc}': {original_total:,.0f}"
                            )
                        return original_total
        return None
    
    # Try to find total row
    total_value = _scan_table_total(date_col)
    if total_value is not None:
        if debug:
            desc_col = dfs_df.columns[0]
            total_rows = dfs_df[
                dfs_df[desc_col].astype(str).str.lower().str.contains('合计|总计|total', regex=True, na=False)
                & ~dfs_df[desc_col].astype(str).str.lower().str.contains('小计|subtotal|sub-total|sub total', regex=True, na=False)
            ]
            if not total_rows.empty:
                print(f"      Found total row: '{total_rows.iloc[0][desc_col]}' → value: {total_value:,.0f}")
        return total_value
    
    # Fallback: when no explicit total row exists, the schedule may use a
    # parent-first structure where the first non-breakdown row IS the total.
    # Use projection_totals_by_date if available (populated from sum-detected rows
    # marked as total via _detect_implicit_breakdowns_from_sum), otherwise try
    # the first row with a non-zero value as the best-effort total.
    if date_col in dfs_df.columns:
        desc_col = dfs_df.columns[0]
        row_types = dfs_df.attrs.get("row_types_by_description") or {}
        for _, row in dfs_df.iterrows():
            desc = str(row[desc_col])
            # Skip rows that are breakdowns or explicitly subtotal/total (already checked above)
            if row_types.get(desc) in ("breakdown", "subtotal"):
                continue
            candidate = row[date_col]
            if isinstance(candidate, (int, float)) and abs(candidate) > 0:
                if debug:
                    print(f"      Using first-row fallback total for '{date_col}': {candidate:,.0f} from '{desc}'")
                return float(candidate)

    if debug:
        desc_col = dfs_df.columns[0]
        print(f"      ❌ No total row found (no '合计', '总计', or 'Total' in descriptions)")
        print(f"      Available descriptions: {dfs_df[desc_col].tolist()}")

    return None


def _should_compare_income_statement_as_absolute(account_name: str, category: Optional[str]) -> bool:
    category_text = str(category or "").strip().lower()
    account_text = str(account_name or "").strip().lower()
    if any(keyword in category_text for keyword in ("expense", "loss", "cost")):
        return True
    return any(keyword in account_text for keyword in ("loss", "expense", "cost", "损失", "费用", "成本"))


def _integrity_metadata(dfs_df: Optional[pd.DataFrame]) -> Dict[str, str]:
    if dfs_df is None:
        return {
            'Projection_Stage': '-',
            'Projection_Date': '-',
            'Integrity_Flag': '-',
        }

    integrity = dfs_df.attrs.get('integrity') or {}
    preferred_stage = integrity.get('preferred_stage')
    effective_stage = integrity.get('effective_stage')
    effective_date = integrity.get('effective_date')
    non_zero_rows = integrity.get('non_zero_rows')

    flag = '-'
    if preferred_stage and effective_stage and preferred_stage != effective_stage:
        flag = f'Fallback from {preferred_stage} to {effective_stage}'
    elif non_zero_rows == 0:
        flag = 'Zero-value projection'

    return {
        'Projection_Stage': effective_stage or preferred_stage or '-',
        'Projection_Date': effective_date or '-',
        'Integrity_Flag': flag,
    }


def reconcile_financial_statements(
    bs_is_results: Dict,
    dfs: Dict[str, pd.DataFrame],
    mappings_file: str = 'fdd_utils/mappings.yml',
    mappings: Optional[dict] = None,
    tolerance: float = 1.0,
    materiality_threshold: float = 0.005,
    debug: bool = False
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Reconcile Balance Sheet and Income Statement between two data sources.
    Only uses the LATEST date column from BS/IS for comparison.
    
    Args:
        bs_is_results: Results from extract_balance_sheet_and_income_statement
                      with keys 'balance_sheet', 'income_statement', 'project_name'
        dfs: Dictionary of DataFrames from extract_data_from_excel
        mappings_file: Path to mappings.yml file
        mappings: Optional preloaded effective mappings, including dynamic mappings
        tolerance: Absolute tolerance for matching (default: 1.0, allows ±1 difference)
        materiality_threshold: Percentage threshold for immaterial differences (default: 0.005 = 0.5%)
        debug: If True, print debugging information
        
    Returns:
        Tuple of (bs_reconciliation_df, is_reconciliation_df)
        Each DataFrame has columns:
        - Financials_Account: Account name from BS/IS
        - Date: Date column (latest only)
        - Financials_Value: Value from BS/IS (expenses converted to positive)
        - Tab_Account: Actual workbook tab name (e.g., '货币资金', not mapping key 'Cash')
        - Tab_Value: Total value from the schedule tab
        - Match: '✅ Match', '❌ Diff: X', '⚠️ Not Found', or '-' (skipped)
    """
    if debug:
        print("=" * 80)
        print("RECONCILIATION - DEBUG MODE")
        print("=" * 80)
    
    mappings = mappings or load_mappings(mappings_file)
    
    bs_recon_rows = []
    is_recon_rows = []
    
    # Reconcile Balance Sheet
    if bs_is_results.get('balance_sheet') is not None:
        bs_df = bs_is_results['balance_sheet']
        date_cols = [col for col in bs_df.columns if col != 'Description']
        
        # Use only the LATEST date column (LAST one, as dates are typically oldest to newest)
        latest_date = date_cols[-1] if date_cols else None
        
        if debug:
            print(f"\n[RECON] Reconciling Balance Sheet...")
            print(f"[RECON]   Accounts to check: {len(bs_df)}")
            print(f"[RECON]   Available dates: {date_cols}")
            print(f"[RECON]   Using latest date (last column): {latest_date}")
        
        if latest_date:
            # Use the 2 most recent columns for the zero-source check:
            # only skip if both are zero (loosened to keep items with adjacent-period data).
            recent_dates = date_cols[-2:] if len(date_cols) >= 2 else date_cols

            for idx, row in bs_df.iterrows():
                account_name = row['Description']
                source_value = row[latest_date]

                # Skip only when ALL of the most-recent columns are zero
                recent_values = [row[d] for d in recent_dates]
                if all(v == 0 for v in recent_values):
                    integrity_fields = _integrity_metadata(None)
                    bs_recon_rows.append({
                        'Financials_Account': account_name,
                        'Date': latest_date,
                        'Financials_Value': source_value,
                        'Tab_Account': '-',
                        'Tab_Value': '-',
                        'Diff': '-',
                        'Match': '-',
                        'Mapping_Key': '-',
                        'Mapping_Status': 'Zero source',
                        'Mapping_Note': 'Most recent period values are all 0, so schedule mapping was skipped.',
                        **integrity_fields,
                    })
                    continue

                # Flag: latest period is 0 but an adjacent period has data
                zero_with_adjacent = source_value == 0 and any(v != 0 for v in recent_values)

                # Find matching account in dfs (ONLY via mappings.yml)
                dfs_key, dfs_df, category, mapping_key, mapping_status, mapping_note = find_account_in_dfs(account_name, dfs, mappings, debug=debug)

                # Handle skipped accounts (totals/profit lines)
                if dfs_key == 'SKIP':
                    integrity_fields = _integrity_metadata(None)
                    bs_recon_rows.append({
                        'Financials_Account': account_name,
                        'Date': latest_date,
                        'Financials_Value': source_value,
                        'Tab_Account': '-',
                        'Tab_Value': '-',
                        'Diff': '-',
                        'Match': '-',
                        'Mapping_Key': '-',
                        'Mapping_Status': mapping_status,
                        'Mapping_Note': mapping_note,
                        **integrity_fields,
                    })
                    continue

                # Get total value from dfs
                dfs_value = get_total_from_dfs(dfs_df, latest_date, debug) if dfs_df is not None else None

                # When the tab exists but has no column for this date and the
                # source is 0 (zero_with_adjacent), treat the missing value as 0
                # — the schedule simply has no data for that period.
                if dfs_value is None and zero_with_adjacent and dfs_df is not None:
                    dfs_value = 0.0

                # Check match
                if dfs_value is None:
                    match_status = '⚠️ Not Found'
                    difference = None
                else:
                    difference = abs(source_value - dfs_value)

                    # Check if within absolute tolerance
                    if difference <= tolerance:
                        match_status = '⚠️ Match' if zero_with_adjacent else '✅ Match'
                    else:
                        # Check if within materiality threshold (percentage)
                        if source_value != 0:
                            pct_diff = difference / abs(source_value)
                            if pct_diff <= materiality_threshold:
                                match_status = '✅ Immaterial'
                            else:
                                match_status = '❌ Diff'
                        else:
                            match_status = '❌ Diff'

                integrity_fields = _integrity_metadata(dfs_df)
                bs_recon_rows.append({
                    'Financials_Account': account_name,
                    'Date': latest_date,
                    'Financials_Value': source_value,
                    'Tab_Account': dfs_key or 'Not Found',
                    'Tab_Value': dfs_value if dfs_value is not None else 0,
                    'Diff': difference if difference is not None else '-',
                    'Match': match_status,
                    'Mapping_Key': mapping_key or '-',
                    'Mapping_Status': mapping_status,
                    'Mapping_Note': mapping_note,
                    **integrity_fields,
                })
    
    # Reconcile Income Statement
    if bs_is_results.get('income_statement') is not None:
        is_df = bs_is_results['income_statement']
        date_cols = [col for col in is_df.columns if col != 'Description']
        
        # Use only the LATEST date column (LAST one, as dates are typically oldest to newest)
        latest_date = date_cols[-1] if date_cols else None
        
        if debug:
            print(f"\n[RECON] Reconciling Income Statement...")
            print(f"[RECON]   Accounts to check: {len(is_df)}")
            print(f"[RECON]   Available dates: {date_cols}")
            print(f"[RECON]   Using latest date (last column): {latest_date}")
        
        if latest_date:
            # Use the 2 most recent columns for the zero-source check:
            # only skip if both are zero (loosened to keep items with adjacent-period data).
            recent_dates = date_cols[-2:] if len(date_cols) >= 2 else date_cols

            for idx, row in is_df.iterrows():
                account_name = row['Description']
                source_value_raw = row[latest_date]

                # Skip only when ALL of the most-recent columns are zero
                recent_values = [row[d] for d in recent_dates]
                if all(v == 0 for v in recent_values):
                    integrity_fields = _integrity_metadata(None)
                    is_recon_rows.append({
                        'Financials_Account': account_name,
                        'Date': latest_date,
                        'Financials_Value': source_value_raw,
                        'Tab_Account': '-',
                        'Tab_Value': '-',
                        'Diff': '-',
                        'Match': '-',
                        'Mapping_Key': '-',
                        'Mapping_Status': 'Zero source',
                        'Mapping_Note': 'Most recent period values are all 0, so schedule mapping was skipped.',
                        **integrity_fields,
                    })
                    continue

                # Flag: latest period is 0 but an adjacent period has data
                zero_with_adjacent = source_value_raw == 0 and any(v != 0 for v in recent_values)

                # Find matching account in dfs (ONLY via mappings.yml)
                dfs_key, dfs_df, category, mapping_key, mapping_status, mapping_note = find_account_in_dfs(account_name, dfs, mappings, debug=debug)

                # Handle skipped accounts (totals/profit lines)
                if dfs_key == 'SKIP':
                    integrity_fields = _integrity_metadata(None)
                    is_recon_rows.append({
                        'Financials_Account': account_name,
                        'Date': latest_date,
                        'Financials_Value': source_value_raw,
                        'Tab_Account': '-',
                        'Tab_Value': '-',
                        'Diff': '-',
                        'Match': '-',
                        'Mapping_Key': '-',
                        'Mapping_Status': mapping_status,
                        'Mapping_Note': mapping_note,
                        **integrity_fields,
                    })
                    continue

                # Get total value from dfs
                dfs_value = get_total_from_dfs(dfs_df, latest_date, debug) if dfs_df is not None else None

                # When the tab exists but has no column for this date and the
                # source is 0 (zero_with_adjacent), treat the missing value as 0
                # — the schedule simply has no data for that period.
                if dfs_value is None and zero_with_adjacent and dfs_df is not None:
                    dfs_value = 0.0

                # For expense/loss-style lines: keep the original sign in display but compare on absolute value.
                source_for_comparison = source_value_raw
                dfs_for_comparison = dfs_value
                if _should_compare_income_statement_as_absolute(account_name, category):
                    source_for_comparison = abs(source_value_raw)
                    if dfs_for_comparison is not None:
                        dfs_for_comparison = abs(dfs_for_comparison)
                    if debug and (source_value_raw < 0 or (dfs_value is not None and dfs_value < 0)):
                        print(
                            f"    [CONVERT] Compare absolute values: source {source_value_raw:,.0f} → {source_for_comparison:,.0f}; "
                            f"dfs {dfs_value if dfs_value is not None else 'None'} → {dfs_for_comparison if dfs_for_comparison is not None else 'None'}"
                        )

                # Check match
                if dfs_value is None:
                    match_status = '⚠️ Not Found'
                    difference = None
                else:
                    difference = abs(source_for_comparison - dfs_for_comparison)

                    # Check if within absolute tolerance
                    if difference <= tolerance:
                        match_status = '⚠️ Match' if zero_with_adjacent else '✅ Match'
                    else:
                        # Check if within materiality threshold (percentage)
                        if source_for_comparison != 0:
                            pct_diff = difference / abs(source_for_comparison)
                            if pct_diff <= materiality_threshold:
                                match_status = '✅ Immaterial'
                            else:
                                match_status = '❌ Diff'
                        else:
                            match_status = '❌ Diff'
                
                integrity_fields = _integrity_metadata(dfs_df)
                is_recon_rows.append({
                    'Financials_Account': account_name,
                    'Date': latest_date,
                    'Financials_Value': source_value_raw,  # Keep original negative value
                    'Tab_Account': dfs_key or 'Not Found',
                    'Tab_Value': dfs_value if dfs_value is not None else 0,
                    'Diff': difference if difference is not None else '-',
                    'Match': match_status,
                    'Mapping_Key': mapping_key or '-',
                    'Mapping_Status': mapping_status,
                    'Mapping_Note': mapping_note,
                    **integrity_fields,
                })
    
    # Create DataFrames
    bs_recon_df = pd.DataFrame(bs_recon_rows) if bs_recon_rows else pd.DataFrame()
    is_recon_df = pd.DataFrame(is_recon_rows) if is_recon_rows else pd.DataFrame()
    
    if debug:
        print("\n" + "=" * 80)
        print("RECONCILIATION SUMMARY")
        print("=" * 80)
        
        if not bs_recon_df.empty:
            matches = (bs_recon_df['Match'] == '✅ Match').sum()
            mismatches = bs_recon_df['Match'].str.contains('❌').sum()
            not_found = (bs_recon_df['Match'] == '⚠️ Not Found').sum()
            print(f"Balance Sheet: {len(bs_recon_df)} comparisons")
            print(f"  ✅ Matches: {matches}")
            print(f"  ❌ Mismatches: {mismatches}")
            print(f"  ⚠️  Not Found: {not_found}")
        
        if not is_recon_df.empty:
            matches = (is_recon_df['Match'] == '✅ Match').sum()
            mismatches = is_recon_df['Match'].str.contains('❌').sum()
            not_found = (is_recon_df['Match'] == '⚠️ Not Found').sum()
            print(f"\nIncome Statement: {len(is_recon_df)} comparisons")
            print(f"  ✅ Matches: {matches}")
            print(f"  ❌ Mismatches: {mismatches}")
            print(f"  ⚠️  Not Found: {not_found}")
    
    return bs_recon_df, is_recon_df


def print_reconciliation_report(bs_recon_df: pd.DataFrame, is_recon_df: pd.DataFrame, 
                                show_only_issues: bool = False):
    """
    Print a formatted reconciliation report.
    
    Args:
        bs_recon_df: Balance Sheet reconciliation DataFrame
        is_recon_df: Income Statement reconciliation DataFrame
        show_only_issues: If True, only show mismatches and not found items
    """
    print("\n" + "=" * 100)
    print("RECONCILIATION REPORT")
    print("=" * 100)
    
    # Balance Sheet
    if not bs_recon_df.empty:
        print("\n📊 BALANCE SHEET RECONCILIATION")
        print("-" * 100)
        
        df_to_show = bs_recon_df.copy()
        if show_only_issues:
            df_to_show = df_to_show[df_to_show['Match'] != '✅ Match']
        
        if not df_to_show.empty:
            # Format for display
            df_display = df_to_show.copy()
            df_display['Financials_Value'] = df_display['Financials_Value'].apply(
                lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x
            )
            df_display['Tab_Value'] = df_display['Tab_Value'].apply(
                lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x
            )
            df_display['Diff'] = df_display['Diff'].apply(
                lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x
            )
            
            print(df_display.to_string(index=False))
        else:
            print("✅ All accounts match perfectly!")
    
    # Income Statement
    if not is_recon_df.empty:
        print("\n\n📈 INCOME STATEMENT RECONCILIATION")
        print("-" * 100)
        
        df_to_show = is_recon_df.copy()
        if show_only_issues:
            df_to_show = df_to_show[df_to_show['Match'] != '✅ Match']
        
        if not df_to_show.empty:
            # Format for display
            df_display = df_to_show.copy()
            df_display['Financials_Value'] = df_display['Financials_Value'].apply(
                lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x
            )
            df_display['Tab_Value'] = df_display['Tab_Value'].apply(
                lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x
            )
            df_display['Diff'] = df_display['Diff'].apply(
                lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x
            )
            
            print(df_display.to_string(index=False))
        else:
            print("✅ All accounts match perfectly!")
    
    print("\n" + "=" * 100)


# Example usage
if __name__ == "__main__":
    print("=" * 80)
    print("RECONCILIATION EXAMPLE")
    print("=" * 80)
    
    example = find_reconciliation_example()
    if not example:
        raise FileNotFoundError(
            "No local reconciliation example workbook with a 'Financials' sheet was found."
        )

    databook_path = example["workbook_path"]
    sheet_name = example["sheet_name"]
    entity_name = example["entity_name"]
    print(f"Using workbook: {databook_path}")
    print(f"Using financial sheet: {sheet_name}")
    
    # Source 1: Financial extraction (BS/IS from single sheet)
    bs_is_results = extract_balance_sheet_and_income_statement(
        workbook_path=databook_path,
        sheet_name=sheet_name,
        debug=False
    )
    
    # Source 2: DFS extraction (account by account)
    dfs, workbook_list, _, language = extract_data_from_excel(
        databook_path=databook_path,
        entity_name=entity_name,
        mode="All"
    )
    
    # Reconcile
    bs_recon, is_recon = reconcile_financial_statements(
        bs_is_results=bs_is_results,
        dfs=dfs,
        tolerance=1.0,
        materiality_threshold=0.005,  # 0.5% materiality
        debug=True
    )
    
    # Print report
    print_reconciliation_report(bs_recon, is_recon, show_only_issues=True)
    
    # Save to Excel
    if not bs_recon.empty:
        with pd.ExcelWriter('reconciliation_report.xlsx') as writer:
            bs_recon.to_excel(writer, sheet_name='Balance Sheet', index=False)
            if not is_recon.empty:
                is_recon.to_excel(writer, sheet_name='Income Statement', index=False)
        print("\n✅ Reconciliation report saved to: reconciliation_report.xlsx")
# --- end workbook/reconcile.py ---

# --- begin workbook/flow.py ---
import contextlib
import io
import logging
import os
import time
from typing import Any, Dict

logger = logging.getLogger(__name__)


def process_workbook_data(
    *,
    temp_path: str,
    entity_name: str,
    selected_sheet: str | None,
    mapping_overrides: Dict[str, str] | None = None,
    debug: bool = False,
) -> Dict[str, Any]:
    process_started = time.perf_counter()
    normalized_results, normalized_workbook_list, _, language, resolution = extract_normalized_data_from_excel(
        databook_path=temp_path,
        entity_name=entity_name,
        mode="All",
        mapping_overrides=mapping_overrides or None,
    )
    logger.debug(
        "Built normalized workbook payload for %s in %.2fs",
        os.path.basename(temp_path),
        time.perf_counter() - process_started,
    )

    dataframe_variants = build_dataframe_variants_from_normalized_results(
        normalized_results=normalized_results,
        workbook_list=normalized_workbook_list,
        report_language=language,
        variant_specs=[
            {
                "name": f"{view_name}_{variant}",
                "variant": variant,
                "filter_details": filter_details,
                "keep_zero_rows": False,
            }
            for view_name, filter_details in (("display", False), ("detail", True))
            for variant in ("original", "default", "analysis")
        ],
    )
    display_dfs_original = dataframe_variants.get("display_original", {}).get("dfs", {})
    display_workbook_list = dataframe_variants.get("display_original", {}).get("workbook_list", [])
    display_dfs = dataframe_variants.get("display_default", {}).get("dfs", {})
    dfs_original = dataframe_variants.get("detail_original", {}).get("dfs", {})
    # Use analysis variant (all Indicative adjusted periods) as primary for AI
    dfs = dataframe_variants.get("detail_analysis", {}).get("dfs", {})
    if not dfs:
        dfs = dataframe_variants.get("detail_default", {}).get("dfs", {})
    workbook_list = dataframe_variants.get("detail_analysis", {}).get("workbook_list", [])
    if not workbook_list:
        workbook_list = dataframe_variants.get("detail_default", {}).get("workbook_list", [])

    debug_buffer = io.StringIO() if debug else None
    debug_ctx = contextlib.redirect_stdout(debug_buffer) if debug_buffer else contextlib.nullcontext()

    bs_is_results = None
    if selected_sheet:
        bs_started = time.perf_counter()
        with debug_ctx:
            bs_is_results = extract_balance_sheet_and_income_statement(
                workbook_path=temp_path,
                sheet_name=selected_sheet,
                debug=debug,
            )
        logger.debug(
            "Extracted financial summary sheet %s in %.2fs",
            selected_sheet,
            time.perf_counter() - bs_started,
        )

    recon_bs, recon_is = None, None
    if dfs_original and bs_is_results:
        recon_started = time.perf_counter()
        effective_mappings = get_effective_mappings(load_mappings(), resolution)
        with debug_ctx:
            recon_bs, recon_is = reconcile_financial_statements(
                bs_is_results=bs_is_results,
                dfs=dfs_original,
                mappings=effective_mappings,
                tolerance=1.0,
                materiality_threshold=0.005,
                debug=debug,
            )
        logger.debug(
            "Reconciled %s account tabs in %.2fs",
            len(dfs),
            time.perf_counter() - recon_started,
        )

    logger.debug(
        "Finished Process Data for %s in %.2fs",
        os.path.basename(temp_path),
        time.perf_counter() - process_started,
    )

    return {
        "dfs": dfs,
        "display_dfs": display_dfs,
        "dfs_variants": {
            "default": dfs,
            "original": dfs_original,
        },
        "display_df_variants": {
            "default": display_dfs,
            "original": display_dfs_original,
        },
        "workbook_list": workbook_list,
        "display_workbook_list": display_workbook_list,
        "language": language,
        "bs_is_results": bs_is_results,
        "reconciliation": (recon_bs, recon_is),
        "resolution": resolution,
        "project_name": bs_is_results.get("project_name") if bs_is_results else None,
        "entity_name": entity_name,
        "display_dfs_original": display_dfs_original,
        "debug_output": debug_buffer.getvalue() if debug_buffer else "",
    }
# --- end workbook/flow.py ---
