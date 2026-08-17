from __future__ import annotations

# re-added: bound by an import in another section of the pre-split module
from typing import Dict, List, Tuple, Optional, Any, Sequence, Union
from functools import lru_cache

"""
Normalize FDD financial schedules into integrity-aware payloads.
"""

from .analysis import build_significant_movements, build_trend_summary
from .inspector import _cell_text, _coerce_numeric, canonical_stage_label, contains_unit_marker, date_row_index, load_workbook_frames, profile_sheet, stage_row_indices
from .statements import _find_description_column, parse_date


from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional
import math
import re

import pandas as pd

from ..financial_common import cell_text, coerce_numeric
from ..keyword_registry import UNIT_THOUSAND_MARKERS


PREFERRED_STAGE = "Indicative adjusted"
INTERNAL_ROW_KEY = "__source_row_idx"
_TOTAL_KEYWORDS = ("total", "总计", "合计", "總計", "合計")
_SUBTOTAL_KEYWORDS = ("subtotal", "sub-total", "sub total", "小计", "小計")
_UNIT_MARKERS = tuple(UNIT_THOUSAND_MARKERS)
_WORKING_REMARK_KEYWORDS = ("check", "对账单", "對賬單", "对帐单", "差异", "差異", "difference", "diff", "recon")
_CARRYING_AMOUNT_LABELS = ("carrying amounts", "net book value", "carrying amount", "账面价值", "賬面價值")


def _forward_fill_stage_row(row: pd.Series) -> Dict[int, Optional[str]]:
    labels: Dict[int, Optional[str]] = {}
    current = None
    for col_idx, value in enumerate(row.tolist()):
        text = _cell_text(value)
        detected = canonical_stage_label(value)
        if detected:
            current = detected
        elif text:
            # Stop the stage span when a non-stage header appears so trailing
            # note/recon columns stay available as remarks but do not get
            # treated as additional financial value columns.
            current = None
        labels[col_idx] = current
    return labels


def _stage_row_indices(df: pd.DataFrame) -> List[int]:
    return stage_row_indices(df, parse_date)


def _block_title_for_stage_row(df: pd.DataFrame, stage_row_idx: int, sheet_name: str) -> str:
    for row_idx in range(stage_row_idx - 1, max(-1, stage_row_idx - 5), -1):
        texts = [_cell_text(value) for value in df.iloc[row_idx].tolist()]
        texts = [text for text in texts if text]
        if not texts:
            continue
        if any(canonical_stage_label(text) for text in texts):
            continue
        if any(parse_date(text) for text in texts):
            continue
        lowered = " ".join(text.lower() for text in texts)
        if contains_unit_marker(lowered, _UNIT_MARKERS):
            continue
        return " ".join(texts)
    return sheet_name


def _extract_entity_name_from_block_title(title: str) -> Optional[str]:
    if not title:
        return None
    parts = re.split(r"\s[-–]\s", title, maxsplit=1)
    if len(parts) != 2:
        return None
    candidate = parts[1].strip()
    if not candidate:
        return None
    if candidate.lower().startswith("project "):
        return None
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", candidate):
        return None
    return candidate


def _is_strict_entity_title_match(block_title: str, block_entity_name: Optional[str], entity_name: str) -> bool:
    normalized_entity = str(entity_name or "").strip().lower()
    if not normalized_entity:
        return False

    normalized_block_entity = str(block_entity_name or "").strip().lower()
    if normalized_block_entity:
        return normalized_block_entity == normalized_entity

    normalized_title = str(block_title or "").strip().lower()
    pattern = rf"(^|[\s\-–_/()]){re.escape(normalized_entity)}($|[\s\-–_/()])"
    return bool(re.search(pattern, normalized_title))


def _select_entity_block(
    df: pd.DataFrame,
    sheet_name: str,
    default_stage_row_idx: int,
    entity_name: Optional[str],
) -> Dict[str, Any]:
    candidates = _stage_row_indices(df)
    if not candidates:
        candidates = [default_stage_row_idx]

    blocks: List[Dict[str, Any]] = []
    for index, candidate_stage_row_idx in enumerate(candidates):
        next_stage_row_idx = candidates[index + 1] if index + 1 < len(candidates) else len(df)
        block_title = _block_title_for_stage_row(df, candidate_stage_row_idx, sheet_name)
        block_entity_name = _extract_entity_name_from_block_title(block_title)
        block_context = " ".join(
            _cell_text(value).lower()
            for value in df.iloc[max(0, candidate_stage_row_idx - 5) : min(len(df), candidate_stage_row_idx + 2)].to_numpy(dtype=object, copy=False).ravel()
        )
        blocks.append(
            {
                "stage_row_idx": candidate_stage_row_idx,
                "date_row_idx": _local_date_row_index(df, candidate_stage_row_idx),
                "data_end_row": next_stage_row_idx,
                "block_title": block_title,
                "block_entity_name": block_entity_name,
                "context": block_context,
            }
        )

    default_block = next((block for block in blocks if block["stage_row_idx"] == default_stage_row_idx), blocks[0])
    if not entity_name:
        return {**default_block, "strict_entity_match": False}

    entity_text = str(entity_name).strip().lower()
    strict_matches = [
        block
        for block in blocks
        if _is_strict_entity_title_match(
            block_title=str(block["block_title"]),
            block_entity_name=block.get("block_entity_name"),
            entity_name=entity_text,
        )
    ]
    if strict_matches:
        working_blocks = strict_matches
        strict_entity_match = True
    else:
        working_blocks = blocks
        strict_entity_match = False

    entity_parts = [part for part in re.split(r"\s+", entity_text) if len(part) >= 3]
    best_block = default_block
    best_score = -1
    for block in working_blocks:
        score = 0
        title_lower = str(block["block_title"]).lower()
        entity_lower = str(block.get("block_entity_name") or "").lower()
        if entity_text and entity_text == entity_lower:
            score += 10
        if entity_text and entity_text in title_lower:
            score += 6
        if entity_text and entity_text in block["context"]:
            score += 4
        score += sum(1 for part in entity_parts if part in title_lower)
        score += sum(1 for part in entity_parts if part in block["context"])
        if score > best_score:
            best_score = score
            best_block = block
    return {**best_block, "strict_entity_match": strict_entity_match}


def _local_date_row_index(df: pd.DataFrame, stage_row_idx: int) -> int:
    # `or stage_row_idx` would incorrectly discard a legitimate result of row 0
    # (falsy in Python) and fall back to the stage row itself -- explicit None
    # check needed since 0 is a valid date-row index (e.g. sheets with no title
    # row above the header, where the date row is the very first row).
    # Bare numbers are refused while CHOOSING the row: a row of 千元 balances
    # scores higher than the real date row otherwise (see parse_date).
    result = date_row_index(
        df, stage_row_idx,
        lambda value: parse_date(value, allow_bare_number=False),
        max_distance=2,
    )
    return result if result is not None else stage_row_idx


def _rollforward_header_row_index(df: pd.DataFrame, desc_col_idx: int) -> Optional[int]:
    for row_idx in range(min(len(df), 12)):
        desc_cell = _cell_text(df.iloc[row_idx, desc_col_idx]) if desc_col_idx < len(df.columns) else ""
        cells = [_cell_text(value) for value in df.iloc[row_idx, desc_col_idx + 1 :].tolist()]
        if not cells:
            continue
        has_unit = contains_unit_marker(desc_cell, _UNIT_MARKERS) or any(
            contains_unit_marker(cell, _UNIT_MARKERS) for cell in cells if cell
        )
        has_component_header = any(
            cell
            and not parse_date(cell)
            and not contains_unit_marker(cell, _UNIT_MARKERS)
            for cell in cells
        )
        if has_unit and has_component_header:
            return row_idx
    return None


def _standardize_rollforward_schedule_df(
    df: pd.DataFrame,
    sheet_name: str,
    profile: Dict[str, Any],
    desc_col_idx: int,
) -> Optional[pd.DataFrame]:
    carrying_row_idx = None
    for row_idx in range(len(df)):
        description = _cell_text(df.iloc[row_idx, desc_col_idx]).lower()
        if description in _CARRYING_AMOUNT_LABELS:
            carrying_row_idx = row_idx
            break
    if carrying_row_idx is None:
        return None

    header_row_idx = _rollforward_header_row_index(df, desc_col_idx)
    if header_row_idx is None:
        return None

    component_columns: List[Dict[str, Any]] = []
    for col_idx in range(desc_col_idx + 1, len(df.columns)):
        header = _cell_text(df.iloc[header_row_idx, col_idx])
        if not header:
            continue
        if parse_date(header):
            continue
        lowered = header.lower()
        if any(marker in lowered for marker in _UNIT_MARKERS):
            continue
        component_columns.append({"col_idx": col_idx, "header": header})
    if not component_columns:
        return None

    date_rows: List[Dict[str, Any]] = []
    for row_idx in range(carrying_row_idx + 1, len(df)):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        lowered = description.lower()
        if lowered.startswith("check"):
            break
        parsed_date = parse_date(description)
        if not parsed_date:
            continue
        date_rows.append(
            {
                "row_idx": row_idx,
                "date": parsed_date.strftime("%Y-%m-%d"),
            }
        )
    if len(date_rows) < 2:
        return None

    usable_components: List[Dict[str, Any]] = []
    for component in component_columns:
        if any(_coerce_numeric(df.iloc[item["row_idx"], component["col_idx"]]) is not None for item in date_rows):
            usable_components.append(component)
    if not usable_components:
        return None

    block_title = str(profile.get("title") or sheet_name).strip() or sheet_name
    temp_rows: List[List[Any]] = []
    temp_rows.append([block_title])
    temp_rows.append(["", *([PREFERRED_STAGE] * len(date_rows))])
    temp_rows.append([block_title, *[item["date"] for item in date_rows]])
    for component in usable_components:
        temp_rows.append(
            [
                component["header"],
                *[_coerce_numeric(df.iloc[item["row_idx"], component["col_idx"]]) for item in date_rows],
            ]
        )
    return pd.DataFrame(temp_rows)


def _is_numeric_enough(df: pd.DataFrame, col_idx: int, data_start_row: int) -> bool:
    values = [_coerce_numeric(value) for value in df.iloc[data_start_row:, col_idx].tolist()]
    non_empty = [value for value in values if value is not None]
    return len(non_empty) >= 1


def _dedupe_columns_by_key(columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen_keys: set[str] = set()
    for column in columns:
        key = str(column.get("key") or "")
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        deduped.append(column)
    return deduped


def _row_type(description: str) -> str:
    lowered = description.lower()
    if lowered.startswith("of which") or lowered.startswith("其中"):
        return "breakdown"
    if any(keyword in lowered for keyword in _SUBTOTAL_KEYWORDS):
        return "subtotal"
    if any(keyword in lowered for keyword in _TOTAL_KEYWORDS):
        return "total"
    if re.match(r"^\d{6,}", description):
        return "breakdown"
    if any(token in description for token in ("有限公司", "股份", "基金", "公司")):
        return "breakdown"
    return "detail"


def _detect_implicit_breakdowns_from_sum(
    row_entries: List[Dict[str, Any]],
    projection_column_key: str,
    tolerance_ratio: float = 0.005,
) -> None:
    """
    Detect rows that are implicit breakdowns of a parent detail row.

    In some Excel schedules the parent total appears first (without a "total"
    keyword) followed by its component sub-rows, e.g.:

        Cash at bank:  1,000        ← parent detail row
          Bank A CNY:    600        ← sub-row (no "其中" / "of which" prefix)
          Bank B USD:    400        ← sub-row

    When the sum of consecutive following detail rows equals the parent's value
    within a small tolerance AND every child is strictly smaller than the parent,
    those following rows are re-classified as "breakdown" so they are filtered
    out of the AI prompt data and reconciliation totals.

    This only runs when there are NO explicitly-typed total/subtotal rows in the
    schedule (i.e. when keyword-based detection has already handled the standard
    case).  Mutates row_entries in-place.
    """
    n = len(row_entries)
    # Only apply when no explicit total/subtotal rows exist
    if any(r["row_type"] in ("total", "subtotal") for r in row_entries):
        return

    marked: set[int] = set()
    for i in range(n - 1):
        if i in marked:
            continue
        parent = row_entries[i]
        if parent["row_type"] != "detail":
            continue
        parent_val = parent["values"].get(projection_column_key)
        if parent_val is None or abs(parent_val) < 1.0:
            continue

        running_sum = 0.0
        child_candidates: List[int] = []
        for j in range(i + 1, min(i + 20, n)):
            if j in marked:
                break
            child = row_entries[j]
            if child["row_type"] != "detail":
                break
            child_val = child["values"].get(projection_column_key) or 0.0
            # Each child must be strictly smaller than the parent
            if abs(child_val) >= abs(parent_val):
                break
            running_sum += child_val
            child_candidates.append(j)
            if len(child_candidates) >= 2:
                tolerance = max(1.0, abs(parent_val) * tolerance_ratio)
                if abs(running_sum - parent_val) <= tolerance:
                    for idx in child_candidates:
                        row_entries[idx]["row_type"] = "breakdown"
                        marked.add(idx)
                    break


def _extract_indent_signal_rows(ws, df: pd.DataFrame, desc_col_idx: int) -> List[Tuple[int, int, str]]:
    """(row_idx, indent_level, stripped_description) for every non-empty
    description cell in one already-open worksheet's description column."""
    rows: List[Tuple[int, int, str]] = []
    for row_idx in range(len(df)):
        cell = ws.cell(row=row_idx + 1, column=desc_col_idx + 1)  # openpyxl is 1-indexed
        raw_value = cell.value
        if raw_value is None or not isinstance(raw_value, str):
            continue
        stripped = raw_value.strip()
        if not stripped:
            continue
        try:
            indent_level = int(cell.alignment.indent or 0)
        except Exception:
            indent_level = 0
        rows.append((row_idx, indent_level, stripped))
    return rows


@lru_cache(maxsize=4)
def _build_indent_signal_index(workbook_path: str) -> Dict[str, List[Tuple[int, int, str]]]:
    """Precomputes the indent-level signal for every sheet's description
    column via exactly ONE openpyxl load, entirely synchronous, done for
    ALL sheets upfront rather than lazily per-sheet. normalize_financial_schedule's
    real caller (extract_normalized_data_from_excel) fans out across a
    ThreadPoolExecutor, and openpyxl's Workbook/Worksheet objects are not
    documented as safe for concurrent multi-threaded construction/access --
    computing this here, once, before any worker thread can touch it, avoids
    that risk by construction rather than relying on incidental GIL behavior.
    Cached per workbook_path, same convention as load_workbook_frames/profile_workbook.
    Only sheets with at least one real indent level (>0) are kept."""
    try:
        from openpyxl import load_workbook as _load_workbook_raw
        wb_raw = _load_workbook_raw(workbook_path, data_only=True)
    except Exception:
        return {}
    try:
        workbook_frames = load_workbook_frames(workbook_path)
    except Exception:
        return {}
    index: Dict[str, List[Tuple[int, int, str]]] = {}
    for sheet_name, df in workbook_frames.items():
        if sheet_name not in wb_raw.sheetnames or df is None or df.empty:
            continue
        desc_col_idx = _find_description_column(df)
        if desc_col_idx is None:
            continue
        rows = _extract_indent_signal_rows(wb_raw[sheet_name], df, desc_col_idx)
        if any(level > 0 for _, level, _ in rows):
            index[sheet_name] = rows
    return index


def _infer_indent_hierarchy(rows: List[Tuple[int, int, str]]) -> Dict[int, List[int]]:
    """rows: (row_idx, indent_level, label) for every non-empty description
    cell in a tab, in sheet order. Returns {parent_row_idx: [direct_child_row_idx, ...]}
    via the standard Excel-outline convention: a row's parent is the NEAREST
    preceding row with a STRICTLY LOWER indent level (a classic stack walk) --
    correctly handles multi-level nesting, not just flat one-level pairs."""
    children_of: Dict[int, List[int]] = {}
    stack: List[Tuple[int, int]] = []  # (indent_level, row_idx)
    for row_idx, indent_level, _label in rows:
        while stack and stack[-1][0] >= indent_level:
            stack.pop()
        if stack:
            parent_row_idx = stack[-1][1]
            children_of.setdefault(parent_row_idx, []).append(row_idx)
        stack.append((indent_level, row_idx))
    return children_of


def _reclassify_indent_rollup_children(
    row_entries: List[Dict[str, Any]],
    projection_column_key: str,
    workbook_path: str,
    sheet_name: str,
) -> None:
    """Excel's own indent level (Format > Cells > Alignment > Indent) is a
    direct structural signal for sub-item ("缩行") rows that pandas'
    read_excel silently discards, so a genuinely-indented child row survives
    only if it happens to also match one of _row_type()'s textual heuristics
    -- otherwise it's classified "detail", identical to a real account, and
    extracted as its own standalone line.

    When a parent row's own value already equals the sum of its indented
    children (within tolerance), those children are already counted in the
    parent's total -- reclassify them "breakdown" so they're excluded from
    becoming standalone accounts, the same treatment an existing "其中"/"of
    which" row already gets. Deliberately conservative: does NOT touch a
    group where the parent has no value of its own (a pure category-label /
    section-header row, not a rollup total -- its children are the real
    leaf-level data) or where the sum doesn't match within tolerance (a
    structurally different relationship, e.g. a contra/provision line, that
    needs individual review rather than a blanket rule). Mutates row_entries
    in place; no-ops fast when the sheet has no real indent signal at all.

    Note: an earlier version cross-checked each precomputed row's text
    against row_entries' resolved description before trusting it (guarding
    against a standardized/rebuilt sheet_df, e.g. a rollforward schedule,
    whose row indices might not line up 1:1 with the original Excel rows).
    That check was removed -- it was rejecting real matches on real data for
    reasons not fully pinned down, and is not the primary safety net anyway:
    the tolerance-based value-match below already guards against acting on a
    wrong row correspondence, since two genuinely different rows coincidentally
    summing to within 0.5% of each other is very unlikely for real financial
    figures. This also now matches inspect_databook.py's check_indent_signals,
    which never had this extra filter and reliably finds the correct matches.

    Second pass (added after real-file evidence): _infer_indent_hierarchy
    only looks BACKWARD for a parent (nearest preceding row with strictly
    lower indent) -- correct for a "total row, then its indented
    breakdown" layout, but some real supporting schedules instead list
    components first and put the actual total row LAST (or the true
    parent is a title/header row outside row_entries entirely, so no
    value is ever found there). A top-level sibling group orphaned this
    way (real values, but their assigned "parent" has no value in
    row_entries) is retried against every row_entries row already
    classified "total"/"subtotal" by _row_type()'s existing, position-
    independent textual heuristics -- matched by value, from anywhere in
    the sheet, not just nearby. Only reclassifies when exactly one such
    row's value matches the group's sum, so an ambiguous or coincidental
    match is left alone rather than guessed at.
    """
    try:
        indent_index = _build_indent_signal_index(workbook_path)
    except Exception:
        return
    all_rows = indent_index.get(sheet_name)
    if not all_rows:
        return  # no real Excel indent anywhere on this tab -- fast no-op

    value_by_row_idx = {
        row["row_idx"]: row["values"].get(projection_column_key) for row in row_entries
    }

    children_of = _infer_indent_hierarchy(all_rows)
    if not children_of:
        return

    row_entry_by_idx = {row["row_idx"]: row for row in row_entries}
    matched_row_idxs: set = set()
    for parent_row_idx, child_row_idxs in children_of.items():
        parent_val = value_by_row_idx.get(parent_row_idx)
        if parent_val is None:
            continue  # pure category label, not a rollup -- leave children as real accounts
        child_vals = [value_by_row_idx.get(c) for c in child_row_idxs]
        if any(v is None for v in child_vals):
            continue  # incomplete data for this group -- not safely checkable
        child_sum = sum(child_vals)
        if abs(parent_val - child_sum) > max(1.0, abs(parent_val) * 0.005):
            continue  # doesn't match -- needs individual review, don't guess
        parent_entry = row_entry_by_idx.get(parent_row_idx)
        parent_desc = str(parent_entry["description"]) if parent_entry else ""
        for child_row_idx in child_row_idxs:
            child_entry = row_entry_by_idx.get(child_row_idx)
            if child_entry is not None:
                child_entry["row_type"] = "breakdown"
                # Which parent this row rolls into, recorded now that the sum
                # has been verified. Nothing downstream could tell a parent
                # from its child otherwise: both end up "breakdown", so the
                # model received a flat list and enumerated across levels --
                # "应付账款-预提管理费用82.2万元、嘉兴市晶美新能源48.8万元、
                # 缪治彬12.5万元" against an 81.7万元 account, where items 2
                # and 3 are inside item 1.
                if parent_desc:
                    child_entry["rollup_parent_desc"] = parent_desc
                matched_row_idxs.add(child_row_idx)

    # Second pass: retry orphaned top-level sibling groups (see docstring)
    # against any already-classified total/subtotal row, by value, from
    # anywhere in the sheet.
    parent_of_row_idx: Dict[int, int] = {
        child_row_idx: parent_row_idx
        for parent_row_idx, child_row_idxs in children_of.items()
        for child_row_idx in child_row_idxs
    }
    siblings_by_parent: Dict[int, List[int]] = {}
    for child_row_idx, parent_row_idx in parent_of_row_idx.items():
        if child_row_idx in matched_row_idxs:
            continue  # already successfully rolled up in the first pass
        if value_by_row_idx.get(child_row_idx) is None:
            continue  # nothing to sum
        siblings_by_parent.setdefault(parent_row_idx, []).append(child_row_idx)

    total_or_subtotal_rows = [row for row in row_entries if row.get("row_type") in ("total", "subtotal")]
    for parent_row_idx, sibling_idxs in siblings_by_parent.items():
        if value_by_row_idx.get(parent_row_idx) is not None:
            continue  # first pass already handled (or correctly declined) this group
        sibling_sum = sum(value_by_row_idx[idx] for idx in sibling_idxs)
        if abs(sibling_sum) < 1.0:
            continue
        candidates = [
            row for row in total_or_subtotal_rows
            if row["row_idx"] not in sibling_idxs
            and row["values"].get(projection_column_key) is not None
            and abs(row["values"][projection_column_key] - sibling_sum) <= max(1.0, abs(sibling_sum) * 0.005)
        ]
        if not candidates:
            continue  # no match at all -- don't guess
        # Multiple candidate ROWS aren't necessarily ambiguous -- e.g. a
        # "小计"/"合计" pair that coincidentally hold the identical value
        # this period (real case: a subtotal-before-provision and a
        # total-after-provision are numerically the same when the provision
        # is 0 this period) both point to the same reclassification decision
        # either way. Only bail when candidates disagree on the VALUE itself.
        distinct_values = {round(row["values"][projection_column_key], 2) for row in candidates}
        if len(distinct_values) != 1:
            continue  # candidates genuinely disagree -- don't guess
        for sibling_idx in sibling_idxs:
            sibling_entry = row_entry_by_idx.get(sibling_idx)
            if sibling_entry is not None:
                sibling_entry["row_type"] = "breakdown"
                # Deliberately NOT setting rollup_parent_desc here. This pass
                # matches a sibling group against a TOTAL row -- which is what
                # every top-level component of a well-formed schedule does, by
                # definition. Recording that as "this row is a child" made the
                # subtable synthesiser, which skips children to avoid
                # double-counting, skip EVERYTHING: measured on a real file,
                # 其他应付款 reported 12 of 12 breakdown rows as rollup children
                # and 0 top-level components, and 固定资产 16 of 16. Only pass 1
                # sets it, where the parent is another COMPONENT carrying its
                # own value -- which is the nesting the skip actually exists for.


def _fallback_description(description: str, title: str, last_label: Optional[str]) -> str:
    if description:
        return description
    if last_label:
        return last_label
    return title


def _looks_like_supporting_note(text: str) -> bool:
    lowered = text.lower()
    if not text or len(text) < 8:
        return False
    if canonical_stage_label(text) or parse_date(text):
        return False
    if any(marker in lowered for marker in _UNIT_MARKERS):
        return False
    return True


def _looks_like_working_remark(text: str) -> bool:
    lowered = str(text or "").strip().lower()
    if not lowered:
        return False
    return any(keyword in lowered for keyword in _WORKING_REMARK_KEYWORDS)


def _is_pure_working_artifact(note: str, block_title: str = "") -> bool:
    """True when a note is nothing but a bookkeeping tie-out marker.

    A blanket "contains 'check'" test would be wrong -- real analyst prose
    can mention it ('待check VAT to be certified是否需要计入RPT' is a genuine
    working question from a real databook). What distinguishes an artefact is
    that once the keyword, the account/block title, and any figures or dates
    are removed, nothing is left ('Check | 55095.750019',
    'Check | Taxes and surcharges'). Those carry no explanatory content and
    should not reach the model as if they were context; on real files they
    were frequently an account's ONLY note."""
    text = str(note or "").strip()
    if not text:
        return False
    lowered = text.lower()
    if not any(keyword in lowered for keyword in _WORKING_REMARK_KEYWORDS):
        return False
    residue = lowered
    for keyword in _WORKING_REMARK_KEYWORDS:
        residue = residue.replace(keyword, " ")
    title = str(block_title or "").strip().lower()
    if title:
        residue = residue.replace(title, " ")
    residue = re.sub(r"\d{4}-\d{2}-\d{2}", " ", residue)
    residue = re.sub(r"-?[\d,]+\.?\d*", " ", residue)
    residue = re.sub(r"[|:;,()\[\]/\\.\-_+*%&#@!?'\"]+", " ", residue)
    return len(residue.strip()) < 2


def _build_working_remark_note(description: str, values: Dict[str, Optional[float]]) -> Optional[str]:
    label = str(description or "").strip()
    if not _looks_like_working_remark(label):
        return None

    non_zero_parts: List[str] = []
    seen_period_values: set[tuple[str, float]] = set()
    for key, value in values.items():
        if value is None:
            continue
        if abs(float(value)) <= 0:
            continue
        period = str(key).split("|", 1)[-1]
        dedupe_key = (period, float(value))
        if dedupe_key in seen_period_values:
            continue
        seen_period_values.add(dedupe_key)
        non_zero_parts.append(f"{period}: {value:,.0f}")

    if non_zero_parts:
        return f"{label} | " + " | ".join(non_zero_parts)
    return label


def _trim_block_end_row(
    df: pd.DataFrame,
    desc_col_idx: int,
    data_start_row: int,
    data_end_row: int,
    stage_row_idx: Optional[int],
    max_col_idx: Optional[int] = None,
) -> int:
    """max_col_idx, when given, bounds the has-numeric/has-text scan to the
    main block's own columns. Without it, a presentation table sitting a
    dozen rows below the main block but a few columns to the right (a real,
    confirmed layout) still has numbers on the sheet's last row, so the
    backward scan hits that row first, sees numeric content, and stops
    immediately -- reporting the main block as running all the way to the
    sheet's own last row and swallowing the presentation table into it
    (confirmed on a real file: two accounts' data_end_row came out exactly
    equal to len(df), which is what let their report tables go undetected)."""
    trimmed_end_row = int(data_end_row or len(df))
    if trimmed_end_row <= data_start_row:
        return trimmed_end_row

    while trimmed_end_row > data_start_row:
        row_idx = trimmed_end_row - 1
        row_values = (df.iloc[row_idx].tolist() if max_col_idx is None
                      else df.iloc[row_idx, :max_col_idx + 1].tolist())
        text_values = [_cell_text(value) for value in row_values if _cell_text(value)]
        if not text_values:
            trimmed_end_row -= 1
            continue

        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        has_numeric = any(_coerce_numeric(value) is not None for value in row_values)
        if has_numeric:
            break

        if stage_row_idx is not None and row_idx == stage_row_idx - 1:
            trimmed_end_row -= 1
            continue

        if description and (_looks_like_supporting_note(description) or _looks_like_working_remark(description)):
            break

        if len(text_values) <= 2 and not any(parse_date(text) or canonical_stage_label(text) for text in text_values):
            trimmed_end_row -= 1
            continue

        break

    return trimmed_end_row


# Row labels that mean a block is NOT a composition breakdown, however
# well-formed it looks. Confirmed against a real databook where the second
# block on a sheet was, variously, a bank-account listing (100204, 100206), a
# fee-rate workpaper (费率, EBITDA/NOI, 测算数), a rollforward (调整前上年年末,
# 本年增加, 本年摊销), a verification-report log, and a set of operating KPIs.
# Putting any of those in a report as "the breakdown" would be wrong.
_NON_COMPOSITION_ROW_MARKERS = (
    "本年增加", "本年摊销", "本年減少", "本年减少", "年初", "年末", "期初", "期末余额",
    "调整前", "調整前", "调整后年初", "费率", "費率", "测算", "測算", "预算", "預算",
    "验资报告", "驗資報告", "工程进度", "工程進度", "占预算", "占預算",
    "ebitda/noi", "ebitda%", "出租率", "单价（", "單價（", "元/平方米",
)
_GL_CODE_RE = re.compile(r"^\d{4,}$")
# A report-ready block is TITLED the way the report titles it -- '示意性调整后
# 管理费用 - <entity>', mirroring the Financials sheet's own
# '经示意性调整后资产负债表 - <entity>'. Checked across every sheet of a real
# databook this marked exactly the four accounts the reference deck tabulates
# (营业成本, 税金及附加, 管理费用, 财务费用) and correctly did NOT mark 营业收入,
# whose block is titled without the proforma prefix and which that deck
# narrates rather than tabulates. A far more direct signal than inferring
# report-worthiness from a block's shape.
_REPORT_TITLE_RE = re.compile(r"(经|經)?示意性(调整|調整)(后|後)")


def _is_composition_label(label: str) -> bool:
    """A label that could name a component of an account balance.

    Excludes GL codes (660202), bare dates, and the movement/ratio/workpaper
    vocabulary above -- a report breakdown names WHAT the balance consists of,
    not how it moved or how a rate was derived."""
    text = str(label or "").strip()
    if not text or _GL_CODE_RE.match(text):
        return False
    if parse_date(text):
        return False
    low = text.lower()
    return not any(m in low for m in _NON_COMPOSITION_ROW_MARKERS)


# A real breakdown can be two-level -- a real 营业成本 table has 折旧与摊销
# broken further into 房屋建筑物/土地使用权/租赁服务费/其他, and 物业管理费 into
# 第三方/上海熙麦 (rendered indented and in a different colour in the deck).
# openpyxl indent/colour don't survive pd.read_excel, and even if they did,
# a preparer's indent choice is a formatting decision, not a structural one
# to depend on. The reliable signal is the same one the block's OVERALL tie-
# out already uses: a parent row's own value is the SUM of the rows right
# after it, across every period it has a value for. That is what nesting
# actually means numerically, and it costs nothing to check since these
# blocks are capped at max_rows to begin with.
_MAX_NEST_WINDOW = 8


def _nest_component_rows(rows: List[Dict[str, Any]], tie_tolerance: float) -> List[Dict[str, Any]]:
    nested: List[Dict[str, Any]] = []
    i = 0
    while i < len(rows):
        parent = rows[i]
        parent_values = parent.get("values") or {}
        children = None
        if parent_values:
            max_window = min(_MAX_NEST_WINDOW, len(rows) - i - 1)
            for window in range(2, max_window + 1):
                candidates = rows[i + 1 : i + 1 + window]
                if all(
                    abs(
                        sum(c["values"].get(period, 0.0) for c in candidates)
                        - parent_values[period]
                    )
                    <= max(1.0, abs(parent_values[period]) * tie_tolerance)
                    for period in parent_values
                ):
                    children = candidates
                    break
        entry = dict(parent)
        if children:
            entry["children"] = children
            i += 1 + len(children)
        else:
            entry["children"] = None
            i += 1
        nested.append(entry)
    return nested


def extract_presentation_detail_table(
    df: pd.DataFrame,
    desc_col_idx: int,
    main_block_end_row: int,
    columns: List[Dict[str, Any]],
    multiplier: int = 1,
    max_rows: int = 40,
    account_totals_by_date: Optional[Dict[str, float]] = None,
    tie_tolerance: float = 0.02,
) -> Optional[Dict[str, Any]]:
    """The report-ready breakdown table an account sheet carries BELOW its
    main schedule, if it has one.

    Real databooks hold two tables per account sheet. The main one is keyed by
    GL codes (660202, 660203, ...) across several reporting stages
    (管理层数 / 审定数 / 示意性调整后), and it is the one normalize_financial_
    schedule parses. Below it sits a second table with the SAME periods but
    only one stage, and human-readable labels -- 会计服务费, 审计费,
    法律咨询费 for 管理费用; 利息支出, 利息收入, 汇兑损益, 银行手续费 for
    财务费用. That second table is what the analyst's own deck references
    when a paragraph ends '明细如下：', and it is why a deck can describe
    composition in real words while commentary generated from the first
    table can only cite account codes.

    _select_entity_block never finds it because it looks for a stage row and
    this table has none -- so it was not mis-mapped, it was never scanned.

    Returns {header, periods, rows: [{label, values}], total_row} or None.
    Deliberately conservative: requires a currency-unit label, at least two
    period headers matching the account's own, and at least two labelled
    numeric rows, so a stray note block is not mistaken for a table.
    """
    # Why a candidate was turned away, collected for the caller. Without this
    # a sheet that visibly HAS a summary just returns None with no way to tell
    # which of the six tests rejected it -- which is exactly what happened on a
    # real file where two of four summaries went missing.
    rejections: List[Dict[str, Any]] = []

    def _reject(row_idx: int, col_idx: Optional[int], why: str) -> None:
        if len(rejections) < 40:
            rejections.append({"row": row_idx + 1, "col": col_idx, "reason": why})

    if df is None or df.empty:
        return None
    if main_block_end_row >= len(df):
        # This guard used to return bare None here too -- indistinguishable,
        # under --explain, from every other silent path. main_block_end_row is
        # the caller's belief about where the main schedule ends; if it is at
        # or past this sheet's own row count, there is no row left for a
        # summary to occupy, which is either genuinely true (short sheet, no
        # second table) or a sign main_block_end_row itself was computed wrong
        # for this account.
        _reject(
            max(len(df) - 1, 0), None,
            f"main_block_end_row ({main_block_end_row}) is at or past this sheet's own "
            f"row count ({len(df)}), leaving no row for a summary below or beside it"
        )
        return {"rows": [], "rejections": rejections}

    def _norm(value) -> str:
        """Both sides to 'YYYY-MM-DD'. The main block stores its dates already
        formatted that way, while parse_date here returns a datetime whose
        str() carries a ' 00:00:00' tail -- comparing the two raw never
        matches, which silently made this function return None for every
        sheet including ones that do have the table."""
        parsed = parse_date(value) if not hasattr(value, "strftime") else value
        if parsed is not None and hasattr(parsed, "strftime"):
            return parsed.strftime("%Y-%m-%d")
        return str(value or "").strip()

    main_periods = {_norm(c.get("date")) for c in (columns or []) if c.get("date")}

    # The summary sits to the RIGHT of the main schedule, so their ROW ranges
    # overlap and scanning only from main_block_end_row onward missed it.
    # Confirmed on a real databook: 税金及附加's summary header is at row 12
    # while its main block runs to row 13, and 财务费用's is at row 21 against a
    # main block ending at 21 -- both were skipped entirely, so only 2 of that
    # file's 4 summaries were found. The right exclusion is by COLUMN: scan
    # every row, but ignore candidates whose label column falls inside the main
    # table's own span.
    # Accept EITHER layout: to the right of the main table's columns, or below
    # its last row in the same columns. Requiring "to the right" alone would
    # have missed a sheet that stacks them, and requiring "below" alone is what
    # caused the original miss.
    main_cols = [c.get("col_idx") for c in (columns or []) if c.get("col_idx") is not None]
    main_col_max = max(main_cols) if main_cols else None

    # Collected up front so a sheet with NO unit-marker cell anywhere can be
    # told apart from one where a marker exists but never lands in an accepted
    # position -- both looked identical (silent None, zero rejections) before
    # this, which is why two of four real summaries vanished from --explain
    # with no reason shown at all.
    marker_hits: List[Tuple[int, int, str]] = []
    for scan_row in range(len(df)):
        for i, v in enumerate(df.iloc[scan_row].tolist()):
            t = _cell_text(v)
            if t and any(m in t.lower() for m in UNIT_THOUSAND_MARKERS):
                marker_hits.append((scan_row, i, t))
    if not marker_hits:
        _reject(0, None, "no currency-unit marker (" + ", ".join(UNIT_THOUSAND_MARKERS) +
                ") found anywhere on this sheet")

    for header_row in range(len(df)):
        cells = [(i, _cell_text(v)) for i, v in enumerate(df.iloc[header_row].tolist())]
        unit_col = None
        for i, t in cells:
            if not t or not any(m in t.lower() for m in UNIT_THOUSAND_MARKERS):
                continue
            to_the_right = main_col_max is not None and i > main_col_max
            below_main = header_row >= main_block_end_row
            if to_the_right or below_main:
                unit_col = i
                break
        if unit_col is None:
            row_hits = [(c, t) for r, c, t in marker_hits if r == header_row]
            if row_hits:
                c, t = row_hits[0]
                _reject(
                    header_row, c,
                    f"unit marker {t!r} at col {c} is neither right of the main table's "
                    f"rightmost column ({main_col_max}) nor at/below its last row "
                    f"(sheet row {main_block_end_row + 1})"
                )
            continue
        # The raw header text (e.g. "2026年1-3月") is kept alongside the
        # parsed date, not just for comparison -- it is the only place a
        # stub/interim period's real label lives; the normalised YYYY-MM-DD
        # alone can't distinguish a fiscal-year-end date from an interim
        # cutoff, so a renderer with only "periods" would have to guess.
        period_cols: List[Tuple[int, Any, str]] = []
        for i, t in cells:
            if i <= unit_col or not t:
                continue
            parsed = parse_date(t)
            if parsed:
                period_cols.append((i, parsed, t))
        if len(period_cols) < 2:
            _reject(header_row, unit_col, f"only {len(period_cols)} parseable period header(s)")
            continue
        # One stage only: the main table repeats its periods once per stage, so
        # a repeated period here means this is another multi-stage block, not
        # the presentation summary.
        seen = [_norm(p) for _i, p, _t in period_cols]
        if len(seen) != len(set(seen)):
            _reject(header_row, unit_col, f"periods repeat ({seen}) -- multi-stage block")
            continue
        if main_periods and not set(seen) & main_periods:
            _reject(header_row, unit_col,
                    f"no period shared with the account ({seen} vs {sorted(main_periods)})")
            continue

        rows: List[Dict[str, Any]] = []
        total_row = None
        for r in range(header_row + 1, min(len(df), header_row + 1 + max_rows)):
            label = _cell_text(df.iloc[r, unit_col])
            if not label:
                # A blank label with blank values ends the table; a blank label
                # with values is a continuation artefact, so keep scanning.
                if not any(_coerce_numeric(df.iloc[r, i]) is not None for i, _p, _t in period_cols):
                    if rows:
                        break
                    continue
            values = {}
            for i, period, _t in period_cols:
                num = _coerce_numeric(df.iloc[r, i])
                if num is not None:
                    values[_norm(period)] = round(num * multiplier, 2)
            if not values:
                continue
            entry = {"label": label, "values": values}
            if any(k in label for k in _TOTAL_KEYWORDS):
                total_row = entry
            elif label:
                rows.append(entry)
            # A row with figures but no label cannot be shown in a table or
            # named in a sentence, so it is skipped rather than emitted with a
            # blank caption -- it is a spacer or a spill-over from a merged
            # cell, not a component.
        if len(rows) < 2:
            _reject(header_row, unit_col, f"only {len(rows)} labelled numeric row(s)")
            continue

        # Every row must plausibly NAME a component. A single GL code, date or
        # rollforward caption is enough to disqualify the block: real
        # breakdowns are uniformly descriptive, and a mixed block is a
        # workpaper that happens to be periodised.
        non_composition = [r["label"] for r in rows if not _is_composition_label(r["label"])]
        if non_composition:
            _reject(header_row, unit_col,
                    f"row label(s) are not component names: {non_composition[:4]}")
            continue

        # The decisive test. A breakdown that belongs in a report SUMS TO the
        # account balance; a fee-rate workpaper, a rollforward or a
        # bank-account listing does not, however descriptive its labels. Only
        # applied when the caller supplies the account's own totals, since
        # without them there is nothing to tie against.
        tie_status = "not checked"
        if account_totals_by_date:
            # Caller keys these by column LABEL, which may carry a stage or an
            # "annualised" suffix; normalise to the same YYYY-MM-DD the block
            # uses so the two can be compared at all.
            normalised_totals = {}
            for label, value in account_totals_by_date.items():
                key = _norm(label)
                if key and key not in normalised_totals:
                    normalised_totals[key] = value
            tied = mismatched = 0
            for period, account_total in normalised_totals.items():
                if not isinstance(account_total, (int, float)) or abs(account_total) < 1e-9:
                    continue
                if total_row and period in (total_row.get("values") or {}):
                    block_total = total_row["values"][period]
                else:
                    block_total = sum(r["values"].get(period, 0.0) for r in rows)
                if abs(block_total - account_total) <= max(1.0, abs(account_total) * tie_tolerance):
                    tied += 1
                else:
                    mismatched += 1
            if tied == 0:
                _reject(header_row, unit_col,
                        f"ties to no period (checked {sorted(normalised_totals)[:5]})")
                continue  # ties to nothing -- not this account's breakdown
            tie_status = f"ties on {tied} period(s), differs on {mismatched}"

        # The block's own title, from the rows just above its header. Where it
        # follows the report-title convention that is a direct statement by
        # the preparer that this block is the reported view.
        title = ""
        for back in (1, 2, 3):
            if header_row - back < 0:
                break
            candidate = _cell_text(df.iloc[header_row - back, unit_col])
            if candidate and not parse_date(candidate):
                title = candidate
                break
        return {
            "header_row": header_row,
            "label_col": unit_col,
            "periods": [_norm(p) for _i, p, _t in period_cols],
            "period_labels": {_norm(p): t for _i, p, t in period_cols},
            "rows": _nest_component_rows(rows, tie_tolerance),
            "total_row": total_row,
            "tie_status": tie_status,
            "title": title,
            "titled_as_report_view": bool(_REPORT_TITLE_RE.search(title)),
            "rejections": rejections,
        }
    return {"rows": [], "rejections": rejections} if rejections else None


def _extract_supporting_notes(
    df: pd.DataFrame,
    desc_col_idx: int,
    columns: List[Dict[str, Any]],
    data_start_row: int,
    data_end_row: int,
) -> List[str]:
    if not columns:
        return []

    notes: List[str] = []
    seen: set[str] = set()
    max_numeric_col_idx = max(column["col_idx"] for column in columns)

    for row_idx in range(data_start_row, data_end_row):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        numeric_hits = [
            _coerce_numeric(df.iloc[row_idx, column["col_idx"]]) is not None
            for column in columns
        ]
        extra_text_cells = [
            _cell_text(value)
            for value in df.iloc[row_idx, max_numeric_col_idx + 1 :].tolist()
            if _looks_like_supporting_note(_cell_text(value))
        ]

        if extra_text_cells:
            note = " | ".join(
                [part for part in [description, *extra_text_cells] if part and not parse_date(part)]
            ).strip()
            if note and note not in seen:
                notes.append(note)
                seen.add(note)
                continue

        if description and not any(numeric_hits) and _looks_like_supporting_note(description):
            if description not in seen:
                notes.append(description)
                seen.add(description)
            continue

        if not any(numeric_hits):
            row_text_cells = [
                _cell_text(value)
                for value in df.iloc[row_idx].tolist()
                if _looks_like_supporting_note(_cell_text(value))
            ]
            if row_text_cells:
                note = " | ".join(dict.fromkeys(row_text_cells))
                if note not in seen:
                    notes.append(note)
                    seen.add(note)

    return notes[:8]


def _extract_auxiliary_check_totals(
    df: pd.DataFrame,
    columns: List[Dict[str, Any]],
    multiplier: int,
) -> Dict[str, float]:
    if not columns:
        return {}

    max_financial_col_idx = max(column["col_idx"] for column in columns)
    trailing_start_col_idx = max_financial_col_idx + 1
    if trailing_start_col_idx >= len(df.columns):
        return {}

    for desc_col_idx in range(trailing_start_col_idx, len(df.columns)):
        for row_idx in range(len(df)):
            label = _cell_text(df.iloc[row_idx, desc_col_idx]).lower()
            if label != "check":
                continue

            header_row_idx = None
            for candidate_row_idx in range(row_idx - 1, -1, -1):
                date_hits = sum(
                    1
                    for col_idx in range(desc_col_idx + 1, len(df.columns))
                    if parse_date(df.iloc[candidate_row_idx, col_idx])
                )
                if date_hits >= 1:
                    header_row_idx = candidate_row_idx
                    break
            if header_row_idx is None:
                continue

            totals_by_date: Dict[str, float] = {}
            for col_idx in range(desc_col_idx + 1, len(df.columns)):
                parsed_date = parse_date(df.iloc[header_row_idx, col_idx])
                if not parsed_date:
                    continue
                value = _coerce_numeric(df.iloc[row_idx, col_idx])
                if value is None:
                    continue
                totals_by_date[parsed_date.strftime("%Y-%m-%d")] = round(value * multiplier, 0)

            if totals_by_date:
                return totals_by_date

    return {}


def _auxiliary_header_context(
    df: pd.DataFrame,
    block_title: str,
    stage_row_idx: int,
    date_row_idx: int,
    col_idx: int,
) -> Dict[str, str]:
    stage_text = _cell_text(df.iloc[stage_row_idx, col_idx]) if stage_row_idx is not None else ""
    date_text = _cell_text(df.iloc[date_row_idx, col_idx]) if date_row_idx is not None else ""

    header_text = ""
    for row_idx in (date_row_idx, stage_row_idx):
        text = _cell_text(df.iloc[row_idx, col_idx])
        lowered = text.lower()
        if not text:
            continue
        if canonical_stage_label(text) or parse_date(text):
            continue
        if any(marker in lowered for marker in _UNIT_MARKERS):
            continue
        header_text = text
        break

    return {
        "table_header": str(block_title or "").strip(),
        "stage_header": stage_text,
        "date_header": date_text,
        "header": header_text or f"Detail {col_idx}",
    }


def _extract_adjacent_detail_columns(
    df: pd.DataFrame,
    block_title: str,
    desc_col_idx: int,
    columns: List[Dict[str, Any]],
    stage_row_idx: int,
    date_row_idx: int,
    data_start_row: int,
    data_end_row: int,
    max_columns: int = 5,
) -> List[Dict[str, Any]]:
    if not columns:
        return []

    numeric_col_indices = {column["col_idx"] for column in columns}
    max_numeric_col_idx = max(numeric_col_indices)
    candidates: List[Dict[str, Any]] = []
    for col_idx in range(max_numeric_col_idx + 1, min(len(df.columns), max_numeric_col_idx + max_columns + 1)):
        if col_idx in numeric_col_indices:
            continue
        text_values = []
        for row_idx in range(data_start_row, data_end_row):
            value = df.iloc[row_idx, col_idx]
            text = _cell_text(value)
            if text:
                text_values.append(text)
        if not text_values:
            continue
        candidates.append(
            {
                "col_idx": col_idx,
                **_auxiliary_header_context(df, block_title, stage_row_idx, date_row_idx, col_idx),
            }
        )
    unique_candidates: List[Dict[str, Any]] = []
    seen_headers: Dict[str, int] = {}
    for candidate in candidates[:max_columns]:
        base_header = str(candidate.get("header") or f"Detail {candidate.get('col_idx')}").strip()
        if not base_header:
            base_header = f"Detail {candidate.get('col_idx')}"
        seen_headers[base_header] = seen_headers.get(base_header, 0) + 1
        header_occurrence = seen_headers[base_header]
        unique_header = base_header if header_occurrence == 1 else f"{base_header} ({header_occurrence})"
        unique_candidates.append(
            {
                **candidate,
                "header": unique_header,
            }
        )
    return unique_candidates


def _build_table_linked_remarks(
    supporting_notes: List[str],
    adjacent_detail_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    linked_remarks: List[Dict[str, Any]] = []
    seen: set[str] = set()

    for note in supporting_notes:
        text = str(note or "").strip()
        if not text:
            continue
        dedupe_key = f"row_note::{text}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        linked_remarks.append(
            {
                "source": "row_note",
                "summary": text,
            }
        )

    for row in adjacent_detail_rows:
        if not isinstance(row, dict):
            continue
        description = str(row.get("Description") or "").strip()
        remarks: List[Dict[str, str]] = []
        for key, value in row.items():
            key_text = str(key)
            if (
                key_text == INTERNAL_ROW_KEY
                or key_text == "Description"
                or key_text.endswith("| table_header")
                or key_text.endswith("| indicative_adjusted_row")
                or key_text.endswith("| date_row")
            ):
                continue
            text = str(value or "").strip()
            if not text:
                continue
            remarks.append(
                {
                    "header": key_text,
                    "value": text,
                    "table_header": str(row.get(f"{key_text} | table_header") or "").strip(),
                    "indicative_adjusted_row": str(row.get(f"{key_text} | indicative_adjusted_row") or "").strip(),
                    "date_row": str(row.get(f"{key_text} | date_row") or "").strip(),
                }
            )
        if not remarks:
            continue
        summary = " | ".join(
            part for part in [description, "; ".join(f"{item['header']}: {item['value']}" for item in remarks)] if part
        ).strip()
        dedupe_key = f"rhs::{summary}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        linked_remarks.append(
            {
                "source": "rhs_columns",
                "description": description,
                "summary": summary,
                "remarks": remarks,
            }
        )

    return linked_remarks


def _build_prompt_analysis_df(
    block_title: str,
    columns: List[Dict[str, Any]],
    row_entries: List[Dict[str, Any]],
    analysis_stage: str,
) -> pd.DataFrame:
    analysis_columns = sorted(
        [column for column in columns if column["stage"] == analysis_stage],
        key=lambda column: column["date"],
    )
    if not analysis_columns:
        return pd.DataFrame(columns=[block_title])

    # "breakdown" rows are a schedule's component lines: an indented child, or
    # a sub-row under a parent that already carries their sum. They are rightly
    # kept out of reconciliation totals (they would double-count) and out of the
    # account list (they are not accounts of their own) -- but they used to be
    # dropped here too, and this table is the ONLY thing the commentary sees.
    #
    # That is what made a schedule arrive as a bare 合计. On one real tab the
    # sheet held 应交增值税 / 土地使用税 / 房产税 / 印花税 with correct values and
    # the model received a single 1,437,383 total, so it could only write "余额
    # 为1,437,383元" where the analyst deliverable writes "主要为应交房产税106.6
    # 万元、应交土地使用税36.8万元以及应交印花税2,938元". No prompt can recover a
    # number it was never given.
    #
    # They are included here and recorded as components. Reconciliation reads
    # row_types_by_description off the main frame, not this one, so totals are
    # untouched. Grounding reads this frame, so enumerating a component no
    # longer reads as a hallucination.
    prompt_rows: List[Dict[str, Any]] = []
    component_descriptions: List[str] = []
    # {parent description: [child descriptions]} for groups whose sum was
    # actually verified against the parent's own value. Only these are stated
    # to the model -- an unverified guess about hierarchy would be worse than
    # no guess at all.
    rollup_groups: Dict[str, List[str]] = {}
    for row in row_entries:
        is_component = row["row_type"] == "breakdown"
        row_values = {
            column["date"]: row["values"].get(column["key"])
            for column in analysis_columns
        }
        has_value = any(value is not None for value in row_values.values())
        if row["row_type"] == "detail" and not has_value:
            continue
        if is_component:
            if not has_value:
                continue                      # dead component, nothing to say
            component_descriptions.append(str(row["description"]))
            parent_desc = str(row.get("rollup_parent_desc") or "")
            if parent_desc:
                rollup_groups.setdefault(parent_desc, []).append(str(row["description"]))
        prompt_rows.append(
            {
                block_title: row["description"],
                INTERNAL_ROW_KEY: row["row_idx"],
                **{key: (0 if value is None else value) for key, value in row_values.items()},
            }
        )

    if not prompt_rows:
        empty = pd.DataFrame(columns=[block_title, INTERNAL_ROW_KEY, *[column["date"] for column in analysis_columns]])
        empty.attrs["component_descriptions"] = []
        empty.attrs["rollup_groups"] = {}
        return empty
    frame = pd.DataFrame(prompt_rows)
    frame.attrs["component_descriptions"] = component_descriptions
    frame.attrs["rollup_groups"] = rollup_groups
    return frame




def _first_table_with_rows(primary, fallback_factory):
    """`primary` when it actually carries rows, otherwise the fallback.

    Exists because extract_presentation_detail_table deliberately returns a
    truthy dict with an empty "rows" list on its rejection paths, so that a
    caller can say WHY a summary was turned away rather than just seeing None.
    That makes `primary or fallback()` wrong: the rejection object is truthy,
    so the fallback is unreachable. Any diagnostic carried on the rejected
    object is preserved on the fallback's result, so --explain still has it.
    """
    if isinstance(primary, dict) and primary.get("rows"):
        return primary
    result = fallback_factory()
    if isinstance(result, dict) and isinstance(primary, dict) and primary.get("rejections"):
        result = dict(result)
        result["rejections"] = primary["rejections"]
    return result if result is not None else primary

def synthesize_detail_table_from_breakdown(
    row_entries: List[Dict[str, Any]],
    columns: List[Dict[str, Any]],
    analysis_stage: Union[str, Sequence[str]],
    block_title: str,
    notes_out: Optional[List[str]] = None,
) -> Optional[Dict[str, Any]]:
    """A detail table built from the schedule's OWN breakdown rows, for the
    sheets that carry no separate report-ready block below the main one.

    extract_presentation_detail_table only finds the second table some
    databooks put under the main schedule. On this client's files almost no
    sheet has one -- the run reports "total only, and NO detail table found
    anywhere on the sheet" for most accounts -- so no subtable could render
    even with the feature switched on, and the model had no named components
    to enumerate either.

    The components exist regardless; they are just inside the main schedule,
    already classified "breakdown" by _row_type and
    _reclassify_indent_rollup_children. This turns them into the same shape
    extract_presentation_detail_table returns, so everything downstream --
    the subtable renderer, _detail_table_guidance, _sublist_text_for_table --
    works without knowing which of the two sources it came from.

    TOP LEVEL ONLY. A row that rolls up into another (rollup_parent_desc, set
    only where the children were verified to sum to the parent) is left out,
    because listing a parent beside its own children is the double-count that
    produced most of a real run's grounding warnings. Returns None rather than
    a one-row table: a "breakdown" with a single line says nothing the total
    does not.
    """
    if not row_entries or not columns:
        return None
    # Try each candidate stage in order and keep the first that yields a real
    # breakdown. The caller passes the PROJECTION stage first -- the one the
    # account's own total actually came from -- because a subtable whose
    # components come from a different stage than the total the commentary
    # states cannot add up to it. Measured on a real databook: analysis_stage
    # alone is "Indicative adjusted" whenever that column merely EXISTS, and on
    # 12 of 18 accounts that stage held no data at all, so every component was
    # dropped as dead and only the 3 accounts carrying a 管理层调整 row (which
    # IS an indicative adjustment, hence non-zero there) produced a table.
    raw_stages = [analysis_stage] if isinstance(analysis_stage, str) else list(analysis_stage)
    stages: List[str] = []
    for stage in raw_stages:          # dedupe, order-preserving: the projection
        if stage and stage not in stages:   # and analysis stage are often the same
            stages.append(stage)
    local_notes: List[str] = []
    for stage in stages:
        built = _synthesize_for_stage(row_entries, columns, stage, block_title, local_notes)
        if built is not None:
            return built
    # Why it declined, written into the caller's own list rather than onto the
    # function object: extraction runs under a thread pool, and a module-level
    # scratch value would be overwritten by whichever account finished last.
    # Two guesses about this threshold have already been wrong, so the answer
    # comes from the real file rather than from another hypothesis.
    if notes_out is not None:
        notes_out.append("; ".join(local_notes) or "no candidate stages")
    return None


def _synthesize_for_stage(
    row_entries: List[Dict[str, Any]],
    columns: List[Dict[str, Any]],
    analysis_stage: str,
    block_title: str,
    notes: Optional[List[str]] = None,
) -> Optional[Dict[str, Any]]:
    def _note(text: str) -> None:
        if notes is not None:
            notes.append(f"[{analysis_stage}] {text}")

    stage_columns = sorted(
        [column for column in columns if column["stage"] == analysis_stage],
        key=lambda column: column["date"],
    )
    if not stage_columns:
        _note("no columns for this stage")
        return None
    n_breakdown = sum(1 for e in row_entries if e.get("row_type") == "breakdown")
    n_children = sum(1 for e in row_entries
                     if e.get("row_type") == "breakdown" and e.get("rollup_parent_desc"))
    _note(f"{n_breakdown} breakdown row(s), {n_children} of them rollup children")
    periods = [column["date"] for column in stage_columns]

    rows: List[Dict[str, Any]] = []
    for entry in row_entries:
        if entry.get("row_type") != "breakdown":
            continue
        if entry.get("rollup_parent_desc"):
            continue                      # a child of a verified rollup
        values = {
            column["date"]: entry["values"].get(column["key"])
            for column in stage_columns
        }
        if not any(isinstance(v, (int, float)) and v != 0 for v in values.values()):
            continue                      # dead component
        label = str(entry.get("description") or "").strip()
        if label:
            rows.append({"label": label, "values": values})

    if len(rows) < 2:
        _note(f"only {len(rows)} top-level component(s) with a non-zero value -- need 2")
        return None
    total_row = None
    for entry in row_entries:
        if str(entry.get("row_type") or "").lower() in ("total", "subtotal"):
            total_row = {
                "label": str(entry.get("description") or "").strip(),
                "values": {
                    column["date"]: entry["values"].get(column["key"])
                    for column in stage_columns
                },
            }
    # Only where the sheet's own header actually differs from the ISO date --
    # a header cell that already IS a date renders identically either way, and
    # a blank one must fall back to the date rather than to "".
    period_labels = {
        column["date"]: str(column.get("label") or "").strip()
        for column in stage_columns
        if str(column.get("label") or "").strip()
        and str(column.get("label") or "").strip() != column["date"]
    }
    return {
        # "title" is what _render_presentation_table draws in the table's own
        # navy band; "header" is not read by anything. Setting only "header"
        # meant every SYNTHESIZED table shipped with a blank title row -- and
        # since these are synthesized whenever the sheet has no report-ready
        # block of its own, that was every table in a real deck. The block
        # title is the account, which is exactly what the band should name.
        "title": block_title,
        "header": block_title,
        "periods": periods,
        "period_labels": period_labels,
        "rows": rows,
        "total_row": total_row,
        "synthesized_from": "main_schedule_breakdown",
    }

def _multiply_factor(profile: Dict[str, Any]) -> int:
    markers = [str(marker).lower() for marker in profile.get("unit_markers") or []]
    if any("cny'000" in marker or "千元" in marker for marker in markers):
        return 1000
    return 1


def _choose_projection(columns: List[Dict[str, Any]], row_entries: List[Dict[str, Any]]) -> Dict[str, Any]:
    def non_zero_score(column: Dict[str, Any]) -> float:
        total = 0.0
        for row in row_entries:
            value = row["values"].get(column["key"])
            if value is not None:
                total += abs(value)
        return total

    stage_priority = [PREFERRED_STAGE, "Audited", "Mgt acc", "Audit adjustment", "Indicative adjustment"]
    sorted_columns = sorted(
        columns,
        key=lambda column: (
            stage_priority.index(column["stage"]) if column["stage"] in stage_priority else len(stage_priority),
            column["date"],
        ),
    )

    preferred_candidates = [column for column in sorted_columns if column["stage"] == PREFERRED_STAGE]
    preferred_candidates.sort(key=lambda column: column["date"], reverse=True)
    for column in preferred_candidates:
        if non_zero_score(column) > 0:
            return {
                "preferred_stage": PREFERRED_STAGE,
                "effective_stage": column["stage"],
                "effective_date": column["date"],
                "column": column,
            }

    fallback_candidates = sorted(sorted_columns, key=lambda column: column["date"], reverse=True)
    for column in fallback_candidates:
        if non_zero_score(column) > 0:
            return {
                "preferred_stage": PREFERRED_STAGE,
                "effective_stage": column["stage"],
                "effective_date": column["date"],
                "column": column,
            }

    column = preferred_candidates[0] if preferred_candidates else fallback_candidates[0]
    return {
        "preferred_stage": PREFERRED_STAGE,
        "effective_stage": column["stage"],
        "effective_date": column["date"],
        "column": column,
    }


def normalize_financial_schedule(
    workbook_path: str,
    sheet_name: str,
    profile: Optional[Dict[str, Any]] = None,
    entity_name: Optional[str] = None,
    sheet_df: Optional[pd.DataFrame] = None,
    statement_type: Optional[str] = None,
) -> Dict[str, Any]:
    df = sheet_df if sheet_df is not None else pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, engine="openpyxl")
    profile = profile or profile_sheet(df, sheet_name)

    desc_col_idx = _find_description_column(df)
    if desc_col_idx is None:
        raise ValueError(f"Unable to detect description column for sheet: {sheet_name}")

    stage_row_idx = profile.get("stage_row_idx")
    date_row_idx = profile.get("date_row_idx")
    if stage_row_idx is None or date_row_idx is None:
        standardized_rollforward_df = _standardize_rollforward_schedule_df(
            df=df,
            sheet_name=sheet_name,
            profile=profile,
            desc_col_idx=desc_col_idx,
        )
        if standardized_rollforward_df is not None:
            standardized_profile = profile_sheet(standardized_rollforward_df, sheet_name)
            standardized_profile["unit_markers"] = profile.get("unit_markers") or standardized_profile.get("unit_markers")
            standardized_profile["sheet_kind"] = "financial_schedule"
            return normalize_financial_schedule(
                workbook_path=workbook_path,
                sheet_name=sheet_name,
                profile=standardized_profile,
                entity_name=entity_name,
                sheet_df=standardized_rollforward_df,
                statement_type=statement_type,
            )
        raise ValueError(f"Unable to detect stage/date rows for sheet: {sheet_name}")
    selected_block = _select_entity_block(df, sheet_name, stage_row_idx, entity_name)
    stage_row_idx = selected_block["stage_row_idx"]
    date_row_idx = selected_block["date_row_idx"]
    block_title = str(selected_block.get("block_title") or profile.get("title") or sheet_name)
    block_entity_name = selected_block.get("block_entity_name")
    strict_entity_match = bool(selected_block.get("strict_entity_match"))

    stage_map = _forward_fill_stage_row(df.iloc[stage_row_idx])
    data_start_row = max(stage_row_idx, date_row_idx) + 1
    raw_data_end_row = int(selected_block.get("data_end_row") or len(df))
    # The rightmost column carrying a real date on the main block's OWN date
    # row. stage_map is unusable for this bound -- it forward-fills, so a
    # presentation table's blank columns on this same row would silently
    # keep reading as "still the last stage" -- but a presentation table's
    # own header sits several rows further down, so its columns are
    # genuinely blank here. Bounds the trim below to the main block's real
    # width so a numeric table elsewhere on the sheet cannot drag
    # data_end_row all the way to the sheet's last row.
    date_row_values = df.iloc[date_row_idx].tolist()
    dated_cols = [i for i, v in enumerate(date_row_values) if i > desc_col_idx and parse_date(v)]
    main_col_bound = max(dated_cols) if dated_cols else None
    data_end_row = _trim_block_end_row(
        df=df,
        desc_col_idx=desc_col_idx,
        data_start_row=data_start_row,
        data_end_row=raw_data_end_row,
        stage_row_idx=raw_data_end_row if raw_data_end_row < len(df) else None,
        max_col_idx=main_col_bound,
    )
    columns: List[Dict[str, Any]] = []
    for col_idx in range(desc_col_idx + 1, len(df.columns)):
        stage = stage_map.get(col_idx)
        parsed_date = parse_date(df.iloc[date_row_idx, col_idx])
        if not stage or not parsed_date:
            continue
        if not _is_numeric_enough(df.iloc[:data_end_row], col_idx, data_start_row):
            continue
        columns.append(
            {
                "col_idx": col_idx,
                "stage": stage,
                "date": parsed_date.strftime("%Y-%m-%d"),
                "key": f"{stage}|{parsed_date.strftime('%Y-%m-%d')}",
                # The sheet's own header text, kept for the same reason
                # extract_presentation_detail_table keeps it (see its
                # period_cols): the normalised date alone cannot say whether
                # a column is a year, a year end or an interim cutoff. A
                # column headed "2023" parses to 2023-01-01, and a real deck
                # printed exactly that in four subtable headers --
                # wrong to a reader and 41.7pt wide against a 40.0pt column,
                # so it also wrapped.
                "label": _cell_text(df.iloc[date_row_idx, col_idx]),
            }
        )
    columns = _dedupe_columns_by_key(columns)

    if not columns:
        raise ValueError(f"No financial value columns detected for sheet: {sheet_name}")

    adjacent_detail_columns = _extract_adjacent_detail_columns(
        df=df,
        block_title=block_title,
        desc_col_idx=desc_col_idx,
        columns=columns,
        stage_row_idx=stage_row_idx,
        date_row_idx=date_row_idx,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
    )

    multiplier = _multiply_factor(profile)
    auxiliary_check_totals_by_date = _extract_auxiliary_check_totals(
        df=df,
        columns=columns,
        multiplier=multiplier,
    )
    row_entries: List[Dict[str, Any]] = []
    working_remark_notes: List[str] = []
    last_label = None
    for row_idx in range(data_start_row, data_end_row):
        description = _cell_text(df.iloc[row_idx, desc_col_idx])
        numeric_values = {column["key"]: _coerce_numeric(df.iloc[row_idx, column["col_idx"]]) for column in columns}
        has_numeric = any(value is not None for value in numeric_values.values())
        if not description and not has_numeric:
            continue
        if description:
            last_label = description
        effective_description = _fallback_description(description, block_title, last_label)
        if not effective_description:
            continue
        values = {
            key: (round(value * multiplier, 0) if value is not None else None)
            for key, value in numeric_values.items()
        }
        working_remark_note = _build_working_remark_note(effective_description, values)
        if working_remark_note:
            working_remark_notes.append(working_remark_note)
            continue
        row_entries.append(
            {
                "row_idx": row_idx,
                "description": effective_description,
                "row_type": _row_type(effective_description),
                "values": values,
            }
        )

    if not row_entries:
        raise ValueError(f"No data rows detected for sheet: {sheet_name}")

    projection = _choose_projection(columns, row_entries)
    # Detect implicit breakdown rows (parent-first structure without total keywords)
    # Must be called after _choose_projection so we know the projection column key.
    _detect_implicit_breakdowns_from_sum(row_entries, projection["column"]["key"])
    _reclassify_indent_rollup_children(
        row_entries=row_entries,
        projection_column_key=projection["column"]["key"],
        workbook_path=workbook_path,
        sheet_name=sheet_name,
    )
    projection_column = projection["column"]
    analysis_stage = PREFERRED_STAGE if any(column["stage"] == PREFERRED_STAGE for column in columns) else projection["effective_stage"]
    # Per-account, so parallel extraction cannot cross-contaminate the note.
    _detail_table_notes: List[str] = []
    prompt_analysis_df = _build_prompt_analysis_df(
        block_title=block_title,
        columns=columns,
        row_entries=row_entries,
        analysis_stage=analysis_stage,
    )
    trend_summary = build_trend_summary(prompt_analysis_df)
    significant_movements = build_significant_movements(prompt_analysis_df)
    supporting_notes = [
        note
        for note in _extract_supporting_notes(df, desc_col_idx, columns, data_start_row, data_end_row)
        if not _is_pure_working_artifact(note, block_title)
    ]
    # Working-remark rows ("Check", "对账单", "差异", ...) are correctly kept out
    # of row_entries above, but they used to be appended to supporting_notes as
    # well -- i.e. handed to the model as explanatory context. Confirmed with
    # the databook owner that these are bookkeeping tie-out figures with no
    # explanatory value. On real files they were frequently an account's ONLY
    # "note", so the model was being given pure noise and nothing else. Kept in
    # their own attrs key so diagnostics can still see them, just not fed to
    # the prompt as if they explained anything.
    annualization = infer_partial_year_annualization(
        statement_type=statement_type or "",
        available_dates=[column["date"] for column in columns],
        effective_date=projection["effective_date"],
    )
    original_column_label = projection_column["date"]
    annualized_column_label = (
        f"{projection_column['date']} annualised"
        if annualization.get("annualized")
        else projection_column["date"]
    )

    projection_rows_original: List[Dict[str, Any]] = []
    projection_rows_annualized: List[Dict[str, Any]] = []
    adjacent_detail_rows: List[Dict[str, Any]] = []
    projection_original_values_by_description: Dict[str, float] = {}
    projection_totals_by_date: Dict[str, float] = {}
    non_zero_rows = 0
    for row in row_entries:
        original_value = row["values"].get(projection_column["key"])
        annualized_value = original_value
        if annualized_value is not None and annualization.get("annualized") and annualization.get("factor"):
            annualized_value = round(annualized_value * float(annualization["factor"]), 0)
        if original_value is None and row["row_type"] == "detail":
            continue
        effective_value = annualized_value if annualization.get("annualized") else original_value
        if effective_value is not None and abs(effective_value) > 0:
            non_zero_rows += 1
        if original_value is not None:
            projection_original_values_by_description[row["description"]] = original_value
        projection_rows_original.append(
            {
                block_title: row["description"],
                INTERNAL_ROW_KEY: row["row_idx"],
                original_column_label: original_value if original_value is not None else 0,
            }
        )
        projection_rows_annualized.append(
            {
                block_title: row["description"],
                INTERNAL_ROW_KEY: row["row_idx"],
                annualized_column_label: annualized_value if annualized_value is not None else 0,
            }
        )
        adjacent_row = {
            INTERNAL_ROW_KEY: row["row_idx"],
            "Description": row["description"],
            annualized_column_label if annualization.get("annualized") else original_column_label: effective_value if effective_value is not None else 0,
        }
        has_adjacent_text = False
        for detail_column in adjacent_detail_columns:
            detail_value = _cell_text(df.iloc[row["row_idx"], detail_column["col_idx"]])
            adjacent_row[f"{detail_column['header']} | table_header"] = detail_column.get("table_header", "")
            adjacent_row[f"{detail_column['header']} | indicative_adjusted_row"] = detail_column.get("stage_header", "")
            adjacent_row[f"{detail_column['header']} | date_row"] = detail_column.get("date_header", "")
            adjacent_row[detail_column["header"]] = detail_value
            if detail_value:
                has_adjacent_text = True
        if has_adjacent_text:
            adjacent_detail_rows.append(adjacent_row)
        if row["row_type"] == "total":
            if original_value is not None:
                projection_totals_by_date[original_column_label] = float(original_value)
            if annualized_value is not None:
                projection_totals_by_date[annualized_column_label] = float(annualized_value)

    table_linked_remarks = _build_table_linked_remarks(
        supporting_notes=supporting_notes,
        adjacent_detail_rows=adjacent_detail_rows,
    )

    projection_df_original = pd.DataFrame(projection_rows_original)
    if projection_df_original.empty:
        projection_df_original = pd.DataFrame(columns=[block_title, original_column_label])

    projection_df_annualized = pd.DataFrame(projection_rows_annualized)
    if projection_df_annualized.empty:
        projection_df_annualized = pd.DataFrame(columns=[block_title, annualized_column_label])

    projection_df = projection_df_annualized.copy() if annualization.get("annualized") else projection_df_original.copy()
    if projection_df.empty:
        projection_df = pd.DataFrame(columns=[block_title, annualized_column_label if annualization.get("annualized") else original_column_label])

    shared_attrs = {
        "preferred_stage": projection["preferred_stage"],
        "effective_stage": projection["effective_stage"],
        "effective_date": projection["effective_date"],
        "statement_type": statement_type or "",
        "sheet_name": sheet_name,
        "non_zero_rows": non_zero_rows,
        "row_type_counts": dict(Counter(row["row_type"] for row in row_entries)),
        "block_title": block_title,
        "block_entity_name": block_entity_name,
        "block_start_row": data_start_row,
        "block_end_row": data_end_row,
        "strict_entity_match": strict_entity_match,
        "annualized": bool(annualization.get("annualized")),
        "annualization_factor": annualization.get("factor"),
        "annualization_months": annualization.get("months"),
        "raw_effective_date": annualization.get("raw_effective_date"),
        "fiscal_year_end_month": annualization.get("fiscal_year_end_month"),
        "fiscal_year_end_day": annualization.get("fiscal_year_end_day"),
    }
    row_types_by_description = {
        row["description"]: row["row_type"] for row in row_entries
    }
    # A row's projection column (the single period downstream filtering
    # normally looks at) can be genuinely 0 while an EARLIER period had real,
    # non-zero activity -- e.g. an account that was active last year but has
    # gone quiet this period. filter_zero_value_rows only ever sees that one
    # projection value, so without this it would silently drop the row and
    # lose the earlier-period history entirely. Computed here while the full
    # multi-period values dict is still available (row["values"] spans every
    # detected column, not just the chosen projection one).
    any_period_nonzero_by_description = {
        row["description"]: any(v is not None and abs(v) >= 0.01 for v in row["values"].values())
        for row in row_entries
    }
    common_attrs = {
        "integrity": shared_attrs,
        "row_types_by_description": row_types_by_description,
        "any_period_nonzero_by_description": any_period_nonzero_by_description,
        "supporting_notes": supporting_notes,
        "working_remark_notes": working_remark_notes,
        # The report-ready breakdown below the main schedule, when the sheet
        # has one -- human-readable labels rather than GL codes. See
        # extract_presentation_detail_table.
        # The sheet's own second table wins when it exists -- it is written for
        # a report and its labels are already human-readable. Falling back to
        # the main schedule's breakdown rows covers the sheets that have no
        # such block, which on some databooks is nearly all of them.
        #
        # The fallback tests for USABLE ROWS, not for None. extract_ returns a
        # truthy {"rows": [], "rejections": [...]} on several paths on purpose,
        # so a caller can report WHY a summary was turned away -- and a plain
        # `A or B` reads that as success and never reaches B. Measured on a real
        # databook: all 18 accounts came back with a table carrying zero rows,
        # so no subtable could render and the synthesised fallback never ran
        # once, which is exactly what "still no subtables" looked like.
        "presentation_detail_table": _first_table_with_rows(
            extract_presentation_detail_table(
                df=df,
                desc_col_idx=desc_col_idx,
                main_block_end_row=data_end_row,
                columns=columns,
                multiplier=multiplier,
                # The account's own period totals, so the candidate block can be
                # required to SUM TO them. Without this the descriptive-label test
                # alone still admits a fee-rate workpaper or a rollforward.
                account_totals_by_date=projection_totals_by_date,
            ),
            lambda: synthesize_detail_table_from_breakdown(
                row_entries=row_entries,
                columns=columns,
                # Projection stage FIRST: it is where this account's own total
                # came from, so the components can actually sum to it.
                analysis_stage=[projection["effective_stage"], analysis_stage],
                block_title=block_title,
                notes_out=_detail_table_notes,
            ),
        ),
        # Why no breakdown table was synthesised, when none was. Read by
        # inspect_databook's 3c so "0 rows" says what stopped it.
        "presentation_detail_table_reason": "; ".join(_detail_table_notes),
        "normalized_columns": columns,
        "source_multiplier": multiplier,
        "sheet_kind": profile.get("sheet_kind"),
        "entity_name": block_entity_name,
        "block_title": block_title,
        "annualization": annualization,
        "projection_original_column_label": original_column_label,
        "projection_annualized_column_label": annualized_column_label,
        "projection_original_values_by_description": projection_original_values_by_description,
        "projection_totals_by_date": projection_totals_by_date,
        "auxiliary_check_totals_by_date": auxiliary_check_totals_by_date,
        "adjacent_detail_columns": adjacent_detail_columns,
        "adjacent_detail_rows": adjacent_detail_rows,
        "table_linked_remarks": table_linked_remarks,
        "prompt_analysis_df": prompt_analysis_df,
        "trend_summary": trend_summary,
        "significant_movements": significant_movements,
        "prompt_analysis_label": (
            "All indicative adjusted periods"
            if analysis_stage == PREFERRED_STAGE
            else f"All {analysis_stage} periods"
        ),
        "prompt_analysis_stage": analysis_stage,
    }
    projection_df.attrs.update(common_attrs)
    projection_df_original.attrs.update(common_attrs)
    projection_df_annualized.attrs.update(common_attrs)
    # Exclude prompt_analysis_df from its own attrs to avoid circular reference
    # (deepcopy during .copy() would recurse infinitely on Python 3.13+)
    prompt_analysis_attrs = {k: v for k, v in common_attrs.items() if k != "prompt_analysis_df"}
    prompt_analysis_df.attrs.update(prompt_analysis_attrs)
    projection_df_original.attrs["selected_variant"] = "original"
    projection_df_annualized.attrs["selected_variant"] = "annualized"
    projection_df.attrs["selected_variant"] = "annualized" if annualization.get("annualized") else "original"
    prompt_analysis_df.attrs["selected_variant"] = "analysis"

    return {
        "sheet_name": sheet_name,
        "title": block_title,
        "entity_name": block_entity_name,
        "block_title": block_title,
        "profile": profile,
        "columns": columns,
        "row_entries": row_entries,
        "projection_df": projection_df,
        "projection_df_original": projection_df_original,
        "projection_df_annualized": projection_df_annualized,
        "prompt_analysis_df": prompt_analysis_df,
        "integrity": projection_df.attrs["integrity"],
    }
def _annualization_factor(column_name: str) -> float | None:
    match = re.match(r"^(\d+)M\d{2}$", str(column_name).strip(), flags=re.IGNORECASE)
    if not match:
        return None
    months = int(match.group(1))
    if months <= 0:
        return None
    return 12.0 / months


def _parse_statement_date_label(value: str) -> datetime | None:
    try:
        return datetime.strptime(str(value), "%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def infer_partial_year_annualization(
    statement_type: str,
    available_dates: list[str],
    effective_date: str,
) -> dict[str, float | int | bool | str | None]:
    metadata: dict[str, float | int | bool | str | None] = {
        "annualized": False,
        "factor": None,
        "months": None,
        "raw_effective_date": effective_date,
        "fiscal_year_end_month": None,
        "fiscal_year_end_day": None,
    }
    if statement_type != "IS":
        return metadata

    period_factor = _annualization_factor(str(effective_date))
    if period_factor is not None:
        months = round(12.0 / period_factor)
        metadata.update({"annualized": months < 12, "factor": period_factor, "months": months})
        return metadata

    parsed_effective = _parse_statement_date_label(effective_date)
    if parsed_effective is None:
        return metadata

    parsed_dates = sorted(
        {
            parsed
            for parsed in (_parse_statement_date_label(value) for value in available_dates)
            if parsed is not None
        }
    )
    if len(parsed_dates) < 2:
        return metadata
    if parsed_dates and parsed_effective != max(parsed_dates):
        return metadata
    previous_date = parsed_dates[-2]
    if (parsed_effective.month, parsed_effective.day) == (previous_date.month, previous_date.day):
        return metadata
    anchor_candidates = [date for date in parsed_dates if (date.month, date.day) != (parsed_effective.month, parsed_effective.day)]
    if anchor_candidates:
        dominant_anchor = max(
            {(date.month, date.day) for date in anchor_candidates},
            key=lambda month_day: sum((candidate.month, candidate.day) == month_day for candidate in anchor_candidates),
        )
    else:
        dominant_anchor = (12, 31)
    metadata.update(
        {
            "fiscal_year_end_month": dominant_anchor[0],
            "fiscal_year_end_day": dominant_anchor[1],
        }
    )

    if (parsed_effective.month, parsed_effective.day) == dominant_anchor:
        return metadata

    months = (parsed_effective.month - dominant_anchor[0]) % 12
    if months <= 0:
        months = parsed_effective.month
    if months >= 12:
        return metadata

    factor = 12.0 / float(months)
    metadata.update({"annualized": True, "factor": factor, "months": months})
    return metadata


# --- end workbook/schedules.py ---
