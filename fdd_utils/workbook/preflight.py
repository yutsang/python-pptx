from __future__ import annotations


from .inspector import _cell_text
import difflib
from functools import lru_cache
import logging
import os
import re
import time
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import load_workbook

from ..financial_common import cell_text

logger = logging.getLogger(__name__)


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _rows_are_blank(rows: Sequence[Sequence[Any]]) -> bool:
    return not any(any(not _is_empty_value(value) for value in row) for row in rows)


def _serialize_rows(rows: Iterable[Sequence[Any]]) -> List[List[Any]]:
    return [list(row) for row in rows]


def _is_likely_entity_name(value: str) -> bool:
    cleaned = value.strip()
    if not 2 < len(cleaned) < 50:
        return False
    if "|" in cleaned:
        return False
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", cleaned):
        return False
    if re.fullmatch(r"[\d\s\-–_/.:]+", cleaned):
        return False
    if cleaned.lower().startswith("project "):
        return False
    lowered = cleaned.lower()
    blocked_phrases = {
        "as at",
        "as of",
        "balance sheet",
        "income statement",
        "statement of financial position",
        "statement of profit or loss",
    }
    if lowered in blocked_phrases or lowered.startswith("as at ") or lowered.startswith("as of "):
        return False
    financial_terms = (
        "vat",
        "tax",
        "receivable",
        "payable",
        "income",
        "expense",
        "asset",
        "liability",
        "增值税",
        "税金",
        "应交",
        "應交",
        "应收",
        "應收",
        "应付",
        "應付",
        "收入",
        "成本",
        "费用",
        "資產",
        "资产",
        "負債",
        "负债",
    )
    if any(term in lowered for term in financial_terms):
        return False
    meta_words = (
        "confidential", "strictly confidential", "draft", "internal",
        "restricted", "private", "proprietary", "classified",
        "do not distribute", "for internal use only",
    )
    if any(m in lowered for m in meta_words):
        return False
    return True


def _looks_like_schedule_title_prefix(value: str) -> bool:
    lowered = value.lower()
    keywords = (
        "cash",
        "receivable",
        "prepayment",
        "payable",
        "capital",
        "property",
        "tax",
        "other current assets",
        "investment",
        "貨幣資金",
        "应收",
        "應收",
        "预付",
        "預付",
        "应付",
        "應付",
        "股本",
        "税",
    )
    return any(keyword in lowered for keyword in keywords)


def _looks_like_generic_entity_prefix(value: str) -> bool:
    cleaned = value.strip()
    lowered = cleaned.lower()
    if not cleaned or any(char.isdigit() for char in cleaned):
        return False
    return lowered in {"project", "entity", "company"}


def _looks_like_financial_schedule_preview(rows: Sequence[Sequence[Any]]) -> bool:
    flattened = " ".join(_cell_text(value).lower() for row in rows for value in row if not _is_empty_value(value))
    if not flattened:
        return False
    if any(token in flattened for token in ("indicative adjusted", "示意性调整后", "示意性調整後")):
        return True
    return bool(re.search(r"\b20\d{2}-\d{2}-\d{2}\b", flattened))


def _strip_leading_date_fragment(value: str) -> str:
    cleaned = value.strip()
    while True:
        updated = re.sub(
            r"^(?:\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}日?)\s*[-–]\s*",
            "",
            cleaned,
        ).strip()
        if updated == cleaned:
            return cleaned
        cleaned = updated


@lru_cache(maxsize=8)
def _build_workbook_preflight_cached(
    workbook_path: str,
    preview_rows: int = 12,
    entity_rows: int = 20,
    file_mtime_ns: int = 0,
    file_size: int = 0,
) -> Dict[str, Any]:
    del file_mtime_ns, file_size
    started = time.perf_counter()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets: List[Dict[str, Any]] = []

    try:
        max_preview_rows = max(preview_rows, entity_rows)
        for worksheet in workbook.worksheets:
            preview = _serialize_rows(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=max_preview_rows,
                    values_only=True,
                )
            )
            preview_slice = preview[:preview_rows]
            sheet_state = getattr(worksheet, "sheet_state", "visible") or "visible"
            max_row = int(getattr(worksheet, "max_row", 0) or 0)
            max_column = int(getattr(worksheet, "max_column", 0) or 0)
            sheets.append(
                {
                    "name": worksheet.title,
                    "sheet_state": sheet_state,
                    "is_hidden": sheet_state != "visible",
                    "is_blank_preview": _rows_are_blank(preview_slice) and max_row <= preview_rows,
                    "max_row": max_row,
                    "max_column": max_column,
                    "preview_rows": preview,
                }
            )
    finally:
        workbook.close()

    logger.debug(
        "Workbook preflight scanned %s sheets from %s in %.2fs",
        len(sheets),
        workbook_path,
        time.perf_counter() - started,
    )
    return {
        "workbook_path": workbook_path,
        "preview_rows": preview_rows,
        "entity_rows": entity_rows,
        "sheets": sheets,
    }


def build_workbook_preflight(
    workbook_path: str,
    preview_rows: int = 12,
    entity_rows: int = 20,
) -> Dict[str, Any]:
    stat = os.stat(workbook_path)
    return _build_workbook_preflight_cached(
        workbook_path,
        preview_rows,
        entity_rows,
        stat.st_mtime_ns,
        stat.st_size,
    )


def _visible_non_blank_sheets(preflight: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [
        sheet
        for sheet in preflight.get("sheets", [])
        if not sheet.get("is_hidden") and not sheet.get("is_blank_preview")
    ]


def extract_entity_names_from_preflight(preflight: Dict[str, Any]) -> List[str]:
    entity_sources: Dict[str, set[str]] = {}
    entity_counts: Dict[str, int] = {}
    visible_sheets = _visible_non_blank_sheets(preflight)

    def add_candidate(name: str, source: str) -> None:
        cleaned_name = _strip_leading_date_fragment(name)
        if not _is_likely_entity_name(cleaned_name):
            return
        cleaned_name = cleaned_name.strip()
        entity_sources.setdefault(cleaned_name, set()).add(source)
        entity_counts[cleaned_name] = entity_counts.get(cleaned_name, 0) + 1

    for sheet in visible_sheets:
        preview_rows = sheet.get("preview_rows", [])[: preflight.get("entity_rows", 20)]
        is_financial_schedule_preview = _looks_like_financial_schedule_preview(preview_rows)

        for row in preview_rows:
            for value in row:
                if _is_empty_value(value):
                    continue
                value_str = _cell_text(value)
                lowered = value_str.lower()
                if (
                    "示意性调整后" in value_str
                    or "balance sheet" in lowered
                    or "利润表" in value_str
                    or "income statement" in lowered
                ):
                    if " - " in value_str:
                        add_candidate(value_str.split(" - ", 1)[1].strip(), "summary_title")
                    elif " – " in value_str:
                        add_candidate(value_str.split(" – ", 1)[1].strip(), "summary_title")
                elif " - " in value_str or " – " in value_str:
                    parts = re.split(r"\s[-–]\s", value_str, maxsplit=1)
                    if len(parts) > 1:
                        prefix = parts[0].strip()
                        candidate = parts[1].strip()
                        if _looks_like_schedule_title_prefix(prefix):
                            source = "financial_schedule_title" if is_financial_schedule_preview else "schedule_title"
                            add_candidate(candidate, source)
                        elif _looks_like_generic_entity_prefix(prefix):
                            add_candidate(candidate, "generic_dash")

    entity_names = [
        name
        for name, count in entity_counts.items()
        if "summary_title" in entity_sources.get(name, set())
        or "financial_schedule_title" in entity_sources.get(name, set())
        or "generic_dash" in entity_sources.get(name, set())
        or count >= 2
    ]
    return sorted(name for name in entity_names if name and name.strip())


_CJK_RUN_RE = re.compile(r"[\u4e00-\u9fff]+")


def split_bilingual_entity_name(name: str) -> tuple[Optional[str], Optional[str]]:
    """A single extracted candidate is sometimes a mixed CJK+English string
    (e.g. "南通通海 Nantong Tonghai" or "无锡项目 (Wuxi Project)") -- the entity
    name selector previously only ever offered that combined form, with no
    way to pick just the Chinese or just the English half. Returns
    (chinese_only, english_only), each None if the name isn't genuinely
    bilingual (pure-CJK or pure-Latin names have nothing to split) or if
    either half would come out empty.
    """
    text = str(name or "").strip()
    if not text:
        return None, None
    has_cjk = bool(_CJK_RUN_RE.search(text))
    has_latin = bool(re.search(r"[A-Za-z]", text))
    if not (has_cjk and has_latin):
        return None, None

    chinese_only = "".join(_CJK_RUN_RE.findall(text)).strip()
    english_only = _CJK_RUN_RE.sub(" ", text)
    english_only = re.sub(r"[()\[\]（）【】\-–—,，·.]", " ", english_only)
    english_only = re.sub(r"\s+", " ", english_only).strip()
    if not chinese_only or not english_only:
        return None, None
    return chinese_only, english_only


def get_financial_sheet_options(preflight: Dict[str, Any]) -> List[str]:
    def sheet_score(sheet: Dict[str, Any]) -> tuple[int, str]:
        lowered = str(sheet.get("name", "")).lower()
        score = 100
        financial_summary_prefix = (
            lowered == "financial"
            or lowered == "financials"
            or lowered.startswith("financials ")
            or lowered.startswith("financials-")
            or lowered.startswith("financials -")
            or lowered.startswith("financial -")
        )
        if "financial" in lowered:
            score += 60
        if "balance" in lowered or "income" in lowered or "profit" in lowered:
            score += 40
        if lowered.startswith("bs") or lowered.startswith("is"):
            score += 30
        if lowered in {"bshn", "bs"}:
            score += 20
        if "-->" in lowered or lowered == "adj":
            score -= 20
        return (0 if financial_summary_prefix else 1, -score, lowered)

    visible_sheets = _visible_non_blank_sheets(preflight)
    return [sheet["name"] for sheet in sorted(visible_sheets, key=sheet_score)]


def _normalize_for_sheet_match(text: str) -> str:
    return re.sub(r"[\s\-_·．.()（）]+", "", str(text or "")).strip().lower()


def suggest_rollup_sheet_for_entity(entity_name: str, sheet_names: List[str]) -> Optional[str]:
    """Fuzzy-suggest which sheet in an uploaded roll-up ("主表") workbook
    belongs to a given entity, for the batch-processing flow's per-entity
    roll-up-sheet dropdown default (still overridable by the user — see the
    batch UI in fdd_app.py). Unlike get_financial_sheet_options above (which
    ranks sheets purely by how "Financials-like" the sheet NAME looks, with
    no entity context at all), this scores each sheet against the entity's
    own name: substring containment first (handles CJK entity names like
    "南通通海" appearing inside "南通通海Financials" with no separator
    between them), falling back to a fuzzy ratio for near-miss spellings,
    with a small tie-breaking bonus for sheet names ending in "Financials"
    (the roll-up file's own naming convention, per the single-file roll-up
    picker's own docstring example). Returns None (no default) when nothing
    clears a hand-picked confidence floor — silently picking the wrong
    entity's sheet is worse than leaving the dropdown unset for the user to
    pick manually.
    """
    entity_norm = _normalize_for_sheet_match(entity_name)
    if not entity_norm or not sheet_names:
        return None

    best_sheet: Optional[str] = None
    best_score = 0.0
    for sheet in sheet_names:
        sheet_norm = _normalize_for_sheet_match(sheet)
        if not sheet_norm:
            continue
        if entity_norm in sheet_norm or sheet_norm in entity_norm:
            score = 0.9
        else:
            score = difflib.SequenceMatcher(None, entity_norm, sheet_norm).ratio()
        if str(sheet).strip().lower().endswith("financials"):
            score += 0.05
        if score > best_score:
            best_score = score
            best_sheet = sheet

    return best_sheet if best_score >= 0.45 else None
# --- end workbook/preflight.py ---
