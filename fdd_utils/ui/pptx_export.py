from __future__ import annotations

# re-added: bound by an import in another section of the pre-split module
from ..ai import (
    FDDConfig,
    WORKBENCH_AVAILABLE_MODELS,
    build_highlighted_commentary_html,
    get_default_config_path,
    get_prompt_engine,
    is_provider_ready,
    load_yaml_config,
    parse_validator_response,
    run_ai_pipeline_with_progress,
    run_generator_reprompt,
)
from ..financial_common import extract_result_text_content, get_pipeline_result_text
import pandas as pd


from .views import derive_reconciliation_matched_keys, detect_statement_mode
from .ai_panel import build_selected_pipeline_dfs, effective_mappings_from_session
import datetime as dt_module
import logging
import os
import re
import time
from typing import Any, Callable, Dict, List, Optional

import streamlit as st

from ..pptx import build_pptx_structured_payloads
from ..workbook import find_mapping_key, get_effective_mappings, load_mappings

logger = logging.getLogger(__name__)


def generate_pptx_presentation(
    *,
    session_state: Any,
    pptx_available: bool,
) -> None:
    if not session_state.ai_results:
        st.error("❌ No AI results available. Generate AI content first.")
        return

    if not pptx_available:
        st.error("❌ PPTX generation not available. Missing required modules.")
        return

    project_name = session_state.get("project_name", "Project")
    entity_name = session_state.get("entity_name", project_name)
    language = session_state.get("language", "Eng")
    mappings = effective_mappings_from_session(session_state)

    template_path = None
    for template in ["fdd_utils/template.pptx", "template.pptx"]:
        if os.path.exists(template):
            template_path = template
            break
    if not template_path:
        st.error("❌ PowerPoint template not found. Please add `fdd_utils/template.pptx` or `template.pptx`.")
        return

    output_dir = "fdd_utils/output"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = dt_module.datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized_entity = re.sub(r"[^\w\-_]", "_", str(entity_name)).strip("_") or "Project"
    selected_pipeline_dfs = build_selected_pipeline_dfs(session_state)

    try:
        combined_output_path = os.path.join(output_dir, f"{sanitized_entity}_{timestamp}.pptx")
        from fdd_utils.pptx import export_pptx_from_structured_data_combined

        structured_payloads = build_pptx_structured_payloads(
            ai_results=session_state.ai_results,
            mappings=mappings,
            bs_is_results=session_state.bs_is_results,
            dfs=selected_pipeline_dfs,
        )
        bs_data = structured_payloads.get("BS", [])
        is_data = structured_payloads.get("IS", [])
        logger.debug("PPTX payload account counts | BS=%s | IS=%s", len(bs_data), len(is_data))

        if not bs_data and not is_data:
            st.error("❌ No content generated for PPTX")
            logger.debug(
                "PPTX payload is empty | ai_results_keys=%s | dfs_keys=%s",
                list(session_state.ai_results.keys())[:10] if session_state.ai_results else "None",
                list(selected_pipeline_dfs.keys())[:10] if selected_pipeline_dfs else "None",
            )
            return

        # Demo mode: skip coSummaryShape AI so export is instant.
        _demo_cfg2 = (FDDConfig().config or {}).get("demo", {})
        _is_demo_pptx = bool(
            _demo_cfg2.get("filename") and
            str(session_state.get("uploaded_filename") or "").strip() == str(_demo_cfg2.get("filename") or "").strip()
        )
        # embed_financial_tables reads (temp_path, selected_sheet) for its
        # currency-unit-label detection and as a fresh-extraction fallback --
        # when this entity's financials came from an uploaded roll-up
        # workbook (the "進階：主表" expander), that source -- not this
        # entity's own file/selected_sheet -- is the one that actually holds
        # them. Mirrors process_workbook_data's own precedence (financials_from
        # or temp_path).
        _rollup_temp_path = session_state.get("rollup_temp_path")
        _financials_workbook_path = _rollup_temp_path or session_state.get("temp_path")
        _financials_sheet_name = (
            session_state.get("rollup_sheet") if _rollup_temp_path else session_state.get("selected_sheet")
        )

        with st.spinner("Generating PPTX…"):
            export_pptx_from_structured_data_combined(
                template_path,
                bs_data,
                is_data,
                combined_output_path,
                entity_name,
                language="chinese" if language == "Chn" else "english",
                temp_path=_financials_workbook_path,
                selected_sheet=_financials_sheet_name,
                is_chinese_databook=(language == "Chn"),
                bs_is_results=session_state.get("bs_is_results"),
                model_type=session_state.get("model_type", "local"),
                model_name=session_state.get("model_name"),
                skip_summary_ai=False,
                pre_generated_summaries=session_state.get("section_summaries") or None,
                mappings=mappings,
            )
        if os.path.exists(combined_output_path):
            with open(combined_output_path, "rb") as handle:
                session_state.pptx_download_data = handle.read()
            session_state.pptx_download_filename = os.path.basename(combined_output_path)
            session_state.pptx_download_mime = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
            session_state.pptx_ready = True

    except Exception as exc:
        st.error(f"❌ PPTX generation failed: {exc}")
        import traceback

        st.code(traceback.format_exc())


def batch_extract_entity_data(
    *,
    temp_path: str,
    entity_name: str,
    selected_sheet: Optional[str] = None,
    financials_from: Optional[str] = None,
    financials_sheet: Optional[str] = None,
    mapping_overrides: Optional[Dict[str, str]] = None,
    language: Optional[str] = None,
) -> Dict[str, Any]:
    """Phase 1 of the batch entity pipeline (process + reconcile -- fast).

    Split out from what used to be one single batch_process_entity() call
    so a checkpoint-based batch UI can st.rerun() BETWEEN this and
    batch_run_ai_for_entity() (phase 2, slow). Streamlit only paints the
    browser once a script run returns control to it: an on_data_ready-style
    callback fired midway through one long blocking call can update
    session_state all it wants, but the switcher/data-view code living
    later in that SAME script run still can't actually render anything
    until the whole call returns -- and it doesn't return until AI
    generation is ALSO done, defeating the entire point of showing data
    before AI finishes. Only an actual rerun between two separately
    checkpointed phases makes the browser paint an intermediate state.

    financials_from/financials_sheet point BS/IS extraction at a sibling
    roll-up ("主表") workbook's named sheet when this entity's own file has
    no Financials-pattern sheet of its own — same mechanism
    process_workbook_data already exposes for the single-file flow.

    Returns {"status": "failed", "entity_name", "error"} on failure, or on
    success {"status": "ok", "entity_name", "data_summary": {the same
    "state"-shaped partial bundle a caller can swap into st.session_state
    and pass to render_data_tables_section for the full recon+breakdown
    view, plus match-count summaries}, "_internal": {everything
    batch_run_ai_for_entity needs to continue without re-deriving it}}.
    """
    from ..workbook import process_workbook_data

    result: Dict[str, Any] = {"entity_name": entity_name, "status": "ok"}

    try:
        state = process_workbook_data(
            temp_path=temp_path,
            entity_name=entity_name,
            selected_sheet=selected_sheet,
            mapping_overrides=mapping_overrides,
            financials_from=financials_from,
            financials_sheet=financials_sheet,
        )
    except Exception as exc:
        result["status"] = "failed"
        result["error"] = f"Processing failed: {exc}"
        return result

    dfs = state.get("dfs") or {}
    if not dfs:
        result["status"] = "failed"
        result["error"] = "No schedule tabs could be extracted from this databook."
        return result

    reconciliation = state.get("reconciliation")
    resolution = state.get("resolution")
    mappings = get_effective_mappings(load_mappings(), resolution)

    # Raw process_workbook_data language is "Eng"/"Chi" (workbook.py's own
    # detection convention); normalise to the UI's "Eng"/"Chn" convention so
    # this matches every == "Chn" check generate_pptx_presentation makes,
    # unless the caller already passed an explicit override in that form.
    if language:
        effective_language = language
    else:
        raw_language = str(state.get("language") or "Eng").strip()
        effective_language = "Chn" if raw_language in ("Chi", "Chn", "chinese", "Chinese") else "Eng"

    statement_mode = detect_statement_mode(reconciliation)
    if statement_mode in ("is_only", "bs_only"):
        target_type = "IS" if statement_mode == "is_only" else "BS"
        matched_mapping_keys = [
            k for k in dfs
            if mappings.get(find_mapping_key(k, mappings) or k, {}).get("type") == target_type
        ]
        if not matched_mapping_keys:
            matched_mapping_keys = list(dfs.keys())
    else:
        matched_mapping_keys = derive_reconciliation_matched_keys(reconciliation, dfs.keys(), resolution, dfs=dfs)
        has_reconciliation_data = bool(
            reconciliation and any(recon_df is not None and not recon_df.empty for recon_df in reconciliation)
        )
        if not has_reconciliation_data:
            matched_mapping_keys = list(dfs.keys())

    if not matched_mapping_keys:
        result["status"] = "failed"
        result["error"] = "No eligible accounts after reconciliation filtering."
        return result

    bs_recon, is_recon = (list(reconciliation) + [None, None])[:2] if reconciliation else (None, None)
    result["data_summary"] = {
        "entity_name": entity_name,
        "accounts_total": len(dfs),
        "accounts_matched": len(matched_mapping_keys),
        "bs_match_counts": bs_recon["Match"].value_counts().to_dict() if bs_recon is not None and not bs_recon.empty else {},
        "is_match_counts": is_recon["Match"].value_counts().to_dict() if is_recon is not None and not is_recon.empty else {},
        # Raw per-account reconciliation breakdowns (same DataFrames
        # render_reconciliation_section uses in the interactive single-file
        # flow) -- so a caller can show the actual account-by-account
        # table, not just the match-status counts.
        "bs_recon_df": bs_recon,
        "is_recon_df": is_recon,
        # Full session_state-shaped (minus ai_results/pptx) partial bundle
        # -- lets a caller swap this into st.session_state and call
        # render_data_tables_section() for the complete per-account
        # breakdown view (cash, investment properties, etc., not just
        # reconciliation), the same rich view a fully-finished entity
        # gets, while AI is still running (or hasn't started yet).
        "state": {
            "dfs": dfs,
            "display_dfs": state.get("display_dfs"),
            "workbook_list": state.get("workbook_list"),
            "display_workbook_list": state.get("display_workbook_list"),
            "language": effective_language,
            "bs_is_results": state.get("bs_is_results"),
            "reconciliation": reconciliation,
            "resolution": resolution,
            "entity_name": entity_name,
        },
    }
    result["_internal"] = {
        "raw_state": state,
        "dfs": dfs,
        "reconciliation": reconciliation,
        "resolution": resolution,
        "mappings": mappings,
        "matched_mapping_keys": matched_mapping_keys,
        "effective_language": effective_language,
        "entity_name": entity_name,
        "temp_path": temp_path,
        "selected_sheet": selected_sheet,
        "financials_from": financials_from,
        "financials_sheet": financials_sheet,
        "mapping_overrides": mapping_overrides,
    }
    return result


def batch_run_ai_for_entity(
    *,
    extracted: Dict[str, Any],
    model_type: str = "local",
    model_name: Optional[str] = None,
    use_multithreading: bool = True,
    max_workers: Optional[int] = None,
    user_comments: Optional[Dict[str, str]] = None,
    template_path: Optional[str] = None,
    output_dir: str = "fdd_utils/output",
    progress_callback: Optional[Callable[..., None]] = None,
) -> Dict[str, Any]:
    """Phase 2 of the batch entity pipeline (AI generation + PPTX export --
    slow). Takes the successful result dict batch_extract_entity_data()
    returned (via its "_internal" bundle) and picks up where extraction
    left off, without re-deriving anything.

    Returns the same shape batch_process_entity's single-call version
    always did: {"status", "output_path", "bs_count", "is_count",
    "accounts_processed", "state": {full session_state-shaped bundle
    including ai_results/pptx_download_data, for swapping into
    st.session_state and reusing render_processed_view unchanged}} on
    success, {"status": "failed", "entity_name", "error"} on failure.
    """
    from ..ai import run_ai_pipeline_with_progress
    from ..pptx import export_pptx_from_structured_data_combined

    internal = extracted["_internal"]
    entity_name = internal["entity_name"]
    dfs = internal["dfs"]
    state = internal["raw_state"]
    reconciliation = internal["reconciliation"]
    resolution = internal["resolution"]
    mappings = internal["mappings"]
    matched_mapping_keys = internal["matched_mapping_keys"]
    effective_language = internal["effective_language"]
    temp_path = internal["temp_path"]
    selected_sheet = internal["selected_sheet"]
    financials_from = internal["financials_from"]
    financials_sheet = internal["financials_sheet"]
    mapping_overrides = internal["mapping_overrides"]

    result: Dict[str, Any] = {"entity_name": entity_name, "status": "ok"}

    ai_results = run_ai_pipeline_with_progress(
        mapping_keys=matched_mapping_keys,
        dfs=dfs,
        model_type=model_type,
        model_name=model_name,
        language=effective_language,
        use_multithreading=use_multithreading,
        max_workers=max_workers,
        progress_callback=progress_callback,
        user_comments=user_comments or {},
    )

    # Executive summary (coSummaryShape) generation -- mirrors what
    # render_ai_generation_section does for the single-file flow right
    # after its own per-account AI pass, which this batch path never had
    # an equivalent of. Without it, export_pptx_from_structured_data_
    # combined's own in-export summary call is SKIPPED ENTIRELY (a
    # deliberate choice there, not a bug -- an in-export LLM call was
    # reported to hang 10+ minutes when the API was flaky), leaving
    # coSummaryShape genuinely blank on every entity's first BS/IS slide.
    # Confirmed via a real batch export's --dump-text output: a literal
    # empty coSummaryShape text frame on both statements.
    from ..pptx import PowerPointGenerator
    is_chinese_db = effective_language == "Chn"
    section_summaries: Dict[str, str] = {}
    try:
        bs_blob: List[str] = []
        is_blob: List[str] = []
        for account_key, ai_result in ai_results.items():
            mapping_key = find_mapping_key(account_key, mappings)
            if not mapping_key or mapping_key not in mappings:
                continue
            atype = mappings[mapping_key].get("type")
            text = extract_result_text_content(
                (ai_result or {}).get("final")
                or (ai_result or {}).get("subagent_4")
                or (ai_result or {}).get("subagent_2")
                or (ai_result or {}).get("subagent_1")
                or ""
            )
            if not text.strip():
                continue
            # A page/section summary needs each account's lead-in theme
            # only, never a table account's per-component "-"/"➢" detail
            # bullets after "明细如下：" -- see strip_table_detail_for_
            # summary's own docstring for the real corrupted-coSummaryShape
            # bug this prevents.
            text = PowerPointGenerator.strip_table_detail_for_summary(text, is_chinese_db)
            if atype == "BS":
                bs_blob.append(text)
            elif atype == "IS":
                is_blob.append(text)
        for stmt, blob in (("BS", bs_blob), ("IS", is_blob)):
            if not blob:
                continue
            try:
                from ..ai import _PIPELINE_BREAKER
                if any(_PIPELINE_BREAKER.is_open(stage) for stage in ("subagent_1", "subagent_2")):
                    continue
            except Exception:
                pass
            summary = PowerPointGenerator.generate_section_summary(
                "\n\n".join(blob),
                is_chinese=is_chinese_db,
                language=("chinese" if is_chinese_db else "english"),
                model_type=model_type,
                model_name=model_name,
            )
            if summary:
                section_summaries[stmt] = summary
    except Exception as exc:
        logger.warning("Batch section summary generation failed for %s (PPTX summary will be blank): %s", entity_name, exc)
        section_summaries = {}

    structured_payloads = build_pptx_structured_payloads(
        ai_results=ai_results,
        mappings=mappings,
        bs_is_results=state.get("bs_is_results"),
        dfs=dfs,
    )
    bs_data = structured_payloads.get("BS", [])
    is_data = structured_payloads.get("IS", [])
    if not bs_data and not is_data:
        result["status"] = "failed"
        result["error"] = "No content generated for PPTX (empty BS and IS payloads)."
        return result

    resolved_template_path = template_path
    if not resolved_template_path:
        for candidate in ["fdd_utils/template.pptx", "template.pptx"]:
            if os.path.exists(candidate):
                resolved_template_path = candidate
                break
    if not resolved_template_path:
        result["status"] = "failed"
        result["error"] = "PowerPoint template not found (fdd_utils/template.pptx)."
        return result

    os.makedirs(output_dir, exist_ok=True)
    timestamp = dt_module.datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized_entity = re.sub(r"[^\w\-_]", "_", str(entity_name)).strip("_") or "Entity"
    output_path = os.path.join(output_dir, f"{sanitized_entity}_{timestamp}.pptx")

    # embed_financial_tables reads (temp_path, selected_sheet) for its
    # currency-unit-label detection and as a fresh-extraction fallback --
    # when financials came from a roll-up workbook, that source (not this
    # entity's own file/blank sheet) is the one that actually holds them.
    # Mirrors process_workbook_data's own precedence (financials_from or
    # temp_path).
    financials_workbook_path = financials_from or temp_path
    financials_sheet_name = financials_sheet if financials_from else selected_sheet

    export_pptx_from_structured_data_combined(
        resolved_template_path,
        bs_data,
        is_data,
        output_path,
        entity_name,
        language="chinese" if effective_language == "Chn" else "english",
        temp_path=financials_workbook_path,
        selected_sheet=financials_sheet_name,
        is_chinese_databook=(effective_language == "Chn"),
        bs_is_results=state.get("bs_is_results"),
        model_type=model_type,
        model_name=model_name,
        skip_summary_ai=False,
        pre_generated_summaries=section_summaries or None,
        mappings=mappings,
    )

    result["output_path"] = output_path
    result["bs_count"] = len(bs_data)
    result["is_count"] = len(is_data)
    result["accounts_processed"] = len(matched_mapping_keys)

    with open(output_path, "rb") as handle:
        pptx_bytes = handle.read()

    # Full session_state-shaped bundle so a caller (the batch UI) can swap
    # this entity's results into st.session_state and reuse the single-file
    # render_processed_view/generate_pptx_presentation UI UNCHANGED, instead
    # of only ever seeing this thin status dict.
    result["state"] = {
        "dfs": state.get("dfs"),
        "display_dfs": state.get("display_dfs"),
        "dfs_variants": state.get("dfs_variants"),
        "display_df_variants": state.get("display_df_variants"),
        "workbook_list": state.get("workbook_list"),
        "display_workbook_list": state.get("display_workbook_list"),
        "language": effective_language,
        "detected_language": effective_language,
        "bs_is_results": state.get("bs_is_results"),
        "reconciliation": reconciliation,
        "resolution": resolution,
        "project_name": state.get("project_name"),
        "entity_name": entity_name,
        # The financials source (not necessarily this entity's own file --
        # see financials_workbook_path/financials_sheet_name above), so a
        # later "Regenerate PPTX" click from within the reused single-file
        # UI still finds the right sheet for the embedded table instead of
        # re-hitting the same blank-selected_sheet bug this export call
        # just worked around.
        "temp_path": financials_workbook_path,
        "selected_sheet": financials_sheet_name,
        "mapping_overrides": mapping_overrides,
        "ai_results": ai_results,
        # So a later "Regenerate PPTX" click from within the reused
        # single-file UI (generate_pptx_presentation, which reads
        # session_state.section_summaries) reuses these instead of
        # falling back to the in-export summary skip.
        "section_summaries": section_summaries,
        "model_type": model_type,
        "model_name": model_name,
        "use_multithreading": use_multithreading,
        "pptx_ready": True,
        "pptx_download_data": pptx_bytes,
        "pptx_download_filename": os.path.basename(output_path),
        "pptx_download_mime": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    }
    return result


def batch_process_entity(
    *,
    temp_path: str,
    entity_name: str,
    selected_sheet: Optional[str] = None,
    financials_from: Optional[str] = None,
    financials_sheet: Optional[str] = None,
    mapping_overrides: Optional[Dict[str, str]] = None,
    model_type: str = "local",
    model_name: Optional[str] = None,
    language: Optional[str] = None,
    use_multithreading: bool = True,
    max_workers: Optional[int] = None,
    user_comments: Optional[Dict[str, str]] = None,
    template_path: Optional[str] = None,
    output_dir: str = "fdd_utils/output",
    progress_callback: Optional[Callable[..., None]] = None,
    on_data_ready: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> Dict[str, Any]:
    """Headless, session_state-free equivalent of the single-file
    process -> reconcile -> AI -> export flow, for driving one entity in
    a single blocking call -- a thin composition of
    batch_extract_entity_data() then batch_run_ai_for_entity(), kept for
    callers that want one-shot headless behavior (e.g. inspect_databook.py
    -style scripts). Mirrors inspect_databook.py's inspect_one() pattern.

    A checkpoint-based batch UI that wants the browser to actually paint
    an intermediate "data ready, AI still pending" state should call the
    two phases separately across two st.rerun()s instead -- see
    fdd_app.py's render_batch_processing_section, and the phase functions'
    own docstrings for why a callback fired midway through this single
    call can't achieve that on its own.

    on_data_ready, if given, fires once (right after data extraction +
    reconciliation complete, before AI generation starts) with the
    extraction phase's "data_summary" dict.
    """
    extracted = batch_extract_entity_data(
        temp_path=temp_path,
        entity_name=entity_name,
        selected_sheet=selected_sheet,
        financials_from=financials_from,
        financials_sheet=financials_sheet,
        mapping_overrides=mapping_overrides,
        language=language,
    )
    if extracted.get("status") != "ok":
        return extracted

    if on_data_ready:
        try:
            on_data_ready(extracted["data_summary"])
        except Exception:
            pass  # a UI-side display glitch should never abort the pipeline

    return batch_run_ai_for_entity(
        extracted=extracted,
        model_type=model_type,
        model_name=model_name,
        use_multithreading=use_multithreading,
        max_workers=max_workers,
        user_comments=user_comments,
        template_path=template_path,
        output_dir=output_dir,
        progress_callback=progress_callback,
    )
# --- end ui/pptx_export.py ---


# --- Bridge Lab (experimental, isolated testing page) ---
# Lets the project team upload a workbook, pick one tab, and auto-detect a
# 'Base'/'Change' bridge helper block (the same convention as the real
# 成都-量价桥图 tab) to generate a NATIVE Excel waterfall chart -- not PPTX,
# because the team's actual downstream tool (UpSlide) links PowerPoint
# charts FROM Excel chart objects; it has no concept of a python-pptx chart,
# so a python-pptx output would be a dead end for their real workflow.
# Entirely separate code path from the main upload/process/AI/PPTX flow --
# gated behind its own session_state flag so it can't interfere with it.
import io as _bridge_lab_io

from openpyxl import Workbook as _bridge_lab_Workbook
from openpyxl import load_workbook as _bridge_lab_load_workbook

from ..bridge_chart_prototype import build_excel_waterfall_chart, find_bridge_blocks
from ..generate_bridge_waterfall_batch import build_bridges_for_ab_tab


def render_bridge_lab_toggle() -> None:
    with st.sidebar:
        st.markdown("---")
        if st.session_state.get("show_bridge_lab"):
            if st.button("← 返回主流程", use_container_width=True, key="bridge_lab_back_btn"):
                st.session_state["show_bridge_lab"] = False
                st.rerun()
        else:
            if st.button("🧪 橋圖測試 (Bridge Lab)", use_container_width=True, key="bridge_lab_enter_btn"):
                st.session_state["show_bridge_lab"] = True
                st.rerun()


def _bridge_lab_show_block(index: int, block, note: str = "") -> bool:
    """Renders one detected bridge block as a table plus a chart, and returns
    whether it's usable (check passed, or unverified -- never a hard mismatch).

    Previously printed one st.write per item, which for 11 items across
    several transitions filled the page with lines nobody reads. A table and
    the chart itself are what actually get checked, so that is all this shows;
    the per-item source notes still travel into the downloaded workbook."""
    header = f"區塊 {index + 1}：{block.items[0].label} → {block.items[-1].label}"
    with st.expander(header, expanded=True):
        if note:
            st.caption(note)

        rows = []
        running = 0.0
        for it in block.items:
            if it.kind == "total":
                running = it.value
                rows.append({"項目": it.label, "類型": "合計", "金額 (千元)": round(it.value, 1),
                             "累計": round(running, 1)})
            else:
                running += it.value
                rows.append({"項目": it.label, "類型": "變動", "金額 (千元)": round(it.value, 1),
                             "累計": round(running, 1)})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        try:
            _bridge_lab_render_preview_chart(block, header)
        except Exception as exc:  # a preview failure must not block the download
            st.caption(f"（圖表預覽無法產生：{exc}）")

        if block.check_ok is True:
            st.success("✅ 核對一致")
            return True
        if block.check_ok is False:
            st.error("❌ 核對不一致 -- 不會為此區塊生成圖表，請檢查來源表格")
            return False
        st.warning("⚠️ 無法核對一致性，仍會生成圖表，請自行核對數字")
        return True


def _bridge_lab_render_preview_chart(block, title: str) -> None:
    """On-screen waterfall preview using the same invisible-base-series maths
    as the downloaded Excel chart, so what is checked here is what ships."""
    from ..bridge_chart_prototype import _compute_waterfall_series

    categories, base_vals, total_vals, inc_vals, dec_vals = _compute_waterfall_series(block)
    frame = pd.DataFrame(
        {
            "（基準）": base_vals,
            "合計": total_vals,
            "增加": inc_vals,
            "減少": dec_vals,
        },
        index=categories,
    )
    st.bar_chart(frame, stack=True, height=340,
                 color=("#00000000", "#00338D", "#6D2077", "#00A3A1"))


def render_bridge_lab() -> None:
    st.title("🧪 橋圖測試 (Bridge Chart Lab)")
    st.caption(
        "實驗性功能，與主流程完全獨立。支援兩類 tab：(1) 已建好 Base/Change 輔助區塊的橋圖表，"
        "(2) AB- 原始數據表（自動計算價/量/天數因子分解）。選擇 tab 後自動偵測、生成原生 Excel "
        "疊加圖，供下載後透過 UpSlide 帶入 PPT。"
    )

    uploaded = st.file_uploader("上傳 Excel 檔案 (.xlsx)", type=["xlsx"], key="bridge_lab_upload")
    if not uploaded:
        st.info("請先上傳一個 .xlsx 檔案。")
        return

    file_bytes = uploaded.getvalue()
    try:
        wb_values = _bridge_lab_load_workbook(_bridge_lab_io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        st.error(f"無法讀取此 Excel 檔案：{exc}")
        return

    sheet_names = wb_values.sheetnames
    selected_sheet = st.selectbox("選擇 tab", sheet_names, key="bridge_lab_sheet")

    if not st.button("偵測並生成", type="primary", key="bridge_lab_detect_btn"):
        return

    ws = wb_values[selected_sheet]

    # Route by tab type. A pre-built Base/Change helper block (like a
    # <entity>-量价桥图 tab) is read directly; otherwise try to treat it as a
    # raw AB-* data tab and COMPUTE the factor decomposition. Both paths
    # converge on a list of BridgeBlock objects rendered identically.
    renderable = []  # list of BridgeBlock that passed (or lack) their check
    prebuilt = find_bridge_blocks(ws)
    if prebuilt:
        st.success(f"偵測到 {len(prebuilt)} 個預建 Base/Change 橋圖區塊。")
        for i, block in enumerate(prebuilt):
            if _bridge_lab_show_block(i, block):
                renderable.append(block)
    else:
        ab_blocks, results = build_bridges_for_ab_tab(ws, selected_sheet)
        if ab_blocks is None:
            st.warning(
                f"「{selected_sheet}」既不是 Base/Change 結構的橋圖表，也不是可識別的 AB- 原始數據表"
                "（找不到 Year/Days 標籤列或分期區塊）。請確認選對了 tab。"
            )
            return
        if not results:
            st.warning(
                f"「{selected_sheet}」偵測為 AB- 原始數據表，但算不出任何年度轉換（可能只有單一年度資料）。"
            )
            return
        st.success(
            f"偵測為 AB- 原始數據表，已計算出 {len(results)} 個年度轉換橋圖"
            "（價/量/天數因子分解，末期採 LTM 滾動12個月口徑）。"
        )
        for i, res in enumerate(results):
            note = "註：末期為不完整年度，已改用 LTM 滾動12個月窗口比較" if res.is_ltm else ""
            if _bridge_lab_show_block(i, res.bridge, note=note):
                renderable.append(res.bridge)

    if not renderable:
        st.error("沒有通過核對、可生成圖表的區塊。")
        return

    # STANDALONE output workbook -- we build a brand-new Workbook and NEVER
    # re-save the user's upload, so their original file's formulas, cached
    # values, and existing native charts are physically untouched (openpyxl
    # round-tripping a workbook silently drops every formula's cached result,
    # which is what made an earlier version appear to "change" old tabs).
    out_wb = _bridge_lab_Workbook()
    out_wb.remove(out_wb.active)
    out_ws = out_wb.create_sheet("Bridge_Output")
    next_row = 1
    for block in renderable:
        title = f"{block.items[0].label} → {block.items[-1].label}"
        next_row = build_excel_waterfall_chart(out_ws, block, title, start_row=next_row)

    out_buffer = _bridge_lab_io.BytesIO()
    out_wb.save(out_buffer)
    st.download_button(
        "下載含橋圖的 Excel",
        data=out_buffer.getvalue(),
        file_name="bridge_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="bridge_lab_download_btn",
    )
    st.caption(
        "此為獨立的新檔案，只含生成的「Bridge_Output」分頁（結構化數據表＋橋圖）；"
        "你上傳的原始檔案完全沒有被修改。"
    )
