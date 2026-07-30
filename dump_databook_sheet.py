#!/usr/bin/env python3
"""Full grid dump of one databook sheet, with Excel row/column coordinates.

Built for the case where a table is identified visually -- from a photo or by
looking at the file -- and then has to be located precisely in the data so it
can be extracted. Existing tools either summarise (inspect_detail_tables.py
reports blocks) or scan for a pattern; this just shows what is in the cells,
addressed the way Excel addresses them, so a table seen on screen can be
matched to exact coordinates.

Every periodised block found is listed first, so the answer to "which block is
the one in the report" can be given as a block number.

Read-only.

Usage:
    python dump_databook_sheet.py "for_test/x.xlsx" --sheet 管理费用
    python dump_databook_sheet.py "for_test/x.xlsx" --sheet 管理费用 --rows 20-45
    python dump_databook_sheet.py "for_test/x.xlsx" --list
"""
import argparse
import re
import sys
import warnings

warnings.filterwarnings("ignore")
sys.path.insert(0, ".")

import pandas as pd
from openpyxl.utils import get_column_letter

_DATE_HDR_RE = re.compile(r"(20\d{2})\s*年|20\d{2}-\d{2}|\b20\d{2}\b|1-\d+月|\d+M\d{2}|"
                          r"\d{4}年\d{1,2}月\d{1,2}日")


def _text(v) -> str:
    s = str(v)
    return "" if s in ("nan", "NaT", "None") else s.strip()


def _row_cells(df, r):
    return [(c, _text(df.iloc[r, c])) for c in range(len(df.columns))
            if _text(df.iloc[r, c])]


def _is_period_header(df, r) -> bool:
    vals = [t for _c, t in _row_cells(df, r)]
    if len(vals) < 3:
        return False
    return sum(1 for v in vals if _DATE_HDR_RE.search(v)) >= 2


def _period_header_spans(df, r):
    """[(label_col, [period_cols])] for every periodised header on this row.

    A row can carry more than one table's header, because real sheets place a
    report-ready summary to the RIGHT of the main schedule rather than below
    it -- confirmed on a real sheet where '人民币千元 | 2021年 | ...' sat in
    columns Y-AC while columns B-J on the same rows held unrelated journal
    detail. A row-only scan reports one block and misses the other, and its
    labels come out as the journal's ('日期', '2022/3/22') instead of the
    table's ('人工成本', '中介费').
    """
    cells = _row_cells(df, r)
    spans = []
    current_label = None
    current_periods = []
    for c, text in cells:
        if _DATE_HDR_RE.search(text):
            if current_label is not None:
                current_periods.append(c)
        else:
            if current_label is not None and len(current_periods) >= 2:
                spans.append((current_label, current_periods))
            current_label, current_periods = c, []
    if current_label is not None and len(current_periods) >= 2:
        spans.append((current_label, current_periods))
    return spans


def find_blocks(df, max_rows=250):
    """Periodised blocks, each anchored to its own COLUMN range so tables laid
    out side by side are reported separately."""
    blocks = []
    n = min(len(df), max_rows)
    for r in range(n):
        for label_col, period_cols in _period_header_spans(df, r):
            data_rows, labels, blank = 0, [], 0
            j = r + 1
            while j < n and blank < 3:
                label = _text(df.iloc[j, label_col])
                nums = [c for c in period_cols
                        if re.fullmatch(r"-?[\d,]+(\.\d+)?",
                                        _text(df.iloc[j, c]).replace("(", "-").replace(")", ""))]
                if not label and not nums:
                    blank += 1
                    j += 1
                    continue
                blank = 0
                if _period_header_spans(df, j) and _text(df.iloc[j, label_col]) and not nums:
                    break
                if nums:
                    data_rows += 1
                    if label:
                        labels.append(label)
                j += 1
            if data_rows >= 2:
                title = ""
                for back in range(1, 3):
                    if r - back >= 0:
                        cand = _text(df.iloc[r - back, label_col])
                        if cand and not _DATE_HDR_RE.search(cand):
                            title = cand
                            break
                blocks.append({
                    "header_row": r, "end_row": j, "data_rows": data_rows,
                    "labels": labels, "title": title,
                    "label_col": label_col, "period_cols": period_cols,
                    "header": [_text(df.iloc[r, c]) for c in [label_col] + period_cols][:8],
                })
    return blocks


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path")
    ap.add_argument("--sheet", default=None, help="sheet to dump")
    ap.add_argument("--rows", default=None, metavar="N-N",
                     help="Excel row range to dump in full (1-indexed, inclusive)")
    ap.add_argument("--list", action="store_true", help="list sheet names and exit")
    ap.add_argument("--max-cols", type=int, default=14, help="columns to show per row")
    args = ap.parse_args()

    xl = pd.ExcelFile(args.path)
    if args.list or not args.sheet:
        print(f"{len(xl.sheet_names)} sheet(s) in {args.path!r}:")
        for name in xl.sheet_names:
            print(f"  {name}")
        if not args.sheet:
            print("\nPass --sheet <name> to dump one.")
        return 0
    if args.sheet not in xl.sheet_names:
        print(f"❌ sheet {args.sheet!r} not found. Use --list to see the names.")
        return 1

    df = pd.read_excel(args.path, sheet_name=args.sheet, header=None)
    print(f"Sheet {args.sheet!r}  {df.shape[0]} rows x {df.shape[1]} cols")
    print("(row numbers below are EXCEL rows, 1-indexed, so they match what you see)\n")

    blocks = find_blocks(df)
    print("=" * 78)
    print(f"PERIODISED BLOCKS FOUND: {len(blocks)}")
    print("=" * 78)
    for i, b in enumerate(blocks, 1):
        cols = [b["label_col"]] + b["period_cols"]
        col_range = f"{get_column_letter(cols[0] + 1)}-{get_column_letter(cols[-1] + 1)}"
        print(f"  BLOCK {i}: Excel rows {b['header_row'] + 1}-{b['end_row']}, "
              f"columns {col_range}   {b['data_rows']} data row(s)")
        if b["title"]:
            print(f"    title : {b['title']}")
        print(f"    header: {b['header']}")
        print(f"    labels: {b['labels'][:12]}")
        print()
    if blocks:
        print("  Tell me which BLOCK number is the one that goes in the report and I can")
        print("  extract exactly that, instead of inferring it from shape.\n")

    if args.rows:
        lo, hi = (int(x) for x in args.rows.split("-"))
        print("=" * 78)
        print(f"FULL DUMP: Excel rows {lo}-{hi}")
        print("=" * 78)
        for excel_row in range(lo, hi + 1):
            r = excel_row - 1
            if r < 0 or r >= len(df):
                continue
            cells = _row_cells(df, r)[:args.max_cols]
            if not cells:
                print(f"  r{excel_row:<4} (empty)")
                continue
            rendered = "  ".join(f"{get_column_letter(c + 1)}={t[:22]}" for c, t in cells)
            print(f"  r{excel_row:<4} {rendered}")
    else:
        print("Pass --rows N-N to dump those Excel rows cell by cell "
              "(e.g. --rows 20-45 to inspect a block closely).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
