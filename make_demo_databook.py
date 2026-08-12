#!/usr/bin/env python3
"""
Turn a real client databook into a demo one: scale every figure by a single
factor, and replace every real company / entity name with a stable fake.

Why one global scale factor rather than per-row noise: a databook's tie-out
structure (detail -> Subtotal -> Total -> Financials, `Check = 0`, BS balance)
is exactly what a demo must keep intact, and a linear transform preserves all
of it for free.  Per-row noise would mean re-deriving every subtotal by hand
from an inferred structure -- and being wrong somewhere.  Scaling also lets the
amounts quoted inside the Chinese remark text ("应收租金 29.1w") move by the
same factor and stay consistent with the table beside them.

FORMULAS: a databook is often mostly formulas -- one real client file measured
22,687 formula cells against 6,937 plain numbers.  Writing a scaled value into
a formula cell would destroy the formula, and leaving the formula alone would
leave openpyxl unable to report a value.  So the demo file is flattened: every
formula cell is replaced by its cached result x factor.  The output is a
pure-value workbook -- which is what the FDD pipeline reads anyway -- and stays
internally consistent by linearity.

Always run --inspect first: it prints a redacted report (safe to share) and
writes the unredacted one to a file, so the extraction can be checked against
this particular workbook's naming habits before anything is generated.

Usage:
    python make_demo_databook.py for_test/<databook>.xlsx --inspect
    python make_demo_databook.py for_test/<databook>.xlsx --dry-run
    python make_demo_databook.py for_test/<databook>.xlsx \
        --rename 上海某某置业有限公司=上海瑞茂置业有限公司
"""

from __future__ import annotations

import argparse
import hashlib
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl

DEFAULT_SCALE = 0.8734

# ---------------------------------------------------------------- name parsing

# Geography is kept as-is: the demo should still read as a real Chinese
# logistics deal, and a scrambled province makes the commentary absurd.
PLACE_PREFIXES = [
    "内蒙古", "黑龙江", "石家庄", "哈尔滨", "乌鲁木齐", "连云港", "张家港",
    "上海", "北京", "深圳", "广州", "天津", "重庆", "成都", "武汉", "南京",
    "杭州", "苏州", "无锡", "常州", "昆山", "宁波", "青岛", "济南", "西安",
    "郑州", "长沙", "合肥", "福州", "厦门", "大连", "沈阳", "佛山", "东莞",
    "珠海", "中山", "惠州", "嘉兴", "绍兴", "台州", "温州", "金华", "湖州",
    "南通", "徐州", "扬州", "镇江", "泰州", "盐城", "淮安", "宿迁", "太仓",
    "义乌", "余姚", "慈溪", "海宁", "平湖", "桐乡", "嘉善", "昆明", "南宁",
    "浙江", "江苏", "山东", "广东", "河北", "河南", "湖北", "湖南", "四川",
    "福建", "安徽", "辽宁", "陕西", "山西", "江西", "云南", "贵州", "广西",
    "海南", "甘肃", "青海", "宁夏", "新疆", "西藏", "吉林",
    "中国", "华东", "华南", "华北", "华中", "西南", "东北",
]
PLACE_PREFIXES.sort(key=len, reverse=True)

# Industry tail, kept as-is so the fake name still describes the same business.
INDUSTRY_TAILS = [
    "投资管理咨询", "企业管理咨询", "供应链管理", "设备安装工程", "房地产经纪",
    "国际贸易", "信息科技", "货物运输", "仓储服务", "物业服务", "物业管理",
    "投资管理", "智能科技", "网络科技", "电子商务", "生物科技", "环保科技",
    "造价咨询", "管理咨询", "安装工程", "财产保险", "供应链", "电动车",
    "新能源", "房地产", "商贸", "贸易", "物流", "电子", "实业", "科技",
    "仓储", "建设", "工程", "咨询", "保险", "运输", "置业", "装饰", "机械",
    "材料", "食品", "医药", "生物", "环保", "能源", "租赁", "销售", "服务",
    "发展", "实体", "工贸", "农业", "化工", "纺织", "家具", "包装", "印刷",
    "国际", "集团", "控股", "投资", "资产", "产业", "地产", "开发",
]
INDUSTRY_TAILS.sort(key=len, reverse=True)

COMPANY_SUFFIXES = [
    "股份有限公司", "有限责任公司", "有限公司", "合伙企业", "分公司",
    "集团", "公司", "中心", "工厂", "厂",
]
COMPANY_SUFFIXES.sort(key=len, reverse=True)

# Generic references that look like company names but name nobody.
GENERIC_NAMES = {
    "目标公司", "本公司", "母公司", "子公司", "关联公司", "总公司", "分公司",
    "公司", "第三方", "该公司", "集团", "关联方", "供应商", "承租人", "出租人",
    "物业公司", "施工单位", "建设单位", "保险公司", "银行", "税务局",
    "上级公司", "同一控制", "关联公司往来", "其他公司",
}

# Footnote markers glued to the front of a name cell ("*某某（浙江）物流有限公司").
# One real workbook puts these on most AR tenant rows; stripping them is what
# turned 998 near-misses into extracted names.
LEADING_MARK_RE = re.compile(r"^['\"`*＊#＃※\s]+")
TRAILING_MARK_RE = re.compile(r"[*＊#＃※\s]+$")

# Account-line noise welded to the front of a name: "应付上海某某有限公司".
# Stripped before the name is taken, so the fake keeps the account word intact.
LEADING_NOISE = [
    "其他应付款", "其他应收款", "其他应付", "其他应收", "预付账款", "预收账款",
    "应付账款", "应收账款", "长期应付", "长期应收", "应付", "应收", "预付",
    "预收", "往来", "关联方", "关联公司", "第三方", "客户", "供应商", "租户",
    "承租人", "出租人", "付", "收",
]
LEADING_NOISE.sort(key=len, reverse=True)

# A company name embedded in a longer cell ("上海某某有限公司往来款").  Used
# only on short cells; long remark sentences are covered by full-text
# substitution of names already found in clean cells.
EMBEDDED_NAME_RE = re.compile(
    r"[一-鿿（）()·、]{2,24}(?:" + "|".join(COMPANY_SUFFIXES) + r")"
)
EMBEDDED_MAX_CELL_LEN = 40

# Characters legitimately found inside a Chinese company name.
NAME_CHAR_RE = re.compile(r"[一-鿿（）()·、\s]")

# A whole cell that is nothing but a company name.  Deliberately anchored: a
# long remark sentence that merely *mentions* a company must NOT match, or the
# extracted "real name" would be a sentence fragment.  Remarks are covered
# instead by full-text substitution of the names found in these clean cells.
COMPANY_CELL_RE = re.compile(
    r"^[一-鿿（）()·、\s]{2,30}(?:" + "|".join(COMPANY_SUFFIXES) + r")$"
)

BRAND_CHARS = (
    "瑞宏恒嘉启泰盛鑫达丰隆通和兴利源正德信顺安佳元茂康联创立明辉阳"
    "卓越弘毅锦程远航睿臻拓新润泽晟昊博轩宁禾岳川岭峰晖朗骏祺"
)


def strip_marks(s: str) -> str:
    return TRAILING_MARK_RE.sub("", LEADING_MARK_RE.sub("", s))


def strip_leading_noise(s: str) -> str:
    """Drop one account-line prefix, so "应付上海某某有限公司" yields the name."""
    for noise in LEADING_NOISE:
        if s.startswith(noise) and len(s) - len(noise) >= 4:
            return s[len(noise):]
    return s


def extract_embedded_names(s: str) -> list[str]:
    """Names inside a longer cell, e.g. "上海某某有限公司往来款".

    90 cells in one real workbook carry a tenant name with a trailing account
    word, which whole-cell matching cannot see -- and those names would
    otherwise never be masked anywhere.  Restricted to short cells and to
    candidates that survive prefix-stripping, because a greedy match on a
    sentence would happily swallow the verb in front of the company.
    """
    if len(s) > EMBEDDED_MAX_CELL_LEN:
        return []
    found = []
    for m in EMBEDDED_NAME_RE.finditer(s):
        cand = strip_leading_noise(strip_marks(m.group(0)))
        if len(cand) >= 5 and cand not in GENERIC_NAMES and COMPANY_CELL_RE.match(cand):
            found.append(cand)
    return found


def normalize_brackets(s: str) -> str:
    """Half-width and full-width brackets name the same company.

    Both spellings occur for the same tenant on different rows; seeding the
    scramble from the normalised form keeps their fakes identical, so the demo
    does not split one tenant into two companies.
    """
    return s.replace("(", "（").replace(")", "）")


def split_company(name: str) -> tuple[str, str, str]:
    """Split into (place prefix, brand core, industry+suffix tail)."""
    rest = name
    prefix = ""
    for p in PLACE_PREFIXES:
        if rest.startswith(p):
            prefix, rest = p, rest[len(p):]
            break

    suffix = ""
    for s in COMPANY_SUFFIXES:
        if rest.endswith(s):
            suffix, rest = s, rest[: -len(s)]
            break

    # Compound name: "中国某某科技股份有限公司上海分公司" still holds a company
    # suffix after the outer one was stripped.  Everything from that inner
    # suffix rightwards is structure, not brand -- scrambling it produces the
    # gibberish that the ⚠ warning was flagging.
    inner_end, inner_suffix = -1, ""
    for s in COMPANY_SUFFIXES:
        idx = rest.rfind(s)
        if idx >= 0 and idx + len(s) > inner_end:
            inner_end, inner_suffix = idx + len(s), s
    if inner_end > 0:
        suffix = rest[inner_end - len(inner_suffix):] + suffix
        rest = rest[: inner_end - len(inner_suffix)]

    # A place can sit after the industry word with no brackets:
    # "某某造价咨询深圳有限公司上海分公司".  Peel it before the industry word,
    # or the brand ends up as "某某造价咨询深圳" and gets scrambled whole.
    # Guarded on length so a brand merely ending in a place character survives.
    for p in PLACE_PREFIXES:
        if rest.endswith(p) and len(rest) - len(p) >= 2:
            suffix, rest = p + suffix, rest[: -len(p)]
            break

    industry = ""
    for t in INDUSTRY_TAILS:
        if rest.endswith(t) and rest != t:
            industry, rest = t, rest[: -len(t)]
            break

    # "某某（浙江）物流有限公司" strips down to a brand of "（浙江）" -- nothing
    # but a preserved bracket, which fake_brand cannot change.  That is the
    # prefix having eaten the actual brand, so give it back.
    if not rest or BRACKET_PLACE_RE.fullmatch(rest):
        rest, prefix = prefix + rest, ""

    return prefix, rest, industry + suffix


# "百盟（上海）供应链有限公司" carries a place name INSIDE the brand.  Blindly
# scrambling those characters turns the bracket pair into gibberish, so the
# bracketed place is carved out and kept.  Both bracket widths occur, sometimes
# for the same company on different rows.
BRACKET_PLACE_RE = re.compile(
    r"[（(](?:" + "|".join(PLACE_PREFIXES) + r")[）)]"
)

# A brand longer than this is not a brand: it is a compound name the splitter
# failed to parse.  Flagged for an explicit --rename rather than scrambled.
LONG_BRAND_WARN = 8


def effective_brand_len(brand: str) -> int:
    """Brand length ignoring a bracketed place name, which is preserved."""
    m = BRACKET_PLACE_RE.search(brand)
    return len(brand) - (m.end() - m.start()) if m else len(brand)


def _scramble(seed: str, length: int, salt: int) -> str:
    if length <= 0:
        return ""
    digest = hashlib.md5(f"{seed}|{salt}".encode()).digest()
    while len(digest) < length:
        digest += hashlib.md5(digest).digest()
    return "".join(BRAND_CHARS[b % len(BRAND_CHARS)] for b in digest[:length])


def fake_brand(real_brand: str, salt: int = 0) -> str:
    """Deterministic same-length replacement brand.

    Same length matters more than it looks: PPTX packing and line-wrap are
    measured per character, so a demo file whose tenant names are a different
    width stops being a valid layout test.
    """
    m = BRACKET_PLACE_RE.search(real_brand)
    if m:
        head, kept, tail = real_brand[: m.start()], m.group(0), real_brand[m.end():]
        return (_scramble(normalize_brackets(head), len(head), salt) + kept
                + _scramble(normalize_brackets(tail), len(tail), salt + 1))
    return _scramble(normalize_brackets(real_brand), len(real_brand), salt)


def brand_warning(name: str) -> str:
    """Why this name needs a human decision, or '' if it parses cleanly."""
    _, brand, _ = split_company(name)
    if not brand:
        return "商號為空，不會被替換"
    if any(suf in brand for suf in COMPANY_SUFFIXES):
        return "商號內還有第二個公司後綴 -> 複合名稱，請用 --rename 指定"
    if effective_brand_len(brand) > LONG_BRAND_WARN:
        return f"商號長達 {effective_brand_len(brand)} 字 -> 可能夾雜其他文字，請用 --rename 指定"
    return ""


def build_name_map(real_names: set[str]) -> dict[str, str]:
    """Map each real company name to a stable fake of identical length."""
    mapping: dict[str, str] = {}
    taken: set[str] = set()
    for name in sorted(real_names):
        prefix, brand, tail = split_company(name)
        if not brand:
            continue
        # Bounded: a brand that no salt can change (all preserved characters)
        # must not spin forever -- fall back to scrambling it whole.
        candidate = ""
        for salt in range(64):
            candidate = prefix + fake_brand(brand, salt) + tail
            if candidate not in taken and candidate != name:
                break
        else:
            candidate = prefix + _scramble(brand, len(brand), 999) + tail
        mapping[name] = candidate
        taken.add(candidate)
    return mapping


def build_abbrev_map(name_map: dict[str, str]) -> dict[str, str]:
    """Short forms used in remark text ("某某" for 上海某某铭电子有限公司).

    Every hit is reported so the substitutions can be eyeballed rather than
    trusted; 2-char forms in particular can collide with ordinary Chinese.
    """
    abbrev: dict[str, str] = {}
    for real, fake in name_map.items():
        _, real_brand, _ = split_company(real)
        _, fake_core, _ = split_company(fake)
        if len(real_brand) >= 2 and real_brand not in name_map:
            abbrev[real_brand] = fake_core
        if len(real_brand) >= 3:
            abbrev[real_brand[:2]] = fake_core[:2]
    return abbrev


# ------------------------------------------------------------- text scaling

# Only numbers carrying an explicit money unit get scaled.  "25.1" in
# "应收25.1一个月的含税租金" is a year-month and "26年2月3日" is a date --
# scaling either would corrupt the narrative rather than anonymise it.
MONEY_UNITS = ["万美元", "亿美元", "美元", "万元", "亿元", "百万", "千元",
               "w元", "W元", "万", "亿", "元", "w", "W"]
MONEY_RE = re.compile(
    r"(?<![\d.年月第])(\d{1,3}(?:,\d{3})+|\d+)(\.\d+)?\s*("
    + "|".join(MONEY_UNITS) + r")"
)


def scale_money_in_text(text: str, factor: float, hits: list[tuple[str, str]]) -> str:
    def repl(m: re.Match) -> str:
        int_part, dec_part, unit = m.group(1), m.group(2) or "", m.group(3)
        try:
            value = float(int_part.replace(",", "") + dec_part)
        except ValueError:
            return m.group(0)
        decimals = len(dec_part) - 1 if dec_part else 0
        grouped = "," if "," in int_part else ""
        new_text = f"{value * factor:{grouped}.{decimals}f}{unit}"
        hits.append((m.group(0), new_text))
        return new_text

    return MONEY_RE.sub(repl, text)


# ------------------------------------------------------------------- scanning


def load_pair(src: Path):
    """Load twice: once for structure/formulas, once for cached results.

    Both are needed because a formula cell's *text* lives in one and its
    *value* in the other, and the demo file needs the value.
    """
    wb_f = openpyxl.load_workbook(src, data_only=False)
    wb_v = openpyxl.load_workbook(src, data_only=True)
    return wb_f, wb_v


def scan_workbook(wb_f, wb_v) -> tuple[set[str], dict, dict]:
    """Collect clean company-name cells and per-column numerics.

    Numerics are read from the cached-value workbook so that formula results
    count too -- in a formula-heavy workbook that is most of the figures.
    """
    names: set[str] = set()
    embedded: set[str] = set()
    numeric_cols: dict = defaultdict(lambda: defaultdict(list))
    stats = {"formula": 0, "stale": 0, "numeric": 0, "text": 0}

    for ws in wb_f.worksheets:
        ws_v = wb_v[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    stats["formula"] += 1
                    cached = ws_v[cell.coordinate].value
                    if cached is None:
                        stats["stale"] += 1
                    elif isinstance(cached, (int, float)) and not isinstance(cached, bool):
                        numeric_cols[ws.title][cell.column_letter].append(float(cached))
                    continue
                if isinstance(v, str):
                    stats["text"] += 1
                    s = strip_marks(v.strip())
                    if s not in GENERIC_NAMES and COMPANY_CELL_RE.match(s):
                        names.add(s)
                    else:
                        embedded.update(extract_embedded_names(v.strip()))
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    stats["numeric"] += 1
                    numeric_cols[ws.title][cell.column_letter].append(float(v))
    embedded -= names
    return names | embedded, embedded, numeric_cols, stats


def detect_ratio_columns(numeric_cols) -> dict[str, set[str]]:
    """Columns that are ratios, not amounts.

    Under a global scale a ratio must NOT move (numerator and denominator both
    scale, so the ratio is unchanged).  Heuristic: every non-zero value in the
    column sits within +/-1 and there are at least three of them -- an amount
    column always carries something bigger.

    A false positive here is the dangerous direction: skipping a real amount
    column would break the tie-out.  That is what verify_sums() is for.
    """
    ratio: dict[str, set[str]] = defaultdict(set)
    for sheet, cols in numeric_cols.items():
        for col, values in cols.items():
            nz = [v for v in values if v != 0]
            if len(nz) >= 3 and all(abs(v) <= 1.0 for v in nz):
                ratio[sheet].add(col)
    return ratio


# ------------------------------------------------------------------ transform


def apply_replacements(text: str, ordered_pairs) -> tuple[str, list[str]]:
    used: list[str] = []
    for real, fake in ordered_pairs:
        if real in text:
            text = text.replace(real, fake)
            used.append(real)
    return text, used


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("source", help="Path to the real databook .xlsx")
    ap.add_argument("--output", help="Output path (default: '[DEMO]<stem>.xlsx' beside the source)")
    ap.add_argument("--scale", type=float, default=DEFAULT_SCALE,
                    help=f"Global factor (default {DEFAULT_SCALE})")
    ap.add_argument("--rename", action="append", default=[],
                    help="Extra literal replacement 'old=new'; repeatable. Use for project/entity names.")
    ap.add_argument("--no-abbrev", action="store_true",
                    help="Do not substitute brand short forms in remark text")
    ap.add_argument("--no-ratio-detect", action="store_true",
                    help="Scale every numeric cell, including detected ratio columns")
    ap.add_argument("--inspect", action="store_true",
                    help="Diagnose the source only. Prints a REDACTED report (safe to share) "
                         "and writes the full one to --inspect-out.")
    ap.add_argument("--inspect-out", default="demo_inspect_FULL.txt",
                    help="Where the unredacted inspect report goes")
    ap.add_argument("--dry-run", action="store_true", help="Report only; write nothing")
    ap.add_argument("--emit-mapping", help="Write the real->fake table here (SENSITIVE: never commit)")
    args = ap.parse_args()

    src = Path(args.source)
    if not src.exists():
        print(f"ERROR: source not found: {src}")
        return 1
    out = Path(args.output) if args.output else src.parent / f"[DEMO]{src.stem}.xlsx"

    with zipfile.ZipFile(src) as z:
        embedded = [n for n in z.namelist() if n.startswith(("xl/media/", "xl/charts/"))]
    if embedded:
        print(f"WARNING: source has {len(embedded)} embedded image/chart part(s); "
              f"openpyxl drops them from the demo file.")

    print("loading (twice: formulas + cached values)...")
    wb_f, wb_v = load_pair(src)
    real_names, embedded_names, numeric_cols, stats = scan_workbook(wb_f, wb_v)
    ratio_cols = {} if args.no_ratio_detect else detect_ratio_columns(numeric_cols)
    name_map = build_name_map(real_names)

    if args.inspect:
        return inspect_only(src, wb_f, wb_v, real_names, embedded_names, ratio_cols,
                            name_map, stats, Path(args.inspect_out))

    abbrev_map = {} if args.no_abbrev else build_abbrev_map(name_map)

    extra_map: dict[str, str] = {}
    for spec in args.rename:
        if "=" not in spec:
            print(f"ERROR: --rename needs 'old=new', got {spec!r}")
            return 1
        old, new = spec.split("=", 1)
        extra_map[old] = new

    # Longest first, so "上海某某贸易有限公司" is consumed before the "某某"
    # abbreviation rule can chew a hole in the middle of it.
    ordered_pairs = sorted({**name_map, **abbrev_map, **extra_map}.items(),
                           key=lambda kv: len(kv[0]), reverse=True)

    print(f"\n=== SOURCE: {src.name} ===")
    print(f"  sheets {len(wb_f.sheetnames)} | formulas {stats['formula']} "
          f"| plain numbers {stats['numeric']} | scale {args.scale}")
    if stats["formula"]:
        print(f"  NOTE: {stats['formula']} formula cell(s) will be FLATTENED to "
              f"(cached value x {args.scale}).")
        if stats["stale"]:
            print(f"  WARNING: {stats['stale']} formula cell(s) have no cached value "
                  f"(never calculated) -> they become blank.")

    print(f"\n--- company names found ({len(name_map)}) ---")
    warned = 0
    for real, fake in sorted(name_map.items()):
        warn = brand_warning(real)
        print(f"  {real}  ->  {fake}" + (f"   ⚠ {warn}" if warn else ""))
        warned += bool(warn)
    if warned:
        print(f"  ⚠ {warned} name(s) need an explicit --rename, or their replacement "
              f"will read as gibberish.")
    if extra_map:
        print(f"\n--- extra --rename rules ({len(extra_map)}) ---")
        for old, new in extra_map.items():
            print(f"  {old}  ->  {new}")
    if abbrev_map:
        print(f"\n--- brand short forms substituted in remark text ({len(abbrev_map)}) ---")
        for old, new in sorted(abbrev_map.items()):
            print(f"  {old}  ->  {new}")
    if ratio_cols:
        n = sum(len(c) for c in ratio_cols.values())
        print(f"\n--- {n} column(s) treated as RATIOS and left unscaled ---")
        for sheet, cols in sorted(ratio_cols.items()):
            print(f"  {sheet}: {', '.join(sorted(cols))}")

    if args.dry_run:
        print("\n[dry run] nothing written.")
        return 0

    scaled = flattened = blanked = skipped_ratio = text_changed = 0
    money_hits: list[tuple[str, str]] = []
    names_hit: set[str] = set()

    def transform_text(s: str) -> str:
        nonlocal names_hit
        new, used = apply_replacements(s, ordered_pairs)
        names_hit.update(used)
        return scale_money_in_text(new, args.scale, money_hits)

    for ws in wb_f.worksheets:
        ws_v = wb_v[ws.title]
        sheet_ratio = ratio_cols.get(ws.title, set())
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or isinstance(v, bool):
                    continue
                is_ratio = cell.column_letter in sheet_ratio

                if isinstance(v, str) and v.startswith("="):
                    cached = ws_v[cell.coordinate].value
                    if cached is None:
                        cell.value = None
                        blanked += 1
                    elif isinstance(cached, (int, float)) and not isinstance(cached, bool):
                        cell.value = cached if is_ratio else cached * args.scale
                        flattened += 1
                        skipped_ratio += is_ratio
                    elif isinstance(cached, str):
                        cell.value = transform_text(cached)
                        flattened += 1
                    else:
                        cell.value = cached
                        flattened += 1
                elif isinstance(v, (int, float)):
                    if is_ratio:
                        skipped_ratio += 1
                        continue
                    cell.value = v * args.scale
                    scaled += 1
                elif isinstance(v, str):
                    new = transform_text(v)
                    if new != v:
                        cell.value = new
                        text_changed += 1

    renamed_sheets = []
    for ws in wb_f.worksheets:
        new_title, _ = apply_replacements(ws.title, ordered_pairs)
        new_title = new_title[:31]
        if new_title != ws.title:
            renamed_sheets.append((ws.title, new_title))
            ws.title = new_title

    print("\n--- transform ---")
    print(f"  plain numbers scaled     : {scaled}")
    print(f"  formulas flattened       : {flattened}")
    print(f"  formulas blanked (stale) : {blanked}")
    print(f"  cells left unscaled(ratio): {skipped_ratio}")
    print(f"  text cells changed       : {text_changed}")
    print(f"  sheets renamed           : {len(renamed_sheets)}")
    for old, new in renamed_sheets:
        print(f"      {old}  ->  {new}")
    if money_hits:
        print(f"  amounts rewritten inside remark text ({len(money_hits)}):")
        for old, new in money_hits[:25]:
            print(f"      {old}  ->  {new}")
        if len(money_hits) > 25:
            print(f"      ... and {len(money_hits) - 25} more")

    unhit = sorted(set(name_map) - names_hit)
    if unhit:
        print(f"\n  NOTE: {len(unhit)} mapped name(s) were never substituted "
              f"(a longer rule consumed the cell first):")
        for n in unhit[:15]:
            print(f"      {n}")

    if out.exists():
        backup = out.with_suffix(out.suffix + ".bak")
        shutil.copy2(out, backup)
        print(f"\n  existing output backed up to {backup.name}")
    wb_f.save(out)
    print(f"\n  written: {out}")

    if args.emit_mapping:
        lines = ["# SENSITIVE - real -> fake mapping. Do NOT commit.", ""]
        lines += [f"{r}\t{f}" for r, f in sorted({**name_map, **extra_map}.items())]
        Path(args.emit_mapping).write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"  mapping written: {args.emit_mapping}  (SENSITIVE - do not commit)")

    return verify(src, out, args.scale, name_map, extra_map, ratio_cols)


# ------------------------------------------------------------------- inspect

CJK_RE = re.compile(r"[一-鿿]")
PURE_CN_RE = re.compile(r"^[一-鿿]{2,8}$")
SUBTOTAL_WORDS = ("subtotal", "total", "check", "小计", "小計", "合计", "合計", "总计", "總計")
# Deliberately loose: the point is to discover units the money regex does NOT
# cover, not to confirm the ones it does.
NUM_UNIT_PROBE = re.compile(r"\d(?:[\d,]*\.?\d*)\s*([一-鿿A-Za-z%]{1,3})")

# Redaction alphabet -- real characters, not circles, so the report reads like
# a databook instead of a wall of placeholders.  Nothing here maps back: the
# substitution is many-to-one by construction.
REDACT_CN = ("德明华兴盛达丰隆通和利源正信顺安佳元茂康联创立辉阳卓越弘毅"
             "锦程远航睿臻拓新润泽晟昊博轩宁禾岳川岭峰晖朗骏祺瑞宏恒嘉启泰鑫")


def redact(s: str) -> str:
    """Structure-preserving redaction with real characters.

    Punctuation is deliberately preserved: a leading '*' or a bracketed place
    name is exactly the structural detail that tells us why a name was missed.
    """
    out = []
    for i, ch in enumerate(s):
        if CJK_RE.match(ch):
            out.append(REDACT_CN[(ord(ch) * 7 + i * 13) % len(REDACT_CN)])
        elif ch.isdigit():
            out.append(str((ord(ch) * 3 + i * 7) % 10))
        elif "a" <= ch <= "z":
            out.append(chr(ord("a") + (ord(ch) + i) % 26))
        elif "A" <= ch <= "Z":
            out.append(chr(ord("A") + (ord(ch) + i) % 26))
        else:
            out.append(ch)
    return "".join(out)


def redact_company(name: str, name_map: dict) -> str:
    """Show the parse, using the actual generated fake as the visible brand."""
    prefix, brand, tail = split_company(name)
    fake = name_map.get(name, "")
    _, fake_core, _ = split_company(fake) if fake else ("", "?" * len(brand), "")
    warn = brand_warning(name)
    return (f"[地名:{prefix or '—'}]"
            f"[商號:{fake_core or '!!空!!'}({effective_brand_len(brand)}字)]"
            f"[尾:{tail or '!!無!!'}]"
            + (f"   ⚠ {warn}" if warn else ""))


def why_not_company(s: str) -> str:
    """Precise reason a suffix-bearing cell was not extracted as a name."""
    if len(s) > 30:
        return f"長度{len(s)}>30(多半是備註句，沒問題)"
    if not any(s.endswith(suf) for suf in COMPANY_SUFFIXES):
        return "公司後綴不在結尾"
    bad = sorted({ch for ch in s if not NAME_CHAR_RE.match(ch)})
    if bad:
        return "含字元 " + " ".join(bad)
    return "未知(請回報)"


def inspect_only(src: Path, wb_f, wb_v, real_names, embedded_names, ratio_cols,
                 name_map, stats, full_path: Path) -> int:
    """Report what the transform WOULD see, twice: redacted here, full to disk."""
    safe: list[str] = []
    full: list[str] = []

    def emit(line: str = "", safe_line: str | None = None):
        full.append(line)
        safe.append(line if safe_line is None else safe_line)

    emit(f"=== INSPECT: {src.name} ===")
    emit(f"sheets: {len(wb_f.sheetnames)}")
    emit("sheet names:", "sheet names (redacted, real chars - NOT the real names):")
    for name in wb_f.sheetnames:
        emit(f"   {name}", f"   {redact(name)}")

    with zipfile.ZipFile(src) as z:
        media = [n for n in z.namelist() if n.startswith(("xl/media/", "xl/charts/"))]
    emit(f"\nembedded media/charts: {len(media)}"
         + ("  (WILL BE LOST by openpyxl)" if media else ""))

    merged = sum(len(ws.merged_cells.ranges) for ws in wb_f.worksheets)
    emit(f"merged cell ranges: {merged}  (preserved)")
    emit(f"formula cells: {stats['formula']}"
         + (f"   -> will be FLATTENED to cached value x factor" if stats["formula"] else ""))
    if stats["stale"]:
        emit(f"formula cells with NO cached value: {stats['stale']}"
             f"   <-- these become blank; open+save in Excel first if this is large")

    money_units: dict[str, int] = defaultdict(int)
    money_samples: list[str] = []
    near_miss: list[tuple[str, str, str]] = []
    short_cn: dict[str, int] = defaultdict(int)
    subtotal_rows: dict[str, list[str]] = defaultdict(list)
    precise = rounded = 0

    for ws in wb_f.worksheets:
        ws_v = wb_v[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    cv = ws_v[cell.coordinate].value
                    if isinstance(cv, float):
                        precise += cv != round(cv, 2)
                        rounded += cv == round(cv, 2)
                    continue
                if isinstance(v, str):
                    s = v.strip()
                    if any(w in s.lower() for w in SUBTOTAL_WORDS) and len(s) <= 20:
                        subtotal_rows[ws.title].append((cell.row, s))
                    if PURE_CN_RE.match(s):
                        short_cn[s] += 1
                    stripped = strip_marks(s)
                    if (stripped not in real_names and stripped not in GENERIC_NAMES
                            and any(suf in s for suf in COMPANY_SUFFIXES)):
                        near_miss.append((f"{ws.title}!{cell.coordinate}", s,
                                          why_not_company(stripped)))
                    for m in NUM_UNIT_PROBE.finditer(s):
                        money_units[m.group(1)] += 1
                    for m in MONEY_RE.finditer(s):
                        money_samples.append(m.group(0))
                elif isinstance(v, float) and not isinstance(v, bool):
                    precise += v != round(v, 2)
                    rounded += v == round(v, 2)

    total_float = precise + rounded or 1
    emit(f"\nfloat precision: {precise} full-precision / {total_float} "
         f"({100 * precise / total_float:.0f}% carry >2dp)")
    emit("  (informational only -- linearity keeps the tie-out either way)")

    clean = sorted(set(real_names) - set(embedded_names))
    emit(f"\n--- company names from a whole-cell match: {len(clean)} ---")
    emit("    left = real, right = the fake that will replace it",
         "    the 商號 shown IS the generated fake, at the real one's length")
    for n in clean:
        emit(f"   {n}   ->   {name_map.get(n, '(不變)')}",
             f"   {redact_company(n, name_map)}")

    emit(f"\n--- names EXTRACTED FROM INSIDE a longer cell: {len(embedded_names)} ---")
    emit("    CHECK THESE: a wrong cut here renames part of an account label.")
    for n in sorted(embedded_names):
        emit(f"   {n}   ->   {name_map.get(n, '(不變)')}",
             f"   {redact_company(n, name_map)}")

    by_reason: dict[str, list] = defaultdict(list)
    for loc, s, reason in near_miss:
        by_reason[reason].append((loc, s))
    emit(f"\n--- NEAR MISSES: {len(near_miss)} cell(s) carry a company suffix "
         f"but were NOT extracted ---")
    emit("    '長度>30' entries are remark sentences: FINE, full-text substitution covers them.")
    emit("    Anything else is a name format the extractor is failing to catch.")
    for reason, items in sorted(by_reason.items(), key=lambda kv: -len(kv[1])):
        emit(f"  [{reason}]  x{len(items)}")
        for loc, s in items[:8]:
            emit(f"      {loc}: {s[:60]}", f"      {loc}: len={len(s)} {redact(s[:45])}")
        if len(items) > 8:
            emit(f"      ... and {len(items) - 8} more")

    emit(f"\n--- number+unit tokens (does the money regex cover them?) ---")
    for unit, count in sorted(money_units.items(), key=lambda kv: -kv[1])[:30]:
        emit(f"   '{unit}' x{count}   [{'SCALED' if unit in MONEY_UNITS else 'not scaled'}]")
    emit(f"   money regex matches {len(money_samples)} token(s); samples:")
    emit(f"      {', '.join(money_samples[:15])}",
         f"      {', '.join(redact(m) for m in money_samples[:15])}")

    n_ratio = sum(len(c) for c in ratio_cols.values())
    emit(f"\n--- {n_ratio} ratio column(s) detected (left unscaled) ---")
    for sheet, cols in sorted(ratio_cols.items()):
        emit(f"   {sheet}: {', '.join(sorted(cols))}",
             f"   {redact(sheet)}: {', '.join(sorted(cols))}")

    emit(f"\n--- subtotal/total/check rows per sheet (the subtable structure) ---")
    for sheet, rows in sorted(subtotal_rows.items()):
        emit(f"   {sheet}: " + "; ".join(f"r{r}:{s}" for r, s in rows[:12]),
             f"   {redact(sheet)}: " + "; ".join(f"r{r}:{redact(s)}" for r, s in rows[:12]))

    emit(f"\n--- short pure-Chinese strings (2-8 chars), {len(short_cn)} distinct ---")
    emit("    Tenant SHORT FORMS and PERSON NAMES hide here. Check the full file.")
    for s, c in sorted(short_cn.items(), key=lambda kv: -kv[1])[:80]:
        full.append(f"   x{c:<4} {s}")

    full_path.write_text("\n".join(full) + "\n", encoding="utf-8")
    print("\n".join(safe))
    print(f"\n\n[full, UNREDACTED report written to: {full_path}]")
    print("[terminal output above is redacted and safe to share; the file is NOT]")
    return 0


# -------------------------------------------------------------- verification

SUM_RE = re.compile(r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$", re.I)


def verify_sums(wb_src_f, wb_out) -> tuple[int, int, list[str]]:
    """Re-add every simple =SUM(range) from the source against the demo file.

    This is the check that does NOT assume linearity, so it is the one that
    catches a ratio column wrongly left unscaled inside a summed range.
    """
    checked = failed = 0
    detail: list[str] = []
    out_by_name = {ws.title: ws for ws in wb_out.worksheets}
    for i, ws in enumerate(wb_src_f.worksheets):
        ws_o = out_by_name.get(ws.title) or wb_out.worksheets[i]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                m = SUM_RE.match(v.replace(" ", ""))
                if not m or m.group(1).upper() != m.group(3).upper():
                    continue
                col, r1, r2 = m.group(1).upper(), int(m.group(2)), int(m.group(4))
                if r2 < r1 or r2 - r1 > 500:
                    continue
                total = ws_o[cell.coordinate].value
                if not isinstance(total, (int, float)) or isinstance(total, bool):
                    continue
                parts = []
                for r in range(r1, r2 + 1):
                    pv = ws_o[f"{col}{r}"].value
                    if isinstance(pv, (int, float)) and not isinstance(pv, bool):
                        parts.append(pv)
                checked += 1
                got = sum(parts)
                if abs(got - total) > max(abs(total) * 1e-6, 1e-6):
                    failed += 1
                    if len(detail) < 8:
                        detail.append(f"{ws.title}!{cell.coordinate}: cell={total!r} "
                                      f"but SUM({col}{r1}:{col}{r2})={got!r}")
    return checked, failed, detail


def verify(src: Path, out: Path, factor: float, name_map, extra_map, ratio_cols) -> int:
    """Two independent checks: linearity, then a real re-addition of subtotals."""
    print("\n=== VERIFY ===")
    wb_src_f = openpyxl.load_workbook(src, data_only=False)
    wb_src_v = openpyxl.load_workbook(src, data_only=True)
    wb_out = openpyxl.load_workbook(out, data_only=False)

    if len(wb_src_f.sheetnames) != len(wb_out.sheetnames):
        print(f"  FAIL: sheet count {len(wb_src_f.sheetnames)} -> {len(wb_out.sheetnames)}")
        return 1

    checked = mismatches = 0
    worst = 0.0
    for i, s_name in enumerate(wb_src_v.sheetnames):
        ws_s, ws_o = wb_src_v[s_name], wb_out.worksheets[i]
        sheet_ratio = ratio_cols.get(s_name, set())
        for row in ws_s.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                got = ws_o[cell.coordinate].value
                expect = v if cell.column_letter in sheet_ratio else v * factor
                checked += 1
                if not isinstance(got, (int, float)) or isinstance(got, bool):
                    mismatches += 1
                    continue
                rel = abs(got - expect) / max(abs(expect), 1e-9)
                worst = max(worst, rel)
                if rel > 1e-9:
                    mismatches += 1
                    if mismatches <= 5:
                        print(f"  MISMATCH {ws_o.title}!{cell.coordinate}: "
                              f"got {got!r}, expected {expect!r}")

    print(f"  [1] linearity: {checked} value(s) checked, {mismatches} mismatch(es), "
          f"worst rel. error {worst:.2e}")

    s_checked, s_failed, s_detail = verify_sums(wb_src_f, wb_out)
    print(f"  [2] subtotal re-addition: {s_checked} =SUM() range(s) re-added, "
          f"{s_failed} do NOT tie")
    for d in s_detail:
        print(f"      {d}")
    if s_failed:
        print("      -> most likely a ratio column was wrongly left unscaled inside a")
        print("         summed range. Re-run with --no-ratio-detect.")

    if mismatches == 0 and s_failed == 0:
        print("  PASS -- every figure is source x factor AND every re-added subtotal ties.")

    leaked = defaultdict(list)
    real_tokens = list(name_map) + list(extra_map)
    for ws in wb_out.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for tok in real_tokens:
                    if tok in cell.value:
                        leaked[tok].append(f"{ws.title}!{cell.coordinate}")
    if leaked:
        print(f"\n  FAIL: {len(leaked)} real name(s) still present in the output:")
        for tok, locs in list(leaked.items())[:10]:
            print(f"      {tok}: {', '.join(locs[:5])}{' ...' if len(locs) > 5 else ''}")
    else:
        print(f"  name leak: none of the {len(real_tokens)} mapped names survive.")

    # Character-level residue: catches typo variants of a real brand
    # (a one-character misspelling of a tenant) that literal substitution misses.
    # A real workbook was found to carry exactly this.
    brand_chars: set[str] = set()
    for real in name_map:
        _, brand, _ = split_company(real)
        brand_chars.update(c for c in brand if CJK_RE.match(c))
    residue = defaultdict(list)
    for ws in wb_out.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or len(val) < 2:
                    continue
                for ch in brand_chars.intersection(val):
                    residue[ch].append(f"{ws.title}!{cell.coordinate}")
    if residue:
        print(f"\n  REVIEW BY EYE: {len(residue)} character(s) from real brands still occur.")
        print("  Most are ordinary Chinese; look for a MISSPELT company name.")
        for ch, locs in sorted(residue.items(), key=lambda kv: -len(kv[1]))[:15]:
            print(f"      '{ch}' x{len(locs)}: {', '.join(locs[:3])}"
                  f"{' ...' if len(locs) > 3 else ''}")

    return 1 if (mismatches or leaked or s_failed) else 0


if __name__ == "__main__":
    raise SystemExit(main())
