#!/usr/bin/env python3
"""Structural consistency check across every 'AB-' prefixed raw-data tab.

Before automating a bridge/waterfall chart FROM the raw AB-* tabs (instead
of reading a pre-built '<entity>-量价桥图' helper tab, which per the
project team likely only exists for one entity so far), we need evidence
of whether every AB-* tab actually follows the same structural convention
confirmed on AB-CD.

Confirmed on TWO real tabs (AB-CD and AB-KS1) so far: both have a clean
English LABEL COLUMN (a low-numbered column like C/D/F/G/I, not fixed to
one letter) with recognizable terms -- 'Year', 'Days in period', 'Occupied
area', 'Revenue', 'Occupancy rate', etc. This is the primary anchor now
(find_labeled_rows / _LABEL_VOCAB), NOT position or repeated-value
structure -- an earlier version of this script tried to infer 'Year'/'Days'
rows from formula patterns and got it wrong on real data (matched a
"=DATE(YEAR(x)+1,...)" formula that merely CONTAINS "YEAR(" before ever
reaching the real year-extraction row). The two tabs also confirmed they
DON'T use the same mechanism for multiple asset types: AB-CD has separate
repeated-text tag rows (find_tag_rows), AB-KS1 folds the type straight into
the metric row's own label (e.g. "Phase 1 - Dry") -- so this script reports
both signals rather than assuming either one always applies.

Usage:
    python inspect_ab_tabs_structure.py "databooks/xx.xlsx"
    python inspect_ab_tabs_structure.py "databooks/xx.xlsx" --tab AB-KS1 --verbose
"""
import argparse
import re
import sys
from collections import Counter
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Excludes period/date-range labels like "LTM Jun 26" / "2024年" -- these
# repeat across many columns exactly like a real type tag does (confirmed:
# 13/17 real tabs flagged row 1 as a "type tag" purely because of this),
# but a genuine asset-type name doesn't contain digits or a month/year word.
_PERIOD_LABEL_RE = re.compile(
    r"\d|年|月|日|LTM|Q[1-4]\b|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec", re.IGNORECASE
)


def _is_short_text(v) -> bool:
    if not (isinstance(v, str) and 1 <= len(v.strip()) <= 8):
        return False
    text = v.strip()
    if text.replace(".", "").isdigit():
        return False
    if _PERIOD_LABEL_RE.search(text):
        return False
    return True


def find_tag_rows(ws, min_repeat: int = 4, max_scan_row: int = 20) -> List[Tuple[int, Dict[str, int]]]:
    """A tag row: mostly-empty except for a small number of DISTINCT short
    text values, each repeated >= min_repeat times across the row (i.e. one
    value tags a contiguous run of month-columns). Returns (row, {label:
    count}) for every row in the first max_scan_row rows that qualifies."""
    hits = []
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan_row, ws.max_row)):
        counts = Counter()
        for cell in row:
            if _is_short_text(cell.value):
                counts[cell.value.strip()] += 1
        qualifying = {label: n for label, n in counts.items() if n >= min_repeat}
        if qualifying:
            hits.append((row[0].row, qualifying))
    return hits


# Real-world confirmed (AB-CD AND AB-KS1 both use the SAME label vocabulary
# in a low-numbered label column, e.g. D or I -- even though they use
# DIFFERENT mechanisms for tagging multiple asset types: AB-CD has separate
# repeated-text tag rows, AB-KS1 folds the type straight into the metric
# row's own label, e.g. "Phase 1 - Dry"). This is a much more reliable
# anchor than guessing from formula patterns or repeated-value structure --
# the formula-pattern approach (previous find_year_days_rows) was CONFIRMED
# WRONG on real data: it matched row 4's "=DATE(YEAR(W4)+1,...)" (a
# "Quarter End + 1 year" formula that merely CONTAINS "YEAR(") before ever
# reaching row 7's real "=YEAR(W3)" year-extraction row, because it wasn't
# anchored to the row's own label text at all.
# (label_text, match_mode, category) -- match_mode "exact" compares the
# whole cell text case-insensitively; "substring" allows it to appear
# anywhere in a longer label (needed for e.g. "Dry rental revenue post VAT"
# to still count as a "revenue" row).
_LABEL_VOCAB: List[Tuple[str, str, str]] = [
    ("Year", "exact", "period_year"),
    ("Days in period", "substring", "period_days"),
    ("Quarter End", "substring", "period_quarter_end"),
    ("Month", "exact", "period_month"),
    ("Quarter", "exact", "period_quarter"),
    ("Beg.", "exact", "period_begin"),
    ("Beginning", "substring", "period_begin"),
    ("End", "exact", "period_end"),
    ("Occupied area", "substring", "metric_area_occupied"),
    ("Gross leasing area", "substring", "metric_area_gross"),
    ("Leased area", "substring", "metric_area_leased"),
    ("Unit rent", "substring", "metric_rent"),
    ("Occupancy rate", "substring", "metric_occupancy"),
    ("Revenue", "substring", "metric_revenue"),
]
_LABEL_SCAN_MAX_COL = 12  # observed label columns (C/D/F/G/I) are all well within this


def find_labeled_rows(ws_values, max_scan_row: int = 60) -> Dict[int, List[Tuple[int, str, str]]]:
    """Returns {row: [(col, matched_vocab_term, category), ...]} for every
    row in the first max_scan_row rows whose label column (a low-index
    column, not a data column) matches something in _LABEL_VOCAB."""
    hits: Dict[int, List[Tuple[int, str, str]]] = {}
    for r in range(1, min(max_scan_row, ws_values.max_row) + 1):
        for c in range(1, min(_LABEL_SCAN_MAX_COL, ws_values.max_column) + 1):
            v = ws_values.cell(row=r, column=c).value
            if not isinstance(v, str) or not v.strip():
                continue
            text = v.strip().lower()
            for term, mode, category in _LABEL_VOCAB:
                matched = (text == term.lower()) if mode == "exact" else (term.lower() in text)
                if matched:
                    hits.setdefault(r, []).append((c, v.strip(), category))
    return hits


def find_year_days_rows(ws_values, ws_formulas, max_scan_row: int = 60) -> Dict[str, Optional[int]]:
    """Label-anchored (not formula-pattern-guessed) -- see find_labeled_rows
    docstring for why the formula-pattern approach this replaced was wrong."""
    labeled = find_labeled_rows(ws_values, max_scan_row)
    year_row = days_row = None
    for r in sorted(labeled):
        cats = {cat for _, _, cat in labeled[r]}
        if year_row is None and "period_year" in cats:
            year_row = r
        if days_row is None and "period_days" in cats:
            days_row = r
    return {"year_row": year_row, "days_row": days_row}


def dump_rows(ws_values, ws_formulas, rows: List[int]):
    """Full raw content (values, formulas where present) for specific rows
    -- for verifying what a flagged row ACTUALLY contains, e.g. confirming
    a candidate tag row is a real type label vs. a period marker."""
    for r in rows:
        print(f"\n--- row {r} ---")
        cells = []
        for c in range(1, ws_values.max_column + 1):
            v = ws_values.cell(row=r, column=c).value
            f = ws_formulas.cell(row=r, column=c).value
            if v is None and f is None:
                continue
            addr = get_column_letter(c)
            if isinstance(f, str) and f.startswith("="):
                cells.append(f"{addr}={f!r} -> {v!r}")
            else:
                cells.append(f"{addr}={v!r}")
        print("  " + " | ".join(cells) if cells else "  (empty)")


def _parse_row_spec(spec: str) -> List[int]:
    rows = []
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            lo, hi = part.split("-")
            rows.extend(range(int(lo), int(hi) + 1))
        else:
            rows.append(int(part))
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the databook .xlsx")
    ap.add_argument("--tab", default=None, help="only inspect this one AB- tab (default: all of them)")
    ap.add_argument("--verbose", "-v", action="store_true", help="print full tag-row/metric-row detail per tab")
    ap.add_argument("--dump-rows", default=None, metavar="SPEC",
                     help="skip the structural scan and just dump full raw content for these rows of "
                          "--tab, e.g. --tab AB-KS1 --dump-rows 1-20,25")
    args = ap.parse_args()

    print(f"Loading {args.path!r}...")
    wb_values = load_workbook(args.path, data_only=True)
    wb_formulas = load_workbook(args.path, data_only=False)

    if args.dump_rows:
        if not args.tab:
            print("❌ --dump-rows requires --tab")
            return 1
        if args.tab not in wb_values.sheetnames:
            print(f"❌ tab {args.tab!r} not found.")
            return 1
        dump_rows(wb_values[args.tab], wb_formulas[args.tab], _parse_row_spec(args.dump_rows))
        return 0

    all_sheets = wb_values.sheetnames
    ab_tabs = [s for s in all_sheets if s.startswith("AB-") or s.startswith("AB")]
    if args.tab:
        ab_tabs = [s for s in ab_tabs if s == args.tab]
        if not ab_tabs:
            print(f"❌ tab {args.tab!r} not found among AB- sheets.")
            return 1

    bridge_like = [s for s in all_sheets if ("桥图" in s or "桥" in s) and s not in ab_tabs]
    print(f"\nPre-built bridge-style tab(s) already in this workbook ({len(bridge_like)}): {bridge_like}")
    print(f"Inspecting {len(ab_tabs)} 'AB-' tab(s) for structural consistency with AB-CD...\n")

    print(f"{'Tab':16s} {'Dims':14s} {'YearRow':8s} {'DaysRow':8s} {'TagRows':10s} {'LabeledMetricRows (row:category)'}")
    print("-" * 130)
    for tab in ab_tabs:
        ws_v = wb_values[tab]
        ws_f = wb_formulas[tab]
        tag_hits = find_tag_rows(ws_v)
        tag_rows = [r for r, _ in tag_hits]
        yd = find_year_days_rows(ws_v, ws_f)
        labeled = find_labeled_rows(ws_v)
        metric_rows = {
            r: sorted({cat for _, _, cat in hits if cat.startswith("metric_")})
            for r, hits in labeled.items()
            if any(cat.startswith("metric_") for _, _, cat in hits)
        }
        metric_summary = ", ".join(f"{r}:{'/'.join(cats)}" for r, cats in sorted(metric_rows.items()))
        print(f"{tab:16s} {ws_v.dimensions:14s} {str(yd['year_row']):8s} {str(yd['days_row']):8s} "
              f"{str(tag_rows):10s} {metric_summary}")
        if args.verbose:
            for r, labels in tag_hits:
                print(f"    tag row {r}: {labels}")
            for r in sorted(labeled):
                for col, text, cat in labeled[r]:
                    print(f"    row {r} col {get_column_letter(col)}: {cat} <- {text!r}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
