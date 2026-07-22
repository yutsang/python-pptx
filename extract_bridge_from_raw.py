#!/usr/bin/env python3
"""Extracts price-volume bridge INPUT data (revenue/area/days per phase per
year) directly from a raw AB-* tab, generically -- no hardcoded row numbers,
using the same label-vocabulary anchors as inspect_ab_tabs_structure.py.

This is the harder "Q2" path: most of the 17 AB-* entities do NOT have a
pre-built '<entity>-量价桥图' helper tab (only 成都/AB-CD does), so a
bridge chart for them has to be computed from this raw layer instead of
just read off an existing tab (see bridge_chart_prototype.py for that
easier path).

Validated structurally against a REAL 17-tab scan: 13/17 tabs have byte-
identical label row numbers, 2/17 (AB-SJ, AB-KM) are offset by exactly 1
row, AB-CD (which DOES have tag rows 9/10/11 unlike the other 16) is
offset differently -- all handled generically via label lookup, not
position.

--validate mode cross-checks this extractor's OUTPUT for AB-CD against the
real, human-built '成都-量价桥图' bridge tab's OWN computed values (the
"answer key") -- confirms the generic extraction logic reproduces what an
analyst already validated by hand, before trusting it on the other 16
entities where there's no existing bridge tab to check against.

Usage:
    python extract_bridge_from_raw.py "databooks/xx.xlsx" --tab AB-KS1
    python extract_bridge_from_raw.py "databooks/xx.xlsx" --tab AB-CD --validate
"""
import argparse
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from inspect_ab_tabs_structure import find_labeled_rows, find_tag_rows


@dataclass
class PhaseBlock:
    label: str
    occupancy_row: Optional[int] = None
    area_row: Optional[int] = None
    rent_row: Optional[int] = None
    revenue_row: Optional[int] = None  # the LAST (post-VAT / excl-VAT) revenue row in the block


def find_phase_blocks(ws_values, max_scan_row: int = 60) -> List[PhaseBlock]:
    """Groups labeled metric rows into phase blocks. A block starts at each
    'metric_occupancy' row (confirmed anchor: every real phase block in the
    17-tab scan starts with a Total occupancy rate row) and runs until the
    next one (or end of scan range).

    Confirmed on real AB-CD data: per-phase rows (occupancy/area/rent/
    revenue) all sit in the SAME label column (D in AB-CD), while the
    aggregate/grand-total rows below the last real phase block ('Total
    Rental Revenue', 'Gross revenue', 'Net revenue', ...) sit in a
    DIFFERENT column (C) despite also being tagged 'metric_revenue' and
    falling inside the last block's row-range fallback window -- matching
    on category alone (the earlier version of this function) silently
    picked up 'Total Rental Revenue' as if it were the last phase's own
    revenue, producing a number that was actually the grand total across
    ALL phases. Restricting every row match to the occupancy row's own
    column fixes this without needing a tighter row-range guess."""
    labeled = find_labeled_rows(ws_values, max_scan_row)
    # {row: {category: column}} -- keeps the column each category was found
    # in, since a row can (rarely) have unrelated labels in other columns.
    by_row: Dict[int, Dict[str, int]] = {}
    for r, hits in labeled.items():
        for col, _text, cat in hits:
            by_row.setdefault(r, {})[cat] = col

    occ_hits = sorted((r, cats["metric_occupancy"]) for r, cats in by_row.items() if "metric_occupancy" in cats)
    if not occ_hits:
        return []

    tag_rows_hits = find_tag_rows(ws_values)
    tag_labels = [labels for _, labels in tag_rows_hits]

    blocks: List[PhaseBlock] = []
    for i, (start, anchor_col) in enumerate(occ_hits):
        end = occ_hits[i + 1][0] - 1 if i + 1 < len(occ_hits) else start + 8
        area_row = rent_row = None
        revenue_rows: List[int] = []
        for r in range(start, end + 1):
            cats = by_row.get(r, {})
            if cats.get("metric_area_occupied") == anchor_col or cats.get("metric_area_leased") == anchor_col:
                area_row = r
            if cats.get("metric_rent") == anchor_col:
                rent_row = r
            if cats.get("metric_revenue") == anchor_col:
                revenue_rows.append(r)
        # AB-CD confirmed: of the 2 (same-column) revenue rows per block
        # (incl-VAT then post-VAT), the bridge tab's own formulas use the
        # LATER (post-VAT) one -- e.g. 干仓 pulled from row 24 ('...post
        # VAT'), not row 22 ('...含税'/incl-VAT), even though both are
        # labeled "revenue".
        revenue_row = revenue_rows[-1] if revenue_rows else None
        label = tag_labels[i] if i < len(tag_labels) else {}
        label_text = "/".join(label.keys()) if label else f"Phase {i + 1}"
        blocks.append(PhaseBlock(label=label_text, occupancy_row=start, area_row=area_row,
                                  rent_row=rent_row, revenue_row=revenue_row))
    return blocks


def _year_col_map(ws_values, year_row: int, max_col: int) -> Dict[int, List[int]]:
    """{year: [columns]} -- from the label-anchored Year row's own values,
    not a guessed column range."""
    mapping: Dict[int, List[int]] = {}
    for c in range(1, max_col + 1):
        v = ws_values.cell(row=year_row, column=c).value
        if isinstance(v, (int, float)) and 2000 <= v <= 2100:
            mapping.setdefault(int(v), []).append(c)
    return mapping


def _phase_start_col(ws_values, block: PhaseBlock, max_col: int) -> Optional[int]:
    """First column (in chronological/left-to-right order) where this
    phase's own area or revenue is ever non-zero -- confirmed against the
    real 干仓/冷库 numbers that AB-CD's tag row (9/10/11) is NOT a per-month
    "was this active" flag; it's a "once this phase commenced, count it in
    every month from here on" flag (days for 干仓 2024 = 214, spanning
    Jun-Dec even though Jul/Aug/Sep read 0 in between -- only reproducible
    by treating the FIRST-ever nonzero column as a permanent start point,
    not by filtering to nonzero months). Entities without an explicit tag
    row (16 of the 17) have no other signal for this, so this is the only
    generic proxy available."""
    for c in range(1, max_col + 1):
        for row in (block.area_row, block.revenue_row):
            if row is None:
                continue
            v = ws_values.cell(row=row, column=c).value
            if isinstance(v, (int, float)) and v != 0:
                return c
    return None


def extract_annual_series(ws_values, block: PhaseBlock, year_row: int, days_row: int) -> Dict[int, Dict[str, float]]:
    """Per-year: revenue (summed), area (averaged over months with area>0),
    days (summed FROM this phase's start column onward -- see
    _phase_start_col), unit_rent (revenue/area/days*1000) -- matches
    AB-CD's own SUMIFS/AVERAGEIFS(...">0") formula conventions confirmed
    earlier."""
    max_col = ws_values.max_column
    years = _year_col_map(ws_values, year_row, max_col)
    start_col = _phase_start_col(ws_values, block, max_col)
    out: Dict[int, Dict[str, float]] = {}
    for year, cols in sorted(years.items()):
        revenue = 0.0
        area_vals = []
        days = 0.0
        for c in cols:
            if block.revenue_row:
                v = ws_values.cell(row=block.revenue_row, column=c).value
                if isinstance(v, (int, float)):
                    revenue += v
            if block.area_row:
                v = ws_values.cell(row=block.area_row, column=c).value
                if isinstance(v, (int, float)) and v > 0:
                    area_vals.append(v)
            if start_col is not None and c >= start_col:
                v = ws_values.cell(row=days_row, column=c).value
                if isinstance(v, (int, float)):
                    days += v
        area = sum(area_vals) / len(area_vals) if area_vals else 0.0
        # AB-CD's own formula is revenue_k/area/days*1000 where revenue_k is
        # ALREADY /1000 -- i.e. raw_revenue/1000/area/days*1000, which
        # simplifies to raw_revenue/area/days. Using raw `revenue` directly
        # (not revenue_k) here, with no separate *1000, is that same
        # simplification -- confirmed against AB-CD's real P6 value
        # (0.1383755...): using revenue_k*1000 like a naive port of the
        # Excel formula produces a result 1000x too large.
        unit_rent = (revenue / area / days) if (area and days) else 0.0
        out[year] = {
            "revenue_k": revenue / 1000,
            "area": area,
            "days": days,
            "unit_rent": unit_rent,
        }
    return out


def find_month_row(ws_values, max_scan_row: int = 60) -> Optional[int]:
    labeled = find_labeled_rows(ws_values, max_scan_row)
    for r in sorted(labeled):
        if any(cat == "period_month" for _, _, cat in labeled[r]):
            return r
    return None


def _month_col_map(ws_values, year_row: int, month_row: int, max_col: int) -> Dict[tuple, int]:
    """{(year, month): column} -- built from the label-anchored Year and
    Month rows' own values, not a guessed column layout."""
    mapping: Dict[tuple, int] = {}
    for c in range(1, max_col + 1):
        y = ws_values.cell(row=year_row, column=c).value
        m = ws_values.cell(row=month_row, column=c).value
        if isinstance(y, (int, float)) and 2000 <= y <= 2100 and isinstance(m, (int, float)) and 1 <= m <= 12:
            mapping[(int(y), int(m))] = c
    return mapping


def extract_ltm_series(ws_values, block: PhaseBlock, year_row: int, month_row: int, days_row: int,
                        end_year: int, end_month: int, window: int = 12) -> Optional[Dict[str, float]]:
    """Same aggregation as extract_annual_series (SUM revenue, AVERAGE
    area>0, SUM days from the phase's own start column onward) but over a
    trailing N-month window ending at (end_year, end_month) instead of a
    calendar year -- e.g. window=12 ending at (2026, 6) covers Jul 2025
    through Jun 2026. This is what the REAL '成都-量价桥图' bridge tab's
    own second transition does for its trailing period ('2025年7月至2026
    年6月收入') instead of comparing a full calendar year against a
    partial one -- confirmed to require no assumption/extrapolation,
    unlike annualizing a partial year. Returns None if the window isn't
    fully covered by actual data (won't fabricate a short window)."""
    max_col = ws_values.max_column
    ym_map = _month_col_map(ws_values, year_row, month_row, max_col)
    if not ym_map:
        return None

    keys = []
    y, m = end_year, end_month
    for _ in range(window):
        keys.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    keys.reverse()
    if any(k not in ym_map for k in keys):
        return None
    cols = [ym_map[k] for k in keys]

    start_col = _phase_start_col(ws_values, block, max_col)
    revenue = 0.0
    area_vals = []
    days = 0.0
    for c in cols:
        if block.revenue_row:
            v = ws_values.cell(row=block.revenue_row, column=c).value
            if isinstance(v, (int, float)):
                revenue += v
        if block.area_row:
            v = ws_values.cell(row=block.area_row, column=c).value
            if isinstance(v, (int, float)) and v > 0:
                area_vals.append(v)
        if start_col is not None and c >= start_col:
            v = ws_values.cell(row=days_row, column=c).value
            if isinstance(v, (int, float)):
                days += v
    area = sum(area_vals) / len(area_vals) if area_vals else 0.0
    unit_rent = (revenue / area / days) if (area and days) else 0.0
    return {"revenue_k": revenue / 1000, "area": area, "days": days, "unit_rent": unit_rent}


def format_ltm_label(end_year: int, end_month: int, window: int = 12) -> str:
    """Matches the real bridge tab's own label convention exactly, e.g.
    (2026, 6) -> '2025年7月至2026年6月收入'."""
    start_month = end_month - window + 1
    start_year = end_year
    while start_month < 1:
        start_month += 12
        start_year -= 1
    return f"{start_year}年{start_month}月至{end_year}年{end_month}月收入"


# Hardcoded from the REAL '成都-量价桥图' bridge tab's own values (already
# confirmed earlier this session against the real file) -- the "answer key"
# for --validate. Only the annual columns (2023/2024/2025), not LTM.
_AB_CD_EXPECTED = {
    "干仓": {
        2023: {"revenue_k": 0.0, "area": 0.0, "days": 0.0, "unit_rent": 0.0},
        2024: {"revenue_k": 1152.26991, "area": 38911.78333333333, "days": 214.0, "unit_rent": 0.1383755348771996},
        2025: {"revenue_k": 7974.30754, "area": 56642.766, "days": 365.0, "unit_rent": 0.38570535057924626},
    },
    "综合楼": {
        2023: {"revenue_k": 0.0, "area": 0.0, "days": 0.0, "unit_rent": 0.0},
        2024: {"revenue_k": 29.75757, "area": 795.6266666666667, "days": 214.0, "unit_rent": 0.1747730075503093},
        2025: {"revenue_k": 70.34938, "area": 339.3733333333333, "days": 365.0, "unit_rent": 0.5679233117822858},
    },
    "冷库": {
        2023: {"revenue_k": 0.0, "area": 0.0, "days": 0.0, "unit_rent": 0.0},
        2024: {"revenue_k": 0.0, "area": 0.0, "days": 0.0, "unit_rent": 0.0},
        2025: {"revenue_k": 3061.83285, "area": 5496.791666666667, "days": 365.0, "unit_rent": 1.526087153994666},
    },
}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the databook .xlsx")
    ap.add_argument("--tab", required=True, help="the AB- tab to extract from")
    ap.add_argument("--validate", action="store_true",
                     help="cross-check output against the real 成都-量价桥图 bridge tab's own "
                          "known values (only meaningful for --tab AB-CD)")
    args = ap.parse_args()

    print(f"Loading {args.path!r}...")
    wb_values = load_workbook(args.path, data_only=True)
    if args.tab not in wb_values.sheetnames:
        print(f"❌ tab {args.tab!r} not found. Available: {wb_values.sheetnames}")
        return 1
    ws = wb_values[args.tab]

    labeled = find_labeled_rows(ws)
    by_row = {r: {cat for _, _, cat in hits} for r, hits in labeled.items()}
    year_row = next((r for r, cats in by_row.items() if "period_year" in cats), None)
    days_row = next((r for r, cats in by_row.items() if "period_days" in cats), None)
    if not year_row or not days_row:
        print(f"❌ Could not find Year/Days rows on {args.tab!r} (year_row={year_row}, days_row={days_row}).")
        return 1
    print(f"Year row: {year_row}, Days row: {days_row}")

    blocks = find_phase_blocks(ws)
    if not blocks:
        print(f"❌ No phase blocks found on {args.tab!r} (no 'Total occupancy rate'-labeled row).")
        return 1
    print(f"Found {len(blocks)} phase block(s).\n")

    all_ok = True
    for block in blocks:
        print(f"--- {block.label} (occupancy row {block.occupancy_row}, area row {block.area_row}, "
              f"revenue row {block.revenue_row}) ---")
        series = extract_annual_series(ws, block, year_row, days_row)
        for year, vals in series.items():
            print(f"  {year}: revenue={vals['revenue_k']:,.2f}k  area={vals['area']:,.2f}  "
                  f"days={vals['days']:.0f}  unit_rent={vals['unit_rent']:.4f}")

        if args.validate:
            expected = _AB_CD_EXPECTED.get(block.label)
            if not expected:
                print(f"  ⚠️  no expected values recorded for block label {block.label!r} -- skipping validation")
                continue
            for year, exp in expected.items():
                got = series.get(year, {})
                for field_name in ("revenue_k", "area", "days", "unit_rent"):
                    exp_v = exp[field_name]
                    got_v = got.get(field_name, None)
                    # unit_rent lives on a 0-2 scale, not the 100s-10000s
                    # scale of the other fields -- the 0.5 absolute floor
                    # below would swallow any real unit_rent mismatch.
                    abs_floor = 0.001 if field_name == "unit_rent" else 0.5
                    tol = max(abs_floor, abs(exp_v) * 0.005)
                    ok = got_v is not None and abs(got_v - exp_v) <= tol
                    all_ok = all_ok and ok
                    status = "✅" if ok else "❌"
                    print(f"    {status} {year} {field_name}: got={got_v} expected={exp_v}")
        print()

    if args.validate:
        print("✅ ALL VALUES MATCH the real bridge tab's own numbers." if all_ok
              else "❌ MISMATCH(ES) found -- do not trust this extractor's output yet.")
        return 0 if all_ok else 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
