#!/usr/bin/env python3
"""Dumps the EXACT month-by-month raw values (year, month, column letter,
area, revenue, days) that extract_bridge_from_raw.py's LTM/annual series
computations actually read, for one phase block on one raw AB-* tab.

Built to trace a specific real anomaly: an LTM-window (trailing 12 months)
area figure that looked suspiciously close to an EARLIER calendar year's
area instead of the expected recent months -- this dumps the underlying
column-by-column data directly so a wrong month->column mapping or a
raw-data issue is visible instead of guessed at from the aggregated
numbers alone.

Usage:
    python inspect_ab_tab_monthly_series.py "databooks/xx.xlsx" --tab AB-NT1
    python inspect_ab_tab_monthly_series.py "databooks/xx.xlsx" --tab AB-NT1 --phase 干仓
"""
import argparse
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from fdd_utils.extract_bridge_from_raw import (
    find_phase_blocks, _year_col_map, _month_col_map, find_month_row, _phase_start_col,
)
from fdd_utils.inspect_ab_tabs_structure import find_labeled_rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the databook .xlsx")
    ap.add_argument("--tab", required=True, help="the AB- tab to inspect")
    ap.add_argument("--phase", default=None, help="only dump this one phase's label (default: all)")
    args = ap.parse_args()

    print(f"Loading {args.path!r}...")
    wb = load_workbook(args.path, data_only=True)
    if args.tab not in wb.sheetnames:
        print(f"❌ tab {args.tab!r} not found. Available: {wb.sheetnames}")
        return 1
    ws = wb[args.tab]

    labeled = find_labeled_rows(ws)
    by_row = {r: {cat for _, _, cat in hits} for r, hits in labeled.items()}
    year_row = next((r for r, cats in by_row.items() if "period_year" in cats), None)
    days_row = next((r for r, cats in by_row.items() if "period_days" in cats), None)
    month_row = find_month_row(ws)
    print(f"year_row={year_row}  month_row={month_row}  days_row={days_row}")
    if not year_row or not days_row or not month_row:
        print("❌ Missing one of year/month/days row -- can't build the (year, month) -> column map.")
        return 1

    max_col = ws.max_column
    year_map = _year_col_map(ws, year_row, max_col)
    print(f"\nYear row values found -> columns: { {y: [get_column_letter(c) for c in cs] for y, cs in year_map.items()} }")

    ym_map = _month_col_map(ws, year_row, month_row, max_col)
    print(f"\n(year, month) -> column mapping ({len(ym_map)} entries):")
    for (y, m), c in sorted(ym_map.items()):
        print(f"  {y}-{m:02d} -> col {get_column_letter(c)} ({c})")

    blocks = find_phase_blocks(ws)
    print(f"\nFound {len(blocks)} phase block(s): {[b.label for b in blocks]}")
    for block in blocks:
        if args.phase and block.label != args.phase:
            continue
        print(f"\n{'=' * 78}")
        print(f"Phase {block.label!r}: occupancy_row={block.occupancy_row} area_row={block.area_row} "
              f"rent_row={block.rent_row} revenue_row={block.revenue_row}")
        start_col = _phase_start_col(ws, block, max_col)
        print(f"_phase_start_col (first ever non-zero area/revenue column) = "
              f"{get_column_letter(start_col) if start_col else None} ({start_col})")
        print(f"{'=' * 78}")
        print(f"{'year-month':10s} {'col':6s} {'area':>14s} {'revenue':>14s} {'days':>8s}")
        for (y, m), c in sorted(ym_map.items()):
            area_v = ws.cell(row=block.area_row, column=c).value if block.area_row else None
            rev_v = ws.cell(row=block.revenue_row, column=c).value if block.revenue_row else None
            days_v = ws.cell(row=days_row, column=c).value
            area_s = f"{area_v:,.2f}" if isinstance(area_v, (int, float)) else str(area_v)
            rev_s = f"{rev_v:,.2f}" if isinstance(rev_v, (int, float)) else str(rev_v)
            days_s = f"{days_v:,.0f}" if isinstance(days_v, (int, float)) else str(days_v)
            in_window_marker = " <-- start_col" if c == start_col else ""
            print(f"{y}-{m:02d}     {get_column_letter(c):6s} {area_s:>14s} {rev_s:>14s} {days_s:>8s}{in_window_marker}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
