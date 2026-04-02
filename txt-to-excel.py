import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

input_txt = Path("databook-Foshan Wanyuan.txt")
output_xlsx = Path("databook-2.xlsx")


def parse_used_range(used_range):
    start_ref, _end_ref = used_range.split(":")
    match = re.match(r"([A-Z]+)(\d+)$", start_ref)
    if not match:
        raise ValueError(f"Invalid USED_RANGE: {used_range}")
    start_col_letters, start_row = match.groups()
    return int(start_row), column_index_from_string(start_col_letters)


def split_pipe_row(line):
    text = line.strip()
    if not text.startswith("|") or not text.endswith("|"):
        return None
    parts = [part.strip() for part in text[1:-1].split("|")]
    return parts


def convert_value(value):
    if value == "":
        return None

    if re.fullmatch(r"-?\d+", value):
        try:
            return int(value)
        except Exception:
            return value

    if re.fullmatch(r"-?\d+\.\d+", value):
        try:
            return float(value)
        except Exception:
            return value

    return value


wb = Workbook()
default_ws = wb.active
wb.remove(default_ws)

current_ws = None
start_row = None
start_col = None
waiting_for_header = False

with input_txt.open("r", encoding="utf-8") as f:
    for raw_line in f:
        line = raw_line.rstrip("\n")

        if line.startswith("===== SHEET: ") and line.endswith(" ====="):
            sheet_name = line[len("===== SHEET: "):-len(" =====")]
            current_ws = wb.create_sheet(title=sheet_name[:31])
            start_row = None
            start_col = None
            waiting_for_header = False
            continue

        if current_ws is None:
            continue

        if line.startswith("USED_RANGE: "):
            used_range = line[len("USED_RANGE: "):].strip()
            start_row, start_col = parse_used_range(used_range)
            waiting_for_header = True
            continue

        if not line.startswith("|"):
            continue

        row_parts = split_pipe_row(line)
        if not row_parts:
            continue

        if waiting_for_header:
            waiting_for_header = False
            continue

        if not row_parts or row_parts[0] == "ExcelRow":
            continue

        try:
            excel_row = int(row_parts[0])
        except ValueError:
            continue

        cell_values = row_parts[1:]

        for offset, value in enumerate(cell_values):
            current_ws.cell(
                row=excel_row,
                column=start_col + offset,
                value=convert_value(value),
            )

wb.save(output_xlsx)
print(f"Saved to: {output_xlsx}")