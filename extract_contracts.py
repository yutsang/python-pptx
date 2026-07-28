#!/usr/bin/env python3
"""Extract lease-contract fields from a folder of PDFs into the summary template.

Default model: Workbench GPT-5.5 (no --model flag).
Nearly all source PDFs are image-only scans → vision path; rare digital PDFs
use text extraction instead.

The local template's already-filled rows are the gold answer key. Use:
    python extract_contracts.py contracts/成都 --gold --validate
to run ONLY those gold PDFs and compare core fields.

Raw model JSON is also written next to the xlsx so you can inspect GPT-5.5
output directly (not only the Excel view).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook

from contract_template_schema import (
    CORE_VALUE_COLUMNS,
    LONG_TEXT_COLUMNS,
    TEMPLATE_COLUMNS,
    column_map,
)
from contract_vision import (
    build_page_data_urls,
    extract_pdf_text,
    image_file_to_jpeg_bytes,
    is_image_file,
    multi_image_byte_budget,
    pdf_is_digital,
    pdf_page_count,
    select_pages,
    to_data_url,
)
from fdd_utils.ai import AIClient
from inspect_contracts import find_template

_DEFAULT_MODEL = "workbench"
_MISSING = "未提及"
_COL_LETTERS = [letter for letter, _ in TEMPLATE_COLUMNS]
# GPT-5.5 spends max_completion_tokens on hidden reasoning first. Contract
# vision + JSON needs a much larger budget than FDD one-liners.
_EXTRACT_MAX_TOKENS = 16000
_EXTRACT_REASONING = "low"
_RETRY_MAX_TOKENS = 32000


def _collect_pdfs(path: Path) -> List[Path]:
    if path.is_file():
        return [path] if path.suffix.lower() == ".pdf" or is_image_file(path) else []
    return sorted(
        p for p in path.rglob("*")
        if p.is_file() and (p.suffix.lower() == ".pdf" or is_image_file(p))
    )


def _group_by_project(files: List[Path], root: Path) -> List[Tuple[str, List[Path]]]:
    groups: Dict[str, List[Path]] = {}
    for f in files:
        try:
            rel = f.relative_to(root)
        except ValueError:
            rel = Path(f.name)
        project = rel.parts[0] if len(rel.parts) > 1 else root.name
        groups.setdefault(project, []).append(f)
    return sorted(groups.items(), key=lambda kv: kv[0])


def _empty_row(filename: str = "") -> Dict[str, str]:
    row = {letter: _MISSING for letter in _COL_LETTERS}
    row["A"] = filename
    row["AC"] = "无"
    return row


def _parse_json_object(text: str) -> Dict[str, Any]:
    raw = (text or "").strip()
    if not raw:
        raise ValueError("empty model response")
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw, re.IGNORECASE)
    if fence:
        raw = fence.group(1).strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("{")
        end = raw.rfind("}")
        if start < 0 or end <= start:
            raise
        data = json.loads(raw[start : end + 1])
    if not isinstance(data, dict):
        raise ValueError("model JSON is not an object")
    return data


def _normalize_row(data: Dict[str, Any], filename: str) -> Dict[str, str]:
    headers = column_map()
    header_to_letter = {h: letter for letter, h in headers.items()}
    loose = {re.sub(r"\s+", "", h): letter for letter, h in headers.items()}

    row = _empty_row(filename)
    for key, value in data.items():
        key_s = str(key).strip()
        letter = None
        upper = key_s.upper()
        if upper in headers:
            letter = upper
        elif key_s in header_to_letter:
            letter = header_to_letter[key_s]
        else:
            letter = loose.get(re.sub(r"\s+", "", key_s))
        if not letter or letter not in row:
            continue
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        row[letter] = text
    row["A"] = filename
    return row


def _schema_prompt_block(letters: Sequence[str]) -> str:
    cmap = column_map()
    lines = [f'  "{letter}": "{cmap[letter]}"' for letter in letters if letter != "A"]
    return "{\n" + ",\n".join(lines) + "\n}"


def _build_extraction_prompt(filename: str, page_note: str, letters: Sequence[str]) -> str:
    core_hint = ""
    if set(letters) <= set(CORE_VALUE_COLUMNS) | {"D", "E", "AC"}:
        core_hint = (
            "特别注意从条款/表格中找：租赁单元、面积、交付日、租期、免租、"
            "租金/物业费合同总额、起始单价、涨幅、保证金、收款账户/账号、含税日单价。\n"
            "字段口径（易错，必须遵守）:\n"
            "- F 租赁单元：短名（如B库一层、C库一层），不要整段地址。\n"
            "- O/P：合同期内租金/物业费总额（含税），通常是较大金额。\n"
            "- Q/R：起始租金/物业费 单价，单位=元/日/平方米（不含税），通常 < 10。\n"
            "- AD/AE：含税日租金/含税日物业费，单位同样是元/日/平方米（含税单价），通常 < 10；"
            "绝不是日租金总价，也绝不是 O/P 合同总额。\n"
            "- AF 合计：= AD + AE（含税日单价之和），不是 O+P。\n"
            "- S/T 涨幅：尽量写成短式，如「每年递增4%」。\n"
            "- L/M 免租：若有多段免租期，写成完整区间文字"
            "（如2026年3月2日至2026年3月31日，…），不要只输出起始日列表。\n"
            "- N：无免租时填「不适用」，不要填「未提及」。\n"
            "- 甲乙方名称注意形近字（如臻/燊），以合同首页/签章为准。\n"
        )
    return (
        "你是租赁合同信息抽取助手。根据提供的合同页面内容，填写租赁台账字段。\n"
        f"源文件名: {filename}\n"
        f"{page_note}\n"
        f"{core_hint}\n"
        "规则:\n"
        "1. 只输出一个 JSON 对象，key 必须是下列 Excel 列字母；不要 markdown，不要解释。\n"
        f"2. 页面上看不到的字段填 \"{_MISSING}\"；备注没有内容时填 \"无\"。\n"
        "3. 日期：单日用 YYYY-MM-DD；免租多段保留中文区间原文风格。数字不要加千分位。\n"
        "4. 长文本字段最多各摘录 120 字；不要整段照抄。\n"
        "5. 不要输出 JSON 以外的任何文字。\n\n"
        "目标字段 (列字母: 含义):\n"
        f"{_schema_prompt_block(letters)}\n"
    )


def _fixup_rate_fields(row: Dict[str, str]) -> Dict[str, str]:
    """Correct common AD/AE/AF unit-rate vs contract-total confusions."""
    out = dict(row)

    def _f(letter: str) -> Optional[float]:
        try:
            return float(str(out.get(letter, "")).strip())
        except Exception:
            return None

    q, r = _f("Q"), _f("R")
    ad, ae, af = _f("AD"), _f("AE"), _f("AF")
    o, p = _f("O"), _f("P")

    # If AF was filled with O+P (contract totals), recompute from unit rates.
    if af is not None and o is not None and p is not None and af > 100:
        if abs(af - (o + p)) <= max(1.0, 0.02 * abs(o + p)):
            if ad is not None and ae is not None and ad < 10 and ae < 10:
                out["AF"] = f"{ad + ae:.6g}"
            elif q is not None and r is not None and q < 10 and r < 10:
                # Fallback: derive 含税单价 from 不含税 × common VAT rates in these leases.
                ad2, ae2 = q * 1.09, r * 1.06
                out["AD"] = f"{ad2:.6g}"
                out["AE"] = f"{ae2:.6g}"
                out["AF"] = f"{ad2 + ae2:.6g}"

    # If AD/AE look like daily totals (>>10) but Q/R are unit rates, derive 含税单价.
    if q is not None and r is not None and q < 10 and r < 10:
        if ad is not None and ad > 10:
            out["AD"] = f"{q * 1.09:.6g}"
        if ae is not None and ae > 10:
            out["AE"] = f"{r * 1.06:.6g}"
        ad2, ae2 = _f("AD"), _f("AE")
        if ad2 is not None and ae2 is not None and ad2 < 10 and ae2 < 10:
            af2 = _f("AF")
            if af2 is None or af2 > 10:
                out["AF"] = f"{ad2 + ae2:.6g}"

    if str(out.get("N", "")).strip() in (_MISSING, "无") and str(out.get("L", "")).strip() in (_MISSING, "无"):
        # No free-rent evidence → 不适用 (ledger convention), not 未提及.
        if str(out.get("M", "")).strip() in (_MISSING, "无"):
            out["N"] = "不适用"
    return out


def _call_model(client: AIClient, user_prompt: Any, *, max_tokens: int) -> Dict[str, Any]:
    return client.get_response(
        user_prompt=user_prompt,
        system_prompt="Return only a single JSON object for the lease ledger fields. No markdown.",
        max_tokens=max_tokens,
        reasoning_effort=_EXTRACT_REASONING,
    )


def _extract_payload(
    client: AIClient,
    user_prompt: Any,
    filename: str,
    *,
    letters: Sequence[str],
) -> Tuple[Dict[str, str], float, str]:
    """Returns (row, duration, raw_json_text)."""
    result = _call_model(client, user_prompt, max_tokens=_EXTRACT_MAX_TOKENS)
    content = str(result.get("content") or "").strip()
    duration = float(result.get("duration") or 0)
    if not content:
        print(
            f"  ⟳ empty response "
            f"(completion_tokens={result.get('completion_tokens')}, "
            f"{duration:.1f}s) — retry with max_tokens={_RETRY_MAX_TOKENS}"
        )
        result = _call_model(client, user_prompt, max_tokens=_RETRY_MAX_TOKENS)
        content = str(result.get("content") or "").strip()
        duration = float(result.get("duration") or 0)
    if not content:
        raise ValueError(
            f"empty model response after retry "
            f"(completion_tokens={result.get('completion_tokens')}, duration={duration}s)"
        )
    parsed = _parse_json_object(content)
    # Keep only requested letters (+A) so a core-pass cannot wipe long fields later.
    filtered = {k: v for k, v in parsed.items() if str(k).upper() in set(letters) | {"A"}}
    row = _normalize_row(filtered, filename)
    # Clear letters not in this pass back to missing so merge is explicit
    keep = set(letters) | {"A"}
    for letter in _COL_LETTERS:
        if letter not in keep:
            row[letter] = _MISSING
    return row, duration, content


def _vision_messages(
    path: Path,
    pages: Sequence[int],
    filename: str,
    letters: Sequence[str],
    page_note_extra: str = "",
) -> List[Any]:
    budget = multi_image_byte_budget(len(pages))
    urls = build_page_data_urls(path, pages, max_bytes_each=budget)
    page_note = (
        f"合同扫描页：第 {', '.join(str(p) for p, _ in urls)} 页。"
        f"{page_note_extra}"
        "只根据这些页面填写；未出现的字段填 "
        f"\"{_MISSING}\"。"
    )
    content: List[Any] = [
        {"type": "text", "text": _build_extraction_prompt(filename, page_note, letters)}
    ]
    for page_num, url in urls:
        content.append({"type": "text", "text": f"[page {page_num}]"})
        content.append({"type": "image_url", "image_url": {"url": url}})
    return content


def _merge_rows(base: Dict[str, str], overlay: Dict[str, str], letters: Sequence[str]) -> Dict[str, str]:
    out = dict(base)
    for letter in letters:
        val = overlay.get(letter, _MISSING)
        if val and val != _MISSING:
            out[letter] = val
    out["A"] = base.get("A") or overlay.get("A", "")
    return out


def _extract_from_digital(client: AIClient, path: Path) -> Tuple[Dict[str, str], float, str]:
    text = extract_pdf_text(path)
    letters = [L for L, _ in TEMPLATE_COLUMNS if L != "A"]
    prompt = _build_extraction_prompt(
        path.name,
        "以下是可直接抽取的电子版 PDF 全文（或截断后的首尾）。",
        letters,
    ) + f"\n\n===== PDF TEXT =====\n{text}\n"
    return _extract_payload(client, prompt, path.name, letters=letters)


def _extract_from_vision(
    client: AIClient,
    path: Path,
    max_pages: int,
) -> Tuple[Dict[str, str], float, str, List[int]]:
    """Two-pass vision: core commercial fields first, then long clause fields."""
    raw_parts: List[str] = []
    total_dur = 0.0

    if path.suffix.lower() != ".pdf":
        raw = image_file_to_jpeg_bytes(path, max_bytes=multi_image_byte_budget(1))
        letters = [L for L, _ in TEMPLATE_COLUMNS if L != "A"]
        content: List[Any] = [
            {"type": "text", "text": _build_extraction_prompt(path.name, "单页图片合同。", letters)},
            {"type": "image_url", "image_url": {"url": to_data_url(raw)}},
        ]
        row, dur, raw_json = _extract_payload(client, content, path.name, letters=letters)
        return row, dur, raw_json, [1]

    n_pages = pdf_page_count(path)
    pages = select_pages(n_pages, max_pages=max_pages)
    print(f"  pages: {pages} / {n_pages}  (JPEG budget ~{multi_image_byte_budget(len(pages))//1024}KB each)")

    # Pass 1 — core ledger numbers/dates (highest value for gold validation)
    core_letters = list(CORE_VALUE_COLUMNS) + ["D", "E", "AC"]
    msg1 = _vision_messages(
        path,
        pages,
        path.name,
        core_letters,
        page_note_extra=f"共 {n_pages} 页。这是第1遍：只抽核心台账字段。",
    )
    row1, dur1, raw1 = _extract_payload(client, msg1, path.name, letters=core_letters)
    row1 = _fixup_rate_fields(row1)
    total_dur += dur1
    raw_parts.append(raw1)
    print(
        f"  pass1 core ({dur1:.1f}s) 开始日={row1.get('I')} 结束日={row1.get('J')} "
        f"面积={row1.get('G')} AD/AE/AF={row1.get('AD')}/{row1.get('AE')}/{row1.get('AF')}"
    )

    # Pass 2 — long free-text clauses (best-effort; may still be 未提及)
    long_letters = list(LONG_TEXT_COLUMNS)
    # Fewer pages for clauses: front 3 + last is usually enough for 支付/违约/续租
    clause_pages = select_pages(n_pages, max_pages=min(5, max_pages))
    msg2 = _vision_messages(
        path,
        clause_pages,
        path.name,
        long_letters,
        page_note_extra=f"共 {n_pages} 页。这是第2遍：只抽长文本条款字段。",
    )
    row2, dur2, raw2 = _extract_payload(client, msg2, path.name, letters=long_letters)
    total_dur += dur2
    raw_parts.append(raw2)
    print(f"  pass2 clauses ({dur2:.1f}s)")

    merged = _merge_rows(row1, row2, long_letters)
    combined_raw = json.dumps({"pass1_core": raw_parts[0], "pass2_clauses": raw_parts[1]}, ensure_ascii=False, indent=2)
    return merged, total_dur, combined_raw, pages


def extract_one(
    client: AIClient,
    path: Path,
    max_pages: int,
) -> Tuple[Dict[str, str], float, str, str]:
    """Returns (row, duration, mode, raw_json)."""
    if path.suffix.lower() == ".pdf" and pdf_is_digital(path):
        row, dur, raw = _extract_from_digital(client, path)
        return row, dur, "digital-text", raw
    row, dur, raw, _pages = _extract_from_vision(client, path, max_pages=max_pages)
    return row, dur, "vision-2pass", raw


def _reference_fill_count(row: Dict[str, str]) -> int:
    """How many core fields look human-filled (not blank / 未提及)."""
    n = 0
    for letter in CORE_VALUE_COLUMNS:
        val = str(row.get(letter, "") or "").strip()
        if val and val not in (_MISSING, "无", "不适用", "不適用"):
            n += 1
    return n


def _load_gold_rows(template_path: Path) -> Dict[str, Dict[str, str]]:
    """filename -> row dict from filled reference rows in the local template.

    These are human-maintained placeholders — useful for calibration, not an
    absolute ground truth (a person may also mis-read a scanned page).
    Rows that only have parties filled (everything else 未提及) are skipped so
    a previous extraction output cannot masquerade as the reference template.
    """
    wb = load_workbook(template_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    gold: Dict[str, Dict[str, str]] = {}
    for row_idx in range(1, (ws.max_row or 0) + 1):
        filename = ws.cell(row=row_idx, column=1).value
        if not filename or not str(filename).lower().endswith((".pdf", ".jpg", ".png")):
            continue
        party_a = ws.cell(row=row_idx, column=2).value
        if party_a is None or str(party_a).strip() == "":
            continue
        row = _empty_row(str(filename).strip())
        for col_idx, letter in enumerate(_COL_LETTERS, start=1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            row[letter] = str(val).strip()
        # Need real commercial fields, not just 甲方/乙方 from a failed extract.
        if _reference_fill_count(row) < 5:
            continue
        gold[row["A"]] = row
    return gold


def _filter_gold_files(files: List[Path], gold_names: Sequence[str]) -> List[Path]:
    wanted = set(gold_names)
    matched = [f for f in files if f.name in wanted]
    return matched


def _norm_cmp(value: str) -> str:
    s = str(value or "").strip()
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # Excel datetime noise
    s = re.sub(r"\s+00:00:00$", "", s)
    s = re.sub(r"\s+", "", s)
    return s.lower()


def _date_tokens(value: str) -> set:
    s = str(value or "")
    tokens = set(re.findall(r"\d{4}-\d{2}-\d{2}", s))
    for y, m, d in re.findall(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", s):
        tokens.add(f"{int(y):04d}-{int(m):02d}-{int(d):02d}")
    return tokens


def _pct_token(value: str) -> Optional[str]:
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", str(value or ""))
    return m.group(1) if m else None


def _soft_equal(letter: str, got_raw: str, exp_raw: str) -> bool:
    got, exp = _norm_cmp(got_raw), _norm_cmp(exp_raw)
    if got == exp:
        return True
    try:
        g, e = float(got), float(exp)
        # Unit rates: allow ~2% relative drift from OCR / rounding.
        tol = max(1e-4, 0.02 * abs(e)) if abs(e) < 10 else max(1.0, 0.002 * abs(e))
        if abs(g - e) <= tol:
            return True
    except Exception:
        pass
    if letter == "F" and (exp in got or got in exp):
        return True
    if letter in ("L", "M"):
        gt, et = _date_tokens(got_raw), _date_tokens(exp_raw)
        if gt and et and gt == et:
            return True
    if letter in ("S", "T"):
        gp, ep = _pct_token(got_raw), _pct_token(exp_raw)
        if gp and ep and gp == ep and ("递增" in got or "增长" in got) and ("递增" in exp or "增长" in exp):
            return True
    if letter == "N" and {got, exp} <= {"未提及", "不适用", "不適用", "无", "n/a", "na"}:
        return True
    # Common OCR near-miss on 臻/燊 in party names.
    if letter in ("B", "C", "X"):
        if got.replace("燊", "臻") == exp.replace("燊", "臻"):
            return True
    return False


def _validate_against_gold(
    extracted: Dict[str, Dict[str, str]],
    gold: Dict[str, Dict[str, str]],
) -> int:
    overlap = sorted(set(extracted) & set(gold))
    if not overlap:
        print("\n(validate) No overlapping filenames vs template reference rows.")
        return 0
    print(
        "\n(validate) Diff vs human reference rows in the template "
        "(reference can also be wrong — use this to spot gaps, not as absolute truth).\n"
        f"Files: {len(overlap)}  |  columns: {', '.join(CORE_VALUE_COLUMNS)}"
    )
    mismatches = 0
    for name in overlap:
        print(f"\n--- {name} ---")
        bad = []
        for letter in CORE_VALUE_COLUMNS:
            got_raw = extracted[name].get(letter, "")
            exp_raw = gold[name].get(letter, "")
            if _soft_equal(letter, got_raw, exp_raw):
                continue
            bad.append(letter)
            mismatches += 1
            print(f"  {letter} differ")
            print(f"    ref:  {exp_raw[:160]}")
            print(f"    gpt:  {got_raw[:160]}")
        if not bad:
            print("  ✅ all core columns agree with reference (incl. soft matches)")
        else:
            ok_n = len(CORE_VALUE_COLUMNS) - len(bad)
            print(f"  → {ok_n}/{len(CORE_VALUE_COLUMNS)} core columns agree with reference")
    print(f"\n(validate) differing core fields: {mismatches}")
    return mismatches


def _write_workbook(
    groups: List[Tuple[str, List[Dict[str, str]]]],
    output_path: Path,
    title: str = "租赁台账",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=2, column=2, value=title)
    for col_idx, (letter, header) in enumerate(TEMPLATE_COLUMNS, start=1):
        if letter == "A":
            continue
        ws.cell(row=3, column=col_idx, value=header)

    r = 4
    for project, rows in groups:
        ws.cell(row=r, column=2, value=f"{project}")
        r += 1
        for row in rows:
            for col_idx, letter in enumerate(_COL_LETTERS, start=1):
                ws.cell(row=r, column=col_idx, value=row.get(letter, _MISSING))
            r += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_debug_json(output_xlsx: Path, payloads: Dict[str, str]) -> Path:
    debug_path = output_xlsx.with_name(output_xlsx.stem + "_raw.json")
    debug_path.write_text(json.dumps(payloads, ensure_ascii=False, indent=2), encoding="utf-8")
    return debug_path


def _resolve_template(path: Path, root: Path, explicit: Optional[str]) -> Optional[Path]:
    if explicit:
        return Path(explicit)
    # Prefer the parent of a project folder (contracts/) over the project
    # folder itself — extraction outputs land in the project folder and used
    # to win over 合同汇总模板.xlsx one level up.
    ordered: List[Path] = []
    for candidate_root in (root.parent, root, path if path.is_dir() else None):
        if candidate_root is None or not candidate_root.exists():
            continue
        if candidate_root not in ordered:
            ordered.append(candidate_root)
    for candidate_root in ordered:
        found = find_template(candidate_root)
        if found:
            return found
    return None


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Extract lease contracts via Workbench GPT-5.5 into the summary template.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python extract_contracts.py contracts/成都 --gold --validate\n"
            "      # only PDFs that have filled human reference rows in 合同汇总模板.xlsx\n"
            "  python extract_contracts.py contracts/成都 --template contracts/合同汇总模板.xlsx --gold --validate\n"
            "  python extract_contracts.py contracts\n"
        ),
    )
    ap.add_argument("path", help="contracts root, one project folder, or a single PDF")
    ap.add_argument("--template", default=None, help="path to 合同汇总模板.xlsx (auto-detected if omitted)")
    ap.add_argument("--out", default=None, help="output xlsx path (default: <path>/合同汇总_extracted.xlsx)")
    ap.add_argument("--max-pages", type=int, default=10, help="max vision pages per PDF (default: 10)")
    ap.add_argument(
        "--gold",
        action="store_true",
        help="only process PDFs that have filled human reference rows in the template",
    )
    ap.add_argument(
        "--validate",
        action="store_true",
        help="diff results against human reference rows in the template (reference can also be wrong)",
    )
    args = ap.parse_args()

    path = Path(args.path)
    if not path.exists():
        print(f"❌ Path not found: {path}")
        return 1

    files = _collect_pdfs(path)
    if not files:
        print(f"❌ No PDF/image files found under {path}")
        return 1

    root = path if path.is_dir() else path.parent
    group_root = root
    template_path = _resolve_template(path, root, args.template)
    out_path = Path(args.out) if args.out else (root / "合同汇总_extracted.xlsx")

    gold: Dict[str, Dict[str, str]] = {}
    if template_path and template_path.exists() and not template_path.name.startswith("~$"):
        gold = _load_gold_rows(template_path)

    if args.gold:
        if not gold:
            print("❌ --gold needs a local template with filled placeholder rows")
            print("   Tip: close Excel if the template is open (avoids ~$ lock files).")
            return 1
        before = len(files)
        files = _filter_gold_files(files, list(gold.keys()))
        print(f"--gold: {len(files)}/{before} file(s) match template placeholder rows")
        if not files:
            print("❌ None of the gold filenames were found under this path.")
            print("   Gold filenames in template:")
            for name in gold:
                print(f"    - {name}")
            return 1
        # Gold calibration always validates.
        args.validate = True

    if args.validate and not gold:
        print("❌ --validate needs a local template with gold rows")
        print("   Tip: close Excel if 合同汇总模板.xlsx is open (avoids ~$ lock files).")
        return 1

    print(f"Model: GPT-5.5 (workbench)  |  max_tokens={_EXTRACT_MAX_TOKENS}  |  reasoning={_EXTRACT_REASONING}")
    print(f"Files: {len(files)} under {path}")
    print(f"Vision pages/PDF: up to {args.max_pages} (2-pass: core then clauses)")
    print(f"Output: {out_path}")
    if template_path:
        print(f"Template: {template_path}")
        if gold:
            print(f"Human reference rows in template: {len(gold)}")
            for name in gold:
                print(f"  - {name}  (core fields filled: {_reference_fill_count(gold[name])})")
    print()

    try:
        client = AIClient(model_type=_DEFAULT_MODEL, agent_name="subagent_2", language="Chi")
    except Exception as exc:
        print(f"❌ Could not initialize AIClient: {exc}")
        return 1

    extracted_by_name: Dict[str, Dict[str, str]] = {}
    raw_by_name: Dict[str, str] = {}
    grouped_files = _group_by_project(files, group_root)
    grouped_rows: List[Tuple[str, List[Dict[str, str]]]] = []
    fail_n = 0

    file_i = 0
    for project, project_files in grouped_files:
        rows: List[Dict[str, str]] = []
        for f in project_files:
            file_i += 1
            print("=" * 78)
            print(f"[{file_i}/{len(files)}] {f.name}")
            print("=" * 78)
            try:
                row, dur, mode, raw = extract_one(client, f, max_pages=args.max_pages)
                extracted_by_name[f.name] = row
                raw_by_name[f.name] = raw
                rows.append(row)
                print(
                    f"✅ {mode} ({dur:.1f}s)  "
                    f"甲方={row.get('B', '')[:32]}  乙方={row.get('C', '')[:32]}  "
                    f"租期={row.get('I')}/{row.get('J')}  租金总额={row.get('O')}"
                )
            except Exception as exc:
                fail_n += 1
                err_row = _empty_row(f.name)
                err_row["AC"] = f"抽取失败: {exc}"
                rows.append(err_row)
                extracted_by_name[f.name] = err_row
                raw_by_name[f.name] = f"ERROR: {exc}"
                print(f"❌ {exc}")
            print()
        grouped_rows.append((project, rows))

    _write_workbook(grouped_rows, out_path)
    debug_path = _write_debug_json(out_path, raw_by_name)
    print("=" * 78)
    print(f"Wrote {out_path}")
    print(f"Raw GPT JSON: {debug_path}")
    print(f"SUMMARY: {len(files) - fail_n} ok / {fail_n} failed / {len(files)} total")
    print("=" * 78)

    mismatch_n = 0
    if args.validate and gold:
        mismatch_n = _validate_against_gold(extracted_by_name, gold)

    if fail_n:
        return 2
    if args.validate and mismatch_n:
        return 3
    return 0


if __name__ == "__main__":
    sys.exit(main())
