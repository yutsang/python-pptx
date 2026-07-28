#!/usr/bin/env python3
"""Dumps the REAL visual formatting of a pre-built bridge tab (e.g.
'成都-量价桥图') -- fonts, fills, borders, number formats, column widths
around its Base/Change block, PLUS any native Excel chart object already
embedded in that sheet (type, series colors, title, position/size) -- so
the auto-generated output for the other 16 entities (which have no
pre-built tab at all) can be made to actually LOOK like this one instead
of the generic invisible-base-series styling generate_bridge_waterfall_batch
currently uses.

This is a read-only inspector -- it doesn't change the source file.

Usage:
    python inspect_bridge_format.py "databooks/xx.xlsx" --sheet "成都-量价桥图"
"""
import argparse
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, ".")
from fdd_utils.bridge_chart_prototype import find_bridge_blocks


def _fmt_color(color) -> str:
    """color.rgb blows up into a confusing descriptor-error STRING (not an
    exception) when the color is actually theme/indexed-based rather than
    plain RGB -- checking color.type first avoids printing that garbage."""
    if color is None:
        return None
    try:
        if color.type == "rgb" and color.rgb and color.rgb not in ("00000000",):
            return f"rgb={color.rgb}"
        if color.type == "theme":
            return f"theme={color.theme} tint={color.tint}"
        if color.type == "indexed":
            return f"indexed={color.indexed}"
    except Exception:
        return None
    return None


def _fmt_font(font) -> str:
    bits = []
    if font.bold:
        bits.append("bold")
    if font.italic:
        bits.append("italic")
    if font.size:
        bits.append(f"size={font.size}")
    color_s = _fmt_color(font.color)
    if color_s:
        bits.append(f"color({color_s})")
    if font.name:
        bits.append(f"name={font.name!r}")
    return ",".join(bits) if bits else "(default)"


def _fmt_fill(fill) -> str:
    if fill is None or not fill.patternType:
        return "(none)"
    color_s = _fmt_color(fill.fgColor)
    return f"pattern={fill.patternType} {color_s or '(no color)'}"


def _fmt_border(border) -> str:
    sides = []
    for name in ("top", "bottom", "left", "right"):
        side = getattr(border, name)
        if side and side.style:
            sides.append(f"{name}={side.style}")
    return ",".join(sides) if sides else "(none)"


def dump_cell(ws, row: int, col: int, tag: str = ""):
    cell = ws.cell(row=row, column=col)
    addr = f"{get_column_letter(col)}{row}"
    print(f"  {addr} {tag}: value={cell.value!r}")
    print(f"      font: {_fmt_font(cell.font)}")
    print(f"      fill: {_fmt_fill(cell.fill)}")
    print(f"      border: {_fmt_border(cell.border)}")
    print(f"      number_format: {cell.number_format!r}")
    print(f"      alignment: horiz={cell.alignment.horizontal} vert={cell.alignment.vertical} "
          f"wrap={cell.alignment.wrap_text} textRotation={cell.alignment.text_rotation}")


def dump_charts(ws):
    charts = getattr(ws, "_charts", None) or []
    print(f"\n=== Native chart objects on this sheet: {len(charts)} ===")
    for i, chart in enumerate(charts):
        print(f"\n--- Chart {i + 1} ---")
        print(f"  type: {type(chart).__name__}")
        print(f"  grouping: {getattr(chart, 'grouping', None)}")
        print(f"  overlap: {getattr(chart, 'overlap', None)}")
        print(f"  gapWidth: {getattr(chart, 'gapWidth', None)}")
        title = getattr(chart, "title", None)
        print(f"  title: {title}")
        legend = getattr(chart, "legend", None)
        print(f"  legend: {'present' if legend is not None else 'None'}"
              + (f" position={legend.position}" if legend is not None else ""))
        anchor = getattr(chart, "anchor", None)
        print(f"  anchor: {anchor}")
        width = getattr(chart, "width", None)
        height = getattr(chart, "height", None)
        print(f"  width={width}cm height={height}cm")
        y_axis = getattr(chart, "y_axis", None)
        if y_axis is not None:
            print(f"  y_axis: title={getattr(y_axis, 'title', None)} "
                  f"numFmt={getattr(y_axis, 'numFmt', None)} "
                  f"min={getattr(y_axis.scaling, 'min', None) if getattr(y_axis, 'scaling', None) else None}")
        series_list = getattr(chart, "series", None) or []
        print(f"  series ({len(series_list)}):")
        for j, s in enumerate(series_list):
            name = None
            tx = getattr(s, "tx", None)
            if tx is not None:
                strRef = getattr(tx, "strRef", None)
                if strRef is not None:
                    name = getattr(strRef, "f", None)
            gp = getattr(s, "graphicalProperties", None) or getattr(s, "spPr", None)
            fill_desc = "?"
            if gp is not None:
                solid = getattr(gp, "solidFill", None)
                no_fill = getattr(gp, "noFill", None)
                if no_fill:
                    fill_desc = "noFill (invisible)"
                elif solid:
                    fill_desc = f"solidFill={solid}"
            dlbls = getattr(s, "dLbls", None)
            print(f"    series[{j}] name_ref={name!r} fill={fill_desc} "
                  f"has_data_labels={dlbls is not None}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the databook .xlsx")
    ap.add_argument("--sheet", required=True, help="the pre-built bridge tab to inspect, e.g. '成都-量价桥图'")
    args = ap.parse_args()

    print(f"Loading {args.path!r} (styles preserved, not data_only)...")
    wb = load_workbook(args.path, data_only=False)
    if args.sheet not in wb.sheetnames:
        print(f"❌ sheet {args.sheet!r} not found. Available: {wb.sheetnames}")
        return 1
    ws = wb[args.sheet]

    wb_values = load_workbook(args.path, data_only=True)
    blocks = find_bridge_blocks(wb_values[args.sheet])
    print(f"\nFound {len(blocks)} Base/Change block(s) via find_bridge_blocks.")

    for bi, block in enumerate(blocks):
        print(f"\n{'=' * 70}\nBlock {bi + 1}: header_row={block.header_row}, "
              f"label_col={get_column_letter(block.label_col)}, "
              f"base_col={get_column_letter(block.base_col)}, "
              f"change_col={get_column_letter(block.change_col)}\n{'=' * 70}")

        # Column widths for label/base/change and a bit either side.
        print("\n--- Column widths ---")
        for c in range(max(1, block.label_col - 2), block.change_col + 3):
            letter = get_column_letter(c)
            dim = ws.column_dimensions.get(letter)
            width = dim.width if dim else None
            print(f"  col {letter}: width={width}")

        # A few rows ABOVE the header (title cell? merged range?).
        print("\n--- Rows above header (title area) ---")
        for r in range(max(1, block.header_row - 3), block.header_row):
            for c in range(max(1, block.label_col - 1), block.change_col + 2):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    dump_cell(ws, r, c, tag="(above header)")

        print("\n--- Header row (Label / Base / Change) ---")
        for c in (block.label_col, block.base_col, block.change_col):
            dump_cell(ws, block.header_row, c, tag="(header)")

        print("\n--- First 2 item rows ---")
        for offset in (1, 2):
            r = block.header_row + offset
            for c in (block.label_col, block.base_col, block.change_col):
                dump_cell(ws, r, c, tag=f"(item row +{offset})")

        print("\n--- Last item row + check row area ---")
        last_row = block.header_row + len(block.items)
        for r in range(last_row, last_row + 3):
            for c in (block.label_col, block.base_col, block.change_col):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    dump_cell(ws, r, c, tag="(tail/check area)")

        # Merged cells overlapping this block's row range (titles are often merged).
        print("\n--- Merged cell ranges overlapping this block ---")
        for mc in ws.merged_cells.ranges:
            if mc.min_row <= last_row + 3 and mc.max_row >= max(1, block.header_row - 3):
                print(f"  {mc}")

    dump_charts(ws)

    print("\n--- Row heights (header_row-3 .. last block's tail+35, deduped) ---")
    if blocks:
        r_start = max(1, blocks[0].header_row - 3)
        r_end = blocks[-1].header_row + len(blocks[-1].items) + 35
        for r in range(r_start, min(r_end, ws.max_row) + 1):
            dim = ws.row_dimensions.get(r)
            if dim and dim.height:
                print(f"  row {r}: height={dim.height}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
