#!/usr/bin/env python3
"""Exploratory inspector for a price-volume bridge/waterfall workings tab.

Answers two structural questions needed before deciding whether a bridge/
waterfall chart can be auto-generated:
  1. Is the bridge tab's B3:AE9 range LIVE (formulas pulling from another
     tab like 'AB-CD') or hardcoded/pasted values? If live, which exact
     cells in the source tab does it actually depend on?
  2. How many 'AB-' prefixed tabs exist (relevant to "what if the number of
     variables/entities grows later" -- one tab per entity/variable set
     suggests a repeatable pattern; a one-off suggests more bespoke work).

Deliberately does NOT dump the entire source tab (could be huge) -- only
the specific cells the bridge range's formulas reference, traced via regex
over the formula strings.

Usage:
    python inspect_bridge_source.py "databooks/xx指標v1.xlsx"
    python inspect_bridge_source.py "databooks/xx指標v1.xlsx" \
        --bridge-sheet "成都-量价桥图" --bridge-range B3:AE9 --source-sheet "AB-CD"
"""
import argparse
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

# Matches 'Some-Sheet'!B3 or 'Some-Sheet'!B3:B9 or PlainSheet!B3 (quoted
# form required whenever the sheet name has a hyphen/space, which is why
# both a quoted and unquoted alternative are here).
_CROSS_SHEET_REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_一-鿿]+))!(\$?[A-Z]{1,3}\$?\d+)(?::(\$?[A-Z]{1,3}\$?\d+))?"
)


def _clean_ref(ref: str) -> str:
    return ref.replace("$", "")


def _dump_range(ws_values, ws_formulas, cell_range: str, label: str):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    print(f"\n--- {label}: {cell_range} "
          f"({max_row - min_row + 1} row(s) x {max_col - min_col + 1} col(s)) ---")
    refs_found = []
    any_cell = False
    for row in range(min_row, max_row + 1):
        row_cells = []
        for col in range(min_col, max_col + 1):
            v = ws_values.cell(row=row, column=col).value
            f = ws_formulas.cell(row=row, column=col).value
            formula = f if isinstance(f, str) and f.startswith("=") else None
            if v is None and formula is None:
                continue
            any_cell = True
            addr = f"{get_column_letter(col)}{row}"
            if formula:
                row_cells.append(f"{addr}={formula!r} -> {v!r}")
                for m in _CROSS_SHEET_REF_RE.finditer(formula):
                    sheet_name = m.group(1) or m.group(2)
                    refs_found.append((sheet_name, _clean_ref(m.group(3)), _clean_ref(m.group(4)) if m.group(4) else None))
            else:
                row_cells.append(f"{addr}={v!r}")
        if row_cells:
            print(f"  row {row}: " + " | ".join(row_cells))
    if not any_cell:
        print("  (empty in this range)")
    return refs_found


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the databook .xlsx")
    ap.add_argument("--bridge-sheet", default="成都-量价桥图", help="the bridge/waterfall workings tab name")
    ap.add_argument("--bridge-range", default="B3:AE9", help="the data-taking range within the bridge tab")
    ap.add_argument("--source-sheet", default="AB-CD", help="the tab the bridge is expected to pull from")
    args = ap.parse_args()

    print("Loading workbook (values pass)...")
    wb_values = load_workbook(args.path, data_only=True)
    print("Loading workbook (formulas pass)...")
    wb_formulas = load_workbook(args.path, data_only=False)

    print("\n" + "=" * 78)
    print("  SHEET INVENTORY")
    print("=" * 78)
    all_sheets = wb_values.sheetnames
    ab_prefixed = [s for s in all_sheets if s.startswith("AB-") or s.startswith("AB")]
    print(f"Total sheets: {len(all_sheets)}")
    print(f"'AB-' prefixed sheets ({len(ab_prefixed)}): {ab_prefixed}")

    if args.bridge_sheet not in all_sheets:
        print(f"\n❌ bridge sheet {args.bridge_sheet!r} not found.")
        print(f"   Available sheets: {all_sheets}")
        return 1

    ws_v = wb_values[args.bridge_sheet]
    ws_f = wb_formulas[args.bridge_sheet]
    print(f"\n{args.bridge_sheet!r} full dimensions: {ws_v.dimensions}")

    print("\n" + "=" * 78)
    print("  BRIDGE TAB CONTENT (values + formulas)")
    print("=" * 78)
    refs = _dump_range(ws_v, ws_f, args.bridge_range, f"BRIDGE [{args.bridge_sheet}]")

    print("\n" + "=" * 78)
    print("  CROSS-SHEET REFERENCES FOUND IN THE BRIDGE RANGE")
    print("=" * 78)
    by_sheet = defaultdict(set)
    for sheet_name, c1, c2 in refs:
        by_sheet[sheet_name].add((c1, c2))
    if not by_sheet:
        print("  None found -- the bridge range's cells are either plain hardcoded")
        print("  values, or formulas that only reference OTHER cells within the same")
        print("  bridge tab (no live link out to a source tab).")
    else:
        for sheet_name, cellset in by_sheet.items():
            print(f"  {sheet_name!r}: {len(cellset)} distinct cell ref(s) -> {sorted(cellset)}")

    target_key = args.source_sheet
    source_refs = by_sheet.get(target_key)
    if source_refs is None:
        for k in by_sheet:
            if k.strip("'") == target_key.strip("'"):
                source_refs = by_sheet[k]
                break

    if args.source_sheet not in all_sheets:
        print(f"\n⚠️ source sheet {args.source_sheet!r} not found in workbook.")
        print(f"   Available sheets: {all_sheets}")
        return 0

    ws_sv = wb_values[args.source_sheet]
    ws_sf = wb_formulas[args.source_sheet]
    print(f"\n{args.source_sheet!r} full dimensions: {ws_sv.dimensions} (NOT dumped in full)")

    print("\n" + "=" * 78)
    print(f"  ONLY THE {args.source_sheet!r} CELLS THE BRIDGE ACTUALLY REFERENCES")
    print("=" * 78)
    if not source_refs:
        print("  (none -- see note above; bridge range has no live formula link to this sheet)")
    else:
        for c1, c2 in sorted(source_refs):
            rng = f"{c1}:{c2}" if c2 else f"{c1}:{c1}"
            _dump_range(ws_sv, ws_sf, rng, f"{args.source_sheet}!{rng}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
