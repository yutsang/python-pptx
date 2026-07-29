#!/usr/bin/env python3
"""Extract utility-fee clauses and per-period payment details from lease contracts.

Outputs per project folder (contracts/<project>/):
  合同_水电约定.xlsx
      One sheet per PDF in the folder. One row: source filename + the
      utility-fee arrangement text found in the contract.
  合同_每期支付明细.xlsx
      One sheet per PDF in the folder. Rows = each payment period with
      期间/日期/金额/备注 columns as stated in the contract.

Uses Workbench GPT-5.5 vision (image-only scans) or text layer (rare digital PDFs).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from contract_vision import (
    build_page_data_urls,
    extract_pdf_text,
    multi_image_byte_budget,
    pdf_is_digital,
    pdf_page_count,
    rasterize_page_jpeg,
    rasterize_page_tile_jpegs,
    select_pages,
    to_data_url,
)
from fdd_utils.ai import AIClient

_DEFAULT_MODEL = "workbench"
_MAX_TOKENS = 16000
_REASONING = "low"

_UTILITY_HEADERS = [
    "源文件名", "水费约定", "电费约定", "抄表/确认方式",
    "公用事业分摊规则", "结算周期", "税费/税率", "其他约定", "原文摘录",
]
_PAYMENT_HEADERS = [
    "期次", "期间开始", "期间结束", "租赁费（含税）", "物业管理服务费（含税）",
    "其他费用", "合计（含税）", "支付截止日/条件", "备注", "原文摘录",
]

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _safe_sheet_name(name: str, used: set) -> str:
    base = _INVALID_SHEET_CHARS.sub("_", name).strip() or "sheet"
    base = base[:28]  # leave room for suffix
    out, i = base, 1
    while out.lower() in used or len(out) > 31:
        suffix = f"_{i}"
        out = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(out.lower())
    return out


def _collect_pdfs(folder: Path) -> List[Path]:
    """PDFs in this folder only — project folders should not inherit code/
    tooling folders that happen to sit beside the contracts."""
    if folder.is_file():
        return [folder] if folder.suffix.lower() == ".pdf" else []
    return sorted(p for p in folder.glob("*.pdf") if p.is_file())


def _project_folders(root: Path) -> List[Path]:
    """Folders that actually contain PDFs at their top level.

    If `root` itself holds PDFs (e.g. contracts/成都/*.pdf), treat root as the
    project — not any code subfolders like fdd_utils/ or font_metrics/ that may
    live alongside the contracts. Otherwise (contracts/ with no loose PDFs),
    use each direct subfolder.
    """
    if root.is_file():
        return [root.parent]
    if any(p.is_file() for p in root.glob("*.pdf")):
        return [root]
    subs = [p for p in sorted(root.iterdir()) if p.is_dir() and any(p.glob("*.pdf"))]
    return subs if subs else [root]


def _parse_json_object(text: str) -> Dict[str, Any]:
    raw = (text or "").strip()
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw, re.IGNORECASE)
    if fence:
        raw = fence.group(1).strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start, end = raw.find("{"), raw.rfind("}")
        if start < 0 or end <= start:
            raise
        data = json.loads(raw[start : end + 1])
    if not isinstance(data, dict):
        raise ValueError("model output is not a JSON object")
    return data


def _vision_pages(client_prompt_text: str, pdf_path: Path, max_pages: int) -> List[Any]:
    n_pages = pdf_page_count(pdf_path)
    pages = select_pages(n_pages, max_pages=max_pages)
    urls = build_page_data_urls(pdf_path, pages, max_bytes_each=multi_image_byte_budget(len(pages)))
    content: List[Any] = [{"type": "text", "text": client_prompt_text + f"\n提供页面: {pages}\n"}]
    for page_num, url in urls:
        content.append({"type": "text", "text": f"[page {page_num}]"})
        content.append({"type": "image_url", "image_url": {"url": url}})
    return content


def _vision_pages_at_dpi(
    client_prompt_text: str,
    pdf_path: Path,
    pages: Sequence[int],
    per_image_bytes: int,
    dpi_start: int,
    max_edge_start: int = 1800,
) -> List[Any]:
    """Higher-fidelity images for dense amount tables (payment details)."""
    content: List[Any] = [{"type": "text", "text": client_prompt_text + f"\n提供页面: {list(pages)}\n"}]
    for page_num in pages:
        raw = rasterize_page_jpeg(
            pdf_path,
            page_num,
            max_bytes=per_image_bytes,
            dpi_start=dpi_start,
            max_edge_start=max_edge_start,
        )
        content.append({"type": "text", "text": f"[page {page_num}]"})
        content.append({"type": "image_url", "image_url": {"url": to_data_url(raw)}})
    return content


def _call_json(client: AIClient, user_prompt: Any) -> Dict[str, Any]:
    result = client.get_response(
        user_prompt=user_prompt,
        system_prompt="Return only one JSON object. No markdown.",
        max_tokens=_MAX_TOKENS,
        reasoning_effort=_REASONING,
    )
    raw = str(result.get("content") or "").strip()
    if not raw:
        raise ValueError(f"empty response (completion_tokens={result.get('completion_tokens')})")
    return _parse_json_object(raw)


def _doc_input(client_prompt_text: str, pdf_path: Path, max_pages: int) -> List[Any] | str:
    if pdf_path.suffix.lower() == ".pdf" and pdf_is_digital(pdf_path):
        text = extract_pdf_text(pdf_path)
        return client_prompt_text + f"\n\n===== PDF TEXT =====\n{text}\n"
    return _vision_pages(client_prompt_text, pdf_path, max_pages)


def _utility_prompt(filename: str) -> str:
    return (
        "你是租赁合同条款抽取助手。只根据提供的内容，抽取该合同里关于水电费的约定。\n"
        f"源文件名: {filename}\n\n"
        "输出一个 JSON 对象（没有就对应字段填空字符串）：\n"
        "{\n"
        '  "水费约定": "",\n'
        '  "电费约定": "",\n'
        '  "抄表/确认方式": "",\n'
        '  "公用事业分摊规则": "",\n'
        '  "结算周期": "",\n'
        '  "税费/税率": "",\n'
        '  "其他约定": "",\n'
        '  "原文摘录": ""\n'
        "}\n"
        "规则:\n"
        "- 各字段为简明短句（可合并多句），没有的填空字符串。\n"
        "- 原文摘录: 抄一句最能代表水电约定的原文（最长约200字）。\n"
        "- 只输出 JSON，不要 markdown。\n"
    )


def _payment_prompt(filename: str) -> str:
    return (
        "你是租赁合同支付明细逐格抄录助手。当前只提供一张已定位的表格页。"
        "请按视觉中的行列逐格抄录，不要结合其他页面推算。\n"
        f"源文件名: {filename}\n\n"
        "输出一个 JSON 对象：\n"
        "{\n"
        '  "periods": [\n'
        '    {\n'
        '      "期次": "",\n'
        '      "期间开始": "",\n'
        '      "期间结束": "",\n'
        '      "租赁费（含税）": "",\n'
        '      "物业管理服务费（含税）": "",\n'
        '      "其他费用": "",\n'
        '      "合计（含税）": "",\n'
        '      "支付截止日/条件": "",\n'
        '      "备注": "",\n'
        '      "原文摘录": ""\n'
        "    }\n"
        "  ]\n"
        "}\n"
        "规则:\n"
        "- 金额必须逐字符抄表；特别检查位数、小数点和每个零，禁止根据相邻行补全。\n"
        "- 每一视觉行只对应一条 periods 记录，不要跨行拼接。\n"
        "- 找不到明细表 → periods = []；找到但数字看不清 → 也不要编，宁可留空该行金额。\n"
        "- 期次：用表内编号或第N期，不要写成日期。\n"
        "- 日期忠实抄录后尽量 YYYY-MM-DD；金额保留原值但不要加千分位符号。\n"
        "- 原文摘录：该行最关键的一句（约120字内），不要整段附件说明。\n"
        "- 只输出 JSON，不要 markdown。\n"
    )


def _payment_scan_prompt() -> str:
    return (
        "逐页检查所提供的合同扫描页，只定位包含付款金额明细的页面。"
        "目标包括：每期应付租赁费/租金/物业服务费明细表，或明确列出期间与应付金额的修订表格。"
        "仅提到『详见附件』、付款规则、银行账户、签字页，不算目标。\n"
        '输出 JSON：{"candidate_pages": [页码整数], "evidence": {"页码": "短证据"}}。\n'
        "页码必须使用每张图片前的 [page N] 标签。没有目标页就返回空数组。"
    )


def _payment_structure_prompt(filename: str, page_num: int) -> str:
    return (
        "你是表格结构识别员。只识别当前付款明细表的结构，不抄数据行。"
        "按页面从左到右返回所有原始列；多层表头用『上层 / 下层』合并，"
        "不得省略税率、单价、面积、天数、不含税金额、税额、含税金额、合计、备注等可见列。"
        f"\n源文件名: {filename}\n页面: {page_num}\n"
        '输出 JSON：{"table_title":"","unit":"","columns":["原表列一","原表列二"]}。'
        "列名必须忠实使用图片文字且保持原顺序，只输出 JSON。"
    )


def _payment_dynamic_rows_prompt(
    filename: str,
    page_num: int,
    tile_num: int,
    tile_count: int,
    columns: List[str],
) -> str:
    return (
        "你是付款表格逐格抄录员。图片是原表的一段高解像横向切片。"
        "只抄图中完整可见的数据行，边界被截断的行不要输出。"
        "每行必须严格按给定原表列顺序输出相同数量的单元格；保留原文、原小数位及数字位数，"
        "不得计算、补全、改写或省略列。\n"
        f"源文件名: {filename}\n页面: {page_num}\n切片: {tile_num}/{tile_count}\n"
        f"原表列（共 {len(columns)} 列）: {json.dumps(columns, ensure_ascii=False)}\n"
        '输出 JSON：{"rows":[["单元格一","单元格二"]]}。只输出 JSON。'
    )


def _dynamic_columns(data: Dict[str, Any]) -> List[str]:
    values = data.get("columns")
    if not isinstance(values, list):
        return []
    columns: List[str] = []
    used: Dict[str, int] = {}
    for index, value in enumerate(values, 1):
        name = re.sub(r"\s+", " ", str(value or "")).strip() or f"未命名列{index}"
        used[name] = used.get(name, 0) + 1
        if used[name] > 1:
            name = f"{name}_{used[name]}"
        columns.append(name)
    return columns


def _dynamic_rows(data: Dict[str, Any], width: int) -> List[List[str]]:
    values = data.get("rows")
    if not isinstance(values, list):
        return []
    rows: List[List[str]] = []
    for value in values:
        if not isinstance(value, list):
            continue
        row = [str(cell).strip() if cell is not None else "" for cell in value[:width]]
        row.extend([""] * (width - len(row)))
        if any(row):
            rows.append(row)
    return rows


def _merge_dynamic_review(first: List[List[str]], reviewed: List[List[str]], width: int) -> List[List[str]]:
    if not reviewed:
        return first
    merged: List[List[str]] = []
    for index in range(max(len(first), len(reviewed))):
        original = first[index] if index < len(first) else [""] * width
        correction = reviewed[index] if index < len(reviewed) else [""] * width
        merged.append([
            correction[col] if correction[col] else original[col]
            for col in range(width)
        ])
    return merged


def _dynamic_row_key(columns: List[str], row: List[str]) -> tuple:
    preferred = [
        index for index, name in enumerate(columns)
        if any(token in name for token in ("期次", "序号", "期间", "开始", "结束", "日期"))
    ]
    indices = preferred[:3] or list(range(min(3, len(row))))
    key = tuple(re.sub(r"[\s年月日./\-]", "", row[index]) for index in indices)
    return key if any(key) else tuple(row)


def _payment_dynamic_review_prompt(columns: List[str], first: Dict[str, Any]) -> str:
    return (
        "逐格复核同一张高解像表格切片及第一次抄录。重点核对每个数字字符、"
        "小数点、零、日期和横向列位置。只能依据图片更正，不得计算。"
        "不得用空白删除第一次已有值。\n"
        f"原表列: {json.dumps(columns, ensure_ascii=False)}\n"
        f"第一次抄录: {json.dumps(first, ensure_ascii=False)}\n"
        '按相同格式输出 JSON：{"rows":[["..."]]}，每行列数必须一致。'
    )


def _payment_review_prompt(filename: str, page_num: int, first_pass: Dict[str, Any]) -> str:
    return (
        "你是金额抄录复核员。对照同一张高解像表格图片，逐行、逐字符核对下面第一次抄录。"
        "重点检查四位/五位数混淆、重复数字、漏零、小数点、跨行错位，以及合计是否与可见分项一致。"
        "只能按图片纠正，禁止推算。复核结果不得用空白删除第一次已有的日期或金额；"
        "若无法确认某个已有值，就保留该值并在备注注明『待人工确认』。"
        f"\n源文件名: {filename}\n页面: {page_num}\n第一次抄录:\n"
        f"{json.dumps(first_pass, ensure_ascii=False)}\n"
        "请按原有 periods JSON schema 返回完整复核结果，只输出 JSON。"
    )


def _get_str(d: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        v = d.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _utility_row(filename: str, data: Dict[str, Any]) -> List[str]:
    return [
        filename,
        _get_str(data, "水费约定"),
        _get_str(data, "电费约定"),
        _get_str(data, "抄表/确认方式", "抄表方式", "确认方式"),
        _get_str(data, "公用事业分摊规则", "分摊规则", "公用事业分摊"),
        _get_str(data, "结算周期", "结算方式"),
        _get_str(data, "税费/税率", "税费", "税率"),
        _get_str(data, "其他约定", "其他"),
        _get_str(data, "原文摘录", "quote"),
    ]


def _payment_rows(data: Dict[str, Any]) -> List[List[str]]:
    periods = data.get("periods")
    if not isinstance(periods, list):
        return []
    rows: List[List[str]] = []
    for item in periods:
        if not isinstance(item, dict):
            continue
        rows.append([
            _get_str(item, "期次", "期数"),
            _get_str(item, "期间开始", "开始日期", "开始"),
            _get_str(item, "期间结束", "结束日期", "结束"),
            _get_str(item, "租赁费（含税）", "租赁费", "租赁费含税"),
            _get_str(item, "物业管理服务费（含税）", "物业管理服务费", "物业服务费"),
            _get_str(item, "其他费用", "其他"),
            _get_str(item, "合计（含税）", "合计", "小计"),
            _get_str(item, "支付截止日/条件", "支付条件", "截止日"),
            _get_str(item, "备注"),
            _get_str(item, "原文摘录", "quote"),
        ])
    return rows


def _amount_decimal(value: str) -> Optional[Decimal]:
    text = re.sub(r"[^\d.\-]", "", str(value or ""))
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _arithmetic_consistent(row: List[str]) -> Optional[bool]:
    if len(row) < 7:
        return None
    rent = _amount_decimal(row[3])
    management = _amount_decimal(row[4])
    total = _amount_decimal(row[6])
    if rent is None or management is None or total is None:
        return None
    other = _amount_decimal(row[5]) or Decimal(0)
    return abs((rent + management + other) - total) <= Decimal("0.02")


def _merge_reviewed_rows(first: List[List[str]], reviewed: List[List[str]]) -> List[List[str]]:
    """Review may correct values, but must never erase populated first-pass cells."""
    if not first:
        return reviewed
    if not reviewed:
        return first
    merged: List[List[str]] = []
    for idx in range(max(len(first), len(reviewed))):
        original = first[idx] if idx < len(first) else [""] * len(_PAYMENT_HEADERS)
        correction = reviewed[idx] if idx < len(reviewed) else [""] * len(_PAYMENT_HEADERS)
        width = max(len(original), len(correction), len(_PAYMENT_HEADERS))
        original = original + [""] * (width - len(original))
        correction = correction + [""] * (width - len(correction))
        merged_row = [
            correction[col].strip() if correction[col].strip() else original[col].strip()
            for col in range(width)
        ]
        # Do not accept a review that breaks a previously valid visible sum.
        if _arithmetic_consistent(original) is True and _arithmetic_consistent(merged_row) is False:
            for col in range(3, 7):
                if original[col].strip():
                    merged_row[col] = original[col].strip()
        merged.append(merged_row)
    return merged


def _amount_coverage(rows: List[List[str]]) -> tuple[int, int]:
    """Rows with at least one visible payment amount versus total rows."""
    return sum(1 for row in rows if any(str(value).strip() for value in row[3:7])), len(rows)


def _payment_row_identity(row: List[str]) -> tuple:
    period = re.sub(r"\D", "", str(row[0])) if row else ""
    start = re.sub(r"\D", "", str(row[1])) if len(row) > 1 else ""
    end = re.sub(r"\D", "", str(row[2])) if len(row) > 2 else ""
    if start or end:
        return ("dates", start, end)
    return ("period", period)


def _merge_zoom_rows(base: List[List[str]], zoomed: List[List[str]]) -> List[List[str]]:
    """Use zoomed values to fill/correct matching rows without creating duplicates."""
    if not zoomed:
        return base
    result = [list(row) for row in base]
    best_zoomed: Dict[tuple, List[str]] = {}
    for row in zoomed:
        identity = _payment_row_identity(row)
        if not any(identity[1:]):
            continue
        existing = best_zoomed.get(identity)
        score = sum(bool(str(value).strip()) for value in row)
        old_score = sum(bool(str(value).strip()) for value in existing) if existing else -1
        if score > old_score:
            best_zoomed[identity] = row
    by_identity: Dict[tuple, int] = {}
    for index, row in enumerate(result):
        identity = _payment_row_identity(row)
        if any(identity[1:]):
            by_identity[identity] = index
    for identity, row in best_zoomed.items():
        identity = _payment_row_identity(row)
        index = by_identity.get(identity)
        if index is None:
            if any(str(value).strip() for value in row[3:7]):
                by_identity[identity] = len(result)
                result.append(list(row))
            continue
        width = min(len(result[index]), len(row))
        for col in range(width):
            if str(row[col]).strip():
                result[index][col] = str(row[col]).strip()
    result.sort(key=lambda row: (
        re.sub(r"\D", "", str(row[1])) if len(row) > 1 else "",
        re.sub(r"\D", "", str(row[0])) if row else "",
    ))
    return result


def extract_utility(client: AIClient, pdf: Path, max_pages: int) -> List[str]:
    payload = _doc_input(_utility_prompt(pdf.name), pdf, max_pages)
    data = _call_json(client, payload)
    return _utility_row(pdf.name, data)


def _scan_payment_pages(client: AIClient, pdf: Path, batch_size: int) -> List[int]:
    """Scan every page at moderate resolution and return actual table pages."""
    n_pages = pdf_page_count(pdf)
    candidates: set[int] = set()
    size = max(2, min(4, batch_size))
    for start in range(1, n_pages + 1, size):
        pages = list(range(start, min(n_pages, start + size - 1) + 1))
        payload = _vision_pages_at_dpi(
            _payment_scan_prompt(),
            pdf,
            pages,
            per_image_bytes=multi_image_byte_budget(len(pages)),
            dpi_start=140,
            max_edge_start=1800,
        )
        data = _call_json(client, payload)
        found = data.get("candidate_pages", [])
        if not isinstance(found, list):
            continue
        for value in found:
            try:
                page_num = int(value)
            except (TypeError, ValueError):
                continue
            if page_num in pages:
                candidates.add(page_num)
    return sorted(candidates)


def _extract_payment_page(client: AIClient, pdf: Path, page_num: int) -> Dict[str, Any]:
    """Recover the source table's full schema and rows at row-level resolution."""
    raw = rasterize_page_jpeg(
        pdf,
        page_num,
        max_bytes=2_450_000,
        dpi_start=240,
        max_edge_start=3400,
    )
    image = {"type": "image_url", "image_url": {"url": to_data_url(raw)}}
    structure_payload: List[Any] = [
        {"type": "text", "text": _payment_structure_prompt(pdf.name, page_num)},
        image,
    ]
    structure = _call_json(client, structure_payload)
    columns = _dynamic_columns(structure)
    if not columns:
        columns = list(_PAYMENT_HEADERS)
    print(f"  (schema page {page_num}) {len(columns)} source column(s)")

    tile_count = 8
    collected: Dict[tuple, List[str]] = {}
    order: List[tuple] = []
    for tile_num, tile_raw in rasterize_page_tile_jpegs(
        pdf,
        page_num,
        tile_count=tile_count,
        dpi=450,
    ):
        tile_image = {"type": "image_url", "image_url": {"url": to_data_url(tile_raw)}}
        prompt = _payment_dynamic_rows_prompt(
            pdf.name,
            page_num,
            tile_num,
            tile_count,
            columns,
        )
        first_payload: List[Any] = [{"type": "text", "text": prompt}, tile_image]
        first = _call_json(client, first_payload)
        first_rows = _dynamic_rows(first, len(columns))
        if not first_rows:
            continue
        review_payload: List[Any] = [
            {"type": "text", "text": _payment_dynamic_review_prompt(columns, first)},
            tile_image,
        ]
        reviewed = _call_json(client, review_payload)
        tile_rows = _merge_dynamic_review(
            first_rows,
            _dynamic_rows(reviewed, len(columns)),
            len(columns),
        )
        for row in tile_rows:
            key = _dynamic_row_key(columns, row)
            existing = collected.get(key)
            if existing is None:
                collected[key] = row
                order.append(key)
            elif sum(bool(cell) for cell in row) > sum(bool(cell) for cell in existing):
                collected[key] = row

    rows = [collected[key] for key in order]
    return {
        "title": _get_str(structure, "table_title", "title"),
        "unit": _get_str(structure, "unit", "金额单位"),
        "columns": columns,
        "rows": rows,
        "source_page": page_num,
    }


def extract_payment(client: AIClient, pdf: Path, max_pages: int) -> List[Dict[str, Any]]:
    pages = _scan_payment_pages(client, pdf, batch_size=max_pages)
    if not pages:
        print(f"  (scan) checked all {pdf_page_count(pdf)} page(s); no payment table found")
        return []
    print(f"  (scan) payment table page(s): {pages}")
    tables: List[Dict[str, Any]] = []
    for page_num in pages:
        table = _extract_payment_page(client, pdf, page_num)
        row_count = len(table.get("rows", []))
        print(f"  (page {page_num}) complete-table transcription: {row_count} row(s)")
        if row_count:
            tables.append(table)
    return tables


def _write_utility_wb(rows_by_folder: List[tuple], out_path: Path) -> None:
    """One global workbook: one sheet per folder, all folder PDFs as rows."""
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for folder_name, rows in rows_by_folder:
        ws = wb.create_sheet(title=_safe_sheet_name(folder_name, used))
        ws.append(_UTILITY_HEADERS)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 40
        for col in "BCDEFGHI":
            ws.column_dimensions[col].width = 26
    if not wb.worksheets:
        ws = wb.create_sheet("水电约定")
        ws.append(_UTILITY_HEADERS)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _write_payment_wb(rows_by_file: List[tuple], out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for filename, tables in rows_by_file:
        ws = wb.create_sheet(title=_safe_sheet_name(Path(filename).stem, used))
        ws.append(["来源文件", filename])
        if not tables:
            ws.append(["状态", "未识别到付款明细表"])
            continue
        for table_index, table in enumerate(tables, 1):
            if table_index > 1:
                ws.append([])
            columns = list(table.get("columns") or [])
            rows = list(table.get("rows") or [])
            ws.append(["表格标题", table.get("title") or f"付款明细表 {table_index}"])
            ws.append(["金额单位", table.get("unit") or "", "来源页码", table.get("source_page") or ""])
            ws.append(columns)
            for row in rows:
                ws.append(row)
            header_row = ws.max_row - len(rows)
            ws.freeze_panes = f"A{header_row + 1}"
            for index, column in enumerate(columns, 1):
                values = [str(column)] + [
                    str(row[index - 1]) for row in rows if index <= len(row)
                ]
                width = min(35, max(10, max(len(value) for value in values) + 2))
                letter = get_column_letter(index)
                ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, width)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _utility_output_path(root: Path) -> Path:
    if root.is_file():
        base = root.parent.parent
    elif any(root.glob("*.pdf")):
        base = root.parent
    else:
        base = root
    return base / "合同_水电约定.xlsx"


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Extract 水电约定 + 每期支付明细 from lease contracts (GPT-5.5).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python extract_contract_clauses.py contracts/成都\n"
            "      # writes one global utility workbook + one folder payment workbook\n"
            "  python extract_contract_clauses.py contracts --max-files 5\n"
            "      # every project folder, limit 5 PDFs each (trial run)\n"
        ),
    )
    ap.add_argument("path", help="contracts root or one project folder")
    ap.add_argument("--max-files", type=int, default=0, help="max PDFs per folder (0 = all)")
    ap.add_argument(
        "--max-pages",
        type=int,
        default=6,
        help="utility-page limit and payment scan batch size (payment still scans every page)",
    )
    ap.add_argument("--skip-utility", action="store_true", help="skip 水电约定 workbook")
    ap.add_argument("--skip-payment", action="store_true", help="skip 每期支付明细 workbook")
    args = ap.parse_args()

    root = Path(args.path)
    if not root.exists():
        print(f"❌ Path not found: {root}")
        return 1

    folders = _project_folders(root)
    print(
        f"Model: GPT-5.5 (workbench)  |  utility pages <= {args.max_pages}"
        f"  |  payment: scan every page in batches of <= {min(4, max(2, args.max_pages))}"
    )
    print(f"Project folder(s) with top-level PDFs: {len(folders)}")
    for f in folders:
        print(f"  - {f}  ({len(_collect_pdfs(f))} pdf(s))")
    print()

    try:
        client = AIClient(model_type=_DEFAULT_MODEL, agent_name="subagent_2", language="Chi")
    except Exception as exc:
        print(f"❌ Could not initialize AIClient: {exc}")
        return 1

    utility_by_folder: List[tuple] = []
    for folder in folders:
        pdfs = _collect_pdfs(folder)
        if args.max_files and args.max_files > 0:
            pdfs = pdfs[: args.max_files]
        if not pdfs:
            print(f"(skip) no PDFs in {folder}")
            continue

        print("=" * 78)
        print(f"FOLDER: {folder}  ({len(pdfs)} pdf(s))")
        print("=" * 78)

        utility_rows: List[tuple] = []
        payment_rows: List[tuple] = []
        fail = 0

        for i, pdf in enumerate(pdfs, 1):
            print(f"\n[{i}/{len(pdfs)}] {pdf.name}")
            if not args.skip_utility:
                try:
                    utility_rows.append(extract_utility(client, pdf, args.max_pages))
                    print("  ✅ 水电约定")
                except Exception as exc:
                    fail += 1
                    utility_rows.append([pdf.name] + [f"抽取失败: {exc}"] + [""] * 7)
                    print(f"  ❌ 水电约定: {exc}")
            if not args.skip_payment:
                try:
                    tables = extract_payment(client, pdf, args.max_pages)
                    payment_rows.append((pdf.name, tables))
                    row_count = sum(len(table.get("rows", [])) for table in tables)
                    print(f"  ✅ 支付明细 ({len(tables)} 表 / {row_count} 行)")
                except Exception as exc:
                    fail += 1
                    payment_rows.append((pdf.name, [{
                        "title": "抽取失败",
                        "unit": "",
                        "columns": ["错误"],
                        "rows": [[str(exc)]],
                        "source_page": "",
                    }]))
                    print(f"  ❌ 支付明细: {exc}")

        if not args.skip_utility:
            utility_by_folder.append((folder.name, utility_rows))
        if not args.skip_payment:
            out_p = folder / "合同_每期支付明细.xlsx"
            _write_payment_wb(payment_rows, out_p)
            print(f"Wrote {out_p}")
        print(f"FOLDER SUMMARY: {len(pdfs) - min(fail, len(pdfs))} ok / {fail} failed")

    if not args.skip_utility:
        out_u = _utility_output_path(root)
        _write_utility_wb(utility_by_folder, out_u)
        print(f"\nWrote global utility workbook: {out_u}")

    print("\nDone. Open the xlsx files above and paste back any rows that look wrong.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
