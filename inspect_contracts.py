#!/usr/bin/env python3
"""Structural inspection tool for a contracts folder -- the required first
step before building a GPT-5.5-based structured contract-data extraction
pipeline, matching this project's established "diagnose real structure
before writing extraction logic" discipline (inspect_databook.py,
inspect_ab_tabs_structure.py, etc.).

Answers exactly the two things needed before any extraction code can be
written, with zero AI calls and zero guessing:
  1. What's actually inside each contract folder -- file type per file,
     page count for PDFs, and (critically) whether each PDF page already
     has a REAL extractable text layer (a genuine digital PDF) or is
     image-only (a scan needing OCR/vision). This directly decides the
     extraction approach per file: a digital PDF needs no OCR at all (just
     extract text directly and hand it to GPT-5.5); an image-only PDF or a
     raw image file needs page images fed to GPT-5.5's vision input
     instead.
  2. What columns/structure the target summary template (an .xlsx file)
     actually has -- dumps sheet names, header rows, and merged-cell
     layout, so any future extraction prompt can target the REAL field
     names instead of a guess.

Fully deterministic -- no AI/OCR calls, just structural inspection. Safe to
run against any real contract folder without touching API budgets.

Usage:
    python inspect_contracts.py "/path/to/contracts/"
        # scans every subfolder + the template, if found automatically
    python inspect_contracts.py "/path/to/contracts/" --folder "SomeFolder"
        # only inspect one specific contract subfolder (try one at a time)
    python inspect_contracts.py "/path/to/contracts/" --template "path/to/template.xlsx"
        # explicit template path, if it's not found automatically
"""
import argparse
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# pypdf's crypto backend emits a CryptographyDeprecationWarning (a UserWarning
# subclass, not a DeprecationWarning -- so a category-based filter doesn't
# catch it) on import on some setups, unrelated to anything this script does.
# Suppressed at the import site so it doesn't clutter output the user is
# meant to paste back.
with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    try:
        from pypdf import PdfReader
    except ImportError:
        PdfReader = None


# Matched case-insensitively against the file STEM (name without extension)
# to auto-find the target summary template without the exact filename
# being hardcoded here -- works for either a Chinese or English name.
_TEMPLATE_NAME_HINTS = ("合同", "汇总", "匯總", "模板", "contract", "summary", "template")
_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".gif", ".webp"}
_MIN_TEXT_CHARS_PER_PAGE = 20  # below this, a PDF page is treated as image-only


def _hr(title: str = "") -> None:
    print("\n" + "=" * 78)
    if title:
        print(f"  {title}")
        print("=" * 78)


def find_template(root: Path) -> Optional[Path]:
    """Looks for a file matching this project's contract-template naming
    convention anywhere under root (not just the top level, since it might
    sit alongside or above the per-contract subfolders)."""
    candidates = []
    for path in root.rglob("*.xlsx"):
        name = path.stem
        if any(hint.lower() in name.lower() for hint in _TEMPLATE_NAME_HINTS):
            candidates.append(path)
    if not candidates:
        return None
    # Prefer the shallowest path (closest to root) if multiple match --
    # avoids accidentally picking a per-contract working copy over the
    # actual top-level template.
    candidates.sort(key=lambda p: len(p.parts))
    return candidates[0]


def inspect_template(template_path: Path) -> None:
    _hr(f"TEMPLATE STRUCTURE: {template_path.name}")
    if load_workbook is None:
        print("❌ openpyxl not available -- cannot inspect the template.")
        return
    try:
        wb = load_workbook(template_path, data_only=True)
    except Exception as exc:
        print(f"❌ Could not open template: {exc}")
        return
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet '{sheet_name}' ---")
        print(f"Dimensions: {ws.dimensions}  (max_row={ws.max_row}, max_col={ws.max_column})")
        merged = list(ws.merged_cells.ranges)
        if merged:
            preview = [str(r) for r in merged[:20]]
            print(f"Merged cell ranges ({len(merged)}): {preview}{' ...' if len(merged) > 20 else ''}")
        print("First 6 rows (raw cell values, first 15 columns):")
        for row_idx in range(1, min(7, ws.max_row + 1)):
            values = []
            for col_idx in range(1, min(16, ws.max_column + 1)):
                v = ws.cell(row=row_idx, column=col_idx).value
                values.append("" if v is None else str(v))
            print(f"  row {row_idx}: {values}")


def _pdf_text_profile(pdf_path: Path) -> str:
    """Returns a short classification: 'digital' (every page has a real
    text layer), 'image-only' (no page does), or 'mixed' -- decides the
    extraction approach per file (direct text extraction vs. feeding page
    images to GPT-5.5's vision input)."""
    if PdfReader is None:
        return "unknown (pypdf not installed -- run `pip install pypdf`)"
    try:
        reader = PdfReader(str(pdf_path))
    except Exception as exc:
        return f"unreadable ({exc})"
    if not reader.pages:
        return "no pages"
    text_pages = 0
    for page in reader.pages:
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        if len(text.strip()) >= _MIN_TEXT_CHARS_PER_PAGE:
            text_pages += 1
    total = len(reader.pages)
    if text_pages == total:
        return f"digital ({total} page(s), all have a real text layer -- no OCR needed)"
    if text_pages == 0:
        return f"image-only ({total} page(s), NO extractable text -- needs OCR/vision)"
    return f"mixed ({text_pages}/{total} page(s) have text -- check individually)"


def inspect_contract_folder(folder: Path) -> None:
    _hr(f"CONTRACT FOLDER: {folder.name}")
    files = sorted(p for p in folder.rglob("*") if p.is_file())
    if not files:
        print("  (empty folder)")
        return
    by_ext: Dict[str, List[Path]] = {}
    for f in files:
        by_ext.setdefault(f.suffix.lower(), []).append(f)
    counts = {(ext or "(no ext)"): len(v) for ext, v in sorted(by_ext.items())}
    print(f"  {len(files)} file(s) total, by type: {counts}")

    for f in files:
        rel = f.relative_to(folder)
        size_kb = f.stat().st_size / 1024
        ext = f.suffix.lower()
        if ext == ".pdf":
            profile = _pdf_text_profile(f)
            print(f"    [PDF]   {rel}  ({size_kb:.0f} KB)  -> {profile}")
        elif ext in _IMAGE_EXTS:
            print(f"    [IMAGE] {rel}  ({size_kb:.0f} KB)  -> needs OCR/vision directly (no text layer possible)")
        elif ext in (".doc", ".docx"):
            print(f"    [WORD]  {rel}  ({size_kb:.0f} KB)  -> not yet handled by this tool, flag if this appears")
        elif ext in (".xls", ".xlsx"):
            print(f"    [EXCEL] {rel}  ({size_kb:.0f} KB)  -> likely a working file, not a scanned contract; check manually")
        else:
            print(f"    [?]     {rel}  ({size_kb:.0f} KB)  -> unrecognized type, check manually")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the contracts root folder")
    ap.add_argument("--folder", default=None,
                     help="only inspect this one subfolder (by name) inside the contracts root, "
                          "instead of every subfolder")
    ap.add_argument("--template", default=None,
                     help="explicit path to the target .xlsx template, if auto-detection "
                          "(searching for a filename containing 合同/汇总/匯總/模板/contract/summary/template) "
                          "doesn't find it")
    args = ap.parse_args()

    root = Path(args.path)
    if not root.exists():
        print(f"❌ Path not found: {root}")
        return 1

    if PdfReader is None:
        print("⚠️  pypdf is not installed -- PDF text-layer detection will be skipped "
              "(run `pip install pypdf` first for full results).\n")

    template_path = Path(args.template) if args.template else find_template(root)
    if template_path and template_path.exists():
        inspect_template(template_path)
    else:
        print("\n⚠️  No template .xlsx found automatically (looked for a filename containing "
              "合同/汇总/匯總/模板/contract/summary/template). Pass --template explicitly if it "
              "has a different name.")

    all_subfolders = sorted(p for p in root.iterdir() if p.is_dir())
    subfolders = all_subfolders
    if args.folder:
        subfolders = [p for p in all_subfolders if p.name == args.folder]
        if not subfolders:
            print(f"\n❌ Folder {args.folder!r} not found under {root}. "
                  f"Available: {[p.name for p in all_subfolders]}")
            return 1

    if not subfolders:
        print(f"\n⚠️  No subfolders found directly under {root} -- if contracts sit as loose files "
              f"directly in this folder instead of per-contract subfolders, this tool doesn't yet "
              f"handle that layout; paste back what `dir` shows so it can be adjusted.")
        return 0

    print(f"\nFound {len(subfolders)} contract folder(s)"
          + (f" (showing only {args.folder!r})" if args.folder else "") + ":")
    for folder in subfolders:
        inspect_contract_folder(folder)

    _hr("SUMMARY")
    print(f"{len(subfolders)} contract folder(s) inspected.")
    print(
        "Next step once you've seen this output: paste it back. The PDF text-layer "
        "classification above (digital / image-only / mixed) decides the extraction "
        "approach per file -- digital PDFs can go straight to GPT-5.5 as text, "
        "image-only PDFs/raw images need page images fed to GPT-5.5's vision input "
        "instead. The template structure above becomes the exact target schema for "
        "the extraction prompt."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
