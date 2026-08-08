#!/usr/bin/env python3
"""Hand-checkable audit trail for ONE bridge transition on ONE AB- tab.

Everything else in this repo either prints the FINAL bridge numbers
(generate_bridge_waterfall_batch) or the raw structure
(inspect_ab_tabs_structure) -- neither lets you follow a single number
from the source cells all the way to the bar in the chart. This does:
for one phase and one year-transition it prints, in order,

  1. WHICH rows were picked for revenue/area/days, and why (the label
     text that matched), so a wrong row shows up immediately;
  2. WHICH columns belong to each year, as Excel column letters, so you
     can select the same range in Excel and compare;
  3. EVERY monthly cell that goes into each aggregate, with the running
     arithmetic spelled out (sum for revenue/days, average-over-nonzero
     for area) -- these are the numbers to tie out by hand;
  4. The unit-rent derivation;
  5. The three factor effects with the formula and the substituted
     values written out, not just the result;
  6. The reconciliation: start total + all factors vs. the end total,
     and the residual.

Usage:
    python verify_bridge_numbers.py "databooks/xx.xlsx" --tab AB-CD
        # audits the LAST transition (usually the LTM one) by default
    python verify_bridge_numbers.py "databooks/xx.xlsx" --tab AB-CD --transition 0
        # audit the first transition instead
    python verify_bridge_numbers.py "databooks/xx.xlsx" --tab AB-CD --phase 干仓
        # restrict the per-cell dump to one phase (default: all of them)
"""

# moved into ad-hoc/ -- put the repo root back on sys.path so
# `import fdd_utils...` still resolves when run from anywhere.
import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parents[2]))
import argparse
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, ".")
from fdd_utils.extract_bridge_from_raw import (
    find_phase_blocks, _year_col_map, _phase_start_col, _month_col_map,
)
from fdd_utils.generate_bridge_waterfall_batch import (
    build_bridges_for_ab_tab, find_year_days_rows, decompose_transition,
)
from fdd_utils.inspect_ab_tabs_structure import find_labeled_rows


def _label_at(labeled, row, col):
    for c, text, _cat in labeled.get(row, []):
        if c == col:
            return text
    for c, text, _cat in labeled.get(row, []):
        return f"{text} (col {get_column_letter(c)})"
    return "(no label found)"


def _dump_cells(ws, row, cols, mode, label):
    """Prints every contributing cell and the running arithmetic. mode is
    'sum' (revenue/days) or 'avg_nonzero' (area -- matches
    extract_annual_series' own AVERAGEIF(...,">0") convention)."""
    if row is None:
        print(f"      {label}: NO ROW FOUND for this phase -- treated as 0")
        return 0.0
    parts, total, used = [], 0.0, 0
    for c in cols:
        v = ws.cell(row=row, column=c).value
        if not isinstance(v, (int, float)):
            continue
        if mode == "avg_nonzero" and v <= 0:
            continue
        parts.append(f"{get_column_letter(c)}{row}={v:,.2f}")
        total += v
        used += 1
    print(f"      {label}: row {row}, {used} contributing cell(s)")
    for i in range(0, len(parts), 4):
        print("        " + "  ".join(parts[i:i + 4]))
    if mode == "sum":
        print(f"        => SUM = {total:,.2f}")
        return total
    if used == 0:
        print("        => no cell > 0, AVERAGE = 0")
        return 0.0
    print(f"        => AVERAGE over {used} cell(s) with value>0 = {total:,.2f} / {used} = {total / used:,.4f}")
    return total / used


def _silent_recompute(ws, block, cols, days_row, start_col):
    """Same aggregation the [3] section prints, without printing -- used by
    --audit-all to re-derive every figure independently and compare."""
    rev = 0.0
    for c in cols:
        v = ws.cell(row=block.revenue_row, column=c).value if block.revenue_row else None
        if isinstance(v, (int, float)):
            rev += v
    area_vals = []
    for c in cols:
        v = ws.cell(row=block.area_row, column=c).value if block.area_row else None
        if isinstance(v, (int, float)) and v > 0:
            area_vals.append(v)
    area = sum(area_vals) / len(area_vals) if area_vals else 0.0
    days = 0.0
    for c in cols:
        if start_col is None or c < start_col:
            continue
        v = ws.cell(row=days_row, column=c).value
        if isinstance(v, (int, float)):
            days += v
    return rev, area, days


def audit_all(wb, path: str, residual_pct_limit: float) -> int:
    """Runs every tab x every transition through the same independent
    re-computation the detailed mode prints, but reports only the verdicts
    -- so all 17 entities can be checked at once instead of reading 17
    separate full dumps. Two independent checks per transition:
    (a) does a fresh re-derivation of revenue/area/days from the raw cells
    match what the extractor produced, and (b) does start + every factor
    reconcile to the end total."""
    ab_tabs = [s for s in wb.sheetnames if s.startswith("AB")]
    print(f"Auditing {len(ab_tabs)} AB- tab(s) in {path!r}\n")
    print(f"{'tab':16s} {'transition':34s} {'recompute':11s} {'residual':>11s}  verdict")
    print("-" * 92)

    problems, checked = [], 0
    for tab in ab_tabs:
        ws = wb[tab]
        blocks, results = build_bridges_for_ab_tab(ws, tab, log=lambda m: None)
        if blocks is None or not results:
            print(f"{tab:16s} {'(no transitions produced)':34s} {'-':11s} {'-':>11s}  ⚠️ skipped")
            problems.append((tab, "produced no bridge transitions"))
            continue

        yd = find_year_days_rows(ws)
        year_row, days_row = yd["year_row"], yd["days_row"]
        max_col = ws.max_column
        year_cols = _year_col_map(ws, year_row, max_col)

        month_row = None
        labeled = find_labeled_rows(ws)
        for r in sorted(labeled):
            if any(cat == "period_month" for _, _, cat in labeled[r]):
                month_row = r
                break

        for res in results:
            checked += 1
            arrow = f"{res.year_a} -> {'LTM' if res.is_ltm else res.year_b}"
            # Rebuild both sides' column sets exactly as the extractor did.
            cols_a = year_cols.get(res.year_a, [])
            if res.is_ltm and month_row:
                ym = _month_col_map(ws, year_row, month_row, max_col)
                keys, y, m = [], res.year_b, res.latest_month
                for _ in range(12):
                    keys.append((y, m))
                    m -= 1
                    if m == 0:
                        m, y = 12, y - 1
                keys.reverse()
                cols_b = [ym[k] for k in keys if k in ym]
            else:
                cols_b = year_cols.get(res.year_b, [])

            mismatches = []
            for block, sa, sb in zip(blocks, res.series_a, res.series_b):
                start_col = _phase_start_col(ws, block, max_col)
                for side, cols, s in (("A", cols_a, sa), ("B", cols_b, sb)):
                    rev, area, days = _silent_recompute(ws, block, cols, days_row, start_col)
                    if (abs(rev - s["revenue_k"] * 1000) > 1.0 or abs(area - s["area"]) > 0.01
                            or abs(days - s["days"]) > 0.5):
                        mismatches.append(f"{block.label}/{side}")

            items = res.bridge.items
            reconstructed = items[0].value + sum(it.value for it in items[1:-1])
            end_v = items[-1].value
            resid_pct = ((reconstructed - end_v) / end_v * 100) if end_v else 0.0

            recompute_s = "✅ match" if not mismatches else f"❌ {len(mismatches)}"
            resid_ok = abs(resid_pct) <= residual_pct_limit
            verdict = "✅" if (not mismatches and resid_ok) else "❌ CHECK"
            print(f"{tab:16s} {arrow:34s} {recompute_s:11s} {resid_pct:>+10.3f}%  {verdict}")
            if mismatches:
                problems.append((tab, f"{arrow}: recompute mismatch on {', '.join(mismatches)}"))
            if not resid_ok:
                problems.append((tab, f"{arrow}: residual {resid_pct:+.3f}% exceeds "
                                      f"±{residual_pct_limit}%"))

    print("-" * 92)
    print(f"\n{checked} transition(s) checked across {len(ab_tabs)} tab(s).")
    if not problems:
        print("✅ No anomalies: every figure re-derives from the raw cells, and every")
        print("   bridge reconciles start + factors to its end total.")
        return 0
    print(f"\n❌ {len(problems)} item(s) need a closer look "
          f"(re-run with --tab <name> for the full cell-level trail):")
    for tab, msg in problems:
        print(f"  - {tab}: {msg}")
    return 1


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the operational-report .xlsx")
    ap.add_argument("--tab", default=None, help="the AB- tab to audit, e.g. AB-CD")
    ap.add_argument("--audit-all", action="store_true",
                     help="check EVERY tab and transition at once, printing only the verdicts "
                          "(independent re-computation + reconciliation) instead of the full "
                          "cell-level trail -- use this to sweep all entities, then drill into "
                          "anything it flags with --tab")
    ap.add_argument("--residual-limit", type=float, default=2.0,
                     help="residual %% of the end total above which a transition is flagged "
                          "(default 2.0; a small residual is expected, see the note in --tab mode)")
    ap.add_argument("--transition", type=int, default=-1,
                     help="which transition to audit (0=first, -1=last/LTM, the default)")
    ap.add_argument("--phase", default=None, help="only dump cells for this phase label")
    args = ap.parse_args()

    print(f"Loading {args.path!r}...")
    wb = load_workbook(args.path, data_only=True)

    if args.audit_all:
        return audit_all(wb, args.path, args.residual_limit)

    if not args.tab:
        print("❌ provide --tab <name>, or --audit-all to sweep every tab at once.")
        return 1
    if args.tab not in wb.sheetnames:
        print(f"❌ tab {args.tab!r} not found. Available: {wb.sheetnames}")
        return 1
    ws = wb[args.tab]

    blocks, results = build_bridges_for_ab_tab(ws, args.tab, log=lambda m: None)
    if blocks is None or not results:
        print(f"❌ {args.tab!r} produced no bridge transitions.")
        return 1

    yd = find_year_days_rows(ws)
    year_row, days_row = yd["year_row"], yd["days_row"]
    labeled = find_labeled_rows(ws)
    max_col = ws.max_column
    year_cols = _year_col_map(ws, year_row, max_col)

    res = results[args.transition]

    print("\n" + "=" * 78)
    print(f"AUDIT TRAIL -- {args.tab}, transition {args.transition} of {len(results)}")
    print(f"  chart title : {res.title}")
    print(f"  from        : {res.start_label}")
    print(f"  to          : {res.end_label}")
    print(f"  LTM window? : {res.is_ltm}")
    print("=" * 78)

    print(f"\n[1] ANCHOR ROWS (found by label text, not by position)")
    print(f"  Year row  : {year_row}  <- {_label_at(labeled, year_row, None)!r}")
    print(f"  Days row  : {days_row}  <- {_label_at(labeled, days_row, None)!r}")
    print(f"\n  Phase blocks detected ({len(blocks)}):")
    for b in blocks:
        print(f"    [{b.label}] occupancy row {b.occupancy_row}, area row {b.area_row}, revenue row {b.revenue_row}")
        if b.area_row:
            print(f"        area row label   : {_label_at(labeled, b.area_row, None)!r}")
        if b.revenue_row:
            print(f"        revenue row label: {_label_at(labeled, b.revenue_row, None)!r}")

    print(f"\n[2] PERIOD COLUMNS")
    if res.is_ltm:
        month_row = None
        for r in sorted(labeled):
            if any(cat == "period_month" for _, _, cat in labeled[r]):
                month_row = r
                break
        ym = _month_col_map(ws, year_row, month_row, max_col)
        keys, y, m = [], res.year_b, res.latest_month
        for _ in range(12):
            keys.append((y, m))
            m -= 1
            if m == 0:
                m, y = 12, y - 1
        keys.reverse()
        cols_b = [ym[k] for k in keys if k in ym]
        print(f"  side A = full year {res.year_a}: columns "
              f"{get_column_letter(year_cols[res.year_a][0])}..{get_column_letter(year_cols[res.year_a][-1])} "
              f"({len(year_cols[res.year_a])} cols)")
        print(f"  side B = LTM 12 months ending {res.year_b}-{res.latest_month:02d}: "
              f"{get_column_letter(cols_b[0])}..{get_column_letter(cols_b[-1])} ({len(cols_b)} cols)")
        print(f"    months: {', '.join(f'{yy}-{mm:02d}' for yy, mm in keys)}")
        cols_a = year_cols[res.year_a]
    else:
        cols_a, cols_b = year_cols[res.year_a], year_cols[res.year_b]
        for lbl, yr, cc in (("A", res.year_a, cols_a), ("B", res.year_b, cols_b)):
            print(f"  side {lbl} = year {yr}: columns "
                  f"{get_column_letter(cc[0])}..{get_column_letter(cc[-1])} ({len(cc)} cols)")

    print(f"\n[3] PER-PHASE CELL-BY-CELL AGGREGATES")
    print("    (tie these out against the same ranges in Excel)")
    for block, sa, sb in zip(blocks, res.series_a, res.series_b):
        if args.phase and block.label != args.phase:
            continue
        start_col = _phase_start_col(ws, block, max_col)
        print(f"\n  --- phase [{block.label}] ---")
        print(f"    this phase's first-ever nonzero column = "
              f"{get_column_letter(start_col) if start_col else 'NONE'} "
              f"(days are summed from here onward only)")
        for side, cols, s in (("A / " + res.start_label, cols_a, sa),
                               ("B / " + res.end_label, cols_b, sb)):
            print(f"\n    [{side}]")
            rev = _dump_cells(ws, block.revenue_row, cols, "sum", "revenue")
            area = _dump_cells(ws, block.area_row, cols, "avg_nonzero", "area")
            day_cols = [c for c in cols if start_col is not None and c >= start_col]
            days = _dump_cells(ws, days_row, day_cols, "sum", "days")
            print(f"      unit rent = revenue / area / days"
                  f" = {rev:,.2f} / {area:,.4f} / {days:,.0f} = "
                  f"{(rev / area / days) if (area and days) else 0:.6f}")
            print(f"      -> matches extractor? revenue_k={s['revenue_k']:,.2f} "
                  f"(raw {s['revenue_k'] * 1000:,.2f}), area={s['area']:,.4f}, "
                  f"days={s['days']:,.0f}, unit_rent={s['unit_rent']:.6f}")
            ok = (abs(rev - s["revenue_k"] * 1000) < 1.0 and abs(area - s["area"]) < 0.01
                  and abs(days - s["days"]) < 0.5)
            print(f"      {'✅ hand-recomputed values MATCH the extractor' if ok else '❌ MISMATCH -- investigate'}")

    print(f"\n[4] FACTOR DECOMPOSITION (formula with values substituted)")
    for block, sa, sb in zip(blocks, res.series_a, res.series_b):
        if args.phase and block.label != args.phase:
            continue
        pa, pb = sa["unit_rent"], sb["unit_rent"]
        aa, ab_ = sa["area"], sb["area"]
        da, db = sa["days"], sb["days"]
        print(f"\n  [{block.label}]")
        print(f"    price effect = (price_B - price_A) * area_A * days_A / 1000")
        print(f"                 = ({pb:.6f} - {pa:.6f}) * {aa:,.2f} * {da:,.0f} / 1000"
              f" = {(pb - pa) * aa * da / 1000:,.2f}k")
        print(f"    area  effect = price_B * (area_B - area_A) * days_A / 1000")
        print(f"                 = {pb:.6f} * ({ab_:,.2f} - {aa:,.2f}) * {da:,.0f} / 1000"
              f" = {pb * (ab_ - aa) * da / 1000:,.2f}k")
        print(f"    days  effect = price_B * area_B * (days_B - days_A) / 1000")
        print(f"                 = {pb:.6f} * {ab_:,.2f} * ({db:,.0f} - {da:,.0f}) / 1000"
              f" = {pb * ab_ * (db - da) / 1000:,.2f}k")

    print(f"\n[5] RECONCILIATION (what the chart's bars must add up to)")
    items = res.bridge.items
    start_v, end_v = items[0].value, items[-1].value
    print(f"  start   {items[0].label:38s} {start_v:>14,.2f}k")
    running = start_v
    for it in items[1:-1]:
        running += it.value
        print(f"    {'+' if it.value >= 0 else '-'} {it.label:36s} {it.value:>14,.2f}k   running {running:>14,.2f}k")
    print(f"  end     {items[-1].label:38s} {end_v:>14,.2f}k  (actual)")
    resid = running - end_v
    print(f"\n  reconstructed end = {running:,.2f}k, actual end = {end_v:,.2f}k, residual = {resid:,.2f}k")
    pct = (resid / end_v * 100) if end_v else 0.0
    print(f"  residual as % of end total = {pct:+.3f}%")
    print("\n  NOTE: a small residual is EXPECTED and is not a bug -- AB-CD's own")
    print("  notes state revenue is all-in while leased area excludes pallet-priced")
    print("  space, so price*area*days can never reconstruct revenue perfectly.")
    print(f"  Extractor's own check flag: {'✅ within tolerance' if res.bridge.check_ok else '⚠️ ABOVE tolerance'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
