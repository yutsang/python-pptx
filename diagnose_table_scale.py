#!/usr/bin/env python3
"""Why does a presentation-detail table show numbers ~1000x too large?

Real symptom: 管理费用's 合计 renders as 877,709 in the new PPTX table, but
the real reference deck (photo-confirmed) shows 878 for the same cell.
23,072 (营业成本, real) vs 23,071,747 (rendered) -- same ~1000x pattern on
every account checked so far.

Checks three things independently so the exact stage is visible:
  1. df.attrs["presentation_detail_table"]["rows"] -- what the table
     extractor stored (this is what pptx.py's renderer reads verbatim,
     with no further scaling of its own).
  2. df.attrs["source_multiplier"] -- what multiplier normalize_financial_
     schedule decided for the account's MAIN block.
  3. The RAW Excel cell value + number_format at the presentation table's
     own header_row/label_col coordinates, read directly via openpyxl,
     bypassing pandas/_coerce_numeric entirely -- this is the ground truth:
     if number_format has a trailing comma (e.g. "#,##0,"), the cell's
     TRUE stored value is the full raw figure and Excel only DIVIDES BY
     1000 FOR DISPLAY, meaning the true value was never in thousands at
     all despite looking like it in Excel.

Read-only.

Usage:
    python diagnose_table_scale.py "for_test/Crescent-databook.xlsx" --account 管理费用
"""
import argparse
import sys
import warnings

warnings.filterwarnings("ignore")
sys.path.insert(0, ".")

import openpyxl

from fdd_utils.workbook import process_workbook_data


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path")
    ap.add_argument("--account", required=True)
    ap.add_argument("--entity", default="x")
    args = ap.parse_args()

    result = process_workbook_data(temp_path=args.path, entity_name=args.entity, selected_sheet=None)
    dfs = result["dfs"]
    df = dfs.get(args.account)
    if df is None:
        print(f"'{args.account}' not found. Available: {sorted(dfs.keys())}")
        return 1

    attrs = df.attrs or {}
    table = attrs.get("presentation_detail_table") or {}
    sheet_name = attrs.get("source_sheet_name") or (attrs.get("integrity") or {}).get("sheet_name")
    source_multiplier = attrs.get("source_multiplier")

    print("=" * 78)
    print(f"{args.account}  (sheet_name={sheet_name!r}, source_multiplier={source_multiplier!r})")
    print("=" * 78)

    print("\n--- STAGE 1: what's stored in presentation_detail_table (what pptx.py reads) ---")
    if not table.get("rows"):
        print("  No rows -- nothing to compare.")
        return 1
    print(f"  header_row={table.get('header_row')}  label_col={table.get('label_col')}  "
          f"periods={table.get('periods')}")
    for row in table["rows"][:5]:
        print(f"  {row['label']:20s} {row['values']}")
        for child in (row.get("children") or [])[:3]:
            print(f"      -> {child['label']:16s} {child['values']}")
    if table.get("total_row"):
        print(f"  {table['total_row']['label']:20s} {table['total_row']['values']}  <-- TOTAL")

    print("\n--- STAGE 2: raw Excel cell values at these exact coordinates (bypasses pandas) ---")
    if sheet_name is None:
        print("  attrs['sheet_name'] is None -- can't open the raw sheet directly.")
        return 1
    try:
        wb = openpyxl.load_workbook(args.path, data_only=True)
    except Exception as exc:
        print(f"  Could not open workbook with openpyxl: {exc}")
        return 1
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet {sheet_name!r} not found. Available: {wb.sheetnames[:20]}")
        return 1
    ws = wb[sheet_name]
    header_row = table["header_row"]  # 0-indexed (pandas convention)
    label_col = table["label_col"]    # 0-indexed
    excel_header_row = header_row + 1
    excel_label_col = label_col + 1
    print(f"  Sheet {sheet_name!r}, header at Excel row {excel_header_row}, "
          f"label col index {label_col} (openpyxl col {excel_label_col})")
    for offset in range(1, 6):
        r = excel_header_row + offset
        label_cell = ws.cell(row=r, column=excel_label_col)
        label = str(label_cell.value or "").strip()
        if not label:
            continue
        print(f"  row {r}: label={label!r}")
        for c_offset in range(1, 5):
            c = excel_label_col + c_offset
            cell = ws.cell(row=r, column=c)
            print(f"      col {c}: raw_value={cell.value!r}  number_format={cell.number_format!r}")

    print("\n--- STAGE 3: does number_format explain a 1000x display-only trick? ---")
    print("  If number_format above ends in a comma (e.g. '#,##0,' or '#,##0.0,'),")
    print("  Excel DIVIDES BY 1000 FOR DISPLAY ONLY -- the raw_value shown above is the")
    print("  TRUE stored number, and what a human sees in Excel (thousands-look) was never")
    print("  actually stored that way. That would mean presentation_detail_table's values")
    print("  need a /1000 correction that isn't currently applied anywhere.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
