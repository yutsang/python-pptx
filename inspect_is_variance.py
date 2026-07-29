#!/usr/bin/env python3
"""Measures the gap between what an IS account's DATA warrants and what
the current prompt rules ALLOW the commentary to say.

Context: the current pipeline applies a FIXED sentence/word cap per
account type (prompts.yml, 2_Auditor) regardless of whether that account
was flat and boring or swung several hundred percent -- and separately
tells the model to DELETE any driver not explicitly stated in the data or
remarks. Together those two rules mean a large, genuinely interesting
movement gets the same terse composition-plus-figures treatment as a
stable one. Nothing currently measures how often that actually bites.

For every IS account this prints:
  * period-over-period movement on the account's own total row, with
    partial (annualised) periods marked, since a 1-month tail period
    otherwise looks like a catastrophic decline against a full year;
  * whether remarks/notes exist that COULD explain the movement;
  * the sentence/word cap that account type will hit;
  * a verdict flagging accounts where a large movement has explanatory
    material available but the cap leaves no room to use it.

It also computes the revenue-vs-expense growth asymmetry across the
whole statement -- an observation no single account's commentary can
currently make, because each account is generated in isolation from its
own DataFrame with no visibility of any other account.

Read-only: loads and analyses, never writes to the databook.

Usage:
    python inspect_is_variance.py "for_test/xxx.xlsx" --entity "Name"
    python inspect_is_variance.py "for_test/xxx.xlsx" --entity "Name" --account "G&A expenses"
"""
import argparse
import sys
import warnings

warnings.filterwarnings("ignore")
sys.path.insert(0, ".")

from fdd_utils.workbook import process_workbook_data, INTERNAL_ROW_KEY


# Mirrors the caps in prompts.yml 2_Auditor (Eng + Chi). Kept as data here
# purely so this tool can REPORT which cap an account will hit -- it does
# not drive any generation.
# Needles must cover BOTH languages: a Chinese databook's mapping keys are
# Chinese (营业收入 / 管理费用 / ...), and an English-only table silently
# matched nothing on them -- every account fell through to "no explicit
# tier", so nothing was ever classified as capped and a real 19-workbook
# sweep reported 0 flagged accounts purely as a matching artefact.
# Ordered MOST-SPECIFIC FIRST: 所得税费用 and 营业外收入 both contain 收入,
# and "income tax expense" contains "income", so a looser revenue rule
# placed earlier would swallow them.
_CAP_TIERS = [
    (("financial expense", "g&a", "general and admin", "s&d", "selling",
      "income tax", "non-operating", "tax and surcharge", "taxes and surcharge",
      "财务费用", "管理费用", "销售费用", "所得税费用", "税金及附加",
      "营业外收入", "营业外支出", "其他收益"),
     "1-3 sentences / 30-80 words (Chi 2-3 句 / 60-130 字)  <-- tightest tier"),
    (("operating income", "revenue", "cogs", "operating cost",
      "营业收入", "营业成本"),
     "2-3 sentences / 60-100 words (Chi 3-5 句 / 100-180 字)"),
    (("investment propert", "other payable", "投资性房地产", "其他应付款"),
     "4-7 sentences / 100-200 words (Chi 4-7 句 / 150-280 字)"),
    (("cash", "receivable", "prepayment", "oci", "reserve", "dta", "ncl",
      "货币资金", "应收账款", "预付款项", "其他综合收益", "盈余公积"),
     "1-3 sentences / 25-80 words (Chi 2-3 句 / 40-90 字)"),
]
_TIGHTEST = _CAP_TIERS[0][1]

# Revenue, for the expense-vs-revenue asymmetry check. Same bilingual
# problem: an English-only match reported "revenue growth n/a" on every
# Chinese workbook. 营业外收入 (non-operating income) is NOT revenue.
_REVENUE_NEEDLES = ("operating income", "revenue", "营业收入")
_NOT_REVENUE_NEEDLES = ("non-operating", "营业外", "cost", "成本")


def _cap_for(account: str) -> str:
    low = account.lower()
    for needles, cap in _CAP_TIERS:
        if any(n in low for n in needles):
            return cap
    return "(no explicit tier -- falls back to general rules)"


def _is_revenue(account: str) -> bool:
    low = account.lower()
    if any(n in low for n in _NOT_REVENUE_NEEDLES):
        return False
    return any(n in low for n in _REVENUE_NEEDLES)


# Strings that occupy a note slot without carrying any explanatory content.
# Confirmed on a real healthy workbook: accounts with no genuine remarks
# still report exactly one supporting_note whose entire text is "Check" -- a
# worksheet tie-out label, not commentary. Counting those as "explanatory
# material available" makes an account look like it has context to draw on
# when it has none, which would argue for loosening its length cap when the
# correct conclusion is the opposite: there is nothing there to say.
_NON_SUBSTANTIVE_NOTES = {"check", "checks", "n/a", "na", "-", "tbc", "tba", "ok", "note", "notes"}


def _is_substantive_note(text) -> bool:
    s = str(text or "").strip()
    if not s:
        return False
    if s.lower() in _NON_SUBSTANTIVE_NOTES:
        return False
    return len(s) >= 4


def _substantive_notes(notes) -> list:
    return [n for n in (notes or []) if _is_substantive_note(n)]


def _substantive_linked(linked) -> list:
    """table_linked_remarks entries are dicts like
    {'source': 'row_note', 'summary': 'Check'} -- judge them on the summary."""
    out = []
    for item in linked or []:
        summary = item.get("summary") if isinstance(item, dict) else item
        if _is_substantive_note(summary):
            out.append(item)
    return out


def _total_row_values(df):
    """(period_label, value) pairs for the account's total row -- the same
    row build_trend_summary focuses on. Excludes INTERNAL_ROW_KEY, which
    is a bookkeeping column (the sheet row index), not a period."""
    period_cols = [c for c in df.columns[1:]
                   if str(c) != INTERNAL_ROW_KEY and not str(c).endswith("_formatted")]
    row_types = df.attrs.get("row_types_by_description") or {}
    desc_col = df.columns[0]
    total_idx = None
    for idx, row in df.iterrows():
        if str(row_types.get(str(row[desc_col]), "")).lower() in ("total", "subtotal"):
            total_idx = idx
    if total_idx is None:
        # No labelled total -- sum the detail rows instead.
        return [(str(c), float(df[c].fillna(0).sum())) for c in period_cols]
    return [(str(c), float(df.loc[total_idx, c] or 0)) for c in period_cols]


def _pct(prev, curr, scale=None):
    """Percentage move, or None when the base is too small for one to mean
    anything. A real workbook produced '财务费用 +31,838,208%' -- arithmetically
    correct, but it only says the prior period was a rounding-error residue,
    not that the account moved 30 million percent. Anything under 1% of the
    account's own largest period is treated as a negligible base, so these
    don't dominate the ranking and crowd out genuine movements."""
    if prev == 0:
        return None
    if scale and abs(prev) < abs(scale) * 0.01:
        return None
    return (curr - prev) / abs(prev) * 100


def _series_scale(series):
    """Largest absolute period value -- the yardstick for whether a given
    period's value is a meaningful base to measure a move against."""
    return max((abs(v) for _p, v in series), default=0.0)


def _fmt_pct(p):
    return "n/a (negligible base)" if p is None else f"{p:+,.1f}%"


def _entity_from_filename(path: str) -> str:
    """Best-effort entity name from a databook filename. Real files are named
    like 'Project Mint.Portfolio I.<entity>.xlsx' or
    'Project Mint.Portfolio I.databook. <entity>.xlsx', so the last
    dot-separated component is the entity. entity_name is only a soft filter
    for per-entity files (verified: passing a wrong name still extracts every
    account), but it does matter on a roll-up/master workbook where several
    entities' blocks share one sheet -- pass --entity explicitly for those."""
    import os
    stem = os.path.splitext(os.path.basename(path))[0]
    last = stem.split(".")[-1].strip()
    return last or stem


def analyse_one(path: str, entity: str, threshold: float, sheet=None,
                 financials_from=None, financials_sheet=None):
    """Runs the variance analysis for one databook and returns structured
    results, so the single-file (detailed) and batch (aggregate) modes share
    exactly one implementation of the measurement itself."""
    result = process_workbook_data(temp_path=path, entity_name=entity, selected_sheet=sheet,
                                    financials_from=financials_from,
                                    financials_sheet=financials_sheet)
    dfs = result["dfs"]
    out = {"language": result.get("language"), "accounts": [], "revenue_growth": None}
    for key, df in dfs.items():
        integrity = df.attrs.get("integrity") or {}
        if str(integrity.get("statement_type", "")).upper() != "IS":
            continue
        months = integrity.get("annualization_months")
        series = _total_row_values(df)
        if len(series) < 2:
            continue
        full = series[:-1] if (months and 0 < months < 12) else series
        scale = _series_scale(full)
        biggest, from_nil = 0.0, False
        for (_p0, v0), (_p1, v1) in zip(full, full[1:]):
            p = _pct(v0, v1, scale)
            if p is not None and abs(p) >= threshold:
                biggest = max(biggest, abs(p))
            elif p is None and v1 != 0:
                from_nil = True
        notes = df.attrs.get("supporting_notes") or []
        rhs = df.attrs.get("adjacent_detail_rows") or []
        linked = df.attrs.get("table_linked_remarks") or []
        sub_notes = _substantive_notes(notes)
        sub_linked = _substantive_linked(linked)
        has_expl = bool(sub_notes or rhs or sub_linked)
        cap = _cap_for(key)
        latest_growth = _pct(full[-2][1], full[-1][1], scale) if len(full) >= 2 else None
        if _is_revenue(key):
            out["revenue_growth"] = latest_growth
        out["accounts"].append({
            "key": key, "series": series, "full": full, "months": months,
            "biggest": biggest, "from_nil": from_nil,
            "notes": len(notes), "rhs": len(rhs), "linked": len(linked),
            "sub_notes": len(sub_notes), "sub_linked": len(sub_linked),
            "has_expl": has_expl,
            "cap": cap, "tightest": cap == _TIGHTEST,
            "latest_growth": latest_growth,
            "flagged": biggest >= threshold and cap == _TIGHTEST and has_expl,
            # Big move, tightest cap, but nothing real to explain it with:
            # loosening the cap here would only create room to invent.
            "starved": biggest >= threshold and not has_expl,
        })
    return out


def run_batch(args) -> int:
    """Sweeps every .xlsx in a directory. The point is sample size: deciding
    how far to loosen a length cap from ONE entity risks tuning to an
    outlier, so this reports how often the cap actually binds across a whole
    portfolio, and how much explanatory material is sitting unused."""
    import glob
    import os

    paths = sorted(glob.glob(os.path.join(args.path, "*.xlsx")))
    paths = [p for p in paths if not os.path.basename(p).startswith("~$")]
    if not paths:
        print(f"❌ no .xlsx files found in {args.path!r} (for a single file, drop --batch)")
        return 1

    print(f"Sweeping {len(paths)} workbook(s) in {args.path!r}\n")
    all_flagged, all_starved, per_file, failed, remark_anomalies = [], [], [], [], []
    total_is = 0

    for p in paths:
        name = os.path.basename(p)
        entity = args.entity or _entity_from_filename(p)
        try:
            res = analyse_one(p, entity, args.threshold, sheet=args.sheet,
                               financials_from=args.financials_from,
                               financials_sheet=args.financials_sheet)
        except Exception as exc:
            print(f"--- {name}\n    ⚠️ skipped: {type(exc).__name__}: {str(exc)[:120]}")
            failed.append((name, f"{type(exc).__name__}: {str(exc)[:120]}"))
            continue
        accts = res["accounts"]
        if not accts:
            print(f"--- {name}\n    (no IS accounts -- BS-only file, or a roll-up needing --entity)")
            continue
        total_is += len(accts)
        flagged = [a for a in accts if a["flagged"]]
        all_flagged.extend((name, a) for a in flagged)
        per_file.append((name, len(accts), len(flagged), res["revenue_growth"]))
        rg = res["revenue_growth"]
        rg_s = "n/a" if rg is None else f"{rg:+,.1f}%"

        # A workbook where EVERY account reports the same remark counts is a
        # data-extraction signal, not a coincidence: a healthy databook shows
        # a wide spread (one real file ranged from 1/1/2 up to 3/20/23). All
        # accounts landing on an identical low count usually means the file's
        # remark columns aren't being picked up and the commentary is being
        # written from figures alone.
        starved = [a for a in accts if a["starved"]]
        all_starved.extend((name, a) for a in starved)
        no_material = sum(1 for a in accts if not a["has_expl"])
        if no_material == len(accts) and len(accts) >= 4:
            uniform = (f"   ⚠️ NONE of the {len(accts)} accounts has any substantive remark "
                       f"(only tie-out artefacts like 'Check')")
            remark_anomalies.append((name, len(accts), no_material))
        else:
            uniform = ""
        print(f"--- {name}  [entity={entity!r}, lang={res['language']}]")
        print(f"    {len(accts)} IS account(s), revenue growth {rg_s}, "
              f"{len(flagged)} flagged, {len(starved)} starved{uniform}")
        for a in sorted(accts, key=lambda x: -x["biggest"]):
            if a["biggest"] < args.threshold:
                continue
            if a["flagged"]:
                mark = "  <-- FLAGGED (tightest cap, real material available)"
            elif a["starved"]:
                mark = "  <-- STARVED (big move, NO real material to explain it)"
            else:
                mark = ""
            print(f"      {a['key']:32s} {a['biggest']:>9,.0f}%   "
                  f"raw={a['notes']}/{a['rhs']}/{a['linked']} "
                  f"substantive={a['sub_notes']}/{a['rhs']}/{a['sub_linked']}"
                  f"   {'TIGHT' if a['tightest'] else 'wider'} cap{mark}")
        print()

    print("=" * 88)
    print("AGGREGATE")
    print("=" * 88)
    print(f"  workbooks analysed        : {len(per_file)}")
    if failed:
        print(f"  workbooks skipped (error) : {len(failed)}")
    print(f"  IS accounts total         : {total_is}")
    print(f"  FLAGGED (big move, tightest cap, real material available): {len(all_flagged)}")
    print(f"  STARVED (big move, but NO substantive material at all)   : {len(all_starved)}")
    if total_is:
        print(f"\n  => {len(all_flagged) / total_is * 100:.1f}% would genuinely benefit from a looser cap")
        print(f"  => {len(all_starved) / total_is * 100:.1f}% moved materially with nothing on file to explain it")
        if len(all_starved) > len(all_flagged):
            print("\n  Loosening the caps is NOT the binding constraint here: most big movers")
            print("  have no explanatory material at all, so extra room would only create")
            print("  space to invent. Getting real remarks into these files comes first.")

    if all_flagged:
        print(f"\n  Flagged accounts, largest movement first:")
        for name, a in sorted(all_flagged, key=lambda t: -t[1]["biggest"]):
            print(f"    {a['biggest']:>9,.0f}%  {a['key']:30s}  "
                  f"material={a['notes']}/{a['rhs']}/{a['linked']}  ({name})")
        by_account = {}
        for _name, a in all_flagged:
            by_account.setdefault(a["key"], 0)
            by_account[a["key"]] += 1
        print(f"\n  Which account types hit the cap most often:")
        for key, n in sorted(by_account.items(), key=lambda kv: -kv[1]):
            print(f"    {n:>3d} x  {key}")
    if remark_anomalies:
        print(f"\n  ⚠️  DATABOOK ISSUE -- {len(remark_anomalies)} workbook(s) where NOT ONE IS account")
        print(f"      carries a substantive remark (only tie-out artefacts such as 'Check'):")
        for name, n_acct, _n in remark_anomalies:
            print(f"    {name}  ({n_acct} accounts, 0 with real remarks)")
        print(f"      A healthy workbook mixes both -- real notes like")
        print(f"      '折旧费 | 2023年11月开始计提固定资产折旧' alongside thin accounts.")
        print(f"      Commentary for these files is therefore written from figures alone.")
        print(f"      Either the remark columns aren't being read, or the project team")
        print(f"      hasn't filled them in -- worth checking one file in Excel before")
        print(f"      any prompt work, because no prompt change can compensate for this.")
    if failed:
        print(f"\n  Skipped:")
        for name, err in failed:
            print(f"    {name}: {err}")
    return 0


def dump_remarks(args) -> int:
    """Prints the actual TEXT of every remark the pipeline extracted per IS
    account, so a uniform remark count can be judged rather than guessed at.
    Counts alone are ambiguous: 1/0/1 is a perfectly normal value for a
    genuinely thin account (a real healthy workbook has it for 税金及附加 and
    营业外收入 while 管理费用 has 3/20/23) -- what makes it suspicious is EVERY
    account sharing it. If the single note also turns out to be the same
    boilerplate string for every account, that is conclusive: nothing
    account-specific is reaching the model."""
    entity = args.entity or _entity_from_filename(args.path)
    print(f"Loading {args.path!r} (entity={entity!r})...\n")
    result = process_workbook_data(temp_path=args.path, entity_name=entity,
                                    selected_sheet=args.sheet,
                                    financials_from=args.financials_from,
                                    financials_sheet=args.financials_sheet)
    dfs = result["dfs"]
    first_notes = {}
    for key in sorted(dfs):
        df = dfs[key]
        integrity = df.attrs.get("integrity") or {}
        if str(integrity.get("statement_type", "")).upper() != "IS":
            continue
        if args.account and key != args.account:
            continue
        notes = df.attrs.get("supporting_notes") or []
        rhs = df.attrs.get("adjacent_detail_rows") or []
        linked = df.attrs.get("table_linked_remarks") or []
        print("=" * 78)
        print(f"{key}   (source sheet: {df.attrs.get('source_sheet_name')!r})")
        print("=" * 78)
        print(f"  supporting_notes ({len(notes)}):")
        for i, n in enumerate(notes):
            print(f"    [{i}] {str(n)[:300]}")
        print(f"  adjacent_detail_rows ({len(rhs)}):")
        for i, rrow in enumerate(rhs[:6]):
            print(f"    [{i}] {str(rrow)[:300]}")
        if len(rhs) > 6:
            print(f"    ... and {len(rhs) - 6} more")
        print(f"  table_linked_remarks ({len(linked)}):")
        for i, lk in enumerate(linked[:6]):
            print(f"    [{i}] {str(lk)[:300]}")
        if len(linked) > 6:
            print(f"    ... and {len(linked) - 6} more")
        print()
        if notes:
            first_notes[key] = str(notes[0])[:200]

    if len(first_notes) >= 2:
        distinct = set(first_notes.values())
        print("=" * 78)
        print("VERDICT")
        print("=" * 78)
        if len(distinct) == 1:
            print(f"  ❌ All {len(first_notes)} accounts share the SAME first note:")
            print(f"     {next(iter(distinct))!r}")
            print("  That is boilerplate, not per-account context -- nothing account-specific")
            print("  is reaching the model, so its commentary can only restate the figures.")
        else:
            print(f"  ✅ {len(distinct)} distinct first-notes across {len(first_notes)} accounts --")
            print("     the notes ARE account-specific; a low count reflects genuinely thin")
            print("     source remarks rather than an extraction failure.")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="path to the databook .xlsx")
    ap.add_argument("--entity", default=None, help="entity name, as you'd type it in the app")
    ap.add_argument("--sheet", default=None, help="specific sheet, if the app asks you to pick one")
    ap.add_argument("--account", default=None, help="only report this one account")
    ap.add_argument("--threshold", type=float, default=50.0,
                     help="movement %% above which an account is called high-variance (default 50)")
    ap.add_argument("--list-sheets", action="store_true",
                     help="just list this workbook's sheets and exit -- use this first when you "
                          "don't know what to pass for --entity / --financials-sheet")
    ap.add_argument("--financials-from", default=None,
                     help="path to a separate roll-up/master workbook holding the Financials "
                          "sheet, when this entity's own file doesn't carry one")
    ap.add_argument("--financials-sheet", default=None,
                     help="name of the Financials sheet to source from (needed when it lives in "
                          "a master/roll-up sheet rather than a per-entity tab)")
    ap.add_argument("--batch", action="store_true",
                     help="treat `path` as a DIRECTORY and sweep every .xlsx in it, printing one "
                          "compact line per account plus an aggregate across all files -- use this "
                          "to see how often the cap actually bites across a whole portfolio "
                          "instead of judging from a single entity")
    ap.add_argument("--dump-remarks", action="store_true",
                     help="print the actual TEXT of every extracted remark per IS account, and "
                          "judge whether they're account-specific or one shared boilerplate -- "
                          "use this on any file --batch flags with a uniform remark count")
    args = ap.parse_args()

    if args.batch:
        return run_batch(args)

    if args.dump_remarks:
        return dump_remarks(args)

    if args.list_sheets:
        from openpyxl import load_workbook
        wb = load_workbook(args.path, read_only=True)
        print(f"{len(wb.sheetnames)} sheet(s) in {args.path!r}:")
        for name in wb.sheetnames:
            print(f"  {name}")
        return 0

    if not args.entity:
        args.entity = _entity_from_filename(args.path)
        print(f"(no --entity given; derived {args.entity!r} from the filename)")

    print(f"Loading {args.path!r} (entity={args.entity!r})...")
    result = process_workbook_data(temp_path=args.path, entity_name=args.entity,
                                    selected_sheet=args.sheet,
                                    financials_from=args.financials_from,
                                    financials_sheet=args.financials_sheet)
    dfs = result["dfs"]
    print(f"{len(dfs)} account(s) processed. Language detected: {result.get('language')}\n")

    is_accounts = {}
    for key, df in dfs.items():
        integrity = df.attrs.get("integrity") or {}
        if str(integrity.get("statement_type", "")).upper() == "IS":
            is_accounts[key] = df
    if not is_accounts:
        print("❌ No IS accounts found -- is this a BS-only databook, or did the entity name not match?")
        return 1
    print(f"{len(is_accounts)} income-statement account(s): {', '.join(sorted(is_accounts))}\n")

    flagged, revenue_growth, expense_growth = [], None, {}

    for key in sorted(is_accounts):
        if args.account and key != args.account:
            continue
        df = is_accounts[key]
        integrity = df.attrs.get("integrity") or {}
        months = integrity.get("annualization_months")
        series = _total_row_values(df)
        if len(series) < 2:
            continue

        notes = df.attrs.get("supporting_notes") or []
        rhs = df.attrs.get("adjacent_detail_rows") or []
        linked = df.attrs.get("table_linked_remarks") or []
        sub_notes = _substantive_notes(notes)
        sub_linked = _substantive_linked(linked)
        has_expl = bool(sub_notes or rhs or sub_linked)

        print("=" * 78)
        print(f"{key}")
        print("=" * 78)
        print(f"  periods ({len(series)}):")
        for i, (p, v) in enumerate(series):
            tail = ""
            if i == len(series) - 1 and months and 0 < months < 12:
                tail = f"   <-- PARTIAL PERIOD ({months} month(s)); not comparable to a full year as-is"
            print(f"    {p:14s} {v:>18,.2f}{tail}")

        # Compare the last two FULL periods -- comparing a 1-month tail
        # against a full year would report a ~-92% "collapse" that is
        # purely a period-length artefact, not a real movement.
        full = series[:-1] if (months and 0 < months < 12) else series
        scale = _series_scale(full)
        print(f"\n  period-over-period movement (full periods only):")
        biggest = 0.0          # largest real measurable % move
        from_nil = False       # a nil -> non-nil start, which has no meaningful %
        for (p0, v0), (p1, v1) in zip(full, full[1:]):
            p = _pct(v0, v1, scale)
            mark = ""
            if p is not None and abs(p) >= args.threshold:
                mark = "   <-- HIGH VARIANCE"
                biggest = max(biggest, abs(p))
            elif p is None and v1 != 0:
                # Commonly just the entity commencing operations (a
                # pre-operational zero year), NOT a swing needing
                # explanation -- tracked separately so it never gets
                # reported as if it were a measured percentage move.
                mark = "   <-- from nil (new activity, no meaningful %)"
                from_nil = True
            print(f"    {p0} -> {p1}: {v0:>16,.2f} -> {v1:>16,.2f}  {_fmt_pct(p)}{mark}")

        low = key.lower()
        if _is_revenue(key):
            if len(full) >= 2:
                revenue_growth = _pct(full[-2][1], full[-1][1], scale)
        elif any(n in low for n in ("expense", "cost", "cogs", "tax",
                                     "费用", "成本", "税金", "所得税")):
            if len(full) >= 2:
                expense_growth[key] = _pct(full[-2][1], full[-1][1], scale)

        cap = _cap_for(key)
        print(f"\n  explanatory material available (raw -> substantive):")
        print(f"    supporting notes    : {len(notes)} -> {len(sub_notes)}")
        print(f"    RHS remark rows     : {len(rhs)}")
        print(f"    table-linked remarks: {len(linked)} -> {len(sub_linked)}")
        if notes and not sub_notes:
            print(f"    (every note is a tie-out artefact, e.g. {str(notes[0])[:40]!r} -- "
                  f"no real context)")
        print(f"  prompt length cap for this account type:\n    {cap}")

        if biggest >= args.threshold and cap == _TIGHTEST and has_expl:
            print(f"\n  ⚠️  FLAGGED: moved {biggest:,.0f}% and HAS explanatory material,")
            print(f"      but sits in the tightest cap tier -- current rules leave no room")
            print(f"      to use that material, and 'delete any driver not explicitly stated'")
            print(f"      removes what little analysis survives.")
            flagged.append((key, biggest, cap))
        elif biggest >= args.threshold and not has_expl:
            print(f"\n  ℹ️  moved {biggest:,.0f}% but NO remarks/notes exist to explain it --")
            print(f"      deeper analysis here would have to be invented, so the current")
            print(f"      'facts only' behaviour is arguably correct for this account.")
        elif from_nil:
            print(f"\n  ·  starts from nil (entity likely pre-operational in the first period);")
            print(f"     largest measurable move afterwards is {biggest:,.1f}% -- below the "
                  f"{args.threshold:,.0f}% threshold.")
        print()

    print("=" * 78)
    print("CROSS-ACCOUNT: revenue vs expense growth asymmetry")
    print("=" * 78)
    print("(no single account's commentary can currently observe this -- each is")
    print(" generated in isolation from its own DataFrame)\n")
    if revenue_growth is None:
        print("  Could not determine revenue growth (no operating-income account matched).")
    else:
        print(f"  revenue growth (latest full period): {_fmt_pct(revenue_growth)}\n")
        if abs(revenue_growth) >= 200:
            print("  ⚠️  Revenue moved >200%, which usually means the entity was still")
            print("      ramping up rather than trading at a steady state. Against a")
            print("      baseline that large, essentially EVERY expense line reads as")
            print("      'asymmetric' -- treat the gaps below as not meaningful here, and")
            print("      judge this check on a stabilised entity instead.\n")
        for key, g in sorted(expense_growth.items(), key=lambda kv: -(abs(kv[1]) if kv[1] else 0)):
            if g is None:
                print(f"    {key:34s} n/a (from nil)")
                continue
            gap = g - revenue_growth
            flag = "   <-- ASYMMETRIC" if abs(gap) >= args.threshold else ""
            print(f"    {key:34s} {g:+9,.1f}%   vs revenue: {gap:+9,.1f} pts{flag}")

    print("\n" + "=" * 78)
    print(f"SUMMARY: {len(flagged)} account(s) where the data warrants analysis the current")
    print("         rules structurally prevent")
    print("=" * 78)
    for key, mv, _cap in flagged:
        print(f"  - {key} (moved {mv:,.0f}%)")
    if not flagged:
        print("  (none -- on this databook the current caps aren't the binding constraint)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
