#!/usr/bin/env python3
"""Verify extracted lease rows against the actual contract PDF pages.

Extraction alone is not proof — this re-opens each source PDF, sends page
images to Workbench GPT-5.5, and asks whether each core field value is
supported by what is visible on the pages (with a short quote as evidence).

Usage:
    python verify_contracts.py contracts/成都
        # auto-finds 合同汇总_extracted.xlsx under that folder
    python verify_contracts.py contracts/成都 --extracted contracts/成都/合同汇总_extracted.xlsx
    python verify_contracts.py contracts/成都 --sample 3
        # randomly verify 3 rows (cheaper)
    python verify_contracts.py contracts/成都 --file "RQBJ-Income-0004"
        # only rows whose filename contains this substring
"""
from __future__ import annotations

import argparse
import json
import random
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from contract_template_schema import CORE_VALUE_COLUMNS, TEMPLATE_COLUMNS, column_map
from contract_vision import (
    build_page_data_urls,
    extract_pdf_text,
    multi_image_byte_budget,
    pdf_is_digital,
    pdf_page_count,
    select_pages,
)
from fdd_utils.ai import AIClient

_DEFAULT_MODEL = "workbench"
_MISSING_TOKENS = {"", "未提及", "无", "不适用", "不適用", "n/a", "na"}
_COL_LETTERS = [letter for letter, _ in TEMPLATE_COLUMNS]
_VERIFY_MAX_TOKENS = 12000
_VERIFY_REASONING = "low"


def _load_extracted_rows(xlsx: Path) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: List[Dict[str, str]] = []
    for r in range(1, (ws.max_row or 0) + 1):
        filename = ws.cell(row=r, column=1).value
        if not filename or not str(filename).lower().endswith((".pdf", ".jpg", ".png")):
            continue
        row = {"A": str(filename).strip()}
        for col_idx, letter in enumerate(_COL_LETTERS, start=1):
            if letter == "A":
                continue
            val = ws.cell(row=r, column=col_idx).value
            row[letter] = "" if val is None else str(val).strip()
        rows.append(row)
    return rows


def _find_pdf(contracts_root: Path, filename: str) -> Optional[Path]:
    direct = contracts_root / filename
    if direct.is_file():
        return direct
    matches = list(contracts_root.rglob(filename))
    if matches:
        return matches[0]
    # fuzzy: stem contains
    stem = Path(filename).stem
    for p in contracts_root.rglob("*.pdf"):
        if stem and stem in p.stem:
            return p
    return None


def _find_extracted(path: Path) -> Optional[Path]:
    candidates = [
        path / "合同汇总_extracted.xlsx" if path.is_dir() else path.parent / "合同汇总_extracted.xlsx",
        path.parent / "合同汇总_extracted.xlsx" if path.is_file() else None,
    ]
    for c in candidates:
        if c and c.is_file() and not c.name.startswith("~$"):
            return c
    if path.is_dir():
        hits = sorted(path.rglob("*_extracted.xlsx"))
        hits = [h for h in hits if not h.name.startswith("~$")]
        if hits:
            return hits[0]
    return None


def _parse_json_object(text: str) -> Dict[str, Any]:
    raw = (text or "").strip()
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw, re.IGNORECASE)
    if fence:
        raw = fence.group(1).strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start, end = raw.find("{"), raw.rfind("}")
        if start < 0 or end <= start:
            raise
        data = json.loads(raw[start : end + 1])
    if not isinstance(data, dict):
        raise ValueError("verify response is not a JSON object")
    return data


def _claims_block(row: Dict[str, str], letters: Sequence[str]) -> str:
    cmap = column_map()
    lines = []
    for letter in letters:
        val = str(row.get(letter, "") or "").strip()
        if val in _MISSING_TOKENS:
            continue
        lines.append(f'  "{letter}" ({cmap.get(letter, letter)}): {json.dumps(val, ensure_ascii=False)}')
    return "{\n" + ",\n".join(lines) + "\n}" if lines else "{}"


def _verify_prompt(filename: str, claims_json: str) -> str:
    # No concrete numeric examples — only verification instructions.
    return (
        "你是租赁合同核对助手。下面 JSON 是从该合同抽取的字段主张（claims）。\n"
        "请只根据提供的合同页面内容核对每一项。\n"
        f"源文件名: {filename}\n\n"
        f"Claims:\n{claims_json}\n\n"
        "对每一个 claim 字段输出一个结果对象，放在 JSON 的 \"fields\" 里，key 为列字母。\n"
        "每个字段必须包含:\n"
        '- "verdict": 只能是 supported / contradict / not_on_pages / missing_ok\n'
        "  - supported: 页面能支持该值（允许格式不同但语义相同）\n"
        "  - contradict: 页面有明确不同的值\n"
        "  - not_on_pages: 提供的页面上看不到足以核对的信息\n"
        "  - missing_ok: claim 本身是空/未提及类，且页面也看不到，可接受\n"
        '- "quote": 支持或反驳时，引用页面上的一句原文；否则空字符串\n'
        '- "note": 简短说明（可空）\n'
        "规则:\n"
        "- 只输出一个 JSON 对象，不要 markdown。\n"
        "- 数值/日期/名称必须来自页面；不要用提示里的任何示例值（本提示不含示例值）。\n"
        "- 不要核对未出现在 Claims 里的字段。\n"
        '- 额外给 "summary": {"supported": n, "contradict": n, "not_on_pages": n}\n'
    )


def _call_verify(client: AIClient, user_prompt: Any) -> Dict[str, Any]:
    result = client.get_response(
        user_prompt=user_prompt,
        system_prompt="Return only one JSON object for contract field verification. No markdown.",
        max_tokens=_VERIFY_MAX_TOKENS,
        reasoning_effort=_VERIFY_REASONING,
    )
    content = str(result.get("content") or "").strip()
    if not content:
        raise ValueError(
            f"empty verify response (completion_tokens={result.get('completion_tokens')})"
        )
    return _parse_json_object(content)


def verify_one(
    client: AIClient,
    pdf_path: Path,
    row: Dict[str, str],
    *,
    max_pages: int,
    letters: Sequence[str],
) -> Dict[str, Any]:
    claims = _claims_block(row, letters)
    if claims == "{}":
        return {
            "file": row.get("A"),
            "pdf": str(pdf_path),
            "skipped": True,
            "reason": "no non-empty core claims to verify",
        }

    prompt = _verify_prompt(row.get("A", pdf_path.name), claims)

    if pdf_path.suffix.lower() == ".pdf" and pdf_is_digital(pdf_path):
        text = extract_pdf_text(pdf_path)
        user_prompt: Any = prompt + f"\n\n===== PDF TEXT =====\n{text}\n"
        pages_used: List[int] = []
    else:
        n_pages = pdf_page_count(pdf_path) if pdf_path.suffix.lower() == ".pdf" else 1
        pages_used = select_pages(n_pages, max_pages=max_pages) if pdf_path.suffix.lower() == ".pdf" else [1]
        budget = multi_image_byte_budget(len(pages_used))
        urls = build_page_data_urls(pdf_path, pages_used, max_bytes_each=budget)
        user_prompt = [{"type": "text", "text": prompt + f"\n提供页面: {pages_used}\n"}]
        for page_num, url in urls:
            user_prompt.append({"type": "text", "text": f"[page {page_num}]"})
            user_prompt.append({"type": "image_url", "image_url": {"url": url}})

    parsed = _call_verify(client, user_prompt)
    fields = parsed.get("fields") if isinstance(parsed.get("fields"), dict) else parsed
    summary = parsed.get("summary") if isinstance(parsed.get("summary"), dict) else {}
    # Local recount if model summary missing
    counts = {"supported": 0, "contradict": 0, "not_on_pages": 0, "missing_ok": 0, "other": 0}
    if isinstance(fields, dict):
        for _k, v in fields.items():
            if not isinstance(v, dict):
                continue
            verdict = str(v.get("verdict", "")).strip().lower()
            if verdict in counts:
                counts[verdict] += 1
            else:
                counts["other"] += 1
    return {
        "file": row.get("A"),
        "pdf": str(pdf_path),
        "pages": pages_used,
        "summary": summary or counts,
        "counts": counts,
        "fields": fields,
    }


def _print_report(report: Dict[str, Any]) -> None:
    print("=" * 78)
    print(report.get("file"))
    print("=" * 78)
    if report.get("skipped"):
        print(f"  skipped: {report.get('reason')}")
        return
    counts = report.get("counts") or {}
    print(
        f"  pages={report.get('pages')}  "
        f"supported={counts.get('supported', 0)}  "
        f"contradict={counts.get('contradict', 0)}  "
        f"not_on_pages={counts.get('not_on_pages', 0)}"
    )
    fields = report.get("fields") or {}
    if not isinstance(fields, dict):
        return
    cmap = column_map()
    for letter in CORE_VALUE_COLUMNS:
        item = fields.get(letter) or fields.get(letter.lower())
        if not isinstance(item, dict):
            continue
        verdict = str(item.get("verdict", "")).strip()
        if verdict in ("supported", "missing_ok"):
            continue
        quote = str(item.get("quote") or "")[:120]
        note = str(item.get("note") or "")[:80]
        print(f"  ! {letter} {cmap.get(letter, '')}: {verdict}")
        if quote:
            print(f"      quote: {quote}")
        if note:
            print(f"      note:  {note}")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Verify extracted contract rows against actual PDF page content (GPT-5.5 vision).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("path", help="contracts project folder (contains PDFs + usually the extracted xlsx)")
    ap.add_argument("--extracted", default=None, help="path to 合同汇总_extracted.xlsx")
    ap.add_argument("--max-pages", type=int, default=10, help="max PDF pages to send per file")
    ap.add_argument("--sample", type=int, default=0, help="verify only N random rows (0 = all)")
    ap.add_argument("--file", default=None, help="only rows whose filename contains this substring")
    ap.add_argument("--seed", type=int, default=0, help="random seed for --sample")
    args = ap.parse_args()

    root = Path(args.path)
    if not root.exists():
        print(f"❌ Path not found: {root}")
        return 1

    extracted = Path(args.extracted) if args.extracted else _find_extracted(root)
    if not extracted or not extracted.exists():
        print("❌ No 合同汇总_extracted.xlsx found. Pass --extracted explicitly.")
        return 1

    rows = _load_extracted_rows(extracted)
    if not rows:
        print(f"❌ No data rows in {extracted}")
        return 1

    if args.file:
        rows = [r for r in rows if args.file in r.get("A", "")]
        if not rows:
            print(f"❌ No rows matching --file {args.file!r}")
            return 1

    if args.sample and args.sample < len(rows):
        rng = random.Random(args.seed)
        rows = rng.sample(rows, args.sample)

    print(f"Model: GPT-5.5 (workbench)  |  verify against PDF pages")
    print(f"Extracted: {extracted}")
    print(f"Rows to verify: {len(rows)}")
    print(f"PDF root: {root}")
    print()

    try:
        client = AIClient(model_type=_DEFAULT_MODEL, agent_name="subagent_2", language="Chi")
    except Exception as exc:
        print(f"❌ Could not initialize AIClient: {exc}")
        return 1

    reports: List[Dict[str, Any]] = []
    total = {"supported": 0, "contradict": 0, "not_on_pages": 0, "missing_ok": 0, "other": 0}
    fail_n = 0

    for i, row in enumerate(rows, 1):
        name = row.get("A", "")
        print(f"[{i}/{len(rows)}] locating {name} ...")
        pdf = _find_pdf(root, name)
        if not pdf:
            fail_n += 1
            report = {"file": name, "error": "pdf not found under path"}
            reports.append(report)
            print(f"  ❌ pdf not found")
            continue
        try:
            report = verify_one(
                client,
                pdf,
                row,
                max_pages=args.max_pages,
                letters=CORE_VALUE_COLUMNS,
            )
            reports.append(report)
            _print_report(report)
            for k, v in (report.get("counts") or {}).items():
                total[k] = total.get(k, 0) + int(v)
        except Exception as exc:
            fail_n += 1
            reports.append({"file": name, "pdf": str(pdf), "error": str(exc)})
            print(f"  ❌ {exc}")
        print()

    out_json = extracted.with_name(extracted.stem + "_verify.json")
    out_json.write_text(json.dumps(reports, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=" * 78)
    print(
        f"VERIFY TOTAL: supported={total.get('supported', 0)}  "
        f"contradict={total.get('contradict', 0)}  "
        f"not_on_pages={total.get('not_on_pages', 0)}  "
        f"errors={fail_n}"
    )
    print(f"Wrote {out_json}")
    print("Paste this summary (or the json) back to review contradictions against the real contracts.")
    print("=" * 78)

    if fail_n:
        return 2
    if total.get("contradict", 0):
        return 3
    return 0


if __name__ == "__main__":
    sys.exit(main())
